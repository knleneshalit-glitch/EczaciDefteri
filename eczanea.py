import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
from datetime import datetime, date, timedelta
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import shutil
import os
import sys
import winreg 
from PIL import Image, ImageTk
import winsound
import calendar
import traceback
import re
import urllib.request
import threading
import subprocess
import time
global PYSTRAY_VAR
try:
    import pystray
    from pystray import MenuItem as item
    from PIL import Image, ImageTk
    PYSTRAY_VAR = True
except ImportError:
    PYSTRAY_VAR = False

# --- PROGRAM VERSİYONU VE GÜNCELLEME LİNKLERİ ---
MEVCUT_VERSIYON = "1.0"
# (Bu linkleri bir sonraki adımda GitHub'ı kurduğumuzda kendinize göre değiştireceğiz)
VERSIYON_URL = "https://raw.githubusercontent.com/KULLANICI_ADINIZ/EczaciDefteri/main/version.txt"
EXE_URL = "https://raw.githubusercontent.com/KULLANICI_ADINIZ/EczaciDefteri/main/EczaciDefteri.exe"


# 1. KÜTÜPHANE KONTROLLERİ
try:
    import pandas as pd
    import openpyxl 
    from openpyxl.styles import PatternFill
    PANDAS_VAR = True
except ImportError:
    PANDAS_VAR = False

# --- ORTAK AĞ VERİTABANI YOLU YÖNETİCİSİ ---
def yerel_db_yolunu_getir():
    import os
    ayar_dosyasi = "db_yolu.txt"
    if os.path.exists(ayar_dosyasi):
        try:
            with open(ayar_dosyasi, "r", encoding="utf-8") as f:
                yol = f.read().strip()
                if yol: return yol
        except: pass
    return os.path.join(os.path.expanduser('~'), "EczaneAsistani_Veri")

def yerel_db_yolunu_kaydet(yeni_yol):
    with open("db_yolu.txt", "w", encoding="utf-8") as f:
        f.write(yeni_yol)

def log_error(msg):
    try:
        with open("HATA_KAYDI.txt", "w", encoding="utf-8") as f:
            f.write(f"Zaman: {datetime.now()}\n")
            f.write(f"Hata: {str(msg)}\n")
            f.write(traceback.format_exc())
    except: pass

# 2. ORTAK TEMA YÖNETİCİSİ
class ThemeManager:
    def __init__(self):
        self.is_dark = False
        self.colors = {
            "light": {
                "bg_main": "#f1f5f9", "bg_sidebar": "#1e293b", "fg_text": "#0f172a",
                "card_bg": "#ffffff", "input_bg": "#ffffff", "input_fg": "#000000",
                "header_bg": "#334155", "btn_primary": "#3b82f6", "btn_success": "#10b981", 
                "btn_danger": "#ef4444", "btn_warning": "#f59e0b", "border": "#cbd5e1"
            },
            "dark": {
                "bg_main": "#0f172a", "bg_sidebar": "#020617", "fg_text": "#f1f5f9",
                "card_bg": "#1e293b", "input_bg": "#334155", "input_fg": "#ffffff",
                "header_bg": "#1e293b", "btn_primary": "#60a5fa", "btn_success": "#34d399", 
                "btn_danger": "#f87171", "btn_warning": "#fbbf24", "border": "#475569"
            }
        }
    
    def get_color(self, key):
        mode = "dark" if self.is_dark else "light"
        return self.colors[mode].get(key, "#000000")

    def toggle(self):
        self.is_dark = not self.is_dark

TM = ThemeManager()
FONT_HEAD = ("Segoe UI", 24, "bold")
FONT_BOLD = ("Segoe UI", 10, "bold")
FONT_NORM = ("Segoe UI", 10)


# 3. YARDIMCI ARAÇLAR VE MASKELER (FİNANS VE SKT İÇİN ORTAK)
def mask_tarih_otomatik(event):
    entry = event.widget
    if event.keysym in ("BackSpace", "Delete"): return
    text = entry.get()
    clean = "".join([c for c in text if c.isdigit()])
    formatted = ""
    if len(clean) > 0: formatted = clean[:4]
    if len(clean) >= 5: formatted += "-" + clean[4:6]
    if len(clean) >= 7: formatted += "-" + clean[6:8]
    if text != formatted:
        entry.delete(0, tk.END)
        entry.insert(0, formatted)

def mask_para_birimi(event):
    entry = event.widget
    if event.keysym in ("BackSpace", "Delete", "Left", "Right", "Tab"): return
    try:
        val = entry.get().replace(".", "")
        if "," in val:
            parts = val.split(",")
            int_part = parts[0]
            dec_part = parts[1]
        else:
            int_part = val
            dec_part = None
            
        if int_part.isdigit():
            new_int = "{:,}".format(int(int_part)).replace(",", ".")
        else:
            new_int = int_part
            
        res = new_int
        if dec_part is not None: res += "," + dec_part
        if entry.get() != res:
            entry.delete(0, tk.END)
            entry.insert(0, res)
    except: pass

def temizle_para(text):
    if not text: return 0.0
    try:
        clean = text.replace(".", "").replace(",", ".")
        return float(clean)
    except: return 0.0

class CreateToolTip(object):
    def __init__(self, widget, text='widget info'):
        self.waittime = 500
        self.wraplength = 300
        self.widget = widget
        self.text = text
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.leave)
        self.widget.bind("<ButtonPress>", self.leave)
        self.id = None
        self.tw = None

    def enter(self, event=None): self.schedule()
    def leave(self, event=None): self.unschedule(); self.hidetip()
    def schedule(self): self.unschedule(); self.id = self.widget.after(self.waittime, self.showtip)
    def unschedule(self):
        id = self.id
        self.id = None
        if id: self.widget.after_cancel(id)

    def showtip(self, event=None):
        if self.tw: return
        x = y = 0
        try:
            x, y, cx, cy = self.widget.bbox("insert")
            x += self.widget.winfo_rootx() + 25
            y += self.widget.winfo_rooty() + 20
        except: return
        self.tw = tk.Toplevel(self.widget)
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry("+%d+%d" % (x, y))
        self.tw.wm_attributes("-topmost", True)
        label = tk.Label(self.tw, text=self.text, justify='left', background="#ffffe0", relief='solid', borderwidth=1, wraplength=self.wraplength)
        label.pack(ipadx=5, ipady=5)

    def hidetip(self):
        tw = self.tw
        self.tw = None
        if tw: tw.destroy()

class ModernButton(tk.Canvas):
    def __init__(self, parent, text, command=None, width=120, height=35, corner_radius=10, bg_color="#3b82f6", fg_color="white", hover_color=None, font=FONT_BOLD, **kwargs):
        super().__init__(parent, width=width, height=height, bg=parent["bg"], highlightthickness=0, **kwargs)
        self.command = command; self.text = text; self.bg_color = bg_color
        self.hover_color = hover_color if hover_color else bg_color
        self.fg_color = fg_color; self.corner_radius = corner_radius; self.font = font
        self.bind("<Button-1>", self._on_click); self.bind("<Enter>", self._on_enter); self.bind("<Leave>", self._on_leave)
        self._draw(self.bg_color)
    def _draw(self, color):
        try:
            if not self.winfo_exists(): return
            self.delete("all"); w, h = int(self["width"]), int(self["height"]); r = self.corner_radius
            points = (r, 0, w-r, 0, w, 0, w, r, w, h-r, w, h, w-r, h, r, h, 0, h, 0, h-r, 0, r, 0, 0)
            self.create_polygon(points, fill=color, smooth=True, outline=color)
            self.create_text(w/2, h/2, text=self.text, fill=self.fg_color, font=self.font, justify="center", tag="text")
        except: pass
    def _on_click(self, event):
        if self.command: self.command()
    def _on_enter(self, event): self.configure(cursor="hand2"); self._draw(self.hover_color)
    def _on_leave(self, event): self.configure(cursor="arrow"); self._draw(self.bg_color)
    def update_color(self, color): self.bg_color = color; self._draw(color)

class DraggableNotebook(ttk.Notebook):
    def __init__(self, master, on_reorder=None, **kwargs):
        super().__init__(master, **kwargs)
        self.on_reorder = on_reorder
        self.bind("<Button-1>", self.on_click)
        self.bind("<B1-Motion>", self.on_drag)
        self.bind("<ButtonRelease-1>", self.on_release)
        self._drag_data = {"x": 0, "y": 0, "index": None}

    def on_click(self, event):
        try:
            index = self.index(f"@{event.x},{event.y}")
            self._drag_data["index"] = index
            self._drag_data["x"] = event.x; self._drag_data["y"] = event.y
        except: pass

    def on_drag(self, event):
        if self._drag_data["index"] is None: return
        try:
            target_index = self.index(f"@{event.x},{event.y}")
            if target_index != self._drag_data["index"]:
                source_index = self._drag_data["index"]
                current_tab_widget = self.tabs()[source_index]
                self.insert(target_index, current_tab_widget)
                self.select(target_index)
                self._drag_data["index"] = target_index
        except: pass

    def on_release(self, event):
        self._drag_data["index"] = None
        if self.on_reorder:
            tab_names = [self.tab(i, "text") for i in range(self.index("end"))]
            self.on_reorder(tab_names)

class TakvimPopup(tk.Toplevel):
    def __init__(self, parent, entry_widget):
        super().__init__(parent)
        self.entry = entry_widget
        self.title("Tarih Seç")
        x = parent.winfo_pointerx(); y = parent.winfo_pointery()
        self.geometry(f"+{x}+{y}")
        self.configure(bg="white", padx=10, pady=10)
        self.resizable(False, False)
        self.transient(parent); self.grab_set()

        try:
            mevcut_str = self.entry.get()
            self.secili_tarih = datetime.strptime(mevcut_str, "%Y-%m-%d").date()
        except: self.secili_tarih = date.today()
        self.yil = self.secili_tarih.year; self.ay = self.secili_tarih.month
        self.arayuz_olustur()

    def arayuz_olustur(self):
        for widget in self.winfo_children(): widget.destroy()
        header = tk.Frame(self, bg="white"); header.pack(fill="x", pady=(0, 10))
        btn_stil = {"relief": "flat", "bg": "#f1f5f9", "width": 3, "cursor": "hand2"}
        tk.Button(header, text="<<", command=self.onceki_yil, **btn_stil).pack(side="left")
        tk.Button(header, text="<", command=self.onceki_ay, **btn_stil).pack(side="left", padx=2)
        tr_aylar = ["", "OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN", "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"]
        tk.Label(header, text=f"{tr_aylar[self.ay]} {self.yil}", font=("Segoe UI", 10, "bold"), bg="white", width=16).pack(side="left", padx=5)
        tk.Button(header, text=">", command=self.sonraki_ay, **btn_stil).pack(side="left", padx=2)
        tk.Button(header, text=">>", command=self.sonraki_yil, **btn_stil).pack(side="left")
        cal_frame = tk.Frame(self, bg="white"); cal_frame.pack()
        gunler_tr = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]
        for i, gun in enumerate(gunler_tr): tk.Label(cal_frame, text=gun, font=("Segoe UI", 8, "bold"), bg="white", fg="#94a3b8", width=4).grid(row=0, column=i, pady=(0,5))
        cal = calendar.monthcalendar(self.yil, self.ay); bugun = date.today()
        for r, hafta in enumerate(cal):
            for c, gun in enumerate(hafta):
                if gun == 0: continue
                bg_color = "#ffffff"; fg_color = "#0f172a"; font_weight = "normal"
                if self.yil == bugun.year and self.ay == bugun.month and gun == bugun.day:
                    bg_color = "#dbeafe"; fg_color = "#1d4ed8"; font_weight = "bold"
                btn = tk.Button(cal_frame, text=str(gun), width=4, relief="flat", bg=bg_color, fg=fg_color, font=("Segoe UI", 9, font_weight), cursor="hand2", command=lambda d=gun: self.tarih_sec(d))
                btn.grid(row=r+1, column=c, padx=1, pady=1)

    def onceki_ay(self):
        self.ay -= 1
        if self.ay < 1: self.ay = 12; self.yil -= 1
        self.arayuz_olustur()
    def sonraki_ay(self):
        self.ay += 1
        if self.ay > 12: self.ay = 1; self.yil += 1
        self.arayuz_olustur()
    def onceki_yil(self): self.yil -= 1; self.arayuz_olustur()
    def sonraki_yil(self): self.yil += 1; self.arayuz_olustur()
    def tarih_sec(self, gun):
        secilen = date(self.yil, self.ay, gun).strftime("%Y-%m-%d")
        self.entry.delete(0, tk.END); self.entry.insert(0, secilen); self.destroy()

def tarih_secici_bagla(frame, entry):
    btn = tk.Button(frame, text="📅", font=("Segoe UI", 9), cursor="hand2", relief="flat", bg="#e2e8f0", command=lambda: TakvimPopup(frame.winfo_toplevel(), entry))
    btn.pack(side="left", padx=(0, 0), fill="y")

# =============================================================================
# PDF İÇİN YARDIMCI PENCERE SINIFI (ŞAHSİ HARCAMA VE EKSTRE HESAPLAMALI)
# =============================================================================
class PDFImportWindow(tk.Toplevel):
    def __init__(self, parent, kart_adi, save_callback, depo_listesi, imlec, baglanti):
        super().__init__(parent)
        self.title(f"{kart_adi} - Akıllı Ekstre ve Şahsi Harcama Analizi")
        self.geometry("1200x750")
        self.save_callback = save_callback
        self.depo_listesi = depo_listesi
        self.kart_adi = kart_adi
        self.imlec = imlec
        self.baglanti = baglanti
        
        self.ekstre_toplam = 0.0
        self.ekstre_asgari = 0.0
        
        try:
            import pdfplumber
            import re
            self.pdfplumber = pdfplumber
            self.re = re
        except ImportError:
            messagebox.showerror("Hata", "Lütfen terminale 'pip install pdfplumber' yazarak kütüphaneyi yükleyin.")
            self.destroy()
            return

        # --- SOL PANEL ---
        left_frame = tk.Frame(self, bg="#f1f5f9", width=700)
        left_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        
        self.lbl_info_header = tk.Label(left_frame, text="Lütfen PDF Dosyası Seçin...", font=("Segoe UI", 11, "bold"), bg="#dbeafe", fg="#1e3a8a", padx=10, pady=15, relief="solid", bd=1)
        self.lbl_info_header.pack(fill="x", pady=(0, 10))

        ModernButton(left_frame, text="📂 PDF DOSYASI SEÇ", command=self.pdf_yukle, width=200, bg_color="#3b82f6").pack(pady=5)
        
        cols = ("TIK", "TARIH", "ACIKLAMA", "TUTAR")
        self.tree = ttk.Treeview(left_frame, columns=cols, show="headings", height=20)
        
        self.tumunu_sec_durumu = False
        def toggle_all():
            self.tumunu_sec_durumu = not self.tumunu_sec_durumu
            ikon = "☑" if self.tumunu_sec_durumu else "☐"
            self.tree.heading("TIK", text=ikon)
            for item in self.tree.get_children():
                vals = list(self.tree.item(item, "values"))
                vals[0] = ikon
                self.tree.item(item, values=vals)
            self.hesaplari_guncelle()

        self.tree.heading("TIK", text="☐", command=toggle_all); self.tree.column("TIK", width=40, anchor="center")
        self.tree.heading("TARIH", text="İşlem Tarihi"); self.tree.column("TARIH", width=90, anchor="center")
        self.tree.heading("ACIKLAMA", text="Açıklama (Eczane Harcamaları)"); self.tree.column("ACIKLAMA", width=350)
        self.tree.heading("TUTAR", text="Aylık Tutar"); self.tree.column("TUTAR", width=100, anchor="e")
        
        sc = ttk.Scrollbar(left_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=sc.set); sc.pack(side="right", fill="y"); self.tree.pack(fill="both", expand=True)
        self.tree.bind("<ButtonRelease-1>", self.satir_tiklandi)

        # --- SAĞ PANEL ---
        right_frame = tk.Frame(self, bg="white", width=500, padx=20, pady=15)
        right_frame.pack(side="right", fill="y", padx=10, pady=10)
        right_frame.pack_propagate(False)

        # 1. BÖLÜM: FİNANSAL ÖZET VE ŞAHSİ HARCAMA
        tk.Label(right_frame, text="📊 EKSTRE ÖZETİ", font=("Segoe UI", 12, "bold"), bg="white", fg="#334155").pack(pady=(0, 10))
        
        f_ozet = tk.Frame(right_frame, bg="#f8fafc", bd=1, relief="solid", padx=15, pady=15)
        f_ozet.pack(fill="x", pady=(0, 20))
        
        self.lbl_ekstre_toplam = tk.Label(f_ozet, text="Ekstre Toplam Borcu: 0.00 ₺", font=("Segoe UI", 11, "bold"), bg="#f8fafc", fg="#0f172a")
        self.lbl_ekstre_toplam.pack(anchor="w", pady=2)
        
        self.lbl_ekstre_asgari = tk.Label(f_ozet, text="Asgari Ödeme: 0.00 ₺", font=("Segoe UI", 10), bg="#f8fafc", fg="#475569")
        self.lbl_ekstre_asgari.pack(anchor="w", pady=(0, 10))
        
        tk.Frame(f_ozet, height=1, bg="#cbd5e1").pack(fill="x", pady=5)
        
        self.lbl_secili_eczane = tk.Label(f_ozet, text="İşaretli Eczane Gideri: 0.00 ₺", font=("Segoe UI", 11, "bold"), bg="#f8fafc", fg="#10b981")
        self.lbl_secili_eczane.pack(anchor="w", pady=5)
        
        self.lbl_sahsi_harcama = tk.Label(f_ozet, text="Şahsi Harcamanız: 0.00 ₺", font=("Segoe UI", 14, "bold"), bg="#f8fafc", fg="#ef4444")
        self.lbl_sahsi_harcama.pack(anchor="w", pady=(5, 0))

        # 2. BÖLÜM: KAYIT DETAYLARI VE EKSTRE GİRİŞİ
        tk.Label(right_frame, text="Sisteme Kayıt Detayları", font=("Segoe UI", 12, "bold"), bg="white", fg="#334155").pack(pady=(10, 15))
        
        # YENİ ALAN: EKSTRE TOPLAMI
        tk.Label(right_frame, text="Ekstre Toplam Borcu (Bu Ay):", bg="white", font=("Segoe UI", 10, "bold"), fg="#3b82f6").pack(anchor="w")
        self.ent_ekstre_toplam = tk.Entry(right_frame, font=("Segoe UI", 12, "bold"), relief="solid", bd=1, bg="#eff6ff", fg="#1e3a8a")
        self.ent_ekstre_toplam.pack(fill="x", pady=(2, 10), ipady=3)
        
        def on_ekstre_change(e):
            mask_para_birimi(e)
            self.hesaplari_guncelle()
        self.ent_ekstre_toplam.bind("<KeyRelease>", on_ekstre_change)

        tk.Label(right_frame, text="Son Ödeme Tarihi (Vade):", bg="white", font=("Segoe UI", 10, "bold"), fg="#ef4444").pack(anchor="w")
        self.ent_son_odeme = tk.Entry(right_frame, font=("Segoe UI", 11), relief="solid", bd=1, bg="#fff1f2")
        self.ent_son_odeme.pack(fill="x", pady=(2, 10), ipady=3)

        f_row = tk.Frame(right_frame, bg="white"); f_row.pack(fill="x", pady=5)
        f_taksit = tk.Frame(f_row, bg="white"); f_taksit.pack(side="left", fill="x", expand=True, padx=(0, 5))
        tk.Label(f_taksit, text="Taksit:", bg="white", font=("Segoe UI", 10, "bold"), fg="#475569").pack(anchor="w")
        self.cmb_taksit = ttk.Combobox(f_taksit, values=["1","2","3","4","5","6","9","12"], state="readonly", font=("Segoe UI", 10))
        self.cmb_taksit.pack(fill="x", pady=2, ipady=2); self.cmb_taksit.current(0)

        f_kesim = tk.Frame(f_row, bg="white"); f_kesim.pack(side="right", fill="x", expand=True, padx=(5, 0))
        tk.Label(f_kesim, text="Kesim Günü:", bg="white", font=("Segoe UI", 10, "bold"), fg="#475569").pack(anchor="w")
        self.cmb_kesim = ttk.Combobox(f_kesim, values=[str(i) for i in range(1,32)], state="readonly", font=("Segoe UI", 10))
        self.cmb_kesim.pack(fill="x", pady=2, ipady=2); self.cmb_kesim.set("15")

        ModernButton(right_frame, text="☑ SEÇİLİ ECZANE GİDERLERİNİ KAYDET", command=self.toplu_kaydet, width=350, height=45, bg_color="#10b981").pack(side="bottom", pady=10)

    def hesaplari_guncelle(self):
        secili_eczane_tutari = 0.0
        for item in self.tree.get_children():
            vals = self.tree.item(item, "values")
            if vals[0] == "☑":
                try:
                    t = float(vals[3].replace(".", "").replace(",", "."))
                    secili_eczane_tutari += t
                except: pass
        
        try:
            val_str = self.ent_ekstre_toplam.get().replace(".", "").replace(",", ".")
            toplam_ekstre = float(val_str) if val_str else 0.0
        except:
            toplam_ekstre = 0.0
            
        sahsi_harcama = toplam_ekstre - secili_eczane_tutari
        
        self.lbl_ekstre_toplam.config(text=f"Ekstre Toplam Borcu: {toplam_ekstre:,.2f} ₺")
        self.lbl_ekstre_asgari.config(text=f"Asgari Ödeme: {self.ekstre_asgari:,.2f} ₺")
        self.lbl_secili_eczane.config(text=f"İşaretli Eczane Gideri: {secili_eczane_tutari:,.2f} ₺")
        
        renk = "#ef4444" if sahsi_harcama >= 0 else "#f59e0b"
        self.lbl_sahsi_harcama.config(text=f"Şahsi Harcamanız: {sahsi_harcama:,.2f} ₺", fg=renk)

    def satir_tiklandi(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region == "cell":
            col = self.tree.identify_column(event.x)
            row_id = self.tree.identify_row(event.y)
            if row_id and col == "#1":
                vals = list(self.tree.item(row_id, "values"))
                vals[0] = "☑" if vals[0] == "☐" else "☐"
                self.tree.item(row_id, values=vals)
                self.hesaplari_guncelle()

    def standart_depo_adi_bul(self, ham_metin):
        kontrol_metin = ham_metin.upper().replace("i", "İ").replace("ı", "I")
        if "SANCAK" in kontrol_metin: return "SANCAK ECZA DEPOSU KK"
        elif "SELCUK" in kontrol_metin or "SELÇUK" in kontrol_metin: return "SELÇUK ECZA DEPOSU KK"
        elif "ISKOOP" in kontrol_metin or "İSKOOP" in kontrol_metin: return "ISKOOP KK"
        elif "NEVZAT" in kontrol_metin: return "NEVZAT ECZA DEPOSU KK"
        elif "AS ECZA" in kontrol_metin: return "AS ECZA DEPOSU KK"
        elif "BEK" in kontrol_metin and "ECZA" in kontrol_metin: return "BEK KK"
        elif "ALLIANCE" in kontrol_metin: return "ALLIANCE HEALTHCARE KK"
        elif "HEDEF" in kontrol_metin: return "HEDEF ECZA DEPOSU KK"
        elif "BAŞKENT" in kontrol_metin or "BASKENT" in kontrol_metin: return "BAŞKENT ECZA DEPOSU KK"
        elif "FARMAZON" in kontrol_metin: return "FARMAZON KK"
        elif "ECZA1" in kontrol_metin: return "ECZA1 KK"
        return ham_metin + " KK"            

    def tutar_ayikla(self, metin):
        # 1. Bankalara özel gereksiz karakterleri temizle
        metin = metin.replace("TL.", "").replace("TL", "").replace(",-", ",00").replace(".-", ".00").strip()
        
        # 2. Satırı sondan başa doğru kelime kelime analiz et
        parcalar = metin.split()
        if parcalar:
            for kelime in reversed(parcalar):
                # Tarihleri ve VakıfBank'ın 2x4000 gibi taksit tutarlarını es geç
                if '/' in kelime or 'x' in kelime.lower() or any(c.isalpha() for c in kelime):
                    continue
                if '-' in kelime and not kelime.startswith('-'):
                    continue
                    
                # İçinde rakam ve ayırıcı (nokta/virgül) varsa
                if any(c.isdigit() for c in kelime) and ('.' in kelime or ',' in kelime):
                    last_dot = kelime.rfind('.')
                    last_comma = kelime.rfind(',')
                    last_sep = max(last_dot, last_comma)
                    
                    rakamlar = "".join([c for c in kelime if c.isdigit()])
                    if not rakamlar: continue
                    
                    # Akıllı Kuruş Algılayıcı (Son noktadan/virgülden sonra tam 2 rakam varsa kuruşludur)
                    is_kuruslu = last_sep != -1 and (len(kelime) - last_sep - 1) == 2
                    
                    if is_kuruslu and len(rakamlar) >= 3:
                        tam = rakamlar[:-2]
                        kurus = rakamlar[-2:]
                    else:
                        tam = rakamlar
                        kurus = "00"
                        
                    try:
                        deneme_float = float(f"{tam}.{kurus}")
                        if deneme_float == 0.0 and kelime == parcalar[-1]: continue # Puan/Bonus sıfırlarını atla
                        tutar_str = f"{deneme_float:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                        return tutar_str, kelime
                    except: pass
                    
        return "0,00", ""

    

    def tarihi_standartlastir(self, tarih_str):
        aylar = {"Ocak": "01", "Şubat": "02", "Mart": "03", "Nisan": "04", "Mayıs": "05", "Haziran": "06",
                 "Temmuz": "07", "Ağustos": "08", "Eylül": "09", "Ekim": "10", "Kasım": "11", "Aralık": "12"}
        try:
            parts = tarih_str.split()
            if len(parts) == 3 and parts[1] in aylar: return f"{parts[2]}-{aylar[parts[1]]}-{parts[0].zfill(2)}"
            for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
                try: return datetime.strptime(tarih_str, fmt).strftime("%Y-%m-%d")
                except: continue
            return tarih_str
        except: return tarih_str

    def toplu_kaydet(self):
        secili_satirlar = [self.tree.item(i, "values") for i in self.tree.get_children() if self.tree.item(i, "values")[0] == "☑"]
        
        son_odeme = self.ent_son_odeme.get().strip()
        taksit_secimi = self.cmb_taksit.get()
        
        if not son_odeme:
            messagebox.showwarning("Eksik", "Lütfen Son Ödeme Tarihini giriniz (Vade hesaplaması için).")
            return
            
        # DÜZELTME: Hiçbir satır seçilmese bile iptal etmiyoruz, kullanıcıya "sadece ekstreyi kaydedeyim mi?" diye soruyoruz.
        if not secili_satirlar:
            if not messagebox.askyesno("Sadece Ekstre", "Sisteme kaydedilecek eczane işlemi seçmediniz.\nSadece 'Ekstre Toplamını' kaydetmek istiyor musunuz?"):
                return
        else:
            if not messagebox.askyesno("Kayıt Onayı", f"Seçilen {len(secili_satirlar)} eczane harcaması '{self.kart_adi}' kartına eklenecek.\n\nTaksit Seçimi: {taksit_secimi}\nSon Ödeme Tarihi: {son_odeme}\n\n(Şahsi harcamalarınız sisteme İŞLENMEYECEKTİR)\n\nOnaylıyor musunuz?"):
                return
                
        try:
            referans_odeme_dt = datetime.strptime(son_odeme, "%Y-%m-%d").date()
            odeme_gunu_rakam = referans_odeme_dt.day
            taksit = int(taksit_secimi)
            donem_str = referans_odeme_dt.strftime("%Y-%m")
            
            # YENİ: Ekstre Toplamını Veritabanına Program Ayarı olarak kaydet
            try:
                ekstre_val = float(self.ent_ekstre_toplam.get().replace(".", "").replace(",", "."))
                self.imlec.execute("INSERT OR REPLACE INTO program_ayarlari (ayar_adi, deger) VALUES (?, ?)", (f"ekstre_toplam_{self.kart_adi}_{donem_str}", str(ekstre_val)))
            except: pass
            
            # Eğer seçili satır varsa onları da tek tek işleyip taksitlendir
            if secili_satirlar:
                for vals in secili_satirlar:
                    islem_tar = vals[1]
                    depo = vals[2]
                    tutar_str = vals[3]
                    
                    try: aylik_tutar = float(tutar_str.replace(".", "").replace(",", ".")) if "." in tutar_str and "," in tutar_str else float(tutar_str.replace(",", "."))
                    except: continue 
                    
                    aylik = aylik_tutar
                    
                    for i in range(taksit):
                        eklenecek_ay = i + 1
                        hedef_yil = referans_odeme_dt.year + (referans_odeme_dt.month + eklenecek_ay - 2) // 12
                        hedef_ay = (referans_odeme_dt.month + eklenecek_ay - 2) % 12 + 1
                        
                        import calendar
                        ayin_son_gunu = calendar.monthrange(hedef_yil, hedef_ay)[1]
                        hedef_gun = min(odeme_gunu_rakam, ayin_son_gunu)
                        
                        from datetime import date, timedelta
                        odeme_tarihi = date(hedef_yil, hedef_ay, hedef_gun)
                        
                        if odeme_tarihi.weekday() == 5: odeme_tarihi += timedelta(days=2)
                        elif odeme_tarihi.weekday() == 6: odeme_tarihi += timedelta(days=1)
                        
                        vade = odeme_tarihi.strftime("%Y-%m-%d")
                        aciklama = f"KART: {self.kart_adi} ({i+1}/{taksit})" if taksit > 1 else f"KART: {self.kart_adi}"
                            
                        self.imlec.execute("INSERT INTO odemeler (fatura_adi, vade_tarihi, tutar, aciklama, durum, alim_tarihi, satir_notu) VALUES (?, ?, ?, ?, 'ODENMEDİ', ?, ?)",
                                           (depo, vade, aylik, "Kredi Kartı", islem_tar, aciklama))
        
            self.baglanti.commit()
            
            # İşlem tipine göre doğru mesajı göster
            if secili_satirlar:
                messagebox.showinfo("Başarılı", "Eczane giderleri başarıyla sisteme işlendi.\nŞahsi harcamalarınız dışarıda bırakıldı.")
            else:
                messagebox.showinfo("Başarılı", "Sadece ekstre toplamı başarıyla güncellendi.")
                
            self.destroy() 
            try: self.save_callback("DUMMY_REFRESH", "PDF_TOPLU", "0", "1", "2000-01-01", "2000-01-01", "1")
            except: pass
            
        except Exception as e:
            self.baglanti.rollback()
            messagebox.showerror("Hata", f"Kayıt sırasında hata:\n{str(e)}")       

# =============================================================================
# SGK ÖZEL FATURA VE KESİNTİ HESAPLAYICI SİHİRBAZI
# =============================================================================
class SgkAylikFaturaSihirbazi(tk.Toplevel):
    def __init__(self, parent, kurum_adi, save_callback):
        super().__init__(parent)
        self.title(f"{kurum_adi} - Detaylı Fatura Hesaplayıcı")
        self.geometry("850x750")
        self.configure(bg="#f8fafc")
        self.save_callback = save_callback
        self.kurum_adi = kurum_adi
        
        self.transient(parent)
        self.grab_set()

        x = parent.winfo_x() + (parent.winfo_width() // 2) - 275
        y = parent.winfo_y() + (parent.winfo_height() // 2) - 375
        self.geometry(f"+{x}+{y}")

        # Üst Kısım: Dönemler
        f_top = tk.Frame(self, bg="white", padx=20, pady=15, bd=1, relief="solid")
        f_top.pack(fill="x", padx=15, pady=15)

        tk.Label(f_top, text="Alım Dönemi (YIL-AY):", font=("Segoe UI", 9, "bold"), bg="white", fg="#64748b").grid(row=0, column=0, sticky="w", padx=5)
        self.ent_alim = tk.Entry(f_top, font=("Segoe UI", 11), relief="solid", bd=1, bg="#f8fafc", width=15)
        self.ent_alim.grid(row=1, column=0, sticky="w", padx=5, pady=(0,5))
        self.ent_alim.insert(0, date.today().strftime("%Y-%m"))
        self.ent_alim.bind("<KeyRelease>", mask_tarih_otomatik)
        
        # YENİ: Alım tarihinden çıkıldığında veya tarih değiştiğinde vadeyi otomatik hesapla
        self.ent_alim.bind("<FocusOut>", self.otomatik_vade_hesapla)

        tk.Label(f_top, text="Ödeme Dönemi (Vade):", font=("Segoe UI", 9, "bold"), bg="white", fg="#64748b").grid(row=0, column=1, sticky="w", padx=20)
        self.ent_vade = tk.Entry(f_top, font=("Segoe UI", 11), relief="solid", bd=1, bg="#f8fafc", width=15)
        self.ent_vade.grid(row=1, column=1, sticky="w", padx=20, pady=(0,5))
        self.ent_vade.bind("<KeyRelease>", mask_tarih_otomatik)

        # YENİ: İlk açılışta vadeyi otomatik olarak 3 ay sonrası (1. ayı 4. ay yapacak şekilde) ayarla
        self.otomatik_vade_hesapla()

        # Orta Kısım: Kaydırılabilir Modül Alanı
        f_mid = tk.Frame(self, bg="#f8fafc")
        f_mid.pack(fill="both", expand=True, padx=15)

        self.canvas = tk.Canvas(f_mid, bg="#f8fafc", highlightthickness=0)
        sc = ttk.Scrollbar(f_mid, orient="vertical", command=self.canvas.yview)
        self.scroll_frame = tk.Frame(self.canvas, bg="#f8fafc")

        self.scroll_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.window_id = self.canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfig(self.window_id, width=e.width))

        self.canvas.configure(yscrollcommand=sc.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        sc.pack(side="right", fill="y")

        def _on_mousewheel(event):
            try: self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except: pass
        self.canvas.bind_all('<MouseWheel>', _on_mousewheel)
        self.bind("<Destroy>", lambda e: self.canvas.unbind_all('<MouseWheel>'))

        # Alt Kısım: Genel Toplam ve Kaydet
        f_bot = tk.Frame(self, bg="#1e293b", pady=15, padx=20)
        f_bot.pack(fill="x", side="bottom")

        self.lbl_genel_toplam = tk.Label(f_bot, text="GENEL TOPLAM: 0.00 ₺", font=("Segoe UI", 16, "bold"), bg="#1e293b", fg="#34d399")
        self.lbl_genel_toplam.pack(side="left")

        tk.Button(f_bot, text="💾 SİSTEME KAYDET", font=("Segoe UI", 10, "bold"), bg="#3b82f6", fg="white", cursor="hand2", relief="flat", padx=10, pady=5, command=self.kaydet).pack(side="right")
        tk.Button(f_bot, text="➕ YENİ MODÜL", font=("Segoe UI", 10, "bold"), bg="#f59e0b", fg="white", cursor="hand2", relief="flat", padx=10, pady=5, command=self.modul_ekle).pack(side="right", padx=10)

        self.moduller = []
        self.modul_ekle() # İlk açılışta 1 tane modül varsayılan eklensin

    # =========================================================================
    # YENİ EKLENEN FONKSİYON: 3 AY SONRASINI HESAPLAR
    # =========================================================================
    def otomatik_vade_hesapla(self, event=None):
        try:
            alim_str = self.ent_alim.get().strip()
            if len(alim_str) >= 7:
                alim_dt = datetime.strptime(alim_str[:7] + "-01", "%Y-%m-%d")
                # Alım ayının üzerine 3 ay ekliyoruz (1. ay -> 4. ay)
                ay = alim_dt.month + 3
                yil = alim_dt.year
                if ay > 12:
                    yil += ay // 12
                    ay = ay % 12
                    if ay == 0: ay = 12; yil -= 1
                vade_str = f"{yil}-{ay:02d}"
                
                self.ent_vade.delete(0, tk.END)
                self.ent_vade.insert(0, vade_str)
        except: pass

    def modul_ekle(self):
        m_idx = len(self.moduller) + 1
        f_mod = tk.Frame(self.scroll_frame, bg="white", bd=1, relief="solid", padx=15, pady=15)
        f_mod.pack(fill="x", pady=(0, 10), padx=5)

        # Modül Başlığı ve Silme Butonu
        f_m_top = tk.Frame(f_mod, bg="white")
        f_m_top.pack(fill="x", pady=(0, 10))

        ent_mod_ad = tk.Entry(f_m_top, font=("Segoe UI", 11, "bold"), bg="#f1f5f9", fg="#334155", relief="flat", width=25)
        ent_mod_ad.pack(side="left", ipady=3, padx=5)
        
        varsayilan_ad = f"A Grubu Reçetesi" if m_idx == 1 else f"B Grubu Reçetesi" if m_idx == 2 else f"C Grubu Reçetesi" if m_idx == 3 else f"Modül {m_idx}"
        ent_mod_ad.insert(0, varsayilan_ad)

        btn_sil = tk.Button(f_m_top, text="🗑️ Sil", bg="white", fg="#ef4444", relief="flat", cursor="hand2", font=("Segoe UI", 9))
        btn_sil.pack(side="right")

        # Girdi Alanları
        f_fields = tk.Frame(f_mod, bg="white")
        f_fields.pack(fill="x")

        def create_field(parent, label, row, fg_color="#0f172a"):
            tk.Label(parent, text=label, bg="white", font=("Segoe UI", 10, "bold"), fg="#475569").grid(row=row, column=0, sticky="w", pady=5)
            ent = tk.Entry(parent, font=("Segoe UI", 11, "bold"), bg="#f8fafc", fg=fg_color, relief="solid", bd=1, justify="right", width=15)
            ent.grid(row=row, column=1, sticky="e", pady=5, padx=10, ipady=3)
            ent.bind("<KeyRelease>", self.hesapla)
            return ent

        ent_fatura = create_field(f_fields, "Fatura Altı Tutar (+):", 0, "#3b82f6")
        ent_muayene = create_field(f_fields, "Elden Hasta Muayene Payı (-):", 1, "#ef4444")
        ent_recete = create_field(f_fields, "Elden Reçete Bedeli (-):", 2, "#ef4444")

        tk.Frame(f_mod, height=1, bg="#e2e8f0").pack(fill="x", pady=10)

        # Modül Net Tutarı
        f_net = tk.Frame(f_mod, bg="white")
        f_net.pack(fill="x")
        tk.Label(f_net, text="Modül Net Geliri:", bg="white", font=("Segoe UI", 11, "bold"), fg="#334155").pack(side="left")
        lbl_net = tk.Label(f_net, text="0.00 ₺", bg="white", font=("Segoe UI", 12, "bold"), fg="#10b981")
        lbl_net.pack(side="right", padx=10)

        modul_dict = {
            "frame": f_mod, "ad": ent_mod_ad, "fatura": ent_fatura,
            "muayene": ent_muayene, "recete": ent_recete, "lbl_net": lbl_net
        }

        def sil_tetik():
            if len(self.moduller) <= 1:
                messagebox.showwarning("Uyarı", "En az bir modül kalmalıdır.", parent=self)
                return
            f_mod.destroy()
            self.moduller.remove(modul_dict)
            self.hesapla()

        btn_sil.config(command=sil_tetik)
        self.moduller.append(modul_dict)

    def hesapla(self, event=None):
        if event: mask_para_birimi(event)

        genel_toplam = 0.0
        for m in self.moduller:
            fatura = temizle_para(m["fatura"].get())
            muayene = temizle_para(m["muayene"].get())
            recete = temizle_para(m["recete"].get())

            net = fatura - muayene - recete
            m["lbl_net"].config(text=f"{net:,.2f} ₺", fg="#10b981" if net >= 0 else "#ef4444")
            genel_toplam += net

        self.lbl_genel_toplam.config(text=f"GENEL TOPLAM: {genel_toplam:,.2f} ₺")

    def kaydet(self):
        alim = self.ent_alim.get().strip()
        if len(alim) == 7: alim += "-01"

        vade = self.ent_vade.get().strip()
        if len(vade) == 7: vade += "-01"

        genel_toplam = 0.0
        detaylar = []
        for m in self.moduller:
            ad = m["ad"].get().strip()
            fatura = temizle_para(m["fatura"].get())
            muayene = temizle_para(m["muayene"].get())
            recete = temizle_para(m["recete"].get())
            
            net = fatura - muayene - recete
            genel_toplam += net
            if net != 0:
                detaylar.append(f"{ad} ({net:,.2f} TL)")

        if genel_toplam == 0:
            messagebox.showwarning("Hata", "Sıfır tutarlı kayıt sisteme işlenemez.", parent=self)
            return

        not_icerigi = " | ".join(detaylar)
        if not not_icerigi: not_icerigi = "SGK Modül Hesaplaması"

        self.save_callback(self.kurum_adi, alim, vade, genel_toplam, not_icerigi, "HAVALE/EFT")
        self.destroy()

# =============================================================================
# MANUEL VERİ GİRİŞ PENCERESİ 
# =============================================================================
class ManualAddWindow(tk.Toplevel):
    def __init__(self, parent, depo_adi, save_callback):
        super().__init__(parent)
        self.title(f"{depo_adi} - Kayıt Ekle")
        self.geometry("450x650") 
        self.configure(bg="white")
        self.save_callback = save_callback
        self.depo_adi = depo_adi
        
        self.transient(parent)
        self.grab_set()
        
        x = parent.winfo_x() + (parent.winfo_width() // 2) - 225
        y = parent.winfo_y() + (parent.winfo_height() // 2) - 325
        self.geometry(f"+{x}+{y}")

        container = tk.Frame(self, bg="white", padx=25, pady=25)
        container.pack(fill="both", expand=True)

        tk.Label(container, text=f"🏢 {depo_adi}", font=("Segoe UI", 16, "bold"), bg="white", fg="#334155").pack(pady=(0, 20))

        tk.Label(container, text="Alım Dönemi (YIL-AY):", font=("Segoe UI", 9, "bold"), bg="white", fg="#64748b").pack(anchor="w")
        self.ent_alim = tk.Entry(container, font=("Segoe UI", 11), relief="solid", bd=1, bg="#f8fafc")
        self.ent_alim.pack(fill="x", ipady=5, pady=(0, 15))
        self.ent_alim.insert(0, date.today().strftime("%Y-%m"))
        self.ent_alim.bind("<KeyRelease>", mask_tarih_otomatik)
        self.ent_alim.bind("<FocusOut>", self.otomatik_vade_hesapla)

        tk.Label(container, text="Ödeme Dönemi (Vade):", font=("Segoe UI", 9, "bold"), bg="white", fg="#64748b").pack(anchor="w")
        self.ent_vade = tk.Entry(container, font=("Segoe UI", 11), relief="solid", bd=1, bg="#f8fafc")
        self.ent_vade.pack(fill="x", ipady=5, pady=(0, 15))
        self.ent_vade.bind("<KeyRelease>", mask_tarih_otomatik)
        
        self.otomatik_vade_hesapla() 

        tk.Label(container, text="Tutar (TL):", font=("Segoe UI", 9, "bold"), bg="white", fg="#64748b").pack(anchor="w")
        self.ent_tutar = tk.Entry(container, font=("Segoe UI", 12, "bold"), relief="solid", bd=1, bg="#fffbeb", fg="#b45309")
        self.ent_tutar.pack(fill="x", ipady=5, pady=(0, 15))
        self.ent_tutar.bind("<KeyRelease>", mask_para_birimi)

        tk.Label(container, text="Ödeme Şekli:", font=("Segoe UI", 9, "bold"), bg="white", fg="#64748b").pack(anchor="w")
        
        self.cmb_tur = ttk.Combobox(container, values=["SENET-HAVALE/EFT", "ÇEK", "SENET-KK"], state="readonly", font=("Segoe UI", 11))
        self.cmb_tur.pack(fill="x", ipady=5, pady=(0, 15))
        self.cmb_tur.current(0) 

        tk.Label(container, text="Not / Açıklama:", font=("Segoe UI", 9, "bold"), bg="white", fg="#64748b").pack(anchor="w")
        self.ent_not = tk.Entry(container, font=("Segoe UI", 11), relief="solid", bd=1, bg="#f8fafc")
        self.ent_not.pack(fill="x", ipady=5, pady=(0, 20))

        ModernButton(container, text="KAYDET", command=self.validate_and_save, width=380, height=45, bg_color="#10b981").pack(side="bottom")

    def otomatik_vade_hesapla(self, event=None):
        try:
            alim_str = self.ent_alim.get().strip()
            if len(alim_str) >= 7:
                alim_dt = datetime.strptime(alim_str[:7] + "-01", "%Y-%m-%d")
                ay = alim_dt.month + 3
                yil = alim_dt.year
                if ay > 12:
                    yil += ay // 12
                    ay = ay % 12
                    if ay == 0: ay = 12; yil -= 1
                vade_str = f"{yil}-{ay:02d}"
                self.ent_vade.delete(0, tk.END)
                self.ent_vade.insert(0, vade_str)
        except: pass

    def validate_and_save(self):
        try:
            alim = self.ent_alim.get().strip()
            if len(alim) == 7: alim += "-01"
            
            vade = self.ent_vade.get().strip()
            if len(vade) == 7: vade += "-01"

            tutar = temizle_para(self.ent_tutar.get())
            not_icerik = self.ent_not.get()
            odeme_turu = self.cmb_tur.get()

            self.save_callback(self.depo_adi, alim, vade, tutar, not_icerik, odeme_turu)
            self.destroy()
        except Exception as e:
            messagebox.showerror("Hata", f"Bilgileri kontrol edin: {e}")
# =============================================================================
# YARDIMCI SINIFLAR (BUNLAR ECZANE ASİSTANI SINIFININ DIŞINDA/ÜSTÜNDE OLMALIDIR)
# =============================================================================

class NoteBubble(tk.Toplevel):
    def __init__(self, parent, x, y, text_content, on_save_callback):
        super().__init__(parent)
        self.on_save = on_save_callback
        self.overrideredirect(True) 
        self.attributes("-topmost", True) 
        self.geometry(f"320x220+{x+20}+{y-10}")
        self.config(bg="#f59e0b", padx=2, pady=2)
        
        container = tk.Frame(self, bg="#fffbeb")
        container.pack(fill="both", expand=True)
        
        header = tk.Frame(container, bg="#fffbeb", cursor="fleur")
        header.pack(fill="x", padx=5, pady=5)
        header.bind("<Button-1>", self.start_move)
        header.bind("<B1-Motion>", self.do_move)
        
        lbl_title = tk.Label(header, text="📝 Not Düzenle", font=("Segoe UI", 10, "bold"), bg="#fffbeb", fg="#92400e", cursor="fleur")
        lbl_title.pack(side="left")
        lbl_title.bind("<Button-1>", self.start_move)
        lbl_title.bind("<B1-Motion>", self.do_move)
        
        close_btn = tk.Label(header, text="✖", font=("Segoe UI", 10, "bold"), bg="#fffbeb", fg="#ef4444", cursor="hand2")
        close_btn.pack(side="right")
        close_btn.bind("<Button-1>", lambda e: self.destroy())
        
        self.txt = tk.Text(container, font=("Segoe UI", 11), height=6, bg="white", bd=0, highlightthickness=1, highlightbackground="#fbbf24")
        self.txt.insert("1.0", text_content)
        self.txt.pack(fill="both", expand=True, padx=10, pady=5)
        self.after(10, lambda: self.txt.focus_force()) 
        
        btn_save = tk.Label(container, text="💾 KAYDET", font=("Segoe UI", 10, "bold"), bg="#f59e0b", fg="white", pady=10, cursor="hand2")
        btn_save.pack(fill="x", side="bottom")
        btn_save.bind("<Button-1>", self.save_and_close)
        self.bind("<Escape>", lambda e: self.destroy())

    def start_move(self, event):
        self.x = event.x
        self.y = event.y

    def do_move(self, event):
        deltax = event.x - self.x
        deltay = event.y - self.y
        x = self.winfo_x() + deltax
        y = self.winfo_y() + deltay
        self.geometry(f"+{x}+{y}")

    def save_and_close(self, event=None):
        self.on_save(self.txt.get("1.0", "end-1c"))
        self.destroy()

class DepotCard(tk.Frame):
    def __init__(self, parent, depo_adi, imlec, baglanti, refresh_callback, swap_callback):
        super().__init__(parent, bg="white", bd=1, relief="solid")
        self.depo_adi = depo_adi
        self.imlec = imlec
        self.baglanti = baglanti
        self.refresh = refresh_callback
        self.swap_callback = swap_callback
        self.is_expanded = False
        self.secili_yil = datetime.now().year

        self.imlec.execute("SELECT anlasma_sarti FROM depo_ayarlari WHERE fatura_adi=?", (depo_adi,))
        res = self.imlec.fetchone()
        self.anlasma_metni = res[0] if res else "Anlaşma şartları girilmedi."

        self.imlec.execute("SELECT SUM(tutar) FROM odemeler WHERE fatura_adi=? AND durum='ODENMEDİ'", (depo_adi,))
        res3 = self.imlec.fetchone()
        self.kalan_borc = res3[0] if res3 and res3[0] else 0.0

        self.header_frame = tk.Frame(self, bg="#f8fafc", padx=20, pady=15, cursor="fleur")
        self.header_frame.pack(fill="x")
        self.bind_drag_events(self.header_frame)
        
        self.lbl_icon = tk.Label(self.header_frame, text="▶", font=("Segoe UI", 16), bg="#f8fafc", fg="#64748b", cursor="hand2")
        self.lbl_icon.pack(side="left", padx=(0, 15))
        self.lbl_icon.bind("<Button-1>", self.toggle)
        
        title_lbl = tk.Label(self.header_frame, text=f"{depo_adi}", font=("Segoe UI", 16, "bold"), bg="#f8fafc", fg="#0f172a", cursor="fleur")
        title_lbl.pack(side="left")
        self.bind_drag_events(title_lbl)
        
        if self.kalan_borc > 0:
            debt_lbl = tk.Label(self.header_frame, text=f"Borç: {self.kalan_borc:,.0f} ₺", font=("Segoe UI", 14, "bold"), bg="#f8fafc", fg="#ef4444", cursor="fleur")
            debt_lbl.pack(side="right")
            self.bind_drag_events(debt_lbl)

        self.content_frame = tk.Frame(self, bg="white", padx=30, pady=20)
        
        f_terms = tk.Frame(self.content_frame, bg="#f1f5f9", padx=15, pady=15, relief="flat")
        f_terms.pack(fill="x", pady=(0, 15))
        tk.Label(f_terms, text="📝 ANLAŞMA ŞARTI:", font=("Segoe UI", 10, "bold"), bg="#f1f5f9", fg="#64748b").pack(anchor="w")
        self.lbl_terms = tk.Label(f_terms, text=self.anlasma_metni, font=("Segoe UI", 12), bg="#f1f5f9", fg="#0f172a", wraplength=800, justify="left")
        self.lbl_terms.pack(anchor="w", pady=(5, 0))
        tk.Button(f_terms, text="DÜZENLE", font=("Segoe UI", 9), command=self.edit_terms, bg="#cbd5e1", relief="flat").pack(anchor="e")

        f_year = tk.Frame(self.content_frame, bg="white", pady=5)
        f_year.pack(fill="x")
        
        tk.Label(f_year, text="📅 İstatistik Yılı Seçiniz:", font=("Segoe UI", 10, "bold"), bg="white", fg="#475569").pack(side="left")
        
        yil_listesi = [str(y) for y in range(datetime.now().year - 4, datetime.now().year + 3)]
        self.cmb_yil = ttk.Combobox(f_year, values=yil_listesi, state="readonly", width=10, font=("Segoe UI", 10))
        self.cmb_yil.set(str(self.secili_yil))
        self.cmb_yil.pack(side="left", padx=10)
        self.cmb_yil.bind("<<ComboboxSelected>>", self.verileri_guncelle)

        tk.Frame(self.content_frame, height=1, bg="#e2e8f0").pack(fill="x", pady=10)

        self.row1_frame = tk.Frame(self.content_frame, bg="white")
        self.row1_frame.pack(fill="x", pady=4)
        self.lbl_row1_title = tk.Label(self.row1_frame, text="Yükleniyor...", font=("Segoe UI", 11), bg="white", fg="#475569")
        self.lbl_row1_title.pack(side="left")
        self.lbl_row1_val = tk.Label(self.row1_frame, text="0.00 ₺", font=("Segoe UI", 11), bg="white", fg="#3b82f6")
        self.lbl_row1_val.pack(side="right")

        self.row2_frame = tk.Frame(self.content_frame, bg="white")
        self.row2_frame.pack(fill="x", pady=4)
        self.lbl_row2_title = tk.Label(self.row2_frame, text="Yükleniyor...", font=("Segoe UI", 11), bg="white", fg="#475569")
        self.lbl_row2_title.pack(side="left")
        self.lbl_row2_val = tk.Label(self.row2_frame, text="0.00 ₺", font=("Segoe UI", 11), bg="white", fg="#f59e0b")
        self.lbl_row2_val.pack(side="right")

        self.row3_frame = tk.Frame(self.content_frame, bg="white")
        self.row3_frame.pack(fill="x", pady=4)
        self.lbl_total_title = tk.Label(self.row3_frame, text="GENEL TOPLAM:", font=("Segoe UI", 11, "bold"), bg="white", fg="#0f172a")
        self.lbl_total_title.pack(side="left")
        self.lbl_total_val = tk.Label(self.row3_frame, text="0.00 ₺", font=("Segoe UI", 11, "bold"), bg="white", fg="#0f172a")
        self.lbl_total_val.pack(side="right")

        tk.Frame(self.content_frame, height=2, bg="#e2e8f0").pack(fill="x", pady=20)
        
        f_debt = tk.Frame(self.content_frame, bg="white")
        f_debt.pack(fill="x")
        tk.Label(f_debt, text="ÖDENMESİ GEREKEN KALAN BORÇ", font=("Segoe UI", 11, "bold"), bg="white", fg="#64748b").pack(anchor="center")
        tk.Label(f_debt, text=f"{self.kalan_borc:,.2f} ₺", font=("Segoe UI", 26, "bold"), bg="white", fg="#ef4444").pack(anchor="center")

        self.verileri_guncelle()

    def verileri_guncelle(self, event=None):
        try:
            yil = self.cmb_yil.get()
        except:
            yil = str(datetime.now().year)

        self.imlec.execute(f"SELECT SUM(tutar) FROM odemeler WHERE fatura_adi=? AND strftime('%Y', alim_tarihi)=? AND strftime('%m', alim_tarihi) BETWEEN '01' AND '06'", (self.depo_adi, yil))
        res1 = self.imlec.fetchone()
        ilk_yari = res1[0] if res1 and res1[0] else 0.0

        self.imlec.execute(f"SELECT SUM(tutar) FROM odemeler WHERE fatura_adi=? AND strftime('%Y', alim_tarihi)=? AND strftime('%m', alim_tarihi) BETWEEN '07' AND '12'", (self.depo_adi, yil))
        res2 = self.imlec.fetchone()
        ikinci_yari = res2[0] if res2 and res2[0] else 0.0
        
        genel_toplam = ilk_yari + ikinci_yari

        self.lbl_row1_title.config(text=f"{yil} Yılı (1-6. Ay) Toplam Alım:")
        self.lbl_row1_val.config(text=f"{ilk_yari:,.2f} ₺")

        self.lbl_row2_title.config(text=f"{yil} Yılı (7-12. Ay) Toplam Alım:")
        self.lbl_row2_val.config(text=f"{ikinci_yari:,.2f} ₺")

        self.lbl_total_title.config(text=f"{yil} Yılı GENEL TOPLAM:")
        self.lbl_total_val.config(text=f"{genel_toplam:,.2f} ₺")

    def bind_drag_events(self, widget):
        widget.bind("<Button-1>", self.start_drag)
        widget.bind("<B1-Motion>", self.do_drag)
        widget.bind("<ButtonRelease-1>", self.stop_drag)

    def start_drag(self, event):
        self.header_frame.config(bg="#e2e8f0")

    def do_drag(self, event):
        pass

    def stop_drag(self, event):
        self.header_frame.config(bg="#f8fafc")
        x, y = self.winfo_pointerxy()
        target = self.winfo_containing(x, y)
        if target:
            current = target
            target_card = None
            while current:
                if isinstance(current, DepotCard):
                    target_card = current
                    break
                current = current.master
            if target_card and target_card != self:
                self.swap_callback(self.depo_adi, target_card.depo_adi)

    def toggle(self, event=None):
        if self.is_expanded:
            self.content_frame.forget()
            self.lbl_icon.config(text="▶")
        else:
            self.content_frame.pack(fill="x")
            self.lbl_icon.config(text="▼")
        self.is_expanded = not self.is_expanded

    def edit_terms(self):
        win = tk.Toplevel(self)
        win.title(f"{self.depo_adi} | Şartları Düzenle")
        win.geometry("600x550") 
        win.configure(bg="#f8fafc") 
        
        win.transient(self.winfo_toplevel())
        win.grab_set()
        
        root_x = self.winfo_toplevel().winfo_x()
        root_y = self.winfo_toplevel().winfo_y()
        root_w = self.winfo_toplevel().winfo_width()
        root_h = self.winfo_toplevel().winfo_height()
        x = root_x + (root_w // 2) - 300
        y = root_y + (root_h // 2) - 275
        win.geometry(f"+{x}+{y}")

        footer = tk.Frame(win, bg="white", pady=15, padx=20, height=80)
        footer.pack(side="bottom", fill="x")
        
        tk.Frame(win, height=1, bg="#e2e8f0").pack(side="bottom", fill="x")

        header = tk.Frame(win, bg="#f8fafc", pady=20, padx=25)
        header.pack(side="top", fill="x")
        
        tk.Label(header, text="Anlaşma ve Çalışma Şartları", font=("Segoe UI", 16, "bold"), bg="#f8fafc", fg="#1e293b").pack(anchor="w")
        
        body = tk.Frame(win, bg="#f8fafc", padx=25, pady=5)
        body.pack(side="top", fill="both", expand=True)
        
        editor_frame = tk.Frame(body, bg="white", bd=1, relief="solid")
        editor_frame.pack(fill="both", expand=True)
        
        sc = ttk.Scrollbar(editor_frame, orient="vertical")
        sc.pack(side="right", fill="y")
        
        txt = tk.Text(editor_frame, font=("Segoe UI", 11), bg="white", fg="#334155",
                      bd=0, highlightthickness=0, padx=15, pady=15,
                      yscrollcommand=sc.set)
        txt.pack(fill="both", expand=True)
        sc.config(command=txt.yview)
        
        mevcut = self.anlasma_metni if self.anlasma_metni else ""
        txt.insert("1.0", mevcut)
        txt.focus_set()
        
        def madde_ekle():
            txt.insert(tk.INSERT, "\n• ")
            txt.focus_set()

        def kaydet_ve_kapat():
            yeni_metin = txt.get("1.0", "end-1c").strip()
            try:
                self.imlec.execute("INSERT OR REPLACE INTO depo_ayarlari (fatura_adi, anlasma_sarti) VALUES (?, ?)", 
                                   (self.depo_adi, yeni_metin))
                self.baglanti_skt.commit()
                
                self.anlasma_metni = yeni_metin
                self.lbl_terms.config(text=yeni_metin)
                win.destroy()
            except Exception as e:
                messagebox.showerror("Hata", f"Kaydedilemedi: {e}")

        btn_madde = tk.Button(footer, text="➕ Madde Ekle", font=("Segoe UI", 10, "bold"), 
                              bg="#eff6ff", fg="#2563eb", activebackground="#dbeafe",
                              relief="flat", cursor="hand2", padx=15, pady=8,
                              command=madde_ekle)
        btn_madde.pack(side="left")

        btn_kaydet = tk.Button(footer, text="KAYDET", font=("Segoe UI", 10, "bold"), 
                               bg="#10b981", fg="white", activebackground="#059669", activeforeground="white",
                               relief="flat", cursor="hand2", padx=20, pady=8,
                               command=kaydet_ve_kapat)
        btn_kaydet.pack(side="right", padx=(10, 0))

        btn_iptal = tk.Button(footer, text="İPTAL", font=("Segoe UI", 10, "bold"), 
                              bg="#f1f5f9", fg="#64748b", activebackground="#e2e8f0",
                              relief="flat", cursor="hand2", padx=15, pady=8,
                              command=win.destroy)
        btn_iptal.pack(side="right")

class OtoGuncelleyici:
    def __init__(self, root, baslat_callback):
        self.root = root
        self.baslat_callback = baslat_callback
        self.guncelleme_var_mi()

    def guncelleme_var_mi(self):
        # Arka planda donmayı engellemek için thread kullanıyoruz
        def kontrol_et():
            try:
                # İnternetteki version.txt dosyasını oku (3 saniye zaman aşımı)
                istek = urllib.request.Request(VERSIYON_URL, headers={'Cache-Control': 'no-cache'})
                cevap = urllib.request.urlopen(istek, timeout=3)
                yeni_versiyon = cevap.read().decode('utf-8').strip()

                if yeni_versiyon != MEVCUT_VERSIYON and len(yeni_versiyon) < 10: # Güvenlik için uzunluk kontrolü
                    self.root.after(0, lambda: self.guncelleme_uyarisi_goster(yeni_versiyon))
                else:
                    self.root.after(0, self.baslat_callback) # Güncelleme yoksa programı normal başlat
            except Exception as e:
                # İnternet yoksa veya siteye ulaşılamadıysa programı normal başlat
                self.root.after(0, self.baslat_callback)
        
        threading.Thread(target=kontrol_et, daemon=True).start()

    def guncelleme_uyarisi_goster(self, yeni_versiyon):
        if messagebox.askyesno("Güncelleme Mevcut", f"Eczacı Defteri'nin yeni bir sürümü (v{yeni_versiyon}) bulundu!\n\nŞu anki sürümünüz: v{MEVCUT_VERSIYON}\n\nŞimdi indirip kurmak ister misiniz? (Program yeniden başlatılacaktır)"):
            self.guncelleme_indir_ve_kur()
        else:
            self.baslat_callback() # İptal ederse normal başlasın

    def guncelleme_indir_ve_kur(self):
        win = tk.Toplevel(self.root)
        win.title("Güncelleniyor...")
        win.geometry("400x150")
        win.transient(self.root)
        win.grab_set()
        
        # Ekranı ortala
        win.geometry(f"+{self.root.winfo_screenwidth()//2 - 200}+{self.root.winfo_screenheight()//2 - 75}")

        tk.Label(win, text="Yeni sürüm indiriliyor, lütfen bekleyin...", font=("Segoe UI", 11, "bold")).pack(pady=20)
        
        progress = ttk.Progressbar(win, orient="horizontal", length=300, mode="determinate")
        progress.pack(pady=10)

        def indir_thread():
            try:
                dosya_adi = "Guncel_EczaciDefteri.exe"
                def raporla(blok_numarasi, blok_boyutu, toplam_boyut):
                    yuzde = int((blok_numarasi * blok_boyutu * 100) / toplam_boyut)
                    if yuzde > 100: yuzde = 100
                    self.root.after(0, lambda: progress.config(value=yuzde))

                # Yeni exe dosyasını indir
                urllib.request.urlretrieve(EXE_URL, dosya_adi, reporthook=raporla)
                
                self.root.after(0, self.kurulumu_baslat)
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Hata", f"İndirme başarısız oldu:\n{e}", parent=win))
                self.root.after(0, win.destroy)
                self.root.after(0, self.baslat_callback)

        threading.Thread(target=indir_thread, daemon=True).start()

    def kurulumu_baslat(self):
        # Arka planda çalışacak bir güncelleme betiği (bat) oluşturuyoruz
        bat_icerik = """@echo off
timeout /t 2 /nobreak > NUL
del "EczaciDefteri.exe"
ren "Guncel_EczaciDefteri.exe" "EczaciDefteri.exe"
start "" "EczaciDefteri.exe"
del "%~f0"
"""
        with open("guncelle.bat", "w") as f:
            f.write(bat_icerik)
        
        # Bat dosyasını çalıştır ve mevcut programı kapat
        subprocess.Popen("guncelle.bat", shell=True)
        sys.exit() # Program kendini kapatır, BAT dosyası yeni exe'yi yerine koyar.

class EczaneAsistani:
    def __init__(self, pencere, aktif_rol="Yönetici"):
        self.pencere = pencere
        self.aktif_rol = aktif_rol # Sisteme kimin girdiğini hafızaya aldık
        
        # --- PERSONEL YETKİLERİNİ VERİTABANINDAN ÇEK ---
        self.personel_yetkileri = {
            "finans_gorsun": False,
            "para_gorsun": False,
            "ayarlar_gorsun": False
        }
        try:
            # Yetkileri daha sonra yükleyebilmek için DB bağlantısından hemen sonraya bir try-except koyacağız.
            pass
        except: pass
        self.pencere.title("Eczacı Defteri")
        self.pencere.geometry("1400x900")
        
        # --- VERİTABANI YOLLARI (AĞ VE AYAR DESTEKLİ) ---
        uygulama_klasoru = yerel_db_yolunu_getir()
        
        if not os.path.exists(uygulama_klasoru):
            try:
                os.makedirs(uygulama_klasoru)
            except:
                print("Ağ klasörüne ulaşılamadı, yerel klasör kullanılıyor...")
                uygulama_klasoru = os.path.join(os.path.expanduser('~'), "EczaneAsistani_Veri")
                if not os.path.exists(uygulama_klasoru): os.makedirs(uygulama_klasoru)
            
        # timeout=15: Ağ üzerinde aynı anda 2 kişi işlem yaparsa kilitlenmeyi önler
        self.db_skt_adi = os.path.join(uygulama_klasoru, "eczane_skt.db")
        self.baglanti_skt = sqlite3.connect(self.db_skt_adi, timeout=15, check_same_thread=False)
        
        kullanici_db_yolu = os.path.join(uygulama_klasoru, "kullanicilar.db")
        self.vt_kullanici = sqlite3.connect(kullanici_db_yolu, timeout=15, check_same_thread=False)
        self.imlec_kullanici = self.vt_kullanici.cursor()
        
        self.imlec = self.baglanti_skt.cursor() 
        
        self.db_finans_adi = os.path.join(uygulama_klasoru, "eczane_finans.db")
        self.baglanti_finans = sqlite3.connect(self.db_finans_adi, timeout=15, check_same_thread=False)
        self.imlec_finans = self.baglanti_finans.cursor()
        
        self.bildirimler = []

        # Logo yolunu sistem tepsisi (tray) için sakla
        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        self.logo_yolu_tray = os.path.join(base_dir, "logo_hd_transparan.png")

        # Çarpı (X) tuşuna basınca programı kapatma, gizle!
        self.pencere.protocol('WM_DELETE_WINDOW', self.kapatma_istegi)
        
        # Kurulumları Başlat
        self.veritabani_kur_skt()
        self.veritabani_kur_finans()
        self.tablo_guncelleme_kontrol()
        # PC Açılış ayarını kontrol et ve zorla uygula
        try:
            self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='pc_acilis_baslat'")
            oto_baslat = self.imlec.fetchone()
            if oto_baslat and oto_baslat[0] == '1':
                self.başlangıç_ayarı_güncelle(True)
        except: pass

        self.sayim_listesi = []; self.karekod_temp_data = []
        self.secili_tumunu = False; self.detayli_gorunum = True
        self.aktif_sekme = 0; self.widgets_to_refresh = []

        self.ana_container = tk.Frame(pencere)
        self.ana_container.pack(fill="both", expand=True)

        # Yetkileri DB'den çek
        try:
            self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='personel_yetkileri'")
            res_yetki = self.imlec.fetchone()
            if res_yetki and res_yetki[0]:
                self.personel_yetkileri = eval(res_yetki[0])
        except: pass

        self.tema_uygula() 
        self.pencere.update()
        
        # --- AĞIR İŞLEMLERİ ERTELE ---
        def verileri_ve_ozeti_yukle():
            # Arayüz burada çiziliyor
            try: self.ana_uygulamaya_gec()
            except: pass
            
            # =================================================================
            # İŞTE SİHİR BURADA: ANA EKRAN %100 ÇİZİLDİ, ŞİMDİ GİRİŞ PERDESİNİ KALDIR!
            # =================================================================
            if hasattr(self.pencere, 'login_perdesi'):
                try: 
                    self.pencere.login_perdesi.destroy()
                    del self.pencere.login_perdesi
                except: pass
            
            # Fareyi normale çevir
            self.pencere.config(cursor="")
            
            # Sessiz görevleri arka planda başlat
            import threading
            try: threading.Thread(target=self.otomatik_aylik_isler, daemon=True).start()
            except: pass
            try: self.otomatik_odeme_motoru()
            except: pass
            try: self.otomatik_kontrol_dongusu()
            except: pass
            
            # YENİ: Ağ bekçisini başlat
            try: self.ag_ve_senkronizasyon_bekcisi()
            except: pass

        # Veri doldurma işlemini iskelet çizildikten 100 milisaniye sonraya bırakıyoruz
        self.pencere.after(100, verileri_ve_ozeti_yukle)

    # =========================================================================
    # MODERN EKSİK / SİPARİŞ SEPETİ (OTO-KOPYALA YAKALAYICI EKLENDİ)
    # =========================================================================
    def toggle_eksik_sepeti(self):
        c = TM  
        
        if hasattr(self, 'eksik_penceresi') and self.eksik_penceresi.winfo_exists():
            self.eksik_kaydet()
            self.eksik_penceresi.destroy()
            return

        self.eksik_penceresi = tk.Toplevel(self.pencere)
        self.eksik_penceresi.overrideredirect(True) 
        self.eksik_penceresi.attributes("-topmost", True) 
        
        x = self.pencere.winfo_rootx() + self.pencere.winfo_width() - 380
        y = self.pencere.winfo_rooty() + 60
        self.eksik_penceresi.geometry(f"360x640+{x}+{y}")
        
        self.eksik_penceresi.configure(bg=c.get_color("card_bg"), highlightthickness=2, highlightbackground=c.get_color("border"))

        # --- SÜRÜKLE BIRAK MANTIĞI ---
        def start_move(event):
            self.eksik_penceresi._drag_start_x = event.x
            self.eksik_penceresi._drag_start_y = event.y
            return "break"

        def do_move(event):
            x = self.eksik_penceresi.winfo_x() - self.eksik_penceresi._drag_start_x + event.x
            y = self.eksik_penceresi.winfo_y() - self.eksik_penceresi._drag_start_y + event.y
            self.eksik_penceresi.geometry(f"+{x}+{y}")
            return "break" 

        # --- ÜST ÇUBUK (HEADER) ---
        header = tk.Frame(self.eksik_penceresi, bg=c.get_color("header_bg"), cursor="fleur", pady=8)
        header.pack(fill="x")
        
        header.bind("<Button-1>", start_move)
        header.bind("<B1-Motion>", do_move)

        lbl_drag = tk.Label(header, text="⠿", font=("Segoe UI", 12), bg=c.get_color("header_bg"), fg="#94a3b8", cursor="fleur")
        lbl_drag.pack(side="left", padx=(10, 5))
        lbl_drag.bind("<Button-1>", start_move)
        lbl_drag.bind("<B1-Motion>", do_move)

        lbl_title = tk.Label(header, text="Eksik Sipariş Sepeti", font=("Segoe UI", 11, "bold"), bg=c.get_color("header_bg"), fg="white", cursor="fleur")
        lbl_title.pack(side="left")
        lbl_title.bind("<Button-1>", start_move)
        lbl_title.bind("<B1-Motion>", do_move)

        def kapat():
            self.eksik_kaydet()
            self.oto_yakala_aktif = False # Kapatınca dinleyiciyi durdur
            self.eksik_penceresi.destroy()

        btn_close = tk.Label(header, text="✖", font=("Segoe UI", 12, "bold"), bg=c.get_color("header_bg"), fg="#ef4444", cursor="hand2")
        btn_close.pack(side="right", padx=10)
        btn_close.bind("<Button-1>", lambda e: kapat())

        # =====================================================================
        # YENİ ÖZELLİK: OTO-KOPYALA YAKALAYICI (PANODAN OTOMATİK ÇEKME)
        # =====================================================================
        self.oto_yakala_aktif = False
        self.son_pano_verisi = ""
        
        lbl_oto = tk.Label(header, text="🧲 KAPALI", font=("Segoe UI", 8, "bold"), bg=c.get_color("header_bg"), fg="#94a3b8", cursor="hand2")
        lbl_oto.pack(side="right", padx=(0, 5))
        
        def oto_dinleyici():
            # Pencere kapandıysa veya özellik kapatıldıysa döngüyü kır
            if not hasattr(self, 'eksik_penceresi') or not self.eksik_penceresi.winfo_exists() or not self.oto_yakala_aktif:
                return
            
            try:
                # Panodaki güncel yazıyı al
                guncel_pano = self.pencere.clipboard_get().strip()
                
                # Eğer panodaki yazı eskisinden farklıysa ve boş değilse (Yeni bir şey kopyalandıysa)
                if guncel_pano and guncel_pano != self.son_pano_verisi:
                    self.son_pano_verisi = guncel_pano
                    
                    # Sondaki boş satırları temizle
                    while self.eksik_listesi and not self.eksik_listesi[-1]["ad"].strip() and not self.eksik_listesi[-1]["adet"].strip():
                        self.eksik_listesi.pop()
                    
                    # Kopyalanan yazıyı (birden fazla satır olabilir) sepete ekle
                    yeni_eklenen = False
                    for satir in guncel_pano.splitlines():
                        if satir.strip():
                            self.eksik_listesi.append({"ad": satir.strip(), "adet": ""})
                            yeni_eklenen = True
                    
                    if yeni_eklenen:
                        # Otomatik olarak son sayfaya git ki kullanıcı eklendiğini görsün
                        max_sayfa = max(1, (len(self.eksik_listesi) // 10) + 1)
                        self.eksik_sayfa = max_sayfa - 1
                        
                        render_page()
                        self.eksik_kaydet_delayed()
            except:
                pass
                
            # Kendini her 800 milisaniyede bir tekrar çağır (Arka planda yormadan dinler)
            self.eksik_penceresi.after(800, oto_dinleyici)

        def toggle_oto(e=None):
            self.oto_yakala_aktif = not self.oto_yakala_aktif
            if self.oto_yakala_aktif:
                lbl_oto.config(fg="#10b981", text="🧲 AÇIK")
                try: 
                    # Açıldığı andaki panoyu hafızaya al ki eskiyi eklemesin
                    self.son_pano_verisi = self.pencere.clipboard_get().strip()
                except: 
                    self.son_pano_verisi = ""
                oto_dinleyici() # Dinlemeyi başlat
                self.goster_bildirim("Oto-Ekle Aktif", "Herhangi bir yerden metin kopyaladığınızda otomatik sepete eklenecek.")
            else:
                lbl_oto.config(fg="#94a3b8", text="🧲 KAPALI")
        
        lbl_oto.bind("<Button-1>", toggle_oto)
        # =====================================================================

        # --- VERİYİ YÜKLE VE PARÇALA ---
        self.eksik_listesi = [] 
        try:
            self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='eksik_sepeti'")
            res = self.imlec.fetchone()
            if res and res[0]:
                satirlar = res[0].split('\n')
                for satir in satirlar:
                    if not satir.strip(): continue
                    if "|" in satir:
                        ad, adet = satir.split("|", 1)
                        self.eksik_listesi.append({"ad": ad.strip(), "adet": adet.strip()})
                    else:
                        self.eksik_listesi.append({"ad": satir.strip(), "adet": ""})
        except: pass
        self.eksik_sayfa = 0

        # --- TABLO BAŞLIKLARI ---
        f_titles = tk.Frame(self.eksik_penceresi, bg=c.get_color("bg_main"), pady=5)
        f_titles.pack(fill="x")
        tk.Label(f_titles, text="No", font=("Segoe UI", 9, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left", padx=(10, 5))
        tk.Label(f_titles, text="İlaç / Ürün Adı", font=("Segoe UI", 9, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left", padx=5)
        tk.Label(f_titles, text="Miktar", font=("Segoe UI", 9, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="right", padx=35) 

        # --- ALT KONTROLLER (Önce Pack Ediyoruz ki Taşmasın) ---
        f_bot = tk.Frame(self.eksik_penceresi, bg=c.get_color("bg_main"), pady=10, padx=10)
        f_bot.pack(side="bottom", fill="x")

        # Sayfalama Butonlarını Tanımlama
        f_page = tk.Frame(f_bot, bg=c.get_color("bg_main"))
        f_page.pack(side="top", fill="x", pady=(0, 10))
        
        btn_prev = tk.Button(f_page, text="◀ ÖNCEKİ", bg=c.get_color("card_bg"), fg=c.get_color("fg_text"), relief="solid", bd=1, cursor="hand2", padx=10, pady=2)
        btn_prev.pack(side="left")
        
        lbl_page = tk.Label(f_page, text="1 / 1", bg=c.get_color("bg_main"), fg=c.get_color("fg_text"), font=("Segoe UI", 10, "bold"))
        lbl_page.pack(side="left", expand=True)
        
        btn_next = tk.Button(f_page, text="SONRAKİ ▶", bg=c.get_color("card_bg"), fg=c.get_color("fg_text"), relief="solid", bd=1, cursor="hand2", padx=10, pady=2)
        btn_next.pack(side="right")

        # --- İÇERİK ALANI ---
        content = tk.Frame(self.eksik_penceresi, bg=c.get_color("card_bg"), padx=5, pady=5)
        content.pack(side="top", fill="both", expand=True)

        self.eksik_satirlari = []

        def update_pagination():
            last_filled = -1
            for idx, item in enumerate(self.eksik_listesi):
                if item["ad"].strip() or item["adet"].strip():
                    last_filled = idx
            
            max_sayfa = (last_filled // 10) + 2 if last_filled >= 0 else 1
            lbl_page.config(text=f"{self.eksik_sayfa + 1} / {max_sayfa}")
            
            if self.eksik_sayfa > 0:
                btn_prev.config(state="normal", bg=c.get_color("card_bg"), fg=c.get_color("fg_text"))
            else:
                btn_prev.config(state="disabled", bg=c.get_color("bg_main"), fg="#94a3b8")
                
            if self.eksik_sayfa < max_sayfa - 1:
                btn_next.config(state="normal", bg="#3b82f6", fg="white")
            else:
                btn_next.config(state="disabled", bg=c.get_color("bg_main"), fg="#94a3b8")

        def sil_satir(row_idx):
            list_idx = (self.eksik_sayfa * 10) + row_idx
            if list_idx < len(self.eksik_listesi):
                self.eksik_listesi.pop(list_idx)
                self.eksik_kaydet()
                render_page()

        def on_yazi_degisti(event, row_idx):
            list_idx = (self.eksik_sayfa * 10) + row_idx
            
            while len(self.eksik_listesi) <= list_idx:
                self.eksik_listesi.append({"ad": "", "adet": ""})
                
            yeni_ad = self.eksik_satirlari[row_idx]['ent_ad'].get()
            yeni_adet = self.eksik_satirlari[row_idx]['ent_adet'].get()
            
            self.eksik_listesi[list_idx] = {"ad": yeni_ad, "adet": yeni_adet}
            
            if yeni_ad.strip() or yeni_adet.strip():
                self.eksik_satirlari[row_idx]['btn_x'].pack(side="right", padx=2)
            else:
                self.eksik_satirlari[row_idx]['btn_x'].pack_forget()
                
            update_pagination()
            self.eksik_kaydet_delayed()

        # 10 Adet Satır Oluştur
        for i in range(10):
            row_f = tk.Frame(content, bg=c.get_color("card_bg"), pady=2)
            row_f.pack(fill="x", pady=2)
            
            lbl_no = tk.Label(row_f, text=f"{i+1}.", font=("Segoe UI", 9, "bold"), bg=c.get_color("card_bg"), fg="#94a3b8", width=3, anchor="e")
            lbl_no.pack(side="left", padx=(0, 5))
            
            btn_x = tk.Label(row_f, text="✖", fg="#ef4444", bg=c.get_color("card_bg"), cursor="hand2", font=("Segoe UI", 10))
            btn_x.bind("<Button-1>", lambda e, idx=i: sil_satir(idx))
            
            ent_adet = tk.Entry(row_f, font=("Segoe UI", 10, "bold"), bg=c.get_color("input_bg"), fg=c.get_color("btn_primary"), bd=1, relief="solid", width=6, justify="center", insertbackground=c.get_color("fg_text"))
            ent_adet.pack(side="right", padx=(5, 5), ipady=4)
            ent_adet.bind("<KeyRelease>", lambda e, idx=i: on_yazi_degisti(e, idx))

            ent_ad = tk.Entry(row_f, font=("Segoe UI", 10), bg=c.get_color("input_bg"), fg=c.get_color("fg_text"), bd=1, relief="solid", insertbackground=c.get_color("fg_text"))
            ent_ad.pack(side="left", fill="x", expand=True, ipady=4)
            ent_ad.bind("<KeyRelease>", lambda e, idx=i: on_yazi_degisti(e, idx))
            
            self.eksik_satirlari.append({
                'lbl_no': lbl_no,
                'ent_ad': ent_ad,
                'ent_adet': ent_adet,
                'btn_x': btn_x
            })

        def render_page():
            while self.eksik_listesi and not self.eksik_listesi[-1]["ad"].strip() and not self.eksik_listesi[-1]["adet"].strip():
                self.eksik_listesi.pop()
                
            update_pagination()
            start_idx = self.eksik_sayfa * 10

            for i in range(10):
                list_idx = start_idx + i
                satir = self.eksik_satirlari[i]
                
                satir['lbl_no'].config(text=f"{list_idx + 1}.")
                satir['ent_ad'].delete(0, tk.END)
                satir['ent_adet'].delete(0, tk.END)
                
                if list_idx < len(self.eksik_listesi):
                    veri = self.eksik_listesi[list_idx]
                    satir['ent_ad'].insert(0, veri["ad"])
                    satir['ent_adet'].insert(0, veri["adet"])
                    
                    if veri["ad"].strip() or veri["adet"].strip():
                        satir['btn_x'].pack(side="right", padx=2)
                    else:
                        satir['btn_x'].pack_forget()
                else:
                    satir['btn_x'].pack_forget()

        def ileri():
            self.eksik_sayfa += 1
            render_page()
            
        def geri():
            if self.eksik_sayfa > 0:
                self.eksik_sayfa -= 1
                render_page()

        btn_prev.config(command=geri)
        btn_next.config(command=ileri)

        def kopyala_ve_temizle():
            kopyalanacak = []
            for item in self.eksik_listesi:
                ad = item["ad"].strip()
                adet = item["adet"].strip()
                if ad:
                    if adet: kopyalanacak.append(f"{adet} x {ad}")
                    else: kopyalanacak.append(f"{ad}")
                    
            if kopyalanacak:
                export_text = "\n".join(kopyalanacak)
                # Kendi kopyaladığımızı geri sepete atmasın diye oto-yakalayıcıyı kandırıyoruz:
                self.son_pano_verisi = export_text 
                
                self.pencere.clipboard_clear()
                self.pencere.clipboard_append(export_text)
                self.goster_bildirim("Sipariş Kopyalandı", f"{len(kopyalanacak)} kalem eksik panoya alındı.")
            else:
                return 
            
            if messagebox.askyesno("Siparişler Verildi mi?", "Listeyi başarıyla kopyaladınız.\nDepoya aktardıysanız tüm listeyi temizleyelim mi?", parent=self.eksik_penceresi):
                self.eksik_listesi.clear()
                self.eksik_sayfa = 0
                render_page()
                self.eksik_kaydet()

        ModernButton(f_bot, text="📋 SİPARİŞ LİSTESİNİ KOPYALA", command=kopyala_ve_temizle, bg_color=c.get_color("btn_primary"), width=300, height=38).pack(anchor="center", pady=(5,0))

        render_page()

    def eksik_kaydet_delayed(self):
        if hasattr(self, '_eksik_timer'):
            self.pencere.after_cancel(self._eksik_timer)
        self._eksik_timer = self.pencere.after(500, self.eksik_kaydet)

    def eksik_kaydet(self):
        if hasattr(self, 'eksik_listesi'):
            temp = list(self.eksik_listesi)
            
            # En alttaki tamamen boş bırakılan satırları temizle
            while temp and not temp[-1].get("ad", "").strip() and not temp[-1].get("adet", "").strip():
                temp.pop()
                
            # Veritabanı için İlaç ve Adet arasına "|" koy
            metin = "\n".join([f'{item.get("ad", "")}|{item.get("adet", "")}' for item in temp])
            try:
                self.imlec.execute("INSERT OR REPLACE INTO ayarlar (anahtar, deger) VALUES ('eksik_sepeti', ?)", (metin,))
                self.baglanti_skt.commit()
            except Exception as e:
                print("Sepet kaydetme hatası:", e)  

    # =========================================================================
    # VERESİYE VE EMANET SEPETİ MOTORU (KAYMA SORUNU ÇÖZÜLDÜ, X BUTONU SABİTLENDİ, KOPYALA EKLENDİ)
    # =========================================================================
    def toggle_ekstra_sepet(self, tur):
        c = TM  
        win_attr = f"sepet_win_{tur.lower()}"
        db_key = f"sepet_{tur.lower()}"
        
        if hasattr(self, win_attr) and getattr(self, win_attr).winfo_exists():
            self.ekstra_sepet_kaydet(tur)
            getattr(self, win_attr).destroy()
            return

        win = tk.Toplevel(self.pencere)
        setattr(self, win_attr, win)
        win.overrideredirect(True) 
        win.attributes("-topmost", True) 
        
        if tur == "EMANET":
            offset = 480
            win_width = 440
        else:
            offset = 400
            win_width = 360
            
        x = self.pencere.winfo_rootx() + self.pencere.winfo_width() - offset
        y = self.pencere.winfo_rooty() + 60
        if x < 0: x = 50
        win.geometry(f"{win_width}x640+{x}+{y}")
        
        win.configure(bg=c.get_color("card_bg"), highlightthickness=2, highlightbackground=c.get_color("border"))

        def start_move(event):
            win._drag_start_x = event.x
            win._drag_start_y = event.y
            return "break"
            
        def do_move(event):
            nx = win.winfo_x() - win._drag_start_x + event.x
            ny = win.winfo_y() - win._drag_start_y + event.y
            win.geometry(f"+{nx}+{ny}")
            return "break"

        header_bg = "#ef4444" if tur == "VERESİYE" else "#f59e0b" 
        header = tk.Frame(win, bg=header_bg, cursor="fleur", pady=8)
        header.pack(fill="x")
        header.bind("<Button-1>", start_move)
        header.bind("<B1-Motion>", do_move)

        lbl_title = tk.Label(header, text=f"{tur} SEPETİ", font=("Segoe UI", 11, "bold"), bg=header_bg, fg="white", cursor="fleur")
        lbl_title.pack(side="left", padx=10)
        lbl_title.bind("<Button-1>", start_move)
        lbl_title.bind("<B1-Motion>", do_move)

        def kapat():
            self.ekstra_sepet_kaydet(tur)
            win.destroy()

        btn_close = tk.Label(header, text="✖", font=("Segoe UI", 12, "bold"), bg=header_bg, fg="white", cursor="hand2")
        btn_close.pack(side="right", padx=10)
        btn_close.bind("<Button-1>", lambda e: kapat())

        if not hasattr(self, 'ekstra_sepet_datalari'):
            self.ekstra_sepet_datalari = {}
        
        self.ekstra_sepet_datalari[tur] = []
        try:
            self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar=?", (db_key,))
            res = self.imlec.fetchone()
            if res and res[0]:
                for satir in res[0].split('\n'):
                    if not satir.strip(): continue
                    parts = satir.split("|")
                    
                    if tur == "EMANET":
                        kisi = parts[0].strip() if len(parts) > 0 else ""
                        ilac = parts[1].strip() if len(parts) > 1 else ""
                        tarih = parts[2].strip() if len(parts) > 2 else ""
                        self.ekstra_sepet_datalari[tur].append({"kisi": kisi, "ilac": ilac, "tarih": tarih})
                    else:
                        ad = parts[0].strip() if len(parts) > 0 else ""
                        adet = parts[1].strip() if len(parts) > 1 else ""
                        self.ekstra_sepet_datalari[tur].append({"ad": ad, "adet": adet})
        except: pass

        setattr(self, f"sepet_sayfa_{tur.lower()}", 0)

        f_titles = tk.Frame(win, bg=c.get_color("bg_main"), pady=5)
        f_titles.pack(fill="x")
        
        tk.Label(f_titles, text="No", font=("Segoe UI", 9, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left", padx=(10, 5))
        
        if tur == "EMANET":
            tk.Label(f_titles, text="Sil", font=("Segoe UI", 9, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="right", padx=(5, 10))
            tk.Label(f_titles, text="Tarih", font=("Segoe UI", 9, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="right", padx=(5, 30))
            tk.Label(f_titles, text="İlaç Adı", font=("Segoe UI", 9, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text"), width=12).pack(side="right", padx=5)
            tk.Label(f_titles, text="Kişi", font=("Segoe UI", 9, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left", padx=5)
        else:
            tk.Label(f_titles, text="Sil", font=("Segoe UI", 9, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="right", padx=(5, 10))
            tk.Label(f_titles, text="Tutar", font=("Segoe UI", 9, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="right", padx=(5, 15))
            tk.Label(f_titles, text="Müşteri / Açıklama", font=("Segoe UI", 9, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left", padx=5)

        f_bot = tk.Frame(win, bg=c.get_color("bg_main"), pady=10, padx=10)
        f_bot.pack(side="bottom", fill="x")

        f_page = tk.Frame(f_bot, bg=c.get_color("bg_main"))
        f_page.pack(side="top", fill="x", pady=(0, 10))
        btn_prev = tk.Button(f_page, text="◀ ÖNCEKİ", bg=c.get_color("card_bg"), fg=c.get_color("fg_text"), relief="solid", bd=1, cursor="hand2", padx=10, pady=2)
        btn_prev.pack(side="left")
        lbl_page = tk.Label(f_page, text="1 / 1", bg=c.get_color("bg_main"), fg=c.get_color("fg_text"), font=("Segoe UI", 10, "bold"))
        lbl_page.pack(side="left", expand=True)
        btn_next = tk.Button(f_page, text="SONRAKİ ▶", bg=c.get_color("card_bg"), fg=c.get_color("fg_text"), relief="solid", bd=1, cursor="hand2", padx=10, pady=2)
        btn_next.pack(side="right")

        content = tk.Frame(win, bg=c.get_color("card_bg"), padx=5, pady=5)
        content.pack(side="top", fill="both", expand=True)

        satirlar = []

        def update_pagination():
            liste = self.ekstra_sepet_datalari[tur]
            sayfa = getattr(self, f"sepet_sayfa_{tur.lower()}")
            last_filled = -1
            for idx, item in enumerate(liste):
                if tur == "EMANET":
                    if item.get("kisi", "").strip() or item.get("ilac", "").strip() or item.get("tarih", "").strip(): last_filled = idx
                else:
                    if item.get("ad", "").strip() or item.get("adet", "").strip(): last_filled = idx
                    
            max_sayfa = (last_filled // 10) + 2 if last_filled >= 0 else 1
            lbl_page.config(text=f"{sayfa + 1} / {max_sayfa}")
            
            btn_prev.config(state="normal" if sayfa > 0 else "disabled", bg=c.get_color("card_bg"), fg=c.get_color("fg_text"))
            btn_next.config(state="normal" if sayfa < max_sayfa - 1 else "disabled", bg="#3b82f6" if sayfa < max_sayfa - 1 else c.get_color("bg_main"), fg="white" if sayfa < max_sayfa - 1 else "#94a3b8")

        def sil_satir(row_idx):
            if satirlar[row_idx]['btn_x'].cget("text") == "": return # Zaten boş satırsa işlem yapma
            liste = self.ekstra_sepet_datalari[tur]
            sayfa = getattr(self, f"sepet_sayfa_{tur.lower()}")
            list_idx = (sayfa * 10) + row_idx
            if list_idx < len(liste):
                liste.pop(list_idx)
                self.ekstra_sepet_kaydet(tur)
                render_page()

        def on_yazi_degisti(event, row_idx):
            liste = self.ekstra_sepet_datalari[tur]
            sayfa = getattr(self, f"sepet_sayfa_{tur.lower()}")
            list_idx = (sayfa * 10) + row_idx
            
            if tur == "EMANET":
                while len(liste) <= list_idx: liste.append({"kisi": "", "ilac": "", "tarih": ""})
                yk = satirlar[row_idx]['ent_kisi'].get()
                yi = satirlar[row_idx]['ent_ilac'].get()
                yt = satirlar[row_idx]['ent_tarih'].get()
                liste[list_idx] = {"kisi": yk, "ilac": yi, "tarih": yt}
                if yk.strip() or yi.strip() or yt.strip(): satirlar[row_idx]['btn_x'].config(text="✖")
                else: satirlar[row_idx]['btn_x'].config(text="")
            else:
                while len(liste) <= list_idx: liste.append({"ad": "", "adet": ""})
                ya = satirlar[row_idx]['ent_ad'].get()
                yd = satirlar[row_idx]['ent_adet'].get()
                liste[list_idx] = {"ad": ya, "adet": yd}
                if ya.strip() or yd.strip(): satirlar[row_idx]['btn_x'].config(text="✖")
                else: satirlar[row_idx]['btn_x'].config(text="")
                
            update_pagination()
            self._ekstra_sepet_kaydet_delayed(tur)

        for i in range(10):
            row_f = tk.Frame(content, bg=c.get_color("card_bg"), pady=2)
            row_f.pack(fill="x", pady=2)
            
            lbl_no = tk.Label(row_f, text=f"{i+1}.", font=("Segoe UI", 9, "bold"), bg=c.get_color("card_bg"), fg="#94a3b8", width=3, anchor="e")
            lbl_no.pack(side="left", padx=(0, 5))
            
            # X Butonu artık silinmiyor, sadece içindeki yazı gizleniyor. Böylece tasarım kaymıyor.
            btn_x = tk.Label(row_f, text="", fg="#ef4444", bg=c.get_color("card_bg"), cursor="hand2", font=("Segoe UI", 10), width=2)
            btn_x.pack(side="right", padx=(2, 0))
            btn_x.bind("<Button-1>", lambda e, idx=i: sil_satir(idx))
            
            if tur == "EMANET":
                ent_tarih = tk.Entry(row_f, font=("Segoe UI", 9), bg=c.get_color("input_bg"), fg=c.get_color("btn_primary"), bd=1, relief="solid", width=10, justify="center", insertbackground=c.get_color("fg_text"))
                ent_tarih.pack(side="right", padx=(5, 5), ipady=4)
                ent_tarih.bind("<KeyRelease>", lambda e, idx=i: on_yazi_degisti(e, idx))
                
                # YENİ: Kutuya ÇİFT TIKLANDIĞINDA takvim penceresi açılır!
                ent_tarih.bind("<Double-Button-1>", lambda e, ent=ent_tarih: TakvimPopup(ent.winfo_toplevel(), ent))

                ent_ilac = tk.Entry(row_f, font=("Segoe UI", 9), bg=c.get_color("input_bg"), fg=c.get_color("fg_text"), bd=1, relief="solid", width=15, insertbackground=c.get_color("fg_text"))
                ent_ilac.pack(side="right", padx=(5, 5), ipady=4)
                ent_ilac.bind("<KeyRelease>", lambda e, idx=i: on_yazi_degisti(e, idx))

                ent_kisi = tk.Entry(row_f, font=("Segoe UI", 9), bg=c.get_color("input_bg"), fg=c.get_color("fg_text"), bd=1, relief="solid", insertbackground=c.get_color("fg_text"))
                ent_kisi.pack(side="left", fill="x", expand=True, ipady=4)
                ent_kisi.bind("<KeyRelease>", lambda e, idx=i: on_yazi_degisti(e, idx))
                
                satirlar.append({'lbl_no': lbl_no, 'ent_kisi': ent_kisi, 'ent_ilac': ent_ilac, 'ent_tarih': ent_tarih, 'btn_x': btn_x})
            else:
                ent_adet = tk.Entry(row_f, font=("Segoe UI", 10, "bold"), bg=c.get_color("input_bg"), fg=c.get_color("btn_primary"), bd=1, relief="solid", width=8, justify="center", insertbackground=c.get_color("fg_text"))
                ent_adet.pack(side="right", padx=(5, 5), ipady=4)
                ent_adet.bind("<KeyRelease>", lambda e, idx=i: on_yazi_degisti(e, idx))

                ent_ad = tk.Entry(row_f, font=("Segoe UI", 10), bg=c.get_color("input_bg"), fg=c.get_color("fg_text"), bd=1, relief="solid", insertbackground=c.get_color("fg_text"))
                ent_ad.pack(side="left", fill="x", expand=True, ipady=4)
                ent_ad.bind("<KeyRelease>", lambda e, idx=i: on_yazi_degisti(e, idx))
                
                satirlar.append({'lbl_no': lbl_no, 'ent_ad': ent_ad, 'ent_adet': ent_adet, 'btn_x': btn_x})

        def render_page():
            liste = self.ekstra_sepet_datalari[tur]
            sayfa = getattr(self, f"sepet_sayfa_{tur.lower()}")
            
            if tur == "EMANET":
                while liste and not liste[-1].get("kisi", "").strip() and not liste[-1].get("ilac", "").strip() and not liste[-1].get("tarih", "").strip(): liste.pop()
            else:
                while liste and not liste[-1].get("ad", "").strip() and not liste[-1].get("adet", "").strip(): liste.pop()
                
            update_pagination()
            start_idx = sayfa * 10

            for i in range(10):
                list_idx = start_idx + i
                satir = satirlar[i]
                
                satir['lbl_no'].config(text=f"{list_idx + 1}.")
                
                if tur == "EMANET":
                    satir['ent_kisi'].delete(0, tk.END)
                    satir['ent_ilac'].delete(0, tk.END)
                    satir['ent_tarih'].delete(0, tk.END)
                    if list_idx < len(liste):
                        veri = liste[list_idx]
                        satir['ent_kisi'].insert(0, veri.get("kisi", ""))
                        satir['ent_ilac'].insert(0, veri.get("ilac", ""))
                        satir['ent_tarih'].insert(0, veri.get("tarih", ""))
                        if veri.get("kisi", "").strip() or veri.get("ilac", "").strip() or veri.get("tarih", "").strip(): satir['btn_x'].config(text="✖")
                        else: satir['btn_x'].config(text="")
                    else:
                        satir['btn_x'].config(text="")
                else:
                    satir['ent_ad'].delete(0, tk.END)
                    satir['ent_adet'].delete(0, tk.END)
                    if list_idx < len(liste):
                        veri = liste[list_idx]
                        satir['ent_ad'].insert(0, veri.get("ad", ""))
                        satir['ent_adet'].insert(0, veri.get("adet", ""))
                        if veri.get("ad", "").strip() or veri.get("adet", "").strip(): satir['btn_x'].config(text="✖")
                        else: satir['btn_x'].config(text="")
                    else:
                        satir['btn_x'].config(text="")

        def ileri():
            setattr(self, f"sepet_sayfa_{tur.lower()}", getattr(self, f"sepet_sayfa_{tur.lower()}") + 1)
            render_page()
            
        def geri():
            sayfa = getattr(self, f"sepet_sayfa_{tur.lower()}")
            if sayfa > 0:
                setattr(self, f"sepet_sayfa_{tur.lower()}", sayfa - 1)
                render_page()

        btn_prev.config(command=geri)
        btn_next.config(command=ileri)

        # KOPYALA FONKSİYONU (Temizle yerine geldi)
        def listeyi_kopyala():
            liste = self.ekstra_sepet_datalari[tur]
            kopyalanacaklar = []
            
            for item in liste:
                if tur == "EMANET":
                    kisi = item.get("kisi", "").strip()
                    ilac = item.get("ilac", "").strip()
                    tarih = item.get("tarih", "").strip()
                    if kisi or ilac or tarih:
                        kopyalanacaklar.append(f"{kisi} - {ilac} - {tarih}")
                else:
                    ad = item.get("ad", "").strip()
                    adet = item.get("adet", "").strip()
                    if ad or adet:
                        kopyalanacaklar.append(f"{ad} - {adet}")
            
            if kopyalanacaklar:
                metin = "\n".join(kopyalanacaklar)
                self.pencere.clipboard_clear()
                self.pencere.clipboard_append(metin)
                messagebox.showinfo("Başarılı", f"Tüm {tur} listesi başarıyla panoya kopyalandı.", parent=win)
            else:
                messagebox.showwarning("Uyarı", "Kopyalanacak herhangi bir kayıt bulunamadı.", parent=win)

        # Yeni Kopyala Butonu
        ModernButton(f_bot, text="📋 TÜM LİSTEYİ KOPYALA", command=listeyi_kopyala, bg_color="#3b82f6", width=300, height=38).pack(anchor="center", pady=(5,0))
        
        render_page()

    def _ekstra_sepet_kaydet_delayed(self, tur):
        timer_attr = f"_sepet_timer_{tur.lower()}"
        if hasattr(self, timer_attr):
            self.pencere.after_cancel(getattr(self, timer_attr))
        setattr(self, timer_attr, self.pencere.after(500, lambda: self.ekstra_sepet_kaydet(tur)))

    def ekstra_sepet_kaydet(self, tur):
        db_key = f"sepet_{tur.lower()}"
        if hasattr(self, 'ekstra_sepet_datalari') and tur in self.ekstra_sepet_datalari:
            temp = list(self.ekstra_sepet_datalari[tur])
            if tur == "EMANET":
                while temp and not temp[-1].get("kisi", "").strip() and not temp[-1].get("ilac", "").strip() and not temp[-1].get("tarih", "").strip(): temp.pop()
                metin = "\n".join([f'{item.get("kisi", "")}|{item.get("ilac", "")}|{item.get("tarih", "")}' for item in temp])
            else:
                while temp and not temp[-1].get("ad", "").strip() and not temp[-1].get("adet", "").strip(): temp.pop()
                metin = "\n".join([f'{item.get("ad", "")}|{item.get("adet", "")}' for item in temp])
                
            try:
                self.imlec.execute("INSERT OR REPLACE INTO ayarlar (anahtar, deger) VALUES (?, ?)", (db_key, metin))
                self.baglanti_skt.commit()
            except Exception as e:
                print(f"{tur} sepet kaydetme hatası:", e)              

    import warnings
    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

    def aylik_dokum_yukle(self, depo_adi):
        if not PANDAS_VAR:
            messagebox.showerror("Hata", "pandas kütüphanesi eksik.")
            return

        file_path = filedialog.askopenfilename(
            title=f"{depo_adi} - Aylık Fatura Dökümü Seç (Excel/CSV)",
            filetypes=[("Excel/CSV Dosyaları", "*.xlsx *.xls *.csv")]
        )
        if not file_path: return

        try:
            df = self.evrensel_dosya_oku(file_path)
            if df.empty: return

            def temizle_float(val):
                if pd.isna(val): return 0.0
                val = str(val).strip().upper().replace('TL', '').replace('₺', '').replace(' ', '')
                if not val or val == 'NAN': return 0.0
                if '.' in val and ',' in val:
                    if val.rfind(',') > val.rfind('.'):
                        val = val.replace('.', '').replace(',', '.')
                    else:
                        val = val.replace(',', '')
                elif ',' in val:
                    val = val.replace(',', '.')
                try: return float(val)
                except: return 0.0

            def sutun_bul(arananlar):
                for col in df.columns:
                    norm_col = str(col).upper()
                    mapping = {'İ': 'I', 'Ğ': 'G', 'Ü': 'U', 'Ş': 'S', 'Ö': 'O', 'Ç': 'C', ' ': ''}
                    for k, v in mapping.items():
                        norm_col = norm_col.replace(k, v)
                    norm_col = norm_col.strip()
                    for a in arananlar:
                        if a in norm_col:
                            return col
                return None

            c_haric  = sutun_bul(['VERGILERHARICTUTAR', 'VERGILERHARIC', 'HARICTUTAR', 'MATRAH'])
            c_dahil  = sutun_bul(['VERGILERDAHILTUTAR', 'VERGILERDAHIL', 'DAHILTUTAR'])
            c_toplam = sutun_bul(['FATURATOPLAMI', 'FATURATOPLAM', 'TOPLAM', 'TUTAR'])
            c_tarih  = sutun_bul(['FATURATARIHI', 'TARIH', 'ISLEMTARIHI'])

            # Fatura dönemini tespit et
            fatura_donemi = date.today().strftime("%Y-%m")
            if c_tarih and not df.empty:
                try:
                    tarihler = pd.to_datetime(df[c_tarih], dayfirst=True, errors='coerce').dropna()
                    if not tarihler.empty:
                        mod_val = tarihler.dt.strftime('%Y-%m').mode()
                        if len(mod_val) > 0:
                            fatura_donemi = str(mod_val.iloc[0])
                except: pass

            if not isinstance(fatura_donemi, str) or '-' not in str(fatura_donemi):
                fatura_donemi = date.today().strftime("%Y-%m")

            # KDV dilimi sözlüğü
            kdv_sinif_toplam = {
                1:  {'matrah': 0.0, 'kdv': 0.0, 'sayi': 0},
                10: {'matrah': 0.0, 'kdv': 0.0, 'sayi': 0},
                20: {'matrah': 0.0, 'kdv': 0.0, 'sayi': 0}
            }
            genel_toplam = 0.0

            if c_haric and c_dahil:
                for _, row in df.iterrows():
                    haric = temizle_float(row[c_haric])
                    dahil = temizle_float(row[c_dahil])
                    kdv_t = dahil - haric

                    if haric > 0:
                        kdv_oran = (kdv_t / haric) * 100
                    else:
                        kdv_oran = 0.0

                    if kdv_oran == 0:
                        sinif = 1
                    elif kdv_oran < 2.0:
                        sinif = 1
                    elif kdv_oran < 15.0:
                        sinif = 10
                    else:
                        sinif = 20

                    kdv_sinif_toplam[sinif]['matrah'] += haric
                    kdv_sinif_toplam[sinif]['kdv']    += kdv_t
                    kdv_sinif_toplam[sinif]['sayi']   += 1
                    genel_toplam += dahil

            elif c_toplam:
                genel_toplam = df[c_toplam].apply(temizle_float).sum()

            if genel_toplam == 0 and all(v['matrah'] == 0 for v in kdv_sinif_toplam.values()):
                messagebox.showwarning("Uyarı", "Dosyada okunabilir KDV sütunları bulunamadı.\nLütfen formatı kontrol edin.")
                return

            self._kdv_ozet_popup(depo_adi, kdv_sinif_toplam, genel_toplam, fatura_donemi)

        except Exception as e:
            messagebox.showerror("Hata", f"Dosya işlenirken hata oluştu: {e}")

    def _kdv_ozet_popup(self, depo_adi, kdv_sinif_toplam, genel_toplam, fatura_donemi):
        top = tk.Toplevel(self.pencere)
        top.title(f"{depo_adi} - Fatura KDV Özeti")
        top.configure(bg="white")
        top.transient(self.pencere)
        top.grab_set()

        # Ekran boyutuna göre pencere boyutu ayarla
        ekran_y = top.winfo_screenheight()
        pencere_h = min(750, ekran_y - 100)
        pencere_w = 800
        x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - (pencere_w // 2)
        y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - (pencere_h // 2)
        top.geometry(f"{pencere_w}x{pencere_h}+{x}+{y}")
        top.resizable(True, True)

        # ── SCROLL ALTYAPISI ──────────────────────────────────────────────────
        canvas = tk.Canvas(top, bg="white", highlightthickness=0)
        scrollbar = ttk.Scrollbar(top, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Scroll edilecek ana çerçeve
        ana = tk.Frame(canvas, bg="white")
        ana_id = canvas.create_window((0, 0), window=ana, anchor="nw")

        def on_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_canvas_configure(e):
            canvas.itemconfig(ana_id, width=e.width)

        ana.bind("<Configure>", on_configure)
        canvas.bind("<Configure>", on_canvas_configure)

        # Mouse scroll
        def on_mousewheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        top.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # ── İÇERİK ────────────────────────────────────────────────────────────
        renk_map = {1: "#10b981", 10: "#3b82f6", 20: "#f59e0b"}
        isim_map = {1: "%1 KDV", 10: "%10 KDV", 20: "%20 KDV"}

        tk.Label(ana, text="📊 Aylık Fatura KDV Dökümü",
                font=("Segoe UI", 14, "bold"), bg="white", fg="#334155").pack(pady=20)

        # ── ÖZET KART ─────────────────────────────────────────────────────────
        f_ozet = tk.Frame(ana, bg="#f8fafc", bd=1, relief="solid", padx=20, pady=15)
        f_ozet.pack(fill="x", padx=30, pady=(0, 10))

        def satir_ekle(parent, sol, sag, renk="#0f172a", bold=False):
            f = tk.Frame(parent, bg="#f8fafc")
            f.pack(fill="x", pady=2)
            font_w = "bold" if bold else "normal"
            tk.Label(f, text=sol, font=("Segoe UI", 10, font_w),
                    bg="#f8fafc", fg="#475569").pack(side="left")
            tk.Label(f, text=sag, font=("Segoe UI", 10, "bold"),
                    bg="#f8fafc", fg=renk).pack(side="right")

        genel_kdv    = 0.0
        genel_matrah = 0.0
        for oran in [1, 10, 20]:
            v = kdv_sinif_toplam[oran]
            if v['matrah'] > 0 or v['kdv'] > 0:
                satir_ekle(f_ozet,
                        f"{isim_map[oran]} Matrahı ({v['sayi']} fatura):",
                        f"{v['matrah']:,.2f} ₺", renk_map[oran])
                satir_ekle(f_ozet,
                        f"{isim_map[oran]} Tutarı:",
                        f"{v['kdv']:,.2f} ₺", renk_map[oran])
                satir_ekle(f_ozet,
                        f"{isim_map[oran]} Dahil Toplam:",
                        f"{v['matrah'] + v['kdv']:,.2f} ₺",
                        renk_map[oran], bold=True)
                tk.Frame(f_ozet, height=1, bg="#e2e8f0").pack(fill="x", pady=5)
                genel_kdv    += v['kdv']
                genel_matrah += v['matrah']

        satir_ekle(f_ozet, "GENEL MATRAH:",
                f"{genel_matrah:,.2f} ₺", "#0f172a", bold=True)
        satir_ekle(f_ozet, "TOPLAM KDV:",
                f"{genel_kdv:,.2f} ₺", "#ef4444", bold=True)
        satir_ekle(f_ozet, "GENEL TOPLAM (KDV Dahil):",
                f"{genel_toplam:,.2f} ₺", "#ef4444", bold=True)

        # ── AKORDİYON ─────────────────────────────────────────────────────────
        tk.Label(ana, text="KDV Dilimi Detayları (başlığa tıklayın):",
                font=("Segoe UI", 10, "bold"), bg="white", fg="#64748b").pack(
                anchor="w", padx=30, pady=(5, 3))

        f_akordion = tk.Frame(ana, bg="white")
        f_akordion.pack(fill="x", padx=30, pady=(0, 5))

        def akordion_olustur(oran):
            v = kdv_sinif_toplam[oran]
            if v['matrah'] == 0 and v['kdv'] == 0:
                return

            renk = renk_map[oran]
            acik = tk.BooleanVar(value=False)

            f_baslik = tk.Frame(f_akordion, bg=renk, cursor="hand2")
            f_baslik.pack(fill="x", pady=2)

            lbl_ok = tk.Label(f_baslik, text="▶",
                            font=("Segoe UI", 10, "bold"),
                            bg=renk, fg="white", cursor="hand2", width=3)
            lbl_ok.pack(side="left", padx=(8, 4), pady=8)

            tk.Label(f_baslik,
                    text=f"{isim_map[oran]}  —  "
                        f"Matrah: {v['matrah']:,.2f} ₺  |  "
                        f"KDV: {v['kdv']:,.2f} ₺  |  "
                        f"Toplam: {v['matrah'] + v['kdv']:,.2f} ₺",
                    font=("Segoe UI", 10, "bold"),
                    bg=renk, fg="white", cursor="hand2").pack(side="left", pady=8)

            ic_bg = "#f0fdf4" if oran == 1 else "#eff6ff" if oran == 10 else "#fffbeb"
            f_ic = tk.Frame(f_akordion, bg=ic_bg, padx=18, pady=10,
                            highlightbackground=renk, highlightthickness=1)

            def ic_satir(sol, sag, renk_val="#0f172a"):
                ff = tk.Frame(f_ic, bg=ic_bg)
                ff.pack(fill="x", pady=3)
                tk.Label(ff, text=sol, font=("Segoe UI", 10),
                        bg=ic_bg, fg="#475569").pack(side="left")
                tk.Label(ff, text=sag, font=("Segoe UI", 10, "bold"),
                        bg=ic_bg, fg=renk_val).pack(side="right")

            ic_satir("Fatura Adedi:", f"{v['sayi']} adet")
            ic_satir("KDV Matrahı (Vergiler Hariç):",
                    f"{v['matrah']:,.2f} ₺", renk)
            ic_satir(f"KDV Tutarı ({isim_map[oran]}):",
                    f"{v['kdv']:,.2f} ₺", "#ef4444")
            ic_satir("KDV Dahil Tutar:",
                    f"{v['matrah'] + v['kdv']:,.2f} ₺", renk)

            def toggle(ic=f_ic, ok=lbl_ok, var=acik):
                if var.get():
                    ic.pack_forget()
                    ok.config(text="▶")
                    var.set(False)
                else:
                    ic.pack(fill="x", pady=(0, 3))
                    ok.config(text="▼")
                    var.set(True)

            for w in [f_baslik, lbl_ok] + list(f_baslik.winfo_children()):
                w.bind("<Button-1>", lambda e, t=toggle: t())

        for oran in [1, 10, 20]:
            akordion_olustur(oran)

        # ── KAYIT FORMU ───────────────────────────────────────────────────────
        f_form = tk.Frame(ana, bg="white", padx=30)
        f_form.pack(fill="x", pady=(10, 0))

        tk.Label(f_form, text="Alım Dönemi (YIL-AY):",
                font=("Segoe UI", 9, "bold"), bg="white", fg="#10b981").pack(
                anchor="w", pady=(5, 0))
        ent_alim = tk.Entry(f_form, font=("Segoe UI", 11, "bold"),
                            relief="solid", bd=1, bg="#f1f5f9", fg="#10b981")
        ent_alim.pack(fill="x", pady=4, ipady=4)
        ent_alim.insert(0, str(fatura_donemi))

        tk.Label(f_form, text="Ödeme (Vade) Dönemi (YIL-AY):",
                font=("Segoe UI", 9, "bold"), bg="white", fg="#64748b").pack(
                anchor="w", pady=(5, 0))
        ent_vade = tk.Entry(f_form, font=("Segoe UI", 11),
                            relief="solid", bd=1, bg="#f1f5f9")
        ent_vade.pack(fill="x", pady=4, ipady=4)
        try:
            yil, ay = map(int, str(fatura_donemi).split('-'))
            vade_dt = date(yil, ay, 1) + timedelta(days=90)
            ent_vade.insert(0, vade_dt.strftime("%Y-%m"))
        except:
            ent_vade.insert(0, (date.today() + timedelta(days=90)).strftime("%Y-%m"))

        def kaydet():
            alim = ent_alim.get().strip()
            if len(alim) == 7: alim += "-01"
            vade = ent_vade.get().strip()
            if len(vade) == 7: vade += "-01"

            eklenen = 0
            ozet_satirlari = []
            for oran in [1, 10, 20]:
                v = kdv_sinif_toplam[oran]
                kdv_dahil = v['matrah'] + v['kdv']
                if kdv_dahil > 0:
                    not_metni = (
                        f"[KDV_DETAY]"
                        f"{isim_map[oran]} Matrah:{v['matrah']:.2f}|"
                        f"{isim_map[oran]} KDV:{v['kdv']:.2f}|"
                        f"{isim_map[oran]} Toplam:{kdv_dahil:.2f}"
                    )
                    self.imlec_finans.execute(
                        "INSERT INTO odemeler "
                        "(fatura_adi, vade_tarihi, tutar, aciklama, durum, alim_tarihi, satir_notu) "
                        "VALUES (?, ?, ?, ?, 'ODENMEDİ', ?, ?)",
                        (depo_adi, vade, kdv_dahil,
                        f"[EXCEL DÖKÜM {isim_map[oran]}]", alim, not_metni)
                    )
                    eklenen += 1
                    ozet_satirlari.append(
                        f"• {isim_map[oran]} Toplam: {kdv_dahil:,.2f} ₺"
                        f"  (KDV: {v['kdv']:,.2f} ₺)"
                    )

            self.baglanti_finans.commit()
            messagebox.showinfo(
                "Başarılı",
                f"Dönem: {fatura_donemi}\n"
                f"{eklenen} adet KDV dilimi sisteme işlendi.\n\n"
                + "\n".join(ozet_satirlari),
                parent=top
            )
            top.destroy()
            self.sekmeleri_guncelle(hedef_sekme_adi=depo_adi)

        ModernButton(ana, text="ONAYLA VE SİSTEME KAYDET",
                    command=kaydet, bg_color="#10b981", width=280, height=45).pack(pady=15)

    # --- VERİTABANI KURULUM FONKSİYONLARI ---
    def veritabani_kur_skt(self):
        self.imlec.execute("""CREATE TABLE IF NOT EXISTS ilaclar (id INTEGER PRIMARY KEY AUTOINCREMENT, barkod TEXT, ad TEXT, parti_no TEXT, seri_no TEXT, adet INTEGER, skt TEXT, raf_yeri TEXT, kayit_tarihi TEXT, yukleme_id INTEGER)""")
        self.imlec.execute("""CREATE TABLE IF NOT EXISTS ilac_kartlari (gtin TEXT PRIMARY KEY, ad TEXT, guncelleme_tarihi TEXT)""")
        self.imlec.execute("CREATE TABLE IF NOT EXISTS ayarlar (anahtar TEXT PRIMARY KEY, deger TEXT)")
        self.imlec.execute("CREATE TABLE IF NOT EXISTS kayitli_listeler (id INTEGER PRIMARY KEY AUTOINCREMENT, liste_adi TEXT, icerik TEXT, tarih TEXT)")
        
        defaults = {
            'gonderen_mail': '', 'uygulama_sifresi': '', 'alici_mail': '', 
            'son_stok_yukleme': '2000-01-01', 'son_mail_tarihi': '2000-01-01', 
            'pc_acilis_baslat': '0', 'yedekleme_konumu': 'Yedekler'
        }
        for k, v in defaults.items(): 
            self.imlec.execute("INSERT OR IGNORE INTO ayarlar (anahtar, deger) VALUES (?, ?)", (k, v))
        self.baglanti_skt.commit()

    def veritabani_kur_finans(self):
        try:
            self.imlec_finans.execute("""CREATE TABLE IF NOT EXISTS odemeler (id INTEGER PRIMARY KEY AUTOINCREMENT, fatura_adi TEXT, vade_tarihi TEXT, tutar REAL, aciklama TEXT, durum TEXT DEFAULT 'ODENMEDİ', alim_tarihi TEXT, satir_notu TEXT)""")
            self.imlec_finans.execute("""CREATE TABLE IF NOT EXISTS depo_ayarlari (fatura_adi TEXT PRIMARY KEY, anlasma_sarti TEXT, sira INTEGER)""")
            self.imlec_finans.execute("""CREATE TABLE IF NOT EXISTS ozel_sekmeler (id INTEGER PRIMARY KEY AUTOINCREMENT, baslik TEXT, icerik TEXT)""")
            self.imlec_finans.execute("""CREATE TABLE IF NOT EXISTS kredi_kartlari (isim TEXT PRIMARY KEY, kesim_gunu INTEGER, sira INTEGER)""")
            self.imlec_finans.execute("CREATE TABLE IF NOT EXISTS sabit_giderler (id INTEGER PRIMARY KEY AUTOINCREMENT, kalem_adi TEXT, varsayilan_tutar REAL)")
            self.imlec_finans.execute("CREATE TABLE IF NOT EXISTS program_ayarlari (ayar_adi TEXT PRIMARY KEY, deger TEXT)")
            self.imlec_finans.execute("INSERT OR IGNORE INTO program_ayarlari VALUES ('oto_odeme_depo', '0')")
            self.imlec_finans.execute("INSERT OR IGNORE INTO program_ayarlari VALUES ('oto_odeme_kart', '0')")
            self.imlec_finans.execute("""
                CREATE TABLE IF NOT EXISTS kasa_defteri (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, tarih TEXT, islem_turu TEXT, kategori TEXT, 
                    aciklama TEXT, tutar REAL, odeme_yontemi TEXT
                )
            """)
            self.imlec_finans.execute("""CREATE TABLE IF NOT EXISTS kurumlar (isim TEXT PRIMARY KEY)""")
            self.baglanti_finans.commit()
        except Exception as e:
            log_error(f"Finans DB Kurulum Hatası: {e}")

    def tablo_guncelleme_kontrol(self):
        try: self.imlec.execute("SELECT raf_yeri FROM ilaclar LIMIT 1")
        except: self.imlec.execute("ALTER TABLE ilaclar ADD COLUMN raf_yeri TEXT"); self.baglanti_skt.commit()
        try: self.imlec.execute("SELECT kayit_tarihi FROM ilaclar LIMIT 1")
        except: self.imlec.execute("ALTER TABLE ilaclar ADD COLUMN kayit_tarihi TEXT"); self.baglanti_skt.commit()
        try: self.imlec.execute("SELECT yukleme_id FROM ilaclar LIMIT 1")
        except: self.imlec.execute("ALTER TABLE ilaclar ADD COLUMN yukleme_id INTEGER"); self.baglanti_skt.commit()

    def modern_baslik_cubugu_ayarla(self, is_dark):
        """Windows 10/11 API kullanarak başlık çubuğunu Karanlık veya Aydınlık temaya uyarlar."""
        try:
            import ctypes
            self.pencere.update_idletasks()
            hwnd = ctypes.windll.user32.GetParent(self.pencere.winfo_id())
            value = ctypes.c_int(1 if is_dark else 0)
            # Windows 11 için Karanlık mod API'si
            ctypes.windll.dwmapi.DwmSetWindowAttribute(hwnd, 20, ctypes.byref(value), ctypes.sizeof(value))
            # Windows 10 için Karanlık mod API'si
            ctypes.windll.dwmapi.DwmSetWindowAttribute(hwnd, 19, ctypes.byref(value), ctypes.sizeof(value))
        except:
            pass
    
    def ana_uygulamaya_gec(self, baslangic_sekmesi=0):
        for w in self.ana_container.winfo_children(): w.destroy()
        c = TM

        # --- SAYFA GEÇMİŞİ (GERİ TUŞU İÇİN) ---
        if not hasattr(self, 'sayfa_gecmisi'):
            self.sayfa_gecmisi = []

        # =====================================================================
        # WINDOWS ÇERÇEVESİNİ GERİ GETİR VE MODERNLEŞTİR
        # =====================================================================
        self.pencere.overrideredirect(False) # Orijinal Windows çerçevesini aktif et
        self.pencere.title("Eczacı Defteri") # Windows başlık çubuğu metni
        
        # Orijinal başlık çubuğunun rengini temamıza göre ayarla
        self.modern_baslik_cubugu_ayarla(c.is_dark)

        # =====================================================================
        # UYGULAMA İÇİ NAVİGASYON ÇUBUĞU (Sadece Menüler İçin)
        # =====================================================================
        self.top_bar = tk.Frame(self.ana_container, bg=c.get_color("header_bg"), height=45)
        self.top_bar.pack(side="top", fill="x")
        self.top_bar.pack_propagate(False)

        nav_style = {"font": ("Segoe UI", 10, "bold"), "bg": c.get_color("header_bg"), "fg": "white", "relief": "flat", "bd": 0, "cursor": "hand2", "activebackground": c.get_color("btn_primary"), "activeforeground": "white"}
        
        # --- GERİ BUTONU ---
        self.btn_geri = tk.Button(self.top_bar, text=" ⬅ GERİ ", command=self.onceki_sayfaya_don, **nav_style)

        # --- LOGO YÜKLEME VE İKON AYARLAMA (HIZLANDIRILDI) ---
        # Eğer logo daha önce yüklendiyse tekrar hesaplayıp programı dondurma!
        if not hasattr(self, 'logo_dashboard'):
            try:
                import os
                import sys
                from PIL import Image, ImageTk
                
                # Dosya klasörünü bul
                if getattr(sys, 'frozen', False):
                    base_dir = os.path.dirname(sys.executable)
                else:
                    base_dir = os.path.dirname(os.path.abspath(__file__))
                    
                logo_yolu = os.path.join(base_dir, "logo_hd_transparan.png")
                orijinal_logo = Image.open(logo_yolu)
                
                # SADECE Görev Çubuğu ve Ana Sayfa(Dashboard) İkonlarını Yükle
                self.icon_photo = ImageTk.PhotoImage(orijinal_logo)
                self.pencere.iconphoto(False, self.icon_photo)

                self.logo_dashboard = ImageTk.PhotoImage(orijinal_logo.resize((130, 130), Image.Resampling.LANCZOS))
                
            except Exception as e:
                print(f"Logo yüklenemedi detayı: {e}")

        # "Geri" butonunun hizasını koruyabilmesi için görünmez/boş bir referans etiketi bırakıyoruz.
        self.lbl_title = tk.Label(self.top_bar, text="", bg=c.get_color("header_bg"))
        self.lbl_title.pack(side="left")

        # --- YENİ: SOL KISIM (Ana Sayfa ve Sepetler) ---
        f_left_menus = tk.Frame(self.top_bar, bg=c.get_color("header_bg"))
        f_left_menus.pack(side="left", padx=10, fill="y")

        tk.Button(f_left_menus, text="🏠 ANA SAYFA", command=lambda: self.sekme_degistir(0), **nav_style).pack(side="left", padx=5, fill="y")
        self.btn_eksik = tk.Button(f_left_menus, text="📝 EKSİKLER", command=self.toggle_eksik_sepeti, **nav_style)
        self.btn_eksik.pack(side="left", padx=5, fill="y")
        self.btn_veresiye = tk.Button(f_left_menus, text="📒 VERESİYE", command=lambda: self.toggle_ekstra_sepet("VERESİYE"), **nav_style)
        self.btn_veresiye.pack(side="left", padx=5, fill="y")
        self.btn_emanet = tk.Button(f_left_menus, text="📦 EMANET", command=lambda: self.toggle_ekstra_sepet("EMANET"), **nav_style)
        self.btn_emanet.pack(side="left", padx=5, fill="y")

        # --- YENİ: SAĞ KISIM (Çıkış, Zil, Sohbet, Tema) ---
        f_win_controls = tk.Frame(self.top_bar, bg=c.get_color("header_bg"))
        f_win_controls.pack(side="right")
        
        self.btn_cikis = tk.Button(f_win_controls, text="🚪 ÇIKIŞ YAP", command=self.kullanici_degistir, **nav_style)
        self.btn_cikis.pack(side="right", padx=15, fill="y")

        self.btn_zil = tk.Button(f_win_controls, text="🔔", command=self.bildirim_panelini_ac, **nav_style)
        self.btn_zil.pack(side="right", padx=5, fill="y")

        self.btn_chat = tk.Button(f_win_controls, text="💬", command=self.chat_panelini_ac, **nav_style)
        self.btn_chat.pack(side="right", padx=5, fill="y")

        # TEMA BUTONU SOHBETİN YANINA EKLENDİ
        self.btn_tema = tk.Button(f_win_controls, text="🌙 / ☀️", command=self.tema_degistir_click, **nav_style)
        self.btn_tema.pack(side="right", padx=10, fill="y")

        # Ekran yüklendikten hemen sonra zildeki sayıyı güncelle
        self.pencere.after(100, self.zil_guncelle)

        # --- İÇERİK ALANI ---
        self.content_area = tk.Frame(self.ana_container, bg=c.get_color("bg_main"))
        self.content_area.pack(side="top", fill="both", expand=True)

        # --- GÜVENLİ ROL VE YETKİ YÜKLEME ---
        self.aktif_rol = str(self.aktif_rol).strip() 
        self.personel_yetkileri = {"finans_gorsun": False, "para_gorsun": False, "ayarlar_gorsun": False}
        
        try:
            self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='personel_yetkileri'")
            res_yetki = self.imlec.fetchone()
            if res_yetki and res_yetki[0]:
                self.personel_yetkileri = eval(res_yetki[0])
        except: 
            pass

        self.sekme_degistir(baslangic_sekmesi)

    # =========================================================================
    # KULLANICI DEĞİŞTİR (OTURUMU KAPAT) MODÜLÜ
    # =========================================================================
    def kullanici_degistir(self):
        if messagebox.askyesno("Oturumu Kapat", "Mevcut oturumu kapatıp Kullanıcı Giriş ekranına dönmek istiyor musunuz?"):
            
            # 1. Ekrandaki her şeyi (Eczane Asistanı arayüzünü) temizle
            for widget in self.pencere.winfo_children():
                widget.destroy()
            
            self.pencere.configure(bg="#f1f5f9") # Giriş ekranı arka plan rengi
            
            # 2. Sürükle bırak özelliğini bozmadan uygulamayı yeniden başlatacak tetiği hazırla
            def uygulamayi_yeniden_baslat(aktif_rol):
                app = EczaneAsistani(self.pencere, aktif_rol)
                try:
                    from tkinterdnd2 import DND_FILES
                    self.pencere.drop_target_register(DND_FILES)
                    self.pencere.dnd_bind('<<Drop>>', app.surukle_birak_yoneticisi)
                except: pass
            
            # 3. Giriş Ekranını (Şifre Ekranı) ekrana geri çağır
            GirisEkrani(self.pencere, uygulamayi_yeniden_baslat)    


    def onceki_sayfaya_don(self):
        """Geri tuşuna basıldığında geçmişteki son sayfaya döner."""
        if hasattr(self, 'sayfa_gecmisi') and len(self.sayfa_gecmisi) > 0:
            onceki_index = self.sayfa_gecmisi.pop()
            self.sekme_degistir(onceki_index, gecmis_kaydet=False)
        else:
            self.sekme_degistir(0, gecmis_kaydet=False)


    def sekme_degistir(self, index, gecmis_kaydet=True):
        """İlgili modülü ekrana basar ve geçmişi günceller."""
        
        # --- GÜVENLİK DUVARI: FİNANS SAYFALARINI KORU ---
        finans_sayfalari = [6, 7, 8, 9, 10, 11, 15, 16]
        if index in finans_sayfalari:
            if self.aktif_rol != "Yönetici" and not self.personel_yetkileri.get("finans_gorsun", False):
                messagebox.showerror("Yetki Hatası", "Bu sayfayı görüntülemek için yetkiniz bulunmuyor.")
                return # Sayfayı değiştirmeden işlemi iptal et
        
        # Eğer Ana Sayfaya dönüyorsak geçmişi sıfırlayalım (Daha sağlıklı bir gezinme için)
        if index == 0:
            self.sayfa_gecmisi = []

        # Eğer yeni bir sekmeye gidiliyorsa ve aynı sekmeye peş peşe basılmadıysa geçmişe kaydet
        elif gecmis_kaydet and hasattr(self, 'aktif_sekme'):
            if self.aktif_sekme != index: 
                if not hasattr(self, 'sayfa_gecmisi'): self.sayfa_gecmisi = []
                self.sayfa_gecmisi.append(self.aktif_sekme)

        self.aktif_sekme = index
        for w in self.content_area.winfo_children(): w.destroy()
        
        # Geri butonunu Göster/Gizle mantığı
        if hasattr(self, 'btn_geri') and hasattr(self, 'lbl_title'):
            # Geçmişte sayfa varsa ve şu an Ana Sayfada (0) değilsek Geri tuşunu göster
            if hasattr(self, 'sayfa_gecmisi') and len(self.sayfa_gecmisi) > 0 and index != 0:
                self.btn_geri.pack(side="left", padx=(10, 0), fill="y", before=self.lbl_title)
            else:
                # Aksi halde gizle
                self.btn_geri.pack_forget()

        # İlgili sayfayı yükle
        try:
            if index == 0: self.arayuz_anasayfa()
            elif index == 1: self.arayuz_stok_takip()
            elif index == 2: self.arayuz_ilac_kartlari()
            elif index == 3: self.arayuz_toplu_karekod()
            elif index == 4: self.arayuz_karekod_olusturucu()
            elif index == 5: self.arayuz_sayim_modu()
            elif index == 6: self.sayfa_depo_odemeleri()
            elif index == 7: self.sayfa_kredi_karti()
            elif index == 8: self.sayfa_kurum_odemeleri()
            elif index == 9: self.sayfa_gelir()
            elif index == 10: self.sayfa_gider()
            elif index == 11: self.sayfa_kasa_defteri()
            elif index == 12: self.arayuz_ayarlar()
            elif index == 13: self.arayuz_geribildirim()
            elif index == 14: self.arayuz_kullanim_kilavuzu()
            elif index == 15: self.arayuz_kredi_hesaplama()
            elif index == 16: self.arayuz_mevcut_krediler()
            elif index == 17: self.arayuz_grafikler()
        except Exception as e:
            print(f"Sayfa yüklenirken hata oluştu: {e}")

    def tema_degistir_click(self):
        TM.toggle()
        self.modern_baslik_cubugu_ayarla(TM.is_dark) # Windows çerçevesinin rengini günceller
        self.tema_uygula()
        self.ana_uygulamaya_gec(self.aktif_sekme)

    # =============================================================================
    # FİNANS MODÜLÜ FONKSİYONLARI (Adım 4.2)
    # =============================================================================
    def gecmise_kaydet(self, islem_tipi, tablo, veri_id, eski_veri_tuple=None):
        pass

    def otomatik_odeme_motoru(self):
        bugun = date.today().strftime("%Y-%m-%d")
        self.imlec_finans.execute("SELECT deger FROM program_ayarlari WHERE ayar_adi='oto_odeme_depo'")
        res_depo = self.imlec_finans.fetchone()
        depo_aktif = True if res_depo and res_depo[0] == '1' else False
        
        if depo_aktif:
            sorgu = """
                UPDATE odemeler SET durum='ODENDİ' 
                WHERE vade_tarihi < ? 
                AND durum='ODENMEDİ'
                AND (satir_notu IS NULL OR (NOT satir_notu LIKE 'KART:%' AND NOT satir_notu LIKE '%KURUM%'))
            """
            self.imlec_finans.execute(sorgu, (bugun,))
            
        self.imlec_finans.execute("SELECT deger FROM program_ayarlari WHERE ayar_adi='oto_odeme_kart'")
        res_kart = self.imlec_finans.fetchone()
        kart_aktif = True if res_kart and res_kart[0] == '1' else False
        
        if kart_aktif:
            sorgu = """
                UPDATE odemeler SET durum='ODENDİ' 
                WHERE vade_tarihi < ? 
                AND durum='ODENMEDİ'
                AND satir_notu LIKE 'KART:%'
            """
            self.imlec_finans.execute(sorgu, (bugun,))
            
        self.baglanti_finans.commit()

    def mouse_scroll_ekle(self, canvas, scrollable_element):
        """Kutuların üzerine gelince scroll'un durması sorununu çözen rekürsif bağlayıcı."""
        def _on_mousewheel(event):
            try: canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except: pass

        def bind_to_all(widget):
            widget.bind('<MouseWheel>', _on_mousewheel)
            for child in widget.winfo_children():
                # Kendi scroll'u olan tablo, liste ve metin kutularını atla (Kendi içlerinde kaysınlar)
                if not isinstance(child, (ttk.Treeview, ttk.Combobox, tk.Listbox, tk.Text)):
                    bind_to_all(child)

        # İlk oluşturulan tüm öğelere scroll özelliğini yapıştır
        bind_to_all(scrollable_element)
        canvas.bind('<MouseWheel>', _on_mousewheel)

        # Yeni bir kutu/kart eklendiğinde onlara da otomatik scroll özelliği ver
        scrollable_element.bind('<Configure>', lambda e: bind_to_all(scrollable_element), add="+")

    def surukle_birak_yoneticisi(self, event):
        """Dışarıdan sürüklenen Excel dosyasını yakalar ve açık olan sekmeye göre işlem yapar."""
        dosya_yolu = event.data.strip('{}') # Gelen yoldaki olası süslü parantezleri temizler

        if not (dosya_yolu.lower().endswith('.xlsx') or dosya_yolu.lower().endswith('.csv') or dosya_yolu.lower().endswith('.xls')):
            messagebox.showwarning("Geçersiz Dosya", "Lütfen sadece Excel veya CSV formatında bir dosya sürükleyin.")
            return

        if self.aktif_sekme == 1:
            # SKT Takibi sekmesindeyiz, stoğa yükle
            self.excel_yukle_stok(otomatik_dosya_yolu=dosya_yolu)
        elif self.aktif_sekme == 4:
            # Karekod Üret sekmesindeyiz, listeye yükle
            self.karekod_dosya_yukle(otomatik_dosya_yolu=dosya_yolu)
        else:
            messagebox.showinfo("Bilgi", "Dosya sürükleme özelliği şu anda 'SKT TAKİBİ' ve 'KAREKOD ÜRET' modüllerinde çalışmaktadır.\n\nLütfen ilgili modülü açıp dosyayı tekrar sürükleyin.")

    def sort_treeview(self, tree, col, reverse):
        l = [(tree.set(k, col), k) for k in tree.get_children('')]
        try:
            if "TUTAR" in col: l.sort(key=lambda t: float(t[0].replace(' ₺','').replace('.','').replace(',','.')), reverse=reverse)
            elif "ALIM" in col or "ODEME" in col:
                def dk(val):
                    if val == "-" or not val: return datetime.min
                    try: return datetime.strptime(val, "%d.%m.%Y")
                    except: return datetime.min
                l.sort(key=lambda t: dk(t[0]), reverse=reverse)
            elif col == "ID": l.sort(key=lambda t: int(t[0]), reverse=reverse)
            else: l.sort(reverse=reverse)
        except: l.sort(reverse=reverse)
        for index, (val, k) in enumerate(l): tree.move(k, '', index)
        tree.heading(col, command=lambda: self.sort_treeview(tree, col, not reverse))

    # =============================================================================
    # 1. DEPO ÖDEMELERİ (TAMAMEN GÜNCELLENMİŞ VE HATASIZ VERSİYON)
    # =============================================================================
    def sayfa_depo_odemeleri(self):
        c = TM
        for w in self.content_area.winfo_children(): w.destroy()
        
        header = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        header.pack(fill="x", pady=(0, 20))
        tk.Label(header, text="Depo Ödeme Yönetimi", font=FONT_HEAD, bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left")
        
        frm_btns = tk.Frame(header, bg=c.get_color("bg_main"))
        frm_btns.pack(side="right")
        
        ModernButton(frm_btns, text="+ DEPO KARTI EKLE", command=self.yeni_sekme_popup, width=200, height=40, bg_color=c.get_color("btn_primary")).pack(side="left", padx=10)
        ModernButton(frm_btns, text="🔄 YENİLE", command=lambda: self.sekmeleri_guncelle(), width=130, height=40, bg_color="#64748b").pack(side="left", padx=10)
        
        self.notebook = DraggableNotebook(self.content_area, on_reorder=self.depo_sirasi_kaydet)
        self.notebook.pack(fill="both", expand=True, pady=10)
        
        self.sekmeleri_guncelle()

    def depo_sirasi_kaydet(self, tab_names):
        # Sekmelerin yeni sırasını veritabanına kaydet
        for index, name in enumerate(tab_names):
            if name == "GENEL ÖZET": continue 
            self.imlec_finans.execute("INSERT OR IGNORE INTO depo_ayarlari (fatura_adi, sira) VALUES (?, ?)", (name, index))
            self.imlec_finans.execute("UPDATE depo_ayarlari SET sira=? WHERE fatura_adi=?", (index, name))
        self.baglanti_finans.commit()

    def sekmeleri_guncelle(self, hedef_sekme_adi=None):
        if hedef_sekme_adi is None:
            try:
                current = self.notebook.select()
                hedef_sekme_adi = self.notebook.tab(current, "text")
            except:
                hedef_sekme_adi = "GENEL ÖZET"

        for tab in self.notebook.tabs(): 
            self.notebook.forget(tab)
        
        # 1. ÖZET SEKME (Sabit)
        tab_ozet = tk.Frame(self.notebook, bg=TM.get_color("bg_main"))
        self.notebook.add(tab_ozet, text="GENEL ÖZET")
        self.ozet_ekranini_olustur(tab_ozet)
        
        # 2. DEPO SEKMELERİ (Manuel Gelir ve Giderler Hariç Tutuldu)
        try:
            sorgu = """
                SELECT DISTINCT fatura_adi FROM odemeler 
                WHERE (aciklama IS NULL OR aciklama != 'Kredi Kartı') 
                AND (satir_notu IS NULL OR (NOT satir_notu LIKE 'KART:%' AND NOT satir_notu LIKE '%KURUM%' AND satir_notu != 'MANUEL_GIDER' AND satir_notu != 'MANUEL_GELIR'))
                UNION 
                SELECT DISTINCT baslik FROM ozel_sekmeler
            """
            self.imlec_finans.execute(sorgu)
            ham_depolar = [r[0] for r in self.imlec_finans.fetchall() if r[0]]
            
            sirali_depolar = []
            for d in ham_depolar:
                if str(d).endswith(" KK"): continue

                self.imlec_finans.execute("SELECT sira FROM depo_ayarlari WHERE fatura_adi=?", (d,))
                res = self.imlec_finans.fetchone()
                sira = res[0] if res and res[0] is not None else 999
                sirali_depolar.append((sira, d))
                
            sirali_depolar.sort(key=lambda x: x[0])
            
            for _, depo in sirali_depolar:
                self.depo_sekmesi_olustur(depo)
                
        except Exception as e:
            print("Sekme güncelleme hatası:", e)
        
        # Hedef sekmeyi seç
        for tab_id in self.notebook.tabs():
            if self.notebook.tab(tab_id, "text") == hedef_sekme_adi:
                self.notebook.select(tab_id)
                break

    def ozet_ekranini_olustur(self, parent):
        canvas = tk.Canvas(parent, bg=TM.get_color("bg_main"), highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=TM.get_color("bg_main"))
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        def on_canvas_configure(e): canvas.itemconfig(window_id, width=e.width)
        canvas.bind("<Configure>", on_canvas_configure)
        
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        scrollbar.pack(side="right", fill="y")
        self.mouse_scroll_ekle(canvas, scrollable_frame)

        try:
            # (Manuel Gelir ve Giderler Hariç Tutuldu)
            sorgu = """
                SELECT DISTINCT fatura_adi FROM odemeler 
                WHERE (aciklama IS NULL OR aciklama != 'Kredi Kartı') 
                AND (satir_notu IS NULL OR (NOT satir_notu LIKE 'KART:%' AND NOT satir_notu LIKE '%KURUM%' AND satir_notu != 'MANUEL_GIDER' AND satir_notu != 'MANUEL_GELIR'))
                UNION 
                SELECT DISTINCT baslik FROM ozel_sekmeler
            """
            self.imlec_finans.execute(sorgu)
            
            ham_depolar = [r[0] for r in self.imlec_finans.fetchall() if r[0]]
            sirali_depolar = []
            
            for d in ham_depolar:
                if str(d).endswith(" KK"): continue 

                self.imlec_finans.execute("SELECT sira FROM depo_ayarlari WHERE fatura_adi=?", (d,))
                res = self.imlec_finans.fetchone()
                sira = res[0] if res and res[0] is not None else 9999
                
                if not res:
                    self.imlec_finans.execute("INSERT OR IGNORE INTO depo_ayarlari (fatura_adi, sira) VALUES (?, ?)", (d, 9999))
                    self.baglanti_finans.commit()
                
                sirali_depolar.append((sira, d))
            
            sirali_depolar.sort(key=lambda x: x[0])
            
            if not sirali_depolar: 
                tk.Label(scrollable_frame, text="Henüz kayıtlı depo yok.", font=("Segoe UI", 12), bg=TM.get_color("bg_main"), fg="#64748b").pack(pady=20)
            
            for sira, depo in sirali_depolar:
                DepotCard(scrollable_frame, depo, self.imlec_finans, self.baglanti_finans, self.sekmeleri_guncelle, self.kartlari_yer_degistir).pack(fill="x", pady=8, padx=5)
        except Exception as e:
            print(f"Özet ekranı yüklenirken hata: {e}")

    def kartlari_yer_degistir(self, kaynak_depo, hedef_depo):
        def get_sira(depo):
            self.imlec_finans.execute("SELECT sira FROM depo_ayarlari WHERE fatura_adi=?", (depo,))
            res = self.imlec_finans.fetchone()
            return res[0] if res and res[0] is not None else 999
        sira1 = get_sira(kaynak_depo); sira2 = get_sira(hedef_depo)
        
        if sira1 == 999 or sira2 == 999 or sira1 == sira2:
            self.imlec_finans.execute("SELECT fatura_adi FROM depo_ayarlari ORDER BY sira")
            mevcutlar = self.imlec_finans.fetchall()
            for idx, (ad,) in enumerate(mevcutlar): 
                self.imlec_finans.execute("UPDATE depo_ayarlari SET sira=? WHERE fatura_adi=?", (idx, ad))
            self.baglanti_finans.commit()
            sira1 = get_sira(kaynak_depo); sira2 = get_sira(hedef_depo)
            
        self.imlec_finans.execute("UPDATE depo_ayarlari SET sira=? WHERE fatura_adi=?", (sira2, kaynak_depo))
        self.imlec_finans.execute("INSERT OR IGNORE INTO depo_ayarlari (fatura_adi, sira) VALUES (?, ?)", (kaynak_depo, sira2))
        self.imlec_finans.execute("UPDATE depo_ayarlari SET sira=? WHERE fatura_adi=?", (sira1, hedef_depo))
        self.imlec_finans.execute("INSERT OR IGNORE INTO depo_ayarlari (fatura_adi, sira) VALUES (?, ?)", (hedef_depo, sira1))
        self.baglanti_finans.commit()
        self.sekmeleri_guncelle()

    def depo_sekmesi_olustur(self, depo_adi):
        c = TM
        tab = tk.Frame(self.notebook, bg=c.get_color("card_bg"))
        self.notebook.add(tab, text=depo_adi)
        
        # --- YENİ EKLENEN BAŞLIK VE DÜZENLE BUTONU ---
        header_strip = tk.Frame(tab, bg=c.get_color("card_bg"))
        header_strip.pack(fill="x", pady=(10, 0), padx=10)
        
        left_box = tk.Frame(header_strip, bg=c.get_color("card_bg"))
        left_box.pack(side="left")
        
        tk.Label(left_box, text=f"🏢 {depo_adi}", font=("Segoe UI", 16, "bold"), bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).pack(side="left")
        
        btn_style = {"font": ("Segoe UI", 9, "bold"), "relief": "flat", "cursor": "hand2", "padx": 10, "pady": 2}
        tk.Button(left_box, text="✏️ DÜZENLE", bg="#f1f5f9", fg="#64748b", **btn_style, command=lambda d=depo_adi: self.depo_duzenle_popup(d)).pack(side="left", padx=(15, 5))
        # ---------------------------------------------

        # ÜST BUTONLAR
        bar = tk.Frame(tab, bg=c.get_color("card_bg"), pady=5, padx=10)
        bar.pack(fill="x")
        
        # TABLO YAPISI
        paned = tk.PanedWindow(tab, orient="vertical", bg=c.get_color("card_bg"), sashwidth=6)
        paned.pack(fill="both", expand=True, padx=5, pady=5)
        
        frm_table = tk.Frame(paned, bg=c.get_color("card_bg"))
        paned.add(frm_table, height=500)
        
        frm_row_ops = tk.Frame(frm_table, bg=c.get_color("card_bg"), pady=5)
        frm_row_ops.pack(side="bottom", fill="x")
        
        # SÜTUNLAR (show="tree headings" yapılarak akordiyon okları aktif edildi)
        cols = ("ID", "TIK", "TUR", "ALIM", "ODEME", "TUTAR", "NOT")
        tree = ttk.Treeview(frm_table, columns=cols, show="tree headings")
        
        # Akordiyon (Aç/Kapat) Oku için en baştaki varsayılan sütun ayarı
        tree.heading("#0", text="▼")
        tree.column("#0", width=40, stretch=False, anchor="center")

        tree.heading("ID", text="ID"); tree.column("ID", width=0, stretch=False)
        
        self.all_checked_state = False 
        def toggle_all_selection():
            self.all_checked_state = not self.all_checked_state
            sembol = "☑" if self.all_checked_state else "☐"
            tree.heading("TIK", text=f"✔ {sembol}")
            # Sadece ana satırları seçer (alt kırılımları seçmez)
            for child in tree.get_children(""):
                vals = list(tree.item(child)['values'])
                vals[1] = sembol
                tree.item(child, values=vals)

        tree.heading("TIK", text="✔", command=toggle_all_selection); tree.column("TIK", width=35, anchor="center")
        tree.heading("TUR", text="ÖDEME TÜRÜ", command=lambda: self.sort_treeview(tree, "TUR", False)); tree.column("TUR", width=110, anchor="center")
        tree.heading("ALIM", text="ALIM DÖNEMİ", command=lambda: self.sort_treeview(tree, "ALIM", False)); tree.column("ALIM", width=140, anchor="center")
        tree.heading("ODEME", text="ÖDEME DÖNEMİ", command=lambda: self.sort_treeview(tree, "ODEME", False)); tree.column("ODEME", width=140, anchor="center")
        tree.heading("TUTAR", text="TUTAR", command=lambda: self.sort_treeview(tree, "TUTAR", False)); tree.column("TUTAR", width=160, anchor="e")
        tree.heading("NOT", text="NOT", command=lambda: self.sort_treeview(tree, "NOT", False)); tree.column("NOT", width=150, anchor="w")
        
        sc = ttk.Scrollbar(frm_table, orient="vertical", command=tree.yview)
        tree.configure(yscroll=sc.set)
        sc.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True)
        
        tree.tag_configure("odendi", background="#f1f5f9", foreground="#94a3b8")
        tree.tag_configure("odenmedi", background="#fee2e2", foreground="#000000")
        tree.tag_configure("alt_kırılım", background="#f8fafc", foreground="#475569", font=("Segoe UI", 9, "italic"))
        
        lbl_borc = tk.Label(frm_table, text="", font=("Segoe UI", 14, "bold"), bg=c.get_color("card_bg"), fg=c.get_color("btn_danger"), anchor="e")
        lbl_borc.pack(side="bottom", fill="x", pady=5)

        def tabloyu_doldur():
            for i in tree.get_children(""): tree.delete(i)
            
            sorgu = """
                SELECT id, alim_tarihi, vade_tarihi, tutar, durum, satir_notu, aciklama 
                FROM odemeler 
                WHERE fatura_adi=? 
                AND (aciklama IS NULL OR aciklama NOT LIKE '%Kredi Kartı%')
                AND (satir_notu IS NULL OR (satir_notu NOT LIKE '%KART%' AND satir_notu != 'MANUEL_GIDER' AND satir_notu != 'MANUEL_GELIR'))
                ORDER BY vade_tarihi ASC
            """
            try:
                self.imlec_finans.execute(sorgu, (depo_adi,))
                veriler = self.imlec_finans.fetchall()
            except Exception as e:
                print("Tablo doldurma hatası:", e)
                return

            toplam_borc = 0
            aylar_tr = {"01": "OCAK", "02": "ŞUBAT", "03": "MART", "04": "NİSAN", "05": "MAYIS", "06": "HAZİRAN", "07": "TEMMUZ", "08": "AĞUSTOS", "09": "EYLÜL", "10": "EKİM", "11": "KASIM", "12": "ARALIK"}

            for v in veriler:
                oid, alim, vade, tutar, durum, not_icerik, aciklama_raw = v
                if durum == "ODENMEDİ": toplam_borc += tutar
                
                odeme_turu = "-"
                temiz_aciklama = str(aciklama_raw).strip() if aciklama_raw else ""
                
                if temiz_aciklama.startswith("[") and "]" in temiz_aciklama:
                    odeme_turu = temiz_aciklama.split("]")[0].replace("[", "")
                elif "Excel" in temiz_aciklama:
                    odeme_turu = "EXCEL"

                try:
                    if vade:
                        v_obj = datetime.strptime(str(vade)[:10], "%Y-%m-%d")
                        vade_str = f"{aylar_tr[v_obj.strftime('%m')]} {v_obj.year}"
                    else: vade_str = "-"
                except: vade_str = vade

                try:
                    if alim and alim != "-":
                        a_obj = datetime.strptime(str(alim)[:10], "%Y-%m-%d")
                        alim_str = f"{aylar_tr[a_obj.strftime('%m')]} {a_obj.year}"
                    else: alim_str = alim
                except: alim_str = alim

                tag = "odendi" if durum == "ODENDİ" else "odenmedi"
                
                # --- AKORDİYON KDV MANTIĞI ---
                if not_icerik and str(not_icerik).startswith("[KDV_DETAY]"):
                    not_sembolu = "KDV Kirilimlari"
                    # Parent Node (Ana Satır)
                    parent_id = tree.insert("", "end", text="📂", values=(oid, "☐", odeme_turu, alim_str, vade_str, f"{tutar:,.2f} ₺", not_sembolu), tags=(tag,))
                    
                    try:
                        kdv_parts = str(not_icerik).replace("[KDV_DETAY]", "").strip().split("|")
                        for p in kdv_parts:
                            if ":" not in p: continue
                            k_isim, k_tutar_str = p.split(":")
                            k_tut = float(k_tutar_str)
                            if k_tut > 0:
                                # Child Node (Alt Kırılım - Akordiyon içi)
                                tree.insert(parent_id, "end", text="↳", values=(f"child_{oid}_{k_isim}", "", "", "", "", f"{k_tut:,.2f} ₺", k_isim.strip()), tags=("alt_kırılım",))
                    except: pass
                else:
                    not_sembolu = "⚠️" if not_icerik and str(not_icerik).strip() not in ["", "None", "MANUEL_GIDER"] else "➕"
                    tree.insert("", "end", text="📄", values=(oid, "☐", odeme_turu, alim_str, vade_str, f"{tutar:,.2f} ₺", not_sembolu), tags=(tag,))
            
            lbl_borc.config(text=f"AÇIK BORÇ: {toplam_borc:,.2f} ₺")

        def on_tree_click(event):
            region = tree.identify("region", event.x, event.y)
            if region != "cell": return
            col = tree.identify_column(event.x)
            row_id = tree.identify_row(event.y)
            if not row_id: return
            
            # Alt satırlara (child) tıklandıysa hiçbir şey yapma (hata vermesini önler)
            if str(row_id).startswith("child_"): return

            item = tree.item(row_id)
            vals = list(item['values'])
            
            if col == "#2": # TIK
                vals[1] = "☑" if vals[1] == "☐" else "☐"
                tree.item(row_id, values=vals)
                if tree.selection(): tree.selection_remove(tree.selection()[0])
            
            elif col == "#7": # NOT SÜTUNU
                oid = vals[0]
                res = self.imlec_finans.execute("SELECT satir_notu FROM odemeler WHERE id=?", (oid,)).fetchone()
                mevcut_not = res[0] if res and res[0] not in ["MANUEL_GIDER", None] else ""
                
                # KDV detaylarının notunu bozmamak için engelliyoruz
                if str(mevcut_not).startswith("[KDV_DETAY]"):
                    messagebox.showinfo("Bilgi", "Bu satır otomatik aylık KDV icmalidir. Notu değiştirilemez.\nDetayları görmek için yanındaki oka tıklayın.")
                    return

                def save_note_callback(new_text):
                    self.imlec_finans.execute("UPDATE odemeler SET satir_notu=? WHERE id=?", (new_text, oid))
                    self.baglanti_finans.commit()
                    tabloyu_doldur()
                
                x_root = tree.winfo_pointerx()
                y_root = tree.winfo_pointery()
                NoteBubble(self.pencere, x_root, y_root, mevcut_not, save_note_callback)
        
        tree.bind("<ButtonRelease-1>", on_tree_click)

        def tiklileri_isle(hedef_durum):
            islem = False
            for child in tree.get_children(""):
                vals = tree.item(child)['values']
                if vals and vals[1] == "☑": 
                    self.imlec_finans.execute("UPDATE odemeler SET durum=? WHERE id=?", (hedef_durum, vals[0]))
                    islem = True
            if islem: 
                self.baglanti_finans.commit()
                tabloyu_doldur()
            else: messagebox.showinfo("Uyarı", "Seçim yapın.")

        def tiklileri_sil(tree):
            ids = [tree.item(c)['values'][0] for c in tree.get_children("") if tree.item(c)['values'] and tree.item(c)['values'][1] == "☑"]
            if not ids: messagebox.showinfo("Uyarı", "Seçili yok."); return
            if messagebox.askyesno("Sil", f"{len(ids)} kaydı silmek istiyor musun?"):
                for oid in ids: self.imlec_finans.execute("DELETE FROM odemeler WHERE id=?", (oid,))
                self.baglanti_finans.commit()
                tabloyu_doldur()

        def ozel_manuel_ekle(depo_adi):
            def veritabanina_ekle_temiz(d_adi, alim, odeme, tutar, not_icerik, odeme_turu):
                aciklama_kayit = f"[{odeme_turu}]"
                satir_notu_kayit = not_icerik if not_icerik else ""
                
                self.imlec_finans.execute("INSERT INTO odemeler (fatura_adi, vade_tarihi, tutar, aciklama, alim_tarihi, satir_notu, durum) VALUES (?, ?, ?, ?, ?, ?, 'ODENMEDİ')", 
                                   (d_adi, odeme, tutar, aciklama_kayit, alim, satir_notu_kayit))
                self.baglanti_finans.commit()
                tabloyu_doldur()

            ManualAddWindow(self.pencere, depo_adi, veritabanina_ekle_temiz)

        # Butonlar
        ModernButton(bar, text="➕ MANUEL EKLE", command=lambda: ozel_manuel_ekle(depo_adi), width=250, height=35, bg_color="#10b981").pack(side="left", padx=5)
        ModernButton(bar, text="📗 TEKLİ FATURA YÜKLE", command=lambda: self.taba_excel_yukle(depo_adi), width=250, height=35, bg_color="#3b82f6").pack(side="left", padx=5)
        
        # YENİ BUTONUMUZ: AYLIK KDV DÖKÜMÜ 
        ModernButton(bar, text="📊 AYLIK DÖKÜM YÜKLE", command=lambda: self.aylik_dokum_yukle(depo_adi), width=250, height=35, bg_color="#8b5cf6").pack(side="left", padx=5)
        
        ModernButton(bar, text="⛔ DEPOYU SİL", command=lambda: self.sekmeyi_komple_sil(depo_adi), width=200, height=35, bg_color="#7f1d1d").pack(side="right", padx=5)
        ModernButton(bar, text="🗑️ SEÇİLENLERİ SİL", command=lambda: tiklileri_sil(tree), width=200, height=35, bg_color="#ef4444").pack(side="right", padx=5)
        
        ModernButton(frm_row_ops, text="✅ ÖDENDİ İŞARETLE", command=lambda: tiklileri_isle("ODENDİ"), width=250, height=40, bg_color=c.get_color("btn_success")).pack(side="left", padx=5)
        ModernButton(frm_row_ops, text="❌ ÖDENMEDİ İŞARETLE", command=lambda: tiklileri_isle("ODENMEDİ"), width=250, height=40, bg_color=c.get_color("btn_danger")).pack(side="left", padx=5)
        
        tabloyu_doldur()

    def taba_excel_yukle(self, hedef_depo_adi):
        if not PANDAS_VAR: messagebox.showerror("Hata", "pandas ve openpyxl kütüphaneleri eksik."); return
        file_path = filedialog.askopenfilename(title=f"{hedef_depo_adi} İçin Excel Seç", filetypes=[("Excel Dosyaları", "*.xlsx *.xls")])
        if not file_path: return
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            eklenen = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                col_alim, col_tutar, col_odeme = None, None, None
                header_row = -1
                
                for r_idx, row in enumerate(ws.iter_rows(max_row=50, values_only=True), 1):
                    row_str = [str(v).upper().strip() if v else "" for v in row]
                    if "TUTAR" in row_str and ("ALIM TARİHİ" in row_str or "ALIM TARIHI" in row_str):
                        header_row = r_idx
                        for c_idx, val in enumerate(row_str):
                            if "ALIM TARİHİ" in val or "ALIM TARIHI" in val: col_alim = c_idx + 1
                            elif "ÖDEME TARİHİ" in val or "ODEME TARIHI" in val or "VADE" in val: col_odeme = c_idx + 1
                            elif "TUTAR" in val: col_tutar = c_idx + 1
                        break
                
                if header_row != -1 and col_tutar and col_alim:
                    for r_idx, row in enumerate(ws.iter_rows(min_row=header_row+1), header_row+1):
                        try:
                            cell_tutar = row[col_tutar-1]
                            cell_alim = row[col_alim-1]
                            cell_odeme = row[col_odeme-1] if col_odeme else None
                            
                            val_tutar = cell_tutar.value
                            if val_tutar is None: continue
                            
                            tutar_str = str(val_tutar).replace("TL", "").strip()
                            tutar = float(tutar_str.replace(",", ".")) if isinstance(val_tutar, str) else float(val_tutar)
                            
                            alim_val = self.tarih_duzelt(cell_alim.value)
                            odeme_val = self.tarih_duzelt(cell_odeme.value) if cell_odeme else None
                            
                            if not odeme_val and alim_val:
                                try:
                                    dt_alim = datetime.strptime(alim_val, "%Y-%m-%d")
                                    import datetime as dt_mod
                                    dt_odeme = dt_alim + dt_mod.timedelta(days=90)
                                    odeme_val = dt_odeme.strftime("%Y-%m-%d")
                                except: pass

                            if not odeme_val: continue
                            
                            durum = "ODENMEDİ"
                            if cell_tutar.fill and cell_tutar.fill.patternType == 'solid':
                                color = cell_tutar.fill.start_color
                                if color.type == 'rgb':
                                    hex_code = str(color.rgb)
                                    if "FF0000" in hex_code or "FFC7CE" in hex_code: durum = "ODENMEDİ"
                                    elif "00FF00" in hex_code or "C6EFCE" in hex_code or "92D050" in hex_code: durum = "ODENDİ"
                            
                            self.imlec_finans.execute("INSERT INTO odemeler (fatura_adi, vade_tarihi, tutar, aciklama, durum, alim_tarihi) VALUES (?, ?, ?, ?, ?, ?)", 
                                               (hedef_depo_adi, odeme_val, tutar, "Excel Import", durum, alim_val))
                            eklenen += 1
                        except: continue
                        
            self.baglanti_finans.commit()
            self.sekmeleri_guncelle(hedef_sekme_adi=hedef_depo_adi)
            
            if eklenen > 0:
                messagebox.showinfo("Başarılı", f"{eklenen} adet kayıt başarıyla eklendi.")
            else:
                messagebox.showwarning("Uyarı", "Excel dosyasında uygun formatta veri bulunamadı.\nLütfen sütun başlıklarının (Alım Tarihi, Ödeme Tarihi, Tutar) doğru olduğundan emin olun.")
                
        except Exception as e:
            messagebox.showerror("Hata", str(e))

    def yeni_sekme_popup(self):
        popup = tk.Toplevel(self.pencere)
        popup.title("Depo Kartı Ekle")
        popup.geometry("360x600")
        popup.configure(bg="#f8fafc")
        popup.transient(self.pencere) 
        popup.grab_set() 
        popup.focus_force()
        
        x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 180
        y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 300
        popup.geometry(f"+{x}+{y}")

        tk.Label(popup, text="Hızlı Ekleme Listesi", font=("Segoe UI", 12, "bold"), bg="#f8fafc", fg="#334155").pack(pady=(20, 10))

        hazir_depolar = [
            "SELÇUK ECZA DEPOSU", "ISKOOP", "BEK", "NEVZAT ECZA DEPOSU",
            "CENCORA (ALLIANCE)", "SANCAK ECZA DEPOSU", "AS ECZA DEPOSU"
        ]

        for depo in hazir_depolar:
            btn = tk.Button(popup, text=depo, font=("Segoe UI", 11), bg="white", fg="#0f172a",
                            relief="solid", bd=1, cursor="hand2", pady=8,
                            activebackground="#e0f2fe",
                            command=lambda d=depo: self._ozel_depo_ekle_islem(d, popup))
            btn.pack(fill="x", padx=30, pady=4)

        tk.Frame(popup, height=2, bg="#cbd5e1").pack(fill="x", padx=30, pady=15)

        btn_diger = tk.Button(popup, text="✍️ DİĞER (Manuel Yaz)", font=("Segoe UI", 11, "bold"), 
                              bg="#3b82f6", fg="white", 
                              relief="flat", cursor="hand2", pady=10,
                              activebackground="#2563eb", activeforeground="white",
                              command=lambda: self._manuel_depo_penceresi(popup))
        btn_diger.pack(fill="x", padx=30, pady=(0, 20))

    def _manuel_depo_penceresi(self, onceki_pencere):
        if onceki_pencere:
            onceki_pencere.destroy()
            
        win = tk.Toplevel(self.pencere)
        win.title("Yeni Depo Ekle")
        win.geometry("400x250")
        win.configure(bg="white")
        win.transient(self.pencere)
        win.grab_set()
        win.focus_force()
        
        x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 200
        y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 125
        win.geometry(f"+{x}+{y}")

        tk.Label(win, text="Yeni Depo / Firma Adı", font=("Segoe UI", 14, "bold"), bg="white", fg="#334155").pack(pady=(25, 10))
        
        ent_isim = tk.Entry(win, font=("Segoe UI", 12), relief="solid", bd=1, bg="#f8fafc")
        ent_isim.pack(fill="x", padx=40, pady=10, ipady=5)
        ent_isim.focus_set()

        def onayla():
            girilen = ent_isim.get().strip()
            if not girilen: return
            isim = girilen.upper().replace("i", "İ").replace("ı", "I")
            self._ozel_depo_ekle_islem(isim, win)

        win.bind('<Return>', lambda e: onayla())
        win.bind('<Escape>', lambda e: win.destroy())

        btn_frame = tk.Frame(win, bg="white")
        btn_frame.pack(fill="x", padx=40, pady=20)
        
        tk.Button(btn_frame, text="İPTAL", font=("Segoe UI", 10, "bold"), bg="#f1f5f9", fg="#64748b", 
                  relief="flat", cursor="hand2", width=12, pady=8,
                  command=win.destroy).pack(side="left")
                  
        tk.Button(btn_frame, text="KAYDET", font=("Segoe UI", 10, "bold"), bg="#10b981", fg="white", 
                  relief="flat", cursor="hand2", width=12, pady=8,
                  command=onayla).pack(side="right")

    def _ozel_depo_ekle_islem(self, isim, pencere_objesi):
        try:
            self.imlec_finans.execute("CREATE TABLE IF NOT EXISTS ozel_sekmeler (id INTEGER PRIMARY KEY AUTOINCREMENT, baslik TEXT, icerik TEXT)")
            self.imlec_finans.execute("CREATE TABLE IF NOT EXISTS depo_ayarlari (fatura_adi TEXT PRIMARY KEY, anlasma_sarti TEXT, sira INTEGER)")
            self.baglanti_finans.commit()

            self.imlec_finans.execute("SELECT id FROM ozel_sekmeler WHERE baslik=?", (isim,))
            var_mi = self.imlec_finans.fetchone()
            
            if var_mi:
                messagebox.showwarning("Mevcut", f"'{isim}' zaten listenizde ekli.")
            else:
                self.imlec_finans.execute("INSERT INTO ozel_sekmeler (baslik, icerik) VALUES (?, '')", (isim,))
                self.imlec_finans.execute("INSERT OR IGNORE INTO depo_ayarlari (fatura_adi, sira) VALUES (?, 999)", (isim,))
                self.baglanti_finans.commit()
                
                self.sekmeleri_guncelle(hedef_sekme_adi=isim)
                
                if pencere_objesi:
                    pencere_objesi.destroy()
                    
        except Exception as e:
            messagebox.showerror("Hata", f"Depo eklenirken hata oluştu: {e}")

    def sekmeyi_komple_sil(self, depo_adi):
        if messagebox.askyesno("ÇOK ÖNEMLİ DİKKAT!", f"'{depo_adi}' deposunu ve içindeki TÜM VERİLERİ kalıcı olarak silmek üzeresin.\n\nBu işlem geri alınamaz!\nOnaylıyor musun?"):
            try:
                self.imlec_finans.execute("DELETE FROM odemeler WHERE fatura_adi=?", (depo_adi,))
                self.imlec_finans.execute("DELETE FROM ozel_sekmeler WHERE baslik=?", (depo_adi,))
                self.imlec_finans.execute("DELETE FROM depo_ayarlari WHERE fatura_adi=?", (depo_adi,))
                self.baglanti_finans.commit()
                messagebox.showinfo("Başarılı", f"'{depo_adi}' başarıyla silindi.")
                self.sekmeleri_guncelle()
            except Exception as e:
                messagebox.showerror("Hata", f"Silme işleminde hata: {e}")

    def depo_duzenle_popup(self, eski_isim):
        win = tk.Toplevel(self.pencere)
        win.title("Depo Adını Düzenle")
        win.geometry("350x200")
        
        try: bg_renk = TM.get_color("bg_main"); fg_renk = TM.get_color("fg_text")
        except: bg_renk = "white"; fg_renk = "black"
        win.configure(bg=bg_renk)
        
        # Ekranı Ortala
        x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 175
        y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 100
        win.geometry(f"+{x}+{y}")
        
        tk.Label(win, text="Yeni Depo Adı:", font=("Segoe UI", 11, "bold"), bg=bg_renk, fg=fg_renk).pack(pady=(20, 5))
        
        e_yeni = ttk.Entry(win, font=("Segoe UI", 11), width=25)
        e_yeni.insert(0, eski_isim)
        e_yeni.pack(pady=5)
        
        def kaydet():
            yeni_isim = e_yeni.get().strip().upper()
            if not yeni_isim or yeni_isim == eski_isim: return
            
            try:
                # 1. Faturaları/Ödemeleri yeni isme aktar
                self.imlec_finans.execute("UPDATE odemeler SET fatura_adi=? WHERE fatura_adi=?", (yeni_isim, eski_isim))
                
                # 2. Ayarları ve Sekmeleri aktar (Eğer yeni isim zaten varsa çakışmayı önler ve birleştirir)
                self.imlec_finans.execute("UPDATE OR IGNORE depo_ayarlari SET fatura_adi=? WHERE fatura_adi=?", (yeni_isim, eski_isim))
                self.imlec_finans.execute("DELETE FROM depo_ayarlari WHERE fatura_adi=?", (eski_isim,))
                
                self.imlec_finans.execute("UPDATE OR IGNORE ozel_sekmeler SET baslik=? WHERE baslik=?", (yeni_isim, eski_isim))
                self.imlec_finans.execute("DELETE FROM ozel_sekmeler WHERE baslik=?", (eski_isim,))
                
                self.baglanti_finans.commit()
                win.destroy()
                
                messagebox.showinfo("Başarılı", f"Depo adı '{yeni_isim}' olarak güncellendi!", parent=self.pencere)
                self.sekmeleri_guncelle(hedef_sekme_adi=yeni_isim)
                
            except Exception as e:
                messagebox.showerror("Hata", f"Güncelleme sırasında hata oluştu:\n{e}", parent=win)
                
        try: ModernButton(win, text="KAYDET", command=kaydet, bg_color="#10b981", width=120, height=35).pack(pady=20)
        except: tk.Button(win, text="KAYDET", command=kaydet, bg="#10b981", fg="white", font=("Segoe UI", 10, "bold"), padx=20, pady=5).pack(pady=20)            

    # --- 2. KREDİ KARTLARI ---
    def sayfa_kredi_karti(self):
        c = TM
        for w in self.content_area.winfo_children(): w.destroy()
        header = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        header.pack(fill="x", pady=(0, 15))
        tk.Label(header, text="Kredi Kartı & Taksit Yönetimi", font=FONT_HEAD, bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left")
        frm_btns = tk.Frame(header, bg=c.get_color("bg_main")); frm_btns.pack(side="right")
        ModernButton(frm_btns, text="+ YENİ KART TANIMLA", command=self.yeni_kart_ekle_popup, width=200, height=40, bg_color=c.get_color("btn_primary")).pack(side="left", padx=10)
        ModernButton(frm_btns, text="🔄 YENİLE", command=self.kredi_karti_arayuzunu_guncelle, width=130, height=40, bg_color="#64748b").pack(side="left", padx=10)
        self.notebook_kart = DraggableNotebook(self.content_area, on_reorder=self.kart_sirasi_kaydet)
        self.notebook_kart.pack(fill="both", expand=True, pady=10)
        self.kredi_karti_arayuzunu_guncelle()

    def kart_sirasi_kaydet(self, tab_names):
        # Kartların yeni sırasını kaydet
        for index, name in enumerate(tab_names):
            if name == "GENEL ÖZET": continue
            self.imlec_finans.execute("UPDATE kredi_kartlari SET sira=? WHERE isim=?", (index, name))
        # HATA BURADAYDI, DÜZELTİLDİ:
        self.baglanti_finans.commit()

    def kredi_karti_arayuzunu_guncelle(self, hedef_sekme=None):
        for tab in self.notebook_kart.tabs(): self.notebook_kart.forget(tab)
        tab_ozet = tk.Frame(self.notebook_kart, bg=TM.get_color("bg_main"))
        self.notebook_kart.add(tab_ozet, text="GENEL ÖZET")
        self.kredi_karti_ozet_sayfasi(tab_ozet)
        self.imlec_finans.execute("SELECT isim FROM kredi_kartlari ORDER BY COALESCE(sira, 999), isim")
        kartlar = [r[0] for r in self.imlec_finans.fetchall()]
        for kart in kartlar:
            try: self.kredi_karti_sekmesi_olustur(kart)
            except Exception as e: print(f"Kart hatası ({kart}): {e}")
        if hedef_sekme:
            for tab_id in self.notebook_kart.tabs():
                if self.notebook_kart.tab(tab_id, "text") == hedef_sekme:
                    self.notebook_kart.select(tab_id); break

    def ozel_kk_depo_secici(self, callback_func):
        popup = tk.Toplevel(self.pencere); popup.title("Firma / Depo Seç"); popup.geometry("380x600"); popup.configure(bg="#f8fafc")
        popup.transient(self.pencere); popup.grab_set()
        x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 190
        y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 300
        popup.geometry(f"+{x}+{y}")
        header_frame = tk.Frame(popup, bg="#f8fafc", pady=10); header_frame.pack(side="top", fill="x")
        tk.Label(header_frame, text="Hızlı Firma Seçimi (KK)", font=("Segoe UI", 11, "bold"), bg="#f8fafc", fg="#334155").pack()
        tk.Label(header_frame, text="(Seçilince sonuna 'KK' eklenir)", font=("Segoe UI", 8), bg="#f8fafc", fg="#64748b").pack()
        footer_frame = tk.Frame(popup, bg="#f8fafc", pady=10); footer_frame.pack(side="bottom", fill="x")
        def manuel_giris_ac():
            popup.destroy()
            input_win = tk.Toplevel(self.pencere); input_win.title("Manuel Giriş"); input_win.geometry("400x250"); input_win.configure(bg="white")
            input_win.transient(self.pencere); input_win.grab_set(); input_win.focus_force()
            ix = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 200
            iy = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 125
            input_win.geometry(f"+{ix}+{iy}")
            tk.Label(input_win, text="Firma / Yer Adı Giriniz", font=("Segoe UI", 12, "bold"), bg="white", fg="#334155").pack(pady=(25, 10))
            ent_isim = tk.Entry(input_win, font=("Segoe UI", 11), relief="solid", bd=1, bg="#f8fafc"); ent_isim.pack(fill="x", padx=30, pady=5, ipady=5); ent_isim.focus_set()
            def onayla():
                yeni_isim = ent_isim.get().strip()
                if yeni_isim:
                    yeni_isim = yeni_isim.upper().replace("i", "İ").replace("ı", "I")
                    if not yeni_isim.endswith(" KK"): yeni_isim += " KK"
                    callback_func(yeni_isim)
                    input_win.destroy()
            input_win.bind('<Return>', lambda e: onayla()); input_win.bind('<Escape>', lambda e: input_win.destroy())
            btn_box = tk.Frame(input_win, bg="white"); btn_box.pack(fill="x", pady=20, padx=30)
            tk.Button(btn_box, text="İPTAL", font=("Segoe UI", 9, "bold"), bg="#f1f5f9", fg="#64748b", relief="flat", command=input_win.destroy, width=12, pady=5).pack(side="left")
            tk.Button(btn_box, text="KAYDET", font=("Segoe UI", 9, "bold"), bg="#3b82f6", fg="white", relief="flat", command=onayla, width=12, pady=5).pack(side="right")
        tk.Button(footer_frame, text="✍️ DİĞER (Manuel Yaz)", font=("Segoe UI", 10, "bold"), bg="#3b82f6", fg="white", relief="flat", cursor="hand2", pady=8, width=30, command=manuel_giris_ac).pack()
        container = tk.Frame(popup, bg="#f8fafc"); container.pack(side="top", fill="both", expand=True, padx=5)
        canvas = tk.Canvas(container, bg="#f8fafc", highlightthickness=0); scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg="#f8fafc")
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        window_id = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.bind('<Configure>', lambda event: canvas.itemconfig(window_id, width=event.width))
        canvas.configure(yscrollcommand=scrollbar.set); canvas.pack(side="left", fill="both", expand=True); scrollbar.pack(side="right", fill="y")
        self.mouse_scroll_ekle(canvas, scroll_frame)
        hazir_depolar = ["SELÇUK ECZA DEPOSU", "ISKOOP", "BEK", "NEVZAT ECZA DEPOSU", "CENCORA (ALLIANCE)", "SANCAK ECZA DEPOSU", "AS ECZA DEPOSU", "HEDEF ECZA DEPOSU", "BAŞKENT ECZA DEPOSU", "FARMAZON", "ECZA1", "GALENOS", "LOKMAN HEKİM"]
        def secim_yap(isim):
            final_isim = f"{isim} KK"; callback_func(final_isim); popup.destroy()
        for depo in hazir_depolar:
            btn = tk.Button(scroll_frame, text=depo, font=("Segoe UI", 10), bg="white", fg="#0f172a", relief="solid", bd=1, cursor="hand2", pady=5, activebackground="#e0f2fe", command=lambda d=depo: secim_yap(d))
            btn.pack(fill="x", padx=15, pady=3)

    def kredi_karti_ozet_sayfasi(self, parent):
        canvas = tk.Canvas(parent, bg=TM.get_color("bg_main"), highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=TM.get_color("bg_main"))
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(window_id, width=e.width))
        canvas.configure(yscrollcommand=scrollbar.set); canvas.pack(side="left", fill="both", expand=True, padx=5, pady=5); scrollbar.pack(side="right", fill="y")
        self.mouse_scroll_ekle(canvas, scrollable_frame)
        bugun = date.today()
        aylar_listesi = ["OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN", "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"]
        tk.Label(scrollable_frame, text="📅 AYLIK KART ÖDEME PLANI", font=("Segoe UI", 16, "bold"), bg=TM.get_color("bg_main"), fg="#334155").pack(pady=(10, 20), anchor="w", padx=10)
        kayit_var = False
        for i in range(12):
            hedef_yil = bugun.year + (bugun.month + i - 1) // 12
            hedef_ay = (bugun.month + i - 1) % 12 + 1
            gruplanmis_veri, ay_genel_toplam = self._kart_verilerini_grupla(hedef_yil, hedef_ay)
            if not gruplanmis_veri: continue
            kayit_var = True
            ay_kutu = tk.Frame(scrollable_frame, bg="white", bd=1, relief="solid"); ay_kutu.pack(fill="x", pady=5, padx=10)
            ay_header = tk.Frame(ay_kutu, bg="#e2e8f0", padx=10, pady=12, cursor="hand2"); ay_header.pack(fill="x")
            lbl_icon_ay = tk.Label(ay_header, text="▶", font=("Segoe UI", 12), bg="#e2e8f0", fg="#475569"); lbl_icon_ay.pack(side="left", padx=(0, 10))
            tk.Label(ay_header, text=f"{aylar_listesi[hedef_ay-1]} {hedef_yil}", font=("Segoe UI", 12, "bold"), bg="#e2e8f0", fg="#1e293b").pack(side="left")
            tk.Label(ay_header, text=f"Toplam: {ay_genel_toplam:,.2f} ₺", font=("Segoe UI", 12, "bold"), bg="#e2e8f0", fg="#ef4444").pack(side="right")
            ay_content = tk.Frame(ay_kutu, bg="white", padx=10, pady=5)
            def toggle_ay(frame=ay_content, icon=lbl_icon_ay):
                if frame.winfo_viewable(): frame.forget(); icon.config(text="▶")
                else: frame.pack(fill="x"); icon.config(text="▼")
            for w in [ay_header, lbl_icon_ay] + ay_header.winfo_children(): w.bind("<Button-1>", lambda e, f=toggle_ay: f())
            for kart_adi, data in gruplanmis_veri.items():
                kart_toplam = data['toplam']; satirlar = data['liste']
                kart_frame = tk.Frame(ay_content, bg="white", bd=0, relief="flat"); kart_frame.pack(fill="x", pady=2)
                kart_header = tk.Frame(kart_frame, bg="#f8fafc", padx=10, pady=8, cursor="hand2", highlightthickness=1, highlightbackground="#cbd5e1"); kart_header.pack(fill="x")
                lbl_icon_kart = tk.Label(kart_header, text="▷", font=("Segoe UI", 10), bg="#f8fafc", fg="#64748b"); lbl_icon_kart.pack(side="left", padx=(5, 10))
                tk.Label(kart_header, text=f"💳 {kart_adi}", font=("Segoe UI", 10, "bold"), bg="#f8fafc", fg="#334155").pack(side="left")
                tk.Label(kart_header, text=f"{kart_toplam:,.2f} ₺", font=("Segoe UI", 10, "bold"), bg="#f8fafc", fg="#059669").pack(side="right")
                kart_detay = tk.Frame(kart_frame, bg="white")
                cols = ("GUN", "FIRMA", "TUTAR", "ACIKLAMA"); tree = ttk.Treeview(kart_detay, columns=cols, show="headings", height=len(satirlar))
                tree.heading("GUN", text="Gün"); tree.column("GUN", width=50, anchor="center")
                tree.heading("FIRMA", text="Firma / Depo"); tree.column("FIRMA", width=200)
                tree.heading("TUTAR", text="Tutar"); tree.column("TUTAR", width=100, anchor="e")
                tree.heading("ACIKLAMA", text="Taksit Detayı"); tree.column("ACIKLAMA", width=200)
                tree.pack(fill="x")
                for row in satirlar: tree.insert("", "end", values=(row[0], row[1], f"{row[2]:,.2f} ₺", row[3]))
                def toggle_kart(frame=kart_detay, icon=lbl_icon_kart):
                    if frame.winfo_viewable(): frame.forget(); icon.config(text="▷")
                    else: frame.pack(fill="x", pady=(0, 10)); icon.config(text="▽")
                for w in [kart_header, lbl_icon_kart] + kart_header.winfo_children(): w.bind("<Button-1>", lambda e, f=toggle_kart: f())
        if not kayit_var: tk.Label(scrollable_frame, text="Gelecek 12 ayda planlanmış ödeme bulunamadı.", font=("Segoe UI", 12), bg=TM.get_color("bg_main")).pack(pady=20)

    def _kart_verilerini_grupla(self, yil, ay):
        ay_str = f"{yil}-{ay:02d}"
        sorgu = """
            SELECT strftime('%d', vade_tarihi), fatura_adi, tutar, satir_notu 
            FROM odemeler 
            WHERE strftime('%Y-%m', vade_tarihi) = ? 
            AND (aciklama LIKE '%Kredi Kartı%' OR satir_notu LIKE 'KART:%')
            AND durum='ODENMEDİ' 
            ORDER BY vade_tarihi ASC
        """
        self.imlec_finans.execute(sorgu, (ay_str,))
        ham_veri = self.imlec_finans.fetchall()
        gruplanmis = {}; genel_toplam = 0.0
        for row in ham_veri:
            gun, firma, tutar, notu = row
            genel_toplam += tutar
            kart_adi = "Diğer Kartlar"
            if notu and "KART:" in notu:
                try:
                    temp = notu.split("KART:")[1].strip()
                    if "(" in temp: kart_adi = temp.split("(")[0].strip()
                    else: kart_adi = temp
                except: pass
            if kart_adi not in gruplanmis: gruplanmis[kart_adi] = {'toplam': 0.0, 'liste': []}
            gruplanmis[kart_adi]['toplam'] += tutar
            gruplanmis[kart_adi]['liste'].append(row)
        return gruplanmis, genel_toplam

    def kredi_karti_sekmesi_olustur(self, kart_adi):
        c = TM
        tab = tk.Frame(self.notebook_kart, bg="white", padx=10, pady=10)
        self.notebook_kart.add(tab, text=kart_adi)
        
        header_strip = tk.Frame(tab, bg="white"); header_strip.pack(fill="x", pady=(0, 10))
        
        left_box = tk.Frame(header_strip, bg="white"); left_box.pack(side="left")
        tk.Label(left_box, text=f"💳 {kart_adi}", font=("Segoe UI", 16, "bold"), bg="white", fg="#334155").pack(side="left")
        
        btn_style = {"font": ("Segoe UI", 9, "bold"), "relief": "flat", "cursor": "hand2", "padx": 10, "pady": 2}
        tk.Button(left_box, text="✏️ DÜZENLE", bg="#f1f5f9", fg="#64748b", **btn_style, command=lambda: self.karti_yeniden_adlandir(kart_adi)).pack(side="left", padx=(15, 5))
        
        # --- YENİ MANUEL EKSTRE GİRİŞİ VE ŞAHSİ HARCAMA ALANI ---
        right_box = tk.Frame(header_strip, bg="white")
        right_box.pack(side="right", padx=(0, 10))
        
        f_ekstre = tk.Frame(right_box, bg="white")
        f_ekstre.pack(anchor="e", pady=(0, 2))
        tk.Label(f_ekstre, text="Bu Ayki Ekstre (TL):", font=("Segoe UI", 10, "bold"), bg="white", fg="#64748b").pack(side="left")
        
        ent_bu_ay_ekstre = tk.Entry(f_ekstre, font=("Segoe UI", 11, "bold"), bg="#fff1f2", fg="#ef4444", width=12, justify="right", relief="solid", bd=1)
        ent_bu_ay_ekstre.pack(side="left", padx=(5, 5))
        ent_bu_ay_ekstre.bind("<KeyRelease>", mask_para_birimi)
        
        def ekstre_kaydet():
            bu_ay = date.today().strftime("%Y-%m")
            val_str = ent_bu_ay_ekstre.get().replace(".", "").replace(",", ".")
            try: ekstre_val = float(val_str) if val_str else 0.0
            except: ekstre_val = 0.0
            
            self.imlec_finans.execute("INSERT OR REPLACE INTO program_ayarlari (ayar_adi, deger) VALUES (?, ?)", (f"ekstre_toplam_{kart_adi}_{bu_ay}", str(ekstre_val)))
            self.baglanti_finans.commit()
            akordiyonlari_yukle() # Rakamları anında yenile
            messagebox.showinfo("Başarılı", "Ekstre tutarı güncellendi ve şahsi harcamalar hesaplandı.")

        ModernButton(f_ekstre, text="💾 KAYDET", command=ekstre_kaydet, width=80, height=24, bg_color="#3b82f6", font=("Segoe UI", 8, "bold")).pack(side="left")

        f_depo = tk.Frame(right_box, bg="white")
        f_depo.pack(anchor="e", pady=(0, 2))
        tk.Label(f_depo, text="Sistemdeki Eczane Gideri:", font=("Segoe UI", 10, "bold"), bg="white", fg="#64748b").pack(side="left")
        lbl_bu_ay_depo = tk.Label(f_depo, text="0.00 ₺", font=("Segoe UI", 11, "bold"), bg="white", fg="#10b981")
        lbl_bu_ay_depo.pack(side="left", padx=(5, 0))
        
        f_sahsi = tk.Frame(right_box, bg="white")
        f_sahsi.pack(anchor="e")
        tk.Label(f_sahsi, text="Eczane Dışı (Şahsi) Harcama:", font=("Segoe UI", 10, "bold"), bg="white", fg="#64748b").pack(side="left")
        lbl_sahsi_harcama = tk.Label(f_sahsi, text="0.00 ₺", font=("Segoe UI", 11, "bold"), bg="white", fg="#f59e0b")
        lbl_sahsi_harcama.pack(side="left", padx=(5, 0))
        # -----------------------------------------------------------------
        
        f_input = tk.LabelFrame(tab, text="Hızlı Ekle", bg="white", font=("Segoe UI", 9, "bold"), padx=5, pady=5)
        f_input.pack(fill="x", pady=(0, 10))
        grid_frame = tk.Frame(f_input, bg="white"); grid_frame.pack(fill="x")
        lbl_font = ("Segoe UI", 8, "bold")
        tk.Label(grid_frame, text="Firma/Depo:", bg="white", fg="#64748b", font=lbl_font).grid(row=0, column=0, padx=5, sticky="w")
        
        # --- YENİ EKLENEN SİL BUTONU VE COMBOBOX KAPSAYICISI ---
        f_depo_secim = tk.Frame(grid_frame, bg="white")
        f_depo_secim.grid(row=1, column=0, padx=5, pady=2, sticky="w")
        
        # Veritabanından listeyi çekerken [GİZLENDİ] olanları hariç tutuyoruz
        self.imlec_finans.execute("SELECT DISTINCT fatura_adi FROM odemeler WHERE fatura_adi NOT LIKE '%[GİZLENDİ]' UNION SELECT DISTINCT baslik FROM ozel_sekmeler WHERE baslik NOT LIKE '%[GİZLENDİ]'")
        
        kk_listesi = sorted([r[0] for r in self.imlec_finans.fetchall() if r[0] and not str(r[0]).endswith("[GİZLENDİ]")])
        if "DİĞER..." in kk_listesi: kk_listesi.remove("DİĞER...")
        kk_listesi.insert(0, "DİĞER...")
        
        cmb_depo = ttk.Combobox(f_depo_secim, values=kk_listesi, state="readonly", width=19)
        cmb_depo.pack(side="left")
        
        def hizli_ekle_listeden_cikar():
            secilen = cmb_depo.get()
            if not secilen or secilen == "DİĞER...": return
            
            if messagebox.askyesno("Listeden Kaldır", f"'{secilen}' kaydını bu hızlı menüden kaldırmak istiyor musunuz?\n\n(Not: Bu isme ait ödenmemiş taksitleriniz veya geçmiş kayıtlarınız silinmez, sadece bu menü temizlenir.)"):
                try:
                    # Ozel sekmelerden tamamen sil
                    self.imlec_finans.execute("DELETE FROM ozel_sekmeler WHERE baslik=?", (secilen,))
                    # Ödemelerdeki adını gizli olarak etiketle ki listede bir daha çıkmasın
                    self.imlec_finans.execute("UPDATE odemeler SET fatura_adi = ? WHERE fatura_adi=?", (f"{secilen} [GİZLENDİ]", secilen))
                    self.baglanti_finans.commit()
                    
                    # Combobox'ı anında güncelle
                    mevcut_degerler = list(cmb_depo["values"])
                    if secilen in mevcut_degerler:
                        mevcut_degerler.remove(secilen)
                        cmb_depo["values"] = mevcut_degerler
                    
                    cmb_depo.set("DİĞER...")
                    akordiyonlari_yukle() # Tabloyu da anında yenile
                    
                    messagebox.showinfo("Başarılı", "Kayıt hızlı ekleme listesinden başarıyla kaldırıldı.")
                except Exception as e:
                    messagebox.showerror("Hata", str(e))

        btn_depo_sil = tk.Button(f_depo_secim, text="🗑️", fg="red", bg="white", font=("Segoe UI", 8), relief="flat", cursor="hand2", command=hizli_ekle_listeden_cikar)
        btn_depo_sil.pack(side="left", padx=(2, 0))
        # --------------------------------------------------------
        
        def depo_secim_event(event):
            if cmb_depo.get() == "DİĞER...":
                def secim_tamam(gelen):
                    mevcut = list(cmb_depo['values'])
                    if gelen not in mevcut: mevcut.append(gelen); cmb_depo['values'] = sorted(mevcut)
                    cmb_depo.set(gelen)
                self.ozel_kk_depo_secici(secim_tamam)
        cmb_depo.bind("<<ComboboxSelected>>", depo_secim_event)
        
        tk.Label(grid_frame, text="Tutar (TL):", bg="white", fg="#64748b", font=lbl_font).grid(row=0, column=1, padx=5, sticky="w")
        ent_tutar = tk.Entry(grid_frame, relief="solid", bd=1, width=10); ent_tutar.grid(row=1, column=1, padx=5, pady=2); ent_tutar.bind("<KeyRelease>", mask_para_birimi)
        
        tk.Label(grid_frame, text="Taksit:", bg="white", fg="#64748b", font=lbl_font).grid(row=0, column=2, padx=5, sticky="w")
        cmb_taksit = ttk.Combobox(grid_frame, values=["1", "2", "3", "4", "5", "6", "9", "12"], state="readonly", width=4); cmb_taksit.grid(row=1, column=2, padx=5, pady=2); cmb_taksit.current(0)
        
        def create_date_wrapper(parent, r, c):
            wrapper = tk.Frame(parent, bg="white", bd=1, relief="solid"); wrapper.grid(row=r, column=c, padx=5, pady=2, sticky="ew")
            entry = tk.Entry(wrapper, relief="flat", width=9); entry.pack(side="left", fill="both", expand=True); entry.bind("<KeyRelease>", mask_tarih_otomatik); tarih_secici_bagla(wrapper, entry)
            return entry
            
        tk.Label(grid_frame, text="İşlem Tar:", bg="white", fg="#64748b", font=lbl_font).grid(row=0, column=3, padx=5, sticky="w")
        ent_islem = create_date_wrapper(grid_frame, 1, 3); ent_islem.insert(0, date.today().strftime("%Y-%m-%d"))
        
        tk.Label(grid_frame, text="Hesap Kesim:", bg="white", fg="#64748b", font=lbl_font).grid(row=0, column=4, padx=5, sticky="w")
        ent_kesim = create_date_wrapper(grid_frame, 1, 4)
        
        tk.Label(grid_frame, text="Son Ödeme:", bg="white", fg="#64748b", font=lbl_font).grid(row=0, column=5, padx=5, sticky="w")
        ent_vade = create_date_wrapper(grid_frame, 1, 5)
        try:
            dt_islem = date.today(); dt_vade = dt_islem + timedelta(days=30); dt_kesim = dt_vade - timedelta(days=10)
            ent_vade.insert(0, dt_vade.strftime("%Y-%m-%d")); ent_kesim.insert(0, dt_kesim.strftime("%Y-%m-%d"))
        except: pass

        lokal_treeviews = []

        def secilenleri_sil():
            ids = []
            for tv in lokal_treeviews:
                for item in tv.get_children():
                    vals = tv.item(item)['values']
                    if vals[1] == "☑": ids.append(vals[0])
            if not ids: 
                messagebox.showwarning("Uyarı", "Silinecek kayıt seçmediniz.")
                return
                
            if messagebox.askyesno("Silme Onayı", f"Seçili {len(ids)} kaydı silmek istiyor musunuz?"):
                for oid in ids: self.imlec_finans.execute("DELETE FROM odemeler WHERE id=?", (oid,))
                self.baglanti_finans.commit()
                akordiyonlari_yukle()

        def kismi_odeme_yap():
            secili_item = None; selected_tv = None
            for tv in lokal_treeviews:
                sel = tv.selection()
                if sel: secili_item = tv.item(sel[0]); selected_tv = tv; break
                
            if not secili_item: 
                messagebox.showwarning("Uyarı", "Lütfen işlem yapılacak satırın üzerine tıklayıp seçin (Mavi olsun).")
                return
                
            vals = secili_item['values']; oid = vals[0]
            res = self.imlec_finans.execute("SELECT tutar, satir_notu, fatura_adi FROM odemeler WHERE id=?", (oid,)).fetchone()
            if not res: return
            
            eski_borc = res[0]; fatura_adi = res[2]
            win = tk.Toplevel(self.pencere); win.title("Kısmi Ödeme"); win.geometry("400x450")
            win.transient(self.pencere); win.grab_set()
            x = self.pencere.winfo_x() + 100; y = self.pencere.winfo_y() + 100; win.geometry(f"+{x}+{y}")
            tk.Label(win, text="Kısmi Ödeme & Devir", font=("Segoe UI", 12, "bold")).pack(pady=20)
            tk.Label(win, text=f"Mevcut Borç: {eski_borc:,.2f} ₺", fg="red", font=("Segoe UI", 11, "bold")).pack()
            tk.Label(win, text="Ödenen Tutar:").pack(pady=(10,0)); e_ode = tk.Entry(win, font=("Segoe UI", 11)); e_ode.pack(pady=5); e_ode.bind("<KeyRelease>", mask_para_birimi)
            tk.Label(win, text="Faiz Oranı (%):").pack(pady=(10,0)); e_faiz = tk.Entry(win, font=("Segoe UI", 11)); e_faiz.pack(pady=5); e_faiz.insert(0, "4.25")
            
            def onayla():
                try:
                    odenen = temizle_para(e_ode.get()); oran = float(e_faiz.get().replace(",", "."))
                    kalan = eski_borc - odenen
                    if odenen > eski_borc: messagebox.showerror("Hata", "Fazla ödeme!"); return
                    yeni_not = f"{res[1]} | [KISMİ: {odenen:,.2f} ödendi]"
                    self.imlec_finans.execute("UPDATE odemeler SET durum='ODENDİ', satir_notu=? WHERE id=?", (yeni_not, oid))
                    if kalan > 0.01:
                        faiz = kalan * (oran/100); toplam_yeni = kalan + faiz
                        yeni_vade = (date.today() + timedelta(days=30)).strftime("%Y-%m-%d")
                        self.imlec_finans.execute("INSERT INTO odemeler (fatura_adi, vade_tarihi, tutar, aciklama, durum, satir_notu) VALUES (?, ?, ?, ?, 'ODENMEDİ', ?)",
                                           (fatura_adi, yeni_vade, toplam_yeni, f"KART: {kart_adi} (DEVİR)", f"Anapara:{kalan:.2f}+Faiz:{faiz:.2f}"))
                    self.baglanti_finans.commit(); win.destroy(); akordiyonlari_yukle()
                except Exception as e: messagebox.showerror("Hata", str(e))
            ModernButton(win, text="ONAYLA", command=onayla, width=120, bg_color="#10b981").pack(pady=20)

        def kaydet_tetik():
            basarili = self.taksit_hesapla_kaydet(kart_adi, cmb_depo.get(), ent_tutar.get(), cmb_taksit.get(), ent_islem.get(), ent_kesim.get(), ent_vade.get())
            if basarili: akordiyonlari_yukle(); ent_tutar.delete(0, tk.END); cmb_taksit.current(0)
            
        ModernButton(grid_frame, text="KAYDET", command=kaydet_tetik, width=90, height=28, bg_color="#10b981").grid(row=1, column=6, padx=15, pady=2)

        list_header = tk.Frame(tab, bg="white"); list_header.pack(fill="x", pady=(5, 5))
        tk.Label(list_header, text=f"📅 {kart_adi} - Ödeme Planı", font=("Segoe UI", 11, "bold"), bg="white", fg="#334155").pack(side="left")
        
        btn_action_frame = tk.Frame(list_header, bg="white"); btn_action_frame.pack(side="right")
        tk.Button(btn_action_frame, text="🗑️ SEÇİLENLERİ SİL", font=("Segoe UI", 9, "bold"), bg="#f59e0b", fg="white", relief="flat", cursor="hand2", padx=10, command=secilenleri_sil).pack(side="right", padx=5)
        tk.Button(btn_action_frame, text="💸 KISMİ ÖDEME", font=("Segoe UI", 9, "bold"), bg="#6b21a8", fg="white", relief="flat", cursor="hand2", padx=10, command=kismi_odeme_yap).pack(side="right", padx=5)
        tk.Button(btn_action_frame, text="⛔ KARTI SİL", font=("Segoe UI", 9, "bold"), bg="#ef4444", fg="white", relief="flat", cursor="hand2", padx=10, command=lambda: self.karti_sil(kart_adi)).pack(side="right", padx=5)

        canvas = tk.Canvas(tab, bg="white", highlightthickness=0); scrollbar = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg="white")
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        cv_id = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.bind('<Configure>', lambda e: canvas.itemconfig(cv_id, width=e.width))
        canvas.configure(yscrollcommand=scrollbar.set); canvas.pack(side="left", fill="both", expand=True, pady=5); scrollbar.pack(side="right", fill="y")
        self.mouse_scroll_ekle(canvas, scroll_frame)

        def akordiyonlari_yukle():
            for w in scroll_frame.winfo_children(): w.destroy()
            lokal_treeviews.clear() 
            
            # --- BU AYIN TOPLAMLARINI HESAPLA VE GÖSTER ---
            bu_ay = date.today().strftime("%Y-%m")
            
            # 1. Tüm Ekstre (Veritabanına manuel kaydedileni çek)
            self.imlec_finans.execute("SELECT deger FROM program_ayarlari WHERE ayar_adi=?", (f"ekstre_toplam_{kart_adi}_{bu_ay}",))
            res_ekstre = self.imlec_finans.fetchone()
            if res_ekstre and res_ekstre[0]:
                bu_ay_ekstre = float(res_ekstre[0])
            else:
                bu_ay_ekstre = 0.0
                
            ent_bu_ay_ekstre.delete(0, tk.END)
            if bu_ay_ekstre > 0:
                ent_bu_ay_ekstre.insert(0, f"{bu_ay_ekstre:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            
            # 2. Sadece Eczane/Depo Ödemeleri (Firma adında KK, DEPO, ECZA, FARMAZON vb. geçenler)
            self.imlec_finans.execute("""
                SELECT SUM(tutar) FROM odemeler 
                WHERE satir_notu LIKE ? AND strftime('%Y-%m', vade_tarihi) = ? AND durum='ODENMEDİ'
                AND (fatura_adi LIKE '%KK%' OR fatura_adi LIKE '%DEPO%' OR fatura_adi LIKE '%ECZA%' OR fatura_adi LIKE '%KOOP%' OR fatura_adi LIKE '%FARMAZON%' OR fatura_adi LIKE '%ALLIANCE%' OR fatura_adi LIKE '%BEK%')
            """, (f"KART: {kart_adi}%", bu_ay))
            res_depo = self.imlec_finans.fetchone()
            bu_ay_depo = res_depo[0] if res_depo and res_depo[0] else 0.0
            lbl_bu_ay_depo.config(text=f"{bu_ay_depo:,.2f} ₺")
            
            # 3. Eczane Dışı (Şahsi) Harcama Hesaplama
            sahsi = bu_ay_ekstre - bu_ay_depo
            if sahsi < 0: sahsi = 0.0
            lbl_sahsi_harcama.config(text=f"{sahsi:,.2f} ₺")
            # -------------------------------------------
            
            self.imlec_finans.execute("""
                SELECT strftime('%Y-%m', vade_tarihi) as donem, SUM(tutar) 
                FROM odemeler WHERE satir_notu LIKE ? AND durum='ODENMEDİ'
                GROUP BY donem ORDER BY donem ASC
            """, (f"KART: {kart_adi}%",))
            donemler = self.imlec_finans.fetchall()
            aylar_tr = {"01": "OCAK", "02": "ŞUBAT", "03": "MART", "04": "NİSAN", "05": "MAYIS", "06": "HAZİRAN", "07": "TEMMUZ", "08": "AĞUSTOS", "09": "EYLÜL", "10": "EKİM", "11": "KASIM", "12": "ARALIK"}
            if not donemler: tk.Label(scroll_frame, text="✅ Gelecek borç bulunmuyor.", font=("Segoe UI", 10, "italic"), bg="white", fg="gray").pack(pady=10)

            def _scroll_canvas(event):
                try: canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                except: pass

            for yil_ay, toplam in donemler:
                if not yil_ay: continue
                yil, ay = yil_ay.split("-"); baslik_str = f"{aylar_tr.get(ay, ay)} {yil}"
                frame_ay = tk.Frame(scroll_frame, bg="white", bd=1, relief="solid"); frame_ay.pack(fill="x", pady=5, padx=5)
                header = tk.Frame(frame_ay, bg="#f1f5f9", padx=10, pady=10, cursor="hand2"); header.pack(fill="x")
                lbl_ok = tk.Label(header, text="▶", font=("Segoe UI", 10), bg="#f1f5f9", fg="#64748b"); lbl_ok.pack(side="left", padx=(0, 10))
                tk.Label(header, text=baslik_str, font=("Segoe UI", 11, "bold"), bg="#f1f5f9", fg="#334155").pack(side="left")
                tk.Label(header, text=f"{toplam:,.2f} ₺", font=("Segoe UI", 11, "bold"), bg="#f1f5f9", fg="#ef4444").pack(side="right")
                
                content = tk.Frame(frame_ay, bg="white", padx=5, pady=5)
                cols = ("ID", "TIK", "GUN", "DEPO", "TUTAR", "ACIKLAMA"); tv = ttk.Treeview(content, columns=cols, show="headings", height=0)
                
                tv.tumunu_sec_durum = False
                def tumunu_sec_toggle(tree_ref=tv):
                    durum = getattr(tree_ref, "tumunu_sec_durum", False)
                    yeni_durum = not durum
                    setattr(tree_ref, "tumunu_sec_durum", yeni_durum)
                    ikon = "☑" if yeni_durum else "☐"
                    tree_ref.heading("TIK", text=ikon)
                    for item in tree_ref.get_children():
                        vals = list(tree_ref.item(item, "values"))
                        vals[1] = ikon
                        tree_ref.item(item, values=vals)

                tv.heading("TIK", text="☐", command=lambda t=tv: tumunu_sec_toggle(t))
                tv.column("TIK", width=30, anchor="center")
                
                tv.heading("GUN", text="Gün"); tv.column("GUN", width=40, anchor="center")
                tv.heading("DEPO", text="Firma"); tv.column("DEPO", width=150)
                tv.heading("TUTAR", text="Tutar"); tv.column("TUTAR", width=100, anchor="e")
                tv.heading("ACIKLAMA", text="Detay"); tv.column("ACIKLAMA", width=250)
                tv.column("ID", width=0, stretch=False); tv.pack(fill="x")
                
                tv.bind('<MouseWheel>', _scroll_canvas)
                
                self.imlec_finans.execute("""
                    SELECT id, strftime('%d', vade_tarihi), fatura_adi, tutar, satir_notu 
                    FROM odemeler WHERE satir_notu LIKE ? AND strftime('%Y-%m', vade_tarihi) = ? AND durum='ODENMEDİ' ORDER BY vade_tarihi ASC
                """, (f"KART: {kart_adi}%", yil_ay))
                items = self.imlec_finans.fetchall()
                tv.configure(height=len(items)) 
                for row in items:
                    not_temiz = row[4].replace(f"KART: {kart_adi}", "").strip()
                    tv.insert("", "end", values=(row[0], "☐", row[1], row[2], f"{row[3]:,.2f} ₺", not_temiz))
                
                lokal_treeviews.append(tv)
                
                def toggle(f=content, i=lbl_ok):
                    if f.winfo_viewable(): f.forget(); i.config(text="▶")
                    else: f.pack(fill="x"); i.config(text="▼")
                for w in [header, lbl_ok] + header.winfo_children(): w.bind("<Button-1>", lambda e, f=content, i=lbl_ok: toggle(f, i))
                
                def on_click(event, tree_ref=tv):
                    region = tree_ref.identify("region", event.x, event.y)
                    if region == "cell" and tree_ref.identify_column(event.x) == "#2":
                        item = tree_ref.identify_row(event.y)
                        if item: 
                            vals = list(tree_ref.item(item, "values"))
                            if len(vals) > 1: 
                                vals[1] = "☑" if vals[1] == "☐" else "☐"
                                tree_ref.item(item, values=vals)
                tv.bind("<Button-1>", on_click) 

            tk.Label(scroll_frame, text="---- GEÇMİŞ DÖNEMLER ----", font=("Segoe UI", 9, "bold"), fg="#cbd5e1", bg="white").pack(pady=20)
            arsiv_frame = tk.Frame(scroll_frame, bg="white", bd=1, relief="solid"); arsiv_frame.pack(fill="x", padx=5, pady=5)
            arsiv_head = tk.Frame(arsiv_frame, bg="#e2e8f0", padx=10, pady=10, cursor="hand2"); arsiv_head.pack(fill="x")
            lbl_ars_icon = tk.Label(arsiv_head, text="▶", font=("Segoe UI", 10), bg="#e2e8f0"); lbl_ars_icon.pack(side="left", padx=5)
            tk.Label(arsiv_head, text="GEÇMİŞ 2 YIL (ÖDENENLER)", font=("Segoe UI", 10, "bold"), bg="#e2e8f0", fg="#475569").pack(side="left")
            arsiv_content = tk.Frame(arsiv_frame, bg="white")
            tv_ars = ttk.Treeview(arsiv_content, columns=("TARIH", "DEPO", "TUTAR"), show="headings", height=10)
            tv_ars.heading("TARIH", text="Vade"); tv_ars.column("TARIH", width=100)
            tv_ars.heading("DEPO", text="Firma"); tv_ars.column("DEPO", width=150)
            tv_ars.heading("TUTAR", text="Tutar"); tv_ars.column("TUTAR", width=100, anchor="e"); tv_ars.pack(fill="x")
            tv_ars.tag_configure("odendi", foreground="#166534")
            
            tv_ars.bind('<MouseWheel>', _scroll_canvas)
            
            iki_yil_once = (date.today() - timedelta(days=730)).strftime("%Y-%m-%d")
            self.imlec_finans.execute("""
                SELECT vade_tarihi, fatura_adi, tutar FROM odemeler 
                WHERE satir_notu LIKE ? AND durum='ODENDİ' AND vade_tarihi > ? ORDER BY vade_tarihi DESC
            """, (f"KART: {kart_adi}%", iki_yil_once))
            for row in self.imlec_finans.fetchall(): tv_ars.insert("", "end", values=(row[0], row[1], f"{row[2]:,.2f} ₺"), tags=("odendi",))
            
            def toggle_arsiv(f=arsiv_content, i=lbl_ars_icon):
                if f.winfo_viewable(): f.forget(); i.config(text="▶")
                else: f.pack(fill="x"); i.config(text="▼")
            for w in [arsiv_head, lbl_ars_icon] + arsiv_head.winfo_children(): w.bind("<Button-1>", lambda e: toggle_arsiv())

        akordiyonlari_yukle()


    def karti_yeniden_adlandir(self, eski_ad):
        yeni_ad = simpledialog.askstring("Kartı Düzenle", f"'{eski_ad}' için yeni isim giriniz:", initialvalue=eski_ad)
        if not yeni_ad or yeni_ad == eski_ad: return 
        yeni_ad = yeni_ad.upper().replace("i", "İ").replace("ı", "I")
        if messagebox.askyesno("Onay", f"Kart adı '{eski_ad}' -> '{yeni_ad}' olarak değiştirilecek.\nTüm eski taksit kayıtları da güncellenecek.\nOnaylıyor musun?"):
            try:
                self.imlec_finans.execute("UPDATE kredi_kartlari SET isim=? WHERE isim=?", (yeni_ad, eski_ad))
                eski_prefix = f"KART: {eski_ad}"
                yeni_prefix = f"KART: {yeni_ad}"
                self.imlec_finans.execute("""
                    UPDATE odemeler SET satir_notu = REPLACE(satir_notu, ?, ?) WHERE satir_notu LIKE ?
                """, (eski_prefix, yeni_prefix, f"{eski_prefix}%"))
                self.baglanti_finans.commit()
                messagebox.showinfo("Başarılı", "Kart ismi ve bağlantılı kayıtlar güncellendi.")
                self.kredi_karti_arayuzunu_guncelle(hedef_sekme=yeni_ad)
            except sqlite3.IntegrityError: messagebox.showerror("Hata", "Bu isimde başka bir kart zaten var. Lütfen farklı bir isim deneyin.")
            except Exception as e: self.baglanti_finans.rollback(); messagebox.showerror("Hata", f"İsim değiştirme hatası:\n{e}")            

    def taksitleri_veritabanina_isle(self, kart_adi, depo, tutar_str, taksit_str, islem_tarihi_str, son_odeme_tarihi_str, kesim_gunu_str):
        # =================================================================
        # HATA ÇÖZÜMÜ: GİZLİ YENİLEME SİNYALİNİ YAKALA (POPUP ÇIKMASINI ÖNLER)
        # =================================================================
        if kart_adi == "DUMMY_REFRESH":
            self.kredi_karti_arayuzunu_guncelle()
            return
            
        if not depo: messagebox.showerror("Hata", "Depo seçilmedi."); return
        try: aylik = float(tutar_str.replace(".", "").replace(",", ".")) if "." in tutar_str and "," in tutar_str else float(tutar_str.replace(",", "."))
        except: messagebox.showerror("Hata", "Tutar hatalı."); return
        try: taksit = int(taksit_str)
        except: return
        try: 
            islem_dt = datetime.strptime(islem_tarihi_str, "%Y-%m-%d").date()
            referans_odeme_dt = datetime.strptime(son_odeme_tarihi_str, "%Y-%m-%d").date()
        except: messagebox.showerror("Hata", "Tarih formatı YIL-AY-GÜN olmalı."); return

        odeme_gunu_rakam = referans_odeme_dt.day
        toplam_borc = aylik * taksit
        
        if not messagebox.askyesno("Onay", f"{kart_adi} - {depo}\n\nİşlem Tarihi: {islem_tarihi_str}\nTaksit Başlangıcı: {islem_dt.month + 1}. Ay (Ayın {odeme_gunu_rakam}. günü)\n\nTutar: {aylik:,.2f} TL x {taksit}\nToplam: {toplam_borc:,.2f} TL\n\nOnaylıyor musun?"): return
        
        for i in range(taksit):
            eklenecek_ay = i + 1
            hedef_yil = islem_dt.year + (islem_dt.month + eklenecek_ay - 1) // 12
            hedef_ay = (islem_dt.month + eklenecek_ay - 1) % 12 + 1
            
            import calendar
            ayin_son_gunu = calendar.monthrange(hedef_yil, hedef_ay)[1]
            hedef_gun = min(odeme_gunu_rakam, ayin_son_gunu)
            
            from datetime import date, timedelta
            odeme_tarihi = date(hedef_yil, hedef_ay, hedef_gun)
            
            if odeme_tarihi.weekday() == 5: odeme_tarihi += timedelta(days=2)
            elif odeme_tarihi.weekday() == 6: odeme_tarihi += timedelta(days=1)
            
            vade = odeme_tarihi.strftime("%Y-%m-%d")
            aciklama = f"KART: {kart_adi} ({i+1}/{taksit})" if taksit > 1 else f"KART: {kart_adi}"
            
            self.imlec_finans.execute("INSERT INTO odemeler (fatura_adi, vade_tarihi, tutar, aciklama, durum, alim_tarihi, satir_notu) VALUES (?, ?, ?, ?, 'ODENMEDİ', ?, ?)",
                                      (depo, vade, aylik, "Kredi Kartı", islem_tarihi_str, aciklama))
                                      
        self.baglanti_finans.commit()
        messagebox.showinfo("Başarılı", "Taksitler işlendi.")
        self.kredi_karti_arayuzunu_guncelle(hedef_sekme=kart_adi)

    def yeni_kart_ekle_popup(self):
        popup = tk.Toplevel(self.pencere); popup.title("Kredi Kartı Ekle"); popup.geometry("380x600"); popup.configure(bg="#f8fafc")
        popup.transient(self.pencere); popup.grab_set(); popup.focus_force()
        x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 190
        y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 300
        popup.geometry(f"+{x}+{y}")
        tk.Label(popup, text="Kart Seçimi", font=("Segoe UI", 12, "bold"), bg="#f8fafc", fg="#334155").pack(pady=(20, 10))
        container = tk.Frame(popup, bg="#f8fafc"); container.pack(side="top", fill="both", expand=True, padx=5)
        canvas = tk.Canvas(container, bg="#f8fafc", highlightthickness=0); scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg="#f8fafc")
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        window_id = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(window_id, width=e.width)); canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True); scrollbar.pack(side="right", fill="y")
        self.mouse_scroll_ekle(canvas, scroll_frame)

        hazir_kartlar = ["BONUS (Garanti)", "AXESS (Akbank)", "WORLD (Yapı Kredi)", "MAXIMUM (İş Bankası)", "CARDFINANS (QNB)", "PARAF (Halkbank)", "BANKKART (Ziraat)", "SAĞLAM KART (Kuveyt Türk)", "ADVANTAGE (HSBC)", "WINGS (Akbank)", "SHOP&FLY (Garanti)", "MILES&SMILES (Garanti)", "TİCARİ KART", "ŞİRKET KARTI"]
        def kart_ekle_islem(isim):
            isim = isim.upper().replace("i", "İ").replace("ı", "I").strip()
            try:
                self.imlec_finans.execute("INSERT INTO kredi_kartlari (isim, kesim_gunu) VALUES (?, ?)", (isim, 1))
                self.baglanti_finans.commit(); popup.destroy()
                self.kredi_karti_arayuzunu_guncelle(hedef_sekme=isim)
                messagebox.showinfo("Başarılı", f"'{isim}' başarıyla eklendi.")
            except sqlite3.IntegrityError: messagebox.showerror("Mevcut", f"'{isim}' adında bir kart zaten ekli!", parent=popup)
            except Exception as e: messagebox.showerror("Hata", str(e), parent=popup)

        for k in hazir_kartlar:
            tk.Button(scroll_frame, text=k, font=("Segoe UI", 10), bg="white", fg="#0f172a", relief="solid", bd=1, cursor="hand2", pady=8, activebackground="#e0f2fe", command=lambda isim=k: kart_ekle_islem(isim)).pack(fill="x", padx=15, pady=3)
        tk.Frame(popup, height=2, bg="#cbd5e1").pack(fill="x", padx=30, pady=10)
        
        def manuel_giris_ac():
            popup.destroy()
            inp_win = tk.Toplevel(self.pencere); inp_win.title("Manuel Kart Ekle"); inp_win.geometry("400x200"); inp_win.configure(bg="white")
            inp_win.transient(self.pencere); inp_win.grab_set()
            ix = self.pencere.winfo_x() + (self.pencere.winfo_width()//2) - 200
            iy = self.pencere.winfo_y() + (self.pencere.winfo_height()//2) - 100
            inp_win.geometry(f"+{ix}+{iy}")
            tk.Label(inp_win, text="Kart Adını Giriniz", font=("Segoe UI", 12, "bold"), bg="white", fg="#334155").pack(pady=20)
            e_isim = tk.Entry(inp_win, font=("Segoe UI", 12), relief="solid", bd=1, bg="#f8fafc"); e_isim.pack(fill="x", padx=30, ipady=5); e_isim.focus_set()
            def onayla():
                val = e_isim.get().strip()
                if val: kart_ekle_islem(val); inp_win.destroy()
            inp_win.bind('<Return>', lambda e: onayla()); ModernButton(inp_win, text="KAYDET", command=onayla, width=150, height=35, bg_color="#10b981").pack(pady=20)

        tk.Button(popup, text="✍️ DİĞER (Manuel Yaz)", font=("Segoe UI", 11, "bold"), bg="#3b82f6", fg="white", relief="flat", cursor="hand2", pady=10, command=manuel_giris_ac).pack(fill="x", padx=30, pady=(0, 20))

    def taksit_hesapla_kaydet(self, kart_adi, depo, tutar_str, taksit_str, islem_tar, kesim_tar, son_odeme_tar):
        try:
            tutar = float(tutar_str.replace(".", "").replace(",", "."))
            taksit_sayisi = int(taksit_str)
            try:
                dt_islem = datetime.strptime(islem_tar, "%Y-%m-%d").date()
                dt_son_odeme = datetime.strptime(son_odeme_tar, "%Y-%m-%d").date()
            except ValueError: messagebox.showerror("Hata", "Tarih formatı hatalı! (Yıl-Ay-Gün)"); return False
            aylik_tutar = tutar / taksit_sayisi
            sutunlar = [row[1] for row in self.imlec_finans.execute("PRAGMA table_info(odemeler)")]
            islem_tarihi_sutunu_var = "islem_tarihi" in sutunlar
            
            for i in range(taksit_sayisi):
                yil = dt_son_odeme.year; ay = dt_son_odeme.month + i
                while ay > 12: ay -= 12; yil += 1
                
                # 1. O ayın kaç gün çektiğini bul (Şubat 30 hatasını engeller)
                ayin_son_gunu = calendar.monthrange(yil, ay)[1]
                hedef_gun = min(dt_son_odeme.day, ayin_son_gunu)
                
                vade_tarihi = date(yil, ay, hedef_gun)
                
                # 2. Hafta sonu kontrolü (5 = Cumartesi, 6 = Pazar)
                if vade_tarihi.weekday() == 5: 
                    vade_tarihi += timedelta(days=2) # Cumartesi ise Pazartesi yap
                elif vade_tarihi.weekday() == 6: 
                    vade_tarihi += timedelta(days=1) # Pazar ise Pazartesi yap
                    
                yeni_vade_tarihi = vade_tarihi.strftime("%Y-%m-%d")
                
                taksit_bilgisi = f"({i+1}/{taksit_sayisi})"
                aciklama = f"KART: {kart_adi} - {taksit_bilgisi}"
                
                if islem_tarihi_sutunu_var:
                    sql = """INSERT INTO odemeler (fatura_adi, tutar, vade_tarihi, islem_tarihi, durum, satir_notu) VALUES (?, ?, ?, ?, 'ODENMEDİ', ?)"""
                    parametreler = (depo, aylik_tutar, yeni_vade_tarihi, islem_tar, aciklama)
                else:
                    aciklama_genis = f"{aciklama} [İşlem: {islem_tar}]"
                    sql = """INSERT INTO odemeler (fatura_adi, tutar, vade_tarihi, durum, satir_notu) VALUES (?, ?, ?, 'ODENMEDİ', ?)"""
                    parametreler = (depo, aylik_tutar, yeni_vade_tarihi, aciklama_genis)
                self.imlec_finans.execute(sql, parametreler)
            self.baglanti_finans.commit()
            messagebox.showinfo("Başarılı", f"{taksit_sayisi} taksit başarıyla kaydedildi.")
            return True
        except Exception as e:
            self.baglanti_finans.rollback()
            messagebox.showerror("Hata", f"Kayıt hatası: {e}")
            return False
            
    def karti_sil(self, kart_adi):
        if messagebox.askyesno("Onay", f"'{kart_adi}' isimli kartı ve ona ait TÜM kayıtları silmek istediğinize emin misiniz?"):
            self.imlec_finans.execute("DELETE FROM kredi_kartlari WHERE isim=?", (kart_adi,))
            self.imlec_finans.execute("DELETE FROM odemeler WHERE satir_notu LIKE ?", (f"KART: {kart_adi}%",))
            self.baglanti_finans.commit()
            self.kredi_karti_arayuzunu_guncelle()

    # --- 3. KURUM ÖDEMELERİ ---
    def sayfa_kurum_odemeleri(self, hedef_sekme=None):
        # --- YENİ: SEKME HAFIZASI (Yenile yaparken sayfayı kaybetmemek için) ---
        if hedef_sekme is None:
            if hasattr(self, 'notebook_kurum') and self.notebook_kurum.winfo_exists():
                try:
                    aktif_id = self.notebook_kurum.select()
                    hedef_sekme = self.notebook_kurum.tab(aktif_id, "text")
                except:
                    pass

        # Ekranı temizle
        for w in self.content_area.winfo_children(): w.destroy()
        c = TM
        
        header = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        header.pack(fill="x", pady=(0, 20))
        tk.Label(header, text="Kurum ve Sabit Gider Yönetimi", font=FONT_HEAD, bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left")
        
        frm_btns = tk.Frame(header, bg=c.get_color("bg_main"))
        frm_btns.pack(side="right")
        
        ModernButton(frm_btns, text="+ YENİ KURUM TANIMLA", command=self.yeni_kurum_ekle_popup, width=220, height=40, bg_color="#f59e0b").pack(side="left", padx=10)
        ModernButton(frm_btns, text="🔄 YENİLE", command=lambda: self.sayfa_kurum_odemeleri(), width=130, height=40, bg_color="#64748b").pack(side="left", padx=10)

        self.notebook_kurum = DraggableNotebook(self.content_area, on_reorder=self.kurum_sirasi_kaydet)
        self.notebook_kurum.pack(fill="both", expand=True, pady=10)
        
        # Hafızadaki hedef sekmeyi aç
        self.kurum_sekmelerini_guncelle(hedef_sekme)

    def kurum_sirasi_kaydet(self, tab_names):
        pass 

    def kurum_sekmelerini_guncelle(self, hedef_sekme=None):
        for tab in self.notebook_kurum.tabs(): self.notebook_kurum.forget(tab)
        tab_ozet = tk.Frame(self.notebook_kurum, bg=TM.get_color("bg_main"))
        self.notebook_kurum.add(tab_ozet, text="GENEL ÖZET")
        self.kurum_ozet_sayfasi(tab_ozet)
        
        # HATA DÜZELTİLDİ: imlec_finans yapıldı
        self.imlec_finans.execute("SELECT isim FROM kurumlar ORDER BY isim")
        kurumlar = [r[0] for r in self.imlec_finans.fetchall()]
        for k in kurumlar: self.kurum_sekmesi_olustur(k)

        if hedef_sekme:
            for tab_id in self.notebook_kurum.tabs():
                if self.notebook_kurum.tab(tab_id, "text") == hedef_sekme:
                    self.notebook_kurum.select(tab_id)
                    return

    def yeni_kurum_ekle_popup(self):
        popup = tk.Toplevel(self.pencere); popup.title("Yeni Kurum Ekle"); popup.geometry("380x600"); popup.configure(bg="#f8fafc")
        popup.transient(self.pencere); popup.grab_set(); popup.focus_force()
        x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 190
        y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 300
        popup.geometry(f"+{x}+{y}")

        tk.Label(popup, text="Hızlı Kurum Seçimi", font=("Segoe UI", 12, "bold"), bg="#f8fafc", fg="#334155").pack(pady=(20, 10))

        container = tk.Frame(popup, bg="#f8fafc"); container.pack(side="top", fill="both", expand=True, padx=5)
        canvas = tk.Canvas(container, bg="#f8fafc", highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg="#f8fafc")

        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        window_id = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(window_id, width=e.width))
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.mouse_scroll_ekle(canvas, scroll_frame)

        hazir_kurumlar = ["SGK (Sosyal Güvenlik Kurumu)", "İŞ BANKASI", "VAKIFBANK", "AKBANK", "OYAK", "ALLIANZ SİGORTA", "MAPFRE SİGORTA", "ANADOLU HAYAT EMEKLİLİK", "TÜRKİYE SİGORTA", "AXA SİGORTA", "ACIBADEM SİGORTA"]

        def ekle_ve_kapat(isim):
            if isim == "DİĞER":
                manuel_giris_ac(); return
            isim = isim.upper().replace("i", "İ").replace("ı", "I").strip()
            try:
                # HATA DÜZELTİLDİ: Tabloyu finans veritabanında oluşturup ekliyoruz
                self.imlec_finans.execute("CREATE TABLE IF NOT EXISTS kurumlar (isim TEXT PRIMARY KEY)")
                self.imlec_finans.execute("INSERT INTO kurumlar (isim) VALUES (?)", (isim,))
                self.baglanti_finans.commit()
                popup.destroy()
                self.sayfa_kurum_odemeleri(hedef_sekme=isim)
            except sqlite3.IntegrityError: messagebox.showerror("Hata", f"'{isim}' zaten listenizde var.", parent=popup)
            except Exception as e: messagebox.showerror("Hata", str(e), parent=popup)

        for k in hazir_kurumlar:
            btn_bg = "#fee2e2" if "SGK" in k else "white"
            btn_fg = "#991b1b" if "SGK" in k else "#0f172a"
            btn = tk.Button(scroll_frame, text=k, font=("Segoe UI", 10), bg=btn_bg, fg=btn_fg, relief="solid", bd=1, cursor="hand2", pady=8, activebackground="#e0f2fe", command=lambda isim=k: ekle_ve_kapat(isim))
            btn.pack(fill="x", padx=15, pady=3)

        tk.Frame(popup, height=2, bg="#cbd5e1").pack(fill="x", padx=30, pady=10)

        def manuel_giris_ac():
            popup.destroy()
            isim = simpledialog.askstring("Manuel Giriş", "Kurum veya Gider Adı:")
            if isim: ekle_ve_kapat(isim)

        tk.Button(popup, text="✍️ DİĞER (Manuel Yaz)", font=("Segoe UI", 11, "bold"), bg="#3b82f6", fg="white", relief="flat", cursor="hand2", pady=10, command=manuel_giris_ac).pack(fill="x", padx=30, pady=(0, 20))

    def kurum_ozet_sayfasi(self, parent):
        # 1. Eski "Gider" olarak kalmış kayıtları otomatik "Gelir"e dönüştüren yama
        try:
            self.imlec_finans.execute("UPDATE odemeler SET satir_notu = 'KURUM_GELIR' WHERE satir_notu LIKE '%KURUM%'")
            self.baglanti_finans.commit()
        except: pass

        c = TM
        stats_frame = tk.Frame(parent, bg=c.get_color("bg_main")); stats_frame.pack(fill="x", pady=20, padx=20)
        
        self.imlec_finans.execute("SELECT SUM(tutar) FROM odemeler WHERE satir_notu = 'KURUM_GELIR' AND durum='ODENMEDİ'")
        res = self.imlec_finans.fetchone()
        toplam_kurum_alacagi = res[0] if res and res[0] else 0.0

        card = tk.Frame(stats_frame, bg="white", bd=1, relief="solid", padx=20, pady=20); card.pack(fill="x")
        tk.Label(card, text="BEKLEYEN TOPLAM KURUM GELİRİ", font=("Segoe UI", 12, "bold"), bg="white", fg="#64748b").pack(anchor="center")
        tk.Label(card, text=f"{toplam_kurum_alacagi:,.2f} ₺", font=("Segoe UI", 24, "bold"), bg="white", fg="#10b981").pack(anchor="center")
        tk.Label(card, text="(SGK ve Diğer Kurum Alacakları)", font=("Segoe UI", 9), bg="white", fg="#94a3b8").pack(anchor="center")

        list_frame = tk.Frame(parent, bg="white", padx=10, pady=10); list_frame.pack(fill="both", expand=True, padx=20, pady=10)
        tk.Label(list_frame, text="Kurum Bazlı Gelir Dağılımı", font=("Segoe UI", 12, "bold"), bg="white", fg="#334155").pack(anchor="w", pady=(0, 10))

        cols = ("KURUM", "ISLEM_ADET", "TOPLAM_TUTAR", "EN_YAKIN_ODEME")
        tree = ttk.Treeview(list_frame, columns=cols, show="headings", height=15)
        tree.heading("KURUM", text="Kurum Adı"); tree.column("KURUM", width=250)
        tree.heading("ISLEM_ADET", text="Bekleyen Fatura"); tree.column("ISLEM_ADET", width=120, anchor="center")
        tree.heading("TOPLAM_TUTAR", text="Toplam Alacak"); tree.column("TOPLAM_TUTAR", width=150, anchor="e")
        tree.heading("EN_YAKIN_ODEME", text="Ödeme Tarihi"); tree.column("EN_YAKIN_ODEME", width=150, anchor="center")
        tree.pack(fill="both", expand=True)

        sorgu = """
            SELECT fatura_adi, COUNT(*), SUM(tutar), MIN(vade_tarihi)
            FROM odemeler WHERE satir_notu = 'KURUM_GELIR' AND durum='ODENMEDİ'
            GROUP BY fatura_adi ORDER BY SUM(tutar) DESC
        """
        self.imlec_finans.execute(sorgu)
        for row in self.imlec_finans.fetchall():
            kurum, adet, tutar, en_yakin = row
            try: tarih_str = datetime.strptime(en_yakin, "%Y-%m-%d").strftime("%d.%m.%Y")
            except: tarih_str = en_yakin
            tree.insert("", "end", values=(kurum, f"{adet} Adet", f"{tutar:,.2f} ₺", tarih_str))

    def kurum_sekmesi_olustur(self, kurum_adi):
        c = TM
        tab = tk.Frame(self.notebook_kurum, bg="white", padx=10, pady=10)
        self.notebook_kurum.add(tab, text=kurum_adi)
        
        bar = tk.Frame(tab, bg="white", pady=5); bar.pack(fill="x", pady=(0, 10))

        cols = ("ID", "TIK", "FATURA", "VADE", "TUTAR", "ACIKLAMA")
        tree = ttk.Treeview(tab, columns=cols, show="headings", height=12)
        
        tree.heading("ID", text="ID"); tree.column("ID", width=0, stretch=False)
        tree.heading("TIK", text="✔"); tree.column("TIK", width=40, anchor="center")
        tree.heading("FATURA", text="Dönem/Fatura"); tree.column("FATURA", width=120, anchor="center")
        tree.heading("VADE", text="Ödeme Tarihi"); tree.column("VADE", width=120, anchor="center")
        tree.heading("TUTAR", text="Tutar"); tree.column("TUTAR", width=150, anchor="e")
        tree.heading("ACIKLAMA", text="Açıklama / Tür"); tree.column("ACIKLAMA", width=300)
        
        tree.tag_configure("gecmis", background="#fee2e2", foreground="#000000")
        tree.tag_configure("guncel", background="#ffffff", foreground="#000000") 
        # --- YENİ EKLENEN YEŞİL TAG (ÖDENDİ RENGİ) ---
        tree.tag_configure("odendi", background="#dcfce7", foreground="#166534", font=("Segoe UI", 10, "bold"))

        sc = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
        tree.configure(yscroll=sc.set); sc.pack(side="right", fill="y"); tree.pack(fill="both", expand=True)

        def tabloyu_yenile():
            for i in tree.get_children(): tree.delete(i)
            
            # DÜZELTME: durum='ODENMEDİ' şartı kaldırıldı!
            # Akıllı Sıralama: Önce Ödenmemişler (1), Sonra Ödenmişler (2) gelir.
            self.imlec_finans.execute("""
                SELECT id, alim_tarihi, vade_tarihi, tutar, aciklama, durum 
                FROM odemeler WHERE fatura_adi=? AND satir_notu = 'KURUM_GELIR' 
                ORDER BY CASE WHEN durum='ODENMEDİ' THEN 1 ELSE 2 END, vade_tarihi ASC
            """, (kurum_adi,))
            
            bugun = date.today()
            for row in self.imlec_finans.fetchall():
                oid, fatura, vade, tutar, ack, durum = row
                try: vade_dt = datetime.strptime(vade, "%Y-%m-%d").date()
                except: vade_dt = bugun
                
                vade_str = vade_dt.strftime("%d.%m.%Y")
                try: fatura_str = datetime.strptime(fatura, "%Y-%m-%d").strftime("%d.%m.%Y")
                except: fatura_str = fatura
                
                # --- DURUMA GÖRE İKON VE RENK BELİRLEME ---
                if durum == 'ODENDİ':
                    tag = "odendi"
                    tik = "✅"
                else:
                    tag = "gecmis" if vade_dt < bugun else "guncel"
                    tik = "☐"
                    
                tree.insert("", "end", values=(oid, tik, fatura_str, vade_str, f"{tutar:,.2f} ₺", ack), tags=(tag,))

        def manuel_ekle_ac():
            def veritabanina_ekle(k_adi, alim, vade, tutar, not_icerik, odeme_turu):
                aciklama_str = f"[{odeme_turu}]"
                if not_icerik: aciklama_str += f" - {not_icerik}"
                
                # SGK için otomatik vade hesaplaması
                if "SGK" in k_adi.upper() or "SOSYAL" in k_adi.upper():
                    try:
                        vade_dt = datetime.strptime(vade, "%Y-%m-%d")
                        vade_dt = vade_dt.replace(day=15)
                        if vade_dt.weekday() == 5:
                            vade_dt += timedelta(days=2) 
                        elif vade_dt.weekday() == 6:
                            vade_dt += timedelta(days=1) 
                        vade = vade_dt.strftime("%Y-%m-%d")
                    except: pass

                self.imlec_finans.execute("""
                    INSERT INTO odemeler (fatura_adi, vade_tarihi, tutar, aciklama, durum, alim_tarihi, satir_notu) 
                    VALUES (?, ?, ?, ?, 'ODENMEDİ', ?, 'KURUM_GELIR')
                """, (k_adi, vade, tutar, aciklama_str, alim))
                self.baglanti_finans.commit()
                tabloyu_yenile()
                messagebox.showinfo("Başarılı", "Kurum geliri sisteme eklendi.")

            # AKILLI YÖNLENDİRME: Kurum adı SGK ise detaylı pencereyi aç, değilse klasiği aç
            if "SGK" in kurum_adi.upper() or "SOSYAL" in kurum_adi.upper():
                SgkAylikFaturaSihirbazi(self.pencere, kurum_adi, veritabanina_ekle)
            else:
                ManualAddWindow(self.pencere, kurum_adi, veritabanina_ekle)

        ModernButton(bar, text="➕ YENİ GELİR EKLE", command=manuel_ekle_ac, width=180, height=35, bg_color="#10b981").pack(side="left", padx=5)
        
        def secilenleri_sil():
            ids = [tree.item(i)['values'][0] for i in tree.get_children() if tree.item(i)['values'][1] == "☑"]
            if not ids: return
            if messagebox.askyesno("Onay", f"{len(ids)} kaydı silmek istiyor musunuz?"):
                for oid in ids: self.imlec_finans.execute("DELETE FROM odemeler WHERE id=?", (oid,))
                self.baglanti_finans.commit()
                tabloyu_yenile()

        ModernButton(bar, text="🗑️ SEÇİLİLERİ SİL", command=secilenleri_sil, width=200, height=35, bg_color="#ef4444").pack(side="right", padx=5)
        ModernButton(bar, text="⛔ KURUMU SİL", command=lambda: self.kurumu_sil(kurum_adi), width=200, height=35, bg_color="#7f1d1d").pack(side="right", padx=5)

        # --- YATAN PARA KONTROL SİHİRBAZI ---
        bottom_bar = tk.Frame(tab, bg="white", pady=10)
        bottom_bar.pack(fill="x")

        def secilenleri_alindi_yap():
            ids = [tree.item(i)['values'][0] for i in tree.get_children() if tree.item(i)['values'][1] == "☑"]
            if not ids: 
                messagebox.showwarning("Uyarı", "Lütfen işlem yapılacak faturayı/faturaları seçin (☑).")
                return
            if messagebox.askyesno("Tahsilat Onayı", "Seçilen faturalar tahsil edildi mi? (Doğrudan Kapat)"):
                for oid in ids: self.imlec_finans.execute("UPDATE odemeler SET durum='ODENDİ' WHERE id=?", (oid,))
                self.baglanti_finans.commit()
                tabloyu_yenile()
                messagebox.showinfo("Başarılı", "Tahsilat kaydedildi.")

        def yatan_para_kontrol_ac():
            ids = [tree.item(i)['values'][0] for i in tree.get_children() if tree.item(i)['values'][1] == "☑"]
            if not ids: 
                messagebox.showwarning("Uyarı", "Lütfen kontrol edilecek kurumu/faturayı seçin (☑).")
                return

            bekleyen_tutar = 0.0
            for i in tree.get_children():
                if tree.item(i)['values'][1] == "☑":
                    try:
                        tut_str = str(tree.item(i)['values'][4]).replace(" ₺", "").strip()
                        if "," in tut_str and "." in tut_str:
                            if tut_str.rfind(",") > tut_str.rfind("."):
                                tut_str = tut_str.replace(".", "").replace(",", ".")
                            else:
                                tut_str = tut_str.replace(",", "")
                        elif "," in tut_str:
                            tut_str = tut_str.replace(",", ".")
                        bekleyen_tutar += float(tut_str)
                    except Exception as e: pass

            win = tk.Toplevel(self.pencere)
            win.title("Parçalı Tahsilat Kontrolü")
            win.geometry("550x800") 
            win.configure(bg="#f8fafc")
            win.transient(self.pencere)
            win.grab_set()
            
            x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 275
            y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 400
            win.geometry(f"+{x}+{y}")

            yatan_paralar = []

            f_bot = tk.Frame(win, bg="#1e293b", pady=15)
            f_bot.pack(fill="x", side="bottom")
            
            def tahsilati_onayla():
                fark = bekleyen_tutar - sum(yatan_paralar)
                if abs(fark) > 0.05:
                    onay = messagebox.askyesno("Eşleşmeyen Tutar Uyarı", f"Bankaya yatan tutar ile beklenen tutar arasında {abs(fark):,.2f} TL fark var.\n\nKesintileri (aidat vb.) kabul edip, bu faturayı tamamen 'ÖDENDİ' olarak işaretleyip kapatmak istiyor musunuz?", parent=win)
                    if not onay: return
                        
                for oid in ids: self.imlec_finans.execute("UPDATE odemeler SET durum='ODENDİ' WHERE id=?", (oid,))
                self.baglanti_finans.commit()
                tabloyu_yenile()
                messagebox.showinfo("Başarılı", "Seçili faturalar tahsil edildi olarak işaretlendi ve yeşile boyandı.", parent=self.pencere)
                win.destroy()

            ModernButton(f_bot, text="✅ SEÇİLİLERİ 'ÖDENDİ' YAP (TAHSİLAT)", command=tahsilati_onayla, bg_color="#10b981", width=420, height=45).pack(anchor="center")

            tk.Label(win, text="🔍 YATAN PARA (TAHSİLAT) KONTROLÜ", font=("Segoe UI", 14, "bold"), bg="#f8fafc", fg="#334155").pack(pady=(20, 10))

            f_main_container = tk.Frame(win, bg="#f8fafc")
            f_main_container.pack(fill="both", expand=True, padx=10, pady=5)
            
            canvas = tk.Canvas(f_main_container, bg="#f8fafc", highlightthickness=0)
            scrollbar = ttk.Scrollbar(f_main_container, orient="vertical", command=canvas.yview)
            f_main = tk.Frame(canvas, bg="#f8fafc")
            
            f_main.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            win_id = canvas.create_window((0, 0), window=f_main, anchor="nw")
            canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))
            
            canvas.configure(yscrollcommand=scrollbar.set)
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            self.mouse_scroll_ekle(canvas, f_main)

            f_ozet = tk.Frame(f_main, bg="white", bd=1, relief="solid", padx=15, pady=8)
            f_ozet.pack(fill="x", padx=10, pady=10)
            
            tk.Label(f_ozet, text="HESAP ÖZETİ", font=("Segoe UI", 11, "bold"), bg="white", fg="#3b82f6").pack(anchor="w", pady=(0, 5))
            
            r1 = tk.Frame(f_ozet, bg="white"); r1.pack(fill="x", pady=2)
            tk.Label(r1, text="Seçilen Faturaların Toplamı (Beklenen):", font=("Segoe UI", 10), bg="white", fg="#475569").pack(side="left")
            lbl_bekleyen = tk.Label(r1, text=f"{bekleyen_tutar:,.2f} ₺", font=("Segoe UI", 10, "bold"), bg="white", fg="#0f172a")
            lbl_bekleyen.pack(side="right")

            r2 = tk.Frame(f_ozet, bg="white"); r2.pack(fill="x", pady=2)
            tk.Label(r2, text="Bankaya Yatan (Eklenen):", font=("Segoe UI", 10), bg="white", fg="#475569").pack(side="left")
            lbl_yatan = tk.Label(r2, text="0.00 ₺", font=("Segoe UI", 10, "bold"), bg="white", fg="#10b981")
            lbl_yatan.pack(side="right")

            tk.Frame(f_ozet, height=1, bg="#e2e8f0").pack(fill="x", pady=6)
            
            r3 = tk.Frame(f_ozet, bg="white"); r3.pack(fill="x", pady=2)
            tk.Label(r3, text="FARK (Eksik / Fazla):", font=("Segoe UI", 11, "bold"), bg="white", fg="#ef4444").pack(side="left")
            lbl_fark = tk.Label(r3, text=f"{bekleyen_tutar:,.2f} ₺", font=("Segoe UI", 14, "bold"), bg="white", fg="#ef4444")
            lbl_fark.pack(side="right")

            lbl_durum = tk.Label(f_main, text="Eksik yatmış, ödeme tamamlanmadı.", font=("Segoe UI", 11, "bold"), bg="#f8fafc", fg="#ef4444")
            lbl_durum.pack(pady=5)

            f_ekle = tk.Frame(f_main, bg="white", bd=1, relief="solid", padx=15, pady=15)
            f_ekle.pack(fill="x", padx=10, pady=10)
            
            tk.Label(f_ekle, text="YATAN PARÇA EKLE", font=("Segoe UI", 11, "bold"), bg="white", fg="#f59e0b").pack(anchor="w", pady=(0, 10))
            
            f_giris = tk.Frame(f_ekle, bg="white")
            f_giris.pack(fill="x")
            
            tk.Label(f_giris, text="Tutar (TL):", font=("Segoe UI", 10, "bold"), bg="white", fg="#334155").pack(side="left")
            e_tutar = tk.Entry(f_giris, font=("Segoe UI", 13, "bold"), justify="right", relief="solid", bd=1, bg="#f8fafc", width=15)
            e_tutar.pack(side="left", padx=10, ipady=4, expand=True, fill="x")
            e_tutar.focus_set()
            
            f_liste = tk.Frame(f_main, bg="white", bd=1, relief="solid", padx=15, pady=8)
            f_liste.pack(fill="both", expand=True, padx=10, pady=10)
            
            tk.Label(f_liste, text="EKLENEN PARÇALAR", font=("Segoe UI", 11, "bold"), bg="white", fg="#8b5cf6").pack(anchor="w", pady=(0, 5))
            
            listbox = tk.Listbox(f_liste, font=("Segoe UI", 11), bg="#f8fafc", fg="#0f172a", relief="flat", highlightthickness=0, height=8)
            listbox.pack(fill="both", expand=True, pady=5)
            tk.Label(f_liste, text="* Eklediğiniz tutarı silmek için üzerine çift tıklayın", font=("Segoe UI", 8), bg="white", fg="#94a3b8").pack(anchor="w")

            def guncelle():
                toplam_yatan = sum(yatan_paralar)
                fark = bekleyen_tutar - toplam_yatan
                
                lbl_yatan.config(text=f"{toplam_yatan:,.2f} ₺")
                lbl_fark.config(text=f"{abs(fark):,.2f} ₺")
                
                if abs(fark) < 0.05: 
                    lbl_durum.config(text="✅ Kuruşu kuruşuna eşleşti! Tahsilat tamam.", fg="#10b981")
                    lbl_fark.config(fg="#10b981")
                elif fark > 0:
                    lbl_durum.config(text="⚠️ Eksik yattı. Kalan tutar bekleniyor.", fg="#ef4444")
                    lbl_fark.config(fg="#ef4444")
                else:
                    lbl_durum.config(text="❓ Fazla yattı! Beklenenden fazla ödeme yapılmış.", fg="#f59e0b")
                    lbl_fark.config(fg="#f59e0b")

            def parca_ekle(event=None):
                ham_giris = e_tutar.get().strip()
                if not ham_giris: return
                
                try:
                    if "," in ham_giris and "." in ham_giris:
                        if ham_giris.rfind(",") > ham_giris.rfind("."):
                            clean_val = ham_giris.replace(".", "").replace(",", ".")
                        else:
                            clean_val = ham_giris.replace(",", "")
                    elif "," in ham_giris:
                        clean_val = ham_giris.replace(",", ".")
                    else:
                        clean_val = ham_giris
                        
                    val = float(clean_val)
                except:
                    val = 0.0

                if val > 0:
                    yatan_paralar.append(val)
                    listbox.insert(tk.END, f"➕ {val:,.2f} ₺ eklendi.")
                    e_tutar.delete(0, tk.END)
                    guncelle()
                    listbox.yview(tk.END)
                
                e_tutar.focus_set()

            ModernButton(f_giris, text="EKLE", command=parca_ekle, bg_color="#3b82f6", width=80, height=32).pack(side="right")
            e_tutar.bind("<Return>", parca_ekle)

            def sil_parca(event=None):
                sel = listbox.curselection()
                if sel:
                    idx = sel[0]
                    yatan_paralar.pop(idx)
                    listbox.delete(idx)
                    guncelle()
            
            listbox.bind("<Delete>", sil_parca)
            listbox.bind("<Double-Button-1>", sil_parca)

        ModernButton(bottom_bar, text="🔍 YATAN PARAYI KONTROL ET", command=yatan_para_kontrol_ac, width=280, height=45, bg_color="#3b82f6").pack(side="left", padx=10)
        ModernButton(bottom_bar, text="✅ DİREKT TAHSİLAT YAPILDI", command=secilenleri_alindi_yap, width=260, height=45, bg_color="#10b981").pack(side="right", padx=10)

        # --- YENİ: SAĞ TIK MENÜSÜ İŞLEMLERİ ---
        sag_tik_menusu = tk.Menu(self.pencere, tearoff=0, bg="white", fg="#0f172a", font=("Segoe UI", 10, "bold"))
        
        def secili_geri_al():
            sel = tree.selection()
            if not sel: return
            oid = tree.item(sel[0], "values")[0]
            self.imlec_finans.execute("UPDATE odemeler SET durum='ODENMEDİ' WHERE id=?", (oid,))
            self.baglanti_finans.commit()
            tabloyu_yenile()
            messagebox.showinfo("Başarılı", "Kayıt geri alındı. Tekrar 'Bekliyor (Ödenmedi)' durumuna geçti.", parent=self.pencere)

        def secili_sil_sag_tik():
            sel = tree.selection()
            if not sel: return
            oid = tree.item(sel[0], "values")[0]
            if messagebox.askyesno("Silme Onayı", "Bu kaydı kalıcı olarak silmek istiyor musunuz?", parent=self.pencere):
                self.imlec_finans.execute("DELETE FROM odemeler WHERE id=?", (oid,))
                self.baglanti_finans.commit()
                tabloyu_yenile()

        sag_tik_menusu.add_command(label="🔄 Geri Al (Ödenmedi Yap)", command=secili_geri_al)
        sag_tik_menusu.add_separator()
        sag_tik_menusu.add_command(label="🗑️ Kaydı Tamamen Sil", command=secili_sil_sag_tik)

        def on_right_click(event):
            iid = tree.identify_row(event.y)
            if iid:
                # Sağ tıklanan satırı mavi (seçili) yap ki kullanıcı neye işlem yaptığını görsün
                tree.selection_set(iid) 
                sag_tik_menusu.tk_popup(event.x_root, event.y_root)

        tree.bind("<Button-3>", on_right_click)

        # --- SOL TIK İŞLEMLERİ ---
        def on_click(event):
            region = tree.identify("region", event.x, event.y)
            if region == "cell":
                if tree.identify_column(event.x) == "#2":
                    item = tree.identify_row(event.y)
                    if item:
                        vals = list(tree.item(item, "values"))
                        if vals[1] == "✅": 
                            # Eğer yeşil (ödendi) satırına sol tıklamaya çalışırsa sağ tıkı hatırlat
                            messagebox.showinfo("Bilgi", "Bu kayıt tahsil edilmiş.\n\nGeri almak veya silmek için farenizle üzerine SAĞ TIKLAYIN.", parent=self.pencere)
                            return
                        vals[1] = "☑" if vals[1] == "☐" else "☐"
                        tree.item(item, values=vals)
                        
        tree.bind("<Button-1>", on_click)
        
        # En son tabloyu ekrana çiz
        tabloyu_yenile()
        
        # --- TABLO ALTI İŞLEM ÇUBUĞU ---
        frm_row_ops = tk.Frame(tab, bg=c.get_color("card_bg"), pady=10, padx=10) 
        frm_row_ops.pack(side="bottom", fill="x")

        bottom_bar = tk.Frame(tab, bg="white", pady=10); bottom_bar.pack(fill="x")
        def secilenleri_ode():
            ids = [tree.item(i)['values'][0] for i in tree.get_children() if tree.item(i)['values'][1] == "☑"]
            if not ids: messagebox.showwarning("Uyarı", "Seçim yapmadınız."); return
            if messagebox.askyesno("Ödeme Onayı", f"{len(ids)} adet kaydı 'ÖDENDİ' olarak işaretleyip kapatmak istiyor musunuz?"):
                for oid in ids: 
                    # HATA DÜZELTİLDİ
                    self.imlec_finans.execute("UPDATE odemeler SET durum='ODENDİ' WHERE id=?", (oid,))
                self.baglanti_finans.commit()
                tabloyu_yenile()
                messagebox.showinfo("Başarılı", "Kayıtlar ödendi olarak işaretlendi.")
                # --- YENİ STOK YÜKLEME TARİHİNİ KAYDET ---
            guncel_tarih = datetime.now().strftime("%d.%m.%Y - %H:%M")
            try:
                self.imlec.execute("SELECT id FROM ayarlar WHERE anahtar='son_stok_yukleme_tarihi'")
                if self.imlec.fetchone():
                    self.imlec.execute("UPDATE ayarlar SET deger=? WHERE anahtar='son_stok_yukleme_tarihi'", (guncel_tarih,))
                else:
                    self.imlec.execute("INSERT INTO ayarlar (anahtar, deger) VALUES ('son_stok_yukleme_tarihi', ?)", (guncel_tarih,))
                self.baglanti_skt.commit()
            except Exception as e:
                print("Stok tarihi güncellenemedi:", e)

    def kurumu_sil(self, isim):
        if messagebox.askyesno("Sil", f"'{isim}' kurumunu ve tüm kayıtlarını silmek istiyor musun?"):
            try:
                # HATA DÜZELTİLDİ
                self.imlec_finans.execute("DELETE FROM kurumlar WHERE isim=?", (isim,))
                self.imlec_finans.execute("DELETE FROM odemeler WHERE fatura_adi=? AND satir_notu LIKE '%KURUM GİDERİ%'", (isim,))
                self.baglanti_finans.commit()
                messagebox.showinfo("Başarılı", "Kurum silindi.")
                self.sayfa_kurum_odemeleri()
            except Exception as e: messagebox.showerror("Hata", f"Silme hatası: {e}")

    # --- 4. GELİR VE GİDER TAKİBİ ---
    def sayfa_gelir(self): self._olustur_aylik_finans_sayfasi("GELİR")

    def sabit_gider_sihirbazi(self, secilen_yil):
        # Tablo kontrolü ve oluşturma
        self.imlec_finans.execute("CREATE TABLE IF NOT EXISTS sabit_giderler (id INTEGER PRIMARY KEY AUTOINCREMENT, kalem_adi TEXT, varsayilan_tutar REAL)")
        self.baglanti_finans.commit()

        # Ana Pencere Ayarları
        win = tk.Toplevel(self.pencere)
        win.title("Sabit Gider Yönetimi")
        win.geometry("600x750") 
        win.transient(self.pencere); win.grab_set()
        win.configure(bg="#f8fafc")
        
        # Ekranı Ortala
        x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 300
        y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 375
        win.geometry(f"+{x}+{y}")

        # --- YENİ ŞABLON EKLEME POPUP ---
        def modern_ekle_popup():
            pop = tk.Toplevel(win)
            pop.title("Yeni Şablon")
            pop.geometry("400x350")
            pop.configure(bg="white")
            pop.transient(win); pop.grab_set()
            
            px = win.winfo_x() + 100; py = win.winfo_y() + 100
            pop.geometry(f"+{px}+{py}")
            
            tk.Label(pop, text="Yeni Sabit Gider Şablonu", font=("Segoe UI", 14, "bold"), bg="white", fg="#334155").pack(pady=20)
            
            tk.Label(pop, text="Gider Adı (Örn: Elektrik, Kira):", font=("Segoe UI", 10, "bold"), bg="white", fg="#64748b").pack(anchor="w", padx=30)
            ent_ad = tk.Entry(pop, font=("Segoe UI", 11), relief="solid", bd=1, bg="#f8fafc")
            ent_ad.pack(fill="x", padx=30, pady=(5, 15), ipady=5)
            ent_ad.focus_set()
            
            tk.Label(pop, text="Varsayılan Tutar (TL):", font=("Segoe UI", 10, "bold"), bg="white", fg="#64748b").pack(anchor="w", padx=30)
            ent_tut = tk.Entry(pop, font=("Segoe UI", 11), relief="solid", bd=1, bg="#f8fafc")
            ent_tut.pack(fill="x", padx=30, pady=(5, 20), ipady=5)
            ent_tut.bind("<KeyRelease>", mask_para_birimi)
            
            def kaydet_kapat():
                ad = ent_ad.get().strip()
                tutar = temizle_para(ent_tut.get())
                
                if not ad:
                    messagebox.showwarning("Eksik", "Lütfen bir isim girin.", parent=pop)
                    return
                
                try:
                    self.imlec_finans.execute("INSERT INTO sabit_giderler (kalem_adi, varsayilan_tutar) VALUES (?, ?)", (ad, tutar))
                    self.baglanti_finans.commit()
                    listeyi_yukle()
                    pop.destroy()
                except Exception as e:
                    messagebox.showerror("Hata", str(e), parent=pop)

            ModernButton(pop, text="KAYDET", command=kaydet_kapat, width=200, height=40, bg_color="#10b981").pack(side="bottom", pady=30)
            pop.bind('<Return>', lambda e: kaydet_kapat())

        # --- ARAYÜZ DÜZENİ ---
        # 1. ÜST PANEL (Dönem)
        frm_top = tk.Frame(win, bg="#f1f5f9", pady=20, padx=20)
        frm_top.pack(side="top", fill="x")
        
        tk.Label(frm_top, text="Hangi aya işlenecek?", bg="#f1f5f9", font=("Segoe UI", 10), fg="#64748b").pack(anchor="w")
        
        bugun = date.today()
        yil_val = secilen_yil if secilen_yil else str(bugun.year)
        secili_ay_str = f"{yil_val}-{bugun.month:02d}"
        
        ent_donem = tk.Entry(frm_top, font=("Segoe UI", 14, "bold"), width=15, justify="center", relief="solid", bd=1, fg="#334155")
        ent_donem.pack(anchor="w", pady=5)
        ent_donem.insert(0, secili_ay_str)

        # 2. ORTA PANEL (LİSTE) - SCROLLABLE
        frm_mid = tk.Frame(win, bg="white", padx=10, pady=10)
        frm_mid.pack(side="top", fill="both", expand=True)
        
        lbl_baslik = tk.Label(frm_mid, text="Şablon Listesi (Tutarları değiştirebilirsiniz)", font=("Segoe UI", 10, "bold"), bg="white", fg="#475569")
        lbl_baslik.pack(anchor="w", pady=(0, 10))
        
        # Scroll Sistemi
        canvas = tk.Canvas(frm_mid, bg="white", highlightthickness=0)
        scrollbar = ttk.Scrollbar(frm_mid, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg="white")
        
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        win_id = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.bind('<Configure>', lambda e: canvas.itemconfig(win_id, width=e.width))
        
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.mouse_scroll_ekle(canvas, scroll_frame)

        # 3. ALT PANEL (Butonlar)
        frm_bot = tk.Frame(win, bg="#e2e8f0", pady=15, padx=20)
        frm_bot.pack(side="bottom", fill="x")

        check_vars = {} 

        def listeyi_yukle():
            for w in scroll_frame.winfo_children(): w.destroy()
            check_vars.clear()
            
            self.imlec_finans.execute("SELECT id, kalem_adi, varsayilan_tutar FROM sabit_giderler")
            kayitlar = self.imlec_finans.fetchall()
            
            if not kayitlar:
                tk.Label(scroll_frame, text="Henüz şablon yok.", font=("Segoe UI", 11), bg="white", fg="#94a3b8").pack(pady=20)
                return

            for row in kayitlar:
                rid, ad, tutar = row
                var = tk.IntVar(value=1)
                
                # Kart Görünümü
                card = tk.Frame(scroll_frame, bg="white", bd=1, relief="solid", padx=5, pady=5)
                card.pack(fill="x", pady=4, padx=2)
                
                cb = tk.Checkbutton(card, variable=var, bg="white", cursor="hand2")
                cb.pack(side="left")
                
                tk.Label(card, text=ad, font=("Segoe UI", 11, "bold"), bg="white", fg="#334155", width=20, anchor="w").pack(side="left", padx=5)
                
                ent_tutar_row = tk.Entry(card, font=("Segoe UI", 11), width=10, justify="right", bd=1, relief="solid", bg="#fffbeb")
                ent_tutar_row.pack(side="left", padx=10)
                tutar_str = f"{tutar:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                ent_tutar_row.insert(0, tutar_str)
                ent_tutar_row.bind("<KeyRelease>", mask_para_birimi)
                
                check_vars[rid] = (var, ad, ent_tutar_row)
                
                tk.Button(card, text="🗑️", font=("Segoe UI", 9), bg="white", fg="#dc2626", 
                          relief="flat", cursor="hand2",
                          command=lambda i=rid: sil(i)).pack(side="right", padx=5)

        def sil(rid):
            if messagebox.askyesno("Onay", "Bu şablonu silmek istiyor musunuz?", parent=win):
                self.imlec_finans.execute("DELETE FROM sabit_giderler WHERE id=?", (rid,))
                self.baglanti_finans.commit()
                listeyi_yukle()

        def aktar():
            tarih_str = ent_donem.get() + "-01"
            secilenler = []
            
            for rid, (var, ad, ent_widget) in check_vars.items():
                if var.get() == 1:
                    guncel_tutar = temizle_para(ent_widget.get())
                    secilenler.append((ad, guncel_tutar))
            
            if not secilenler:
                messagebox.showwarning("Uyarı", "Listeden en az bir gider seçmelisiniz.", parent=win)
                return
            
            toplam_tutar = sum([x[1] for x in secilenler])
            
            if messagebox.askyesno("Onay", f"{len(secilenler)} adet gider kalemi\nToplam: {toplam_tutar:,.2f} TL\n\n{ent_donem.get()} dönemine işlenecek.\nOnaylıyor musunuz?", parent=win):
                eklenen = 0
                for ad, tutar in secilenler:
                    try:
                        aciklama = f"[SABİT] {ad}"
                        self.imlec_finans.execute("INSERT INTO odemeler (fatura_adi, vade_tarihi, tutar, aciklama, durum, satir_notu) VALUES (?, ?, ?, ?, 'ODENMEDİ', ?)", 
                                           (ad, tarih_str, tutar, aciklama, "MANUEL_GIDER"))
                        eklenen += 1
                    except: pass
                
                self.baglanti_finans.commit()
                messagebox.showinfo("Başarılı", f"{eklenen} adet gider eklendi.", parent=win)
                win.destroy()
                self.sayfa_gider()

        ModernButton(frm_bot, text="➕ YENİ ŞABLON EKLE", command=modern_ekle_popup, width=200, bg_color="#64748b").pack(side="left")
        ModernButton(frm_bot, text="⚡ SEÇİLENLERİ İŞLE", command=aktar, width=220, bg_color="#10b981").pack(side="right")
        
        listeyi_yukle()

    def sayfa_gider(self): self._olustur_aylik_finans_sayfasi("GİDER")

    def _olustur_aylik_finans_sayfasi(self, tip, korunacak_yview=None):
        for w in self.content_area.winfo_children(): w.destroy()
        c = TM
        ana_renk = "#10b981" if tip == "GELİR" else "#ef4444"
        # İç boşluklar (padx=20, pady=15) artırıldı.
        header = tk.Frame(self.content_area, bg=c.get_color("bg_main"), padx=20, pady=15)
        header.pack(fill="x")
        
        # Başlığa ikon eklendi ve font boyutu 24'e çıkarıldı.
        ikon = "💰" if tip == "GELİR" else "📉"
        tk.Label(header, text=f"{ikon} {tip} YÖNETİMİ", font=("Segoe UI", 24, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left")
        frm_btns = tk.Frame(header, bg=c.get_color("bg_main")); frm_btns.pack(side="right")
        tk.Label(frm_btns, text="Yıl:", bg=c.get_color("bg_main"), font=("Segoe UI", 10, "bold")).pack(side="left", padx=5)
        # Yıl aralığı 10+ yıllık kredileri kapsayacak şekilde +16'ya çıkarıldı
        yil_listesi = [str(y) for y in range(datetime.now().year - 2, datetime.now().year + 16)]
        cmb_yil = ttk.Combobox(frm_btns, values=yil_listesi, state="readonly", width=6, font=("Segoe UI", 10))
        cmb_yil.set(str(datetime.now().year)); cmb_yil.pack(side="left", padx=(0, 15))
        ModernButton(frm_btns, text=f"➕  {tip} EKLE", command=lambda: self.manuel_finans_popup(tip), width=200, height=35, bg_color=ana_renk).pack(side="left", padx=5)
        
        if tip == "GİDER":
            ModernButton(frm_btns, text="⚡ SABİT GİDERLER", command=lambda: self.sabit_gider_sihirbazi(cmb_yil.get()), width=200, height=35, bg_color="#8b5cf6").pack(side="left", padx=5)

        def yenile(gecici_pos=None): 
            kayitli_pos = gecici_pos if gecici_pos else self.finans_canvas.yview()
            self._aylik_tablolari_doldur(scrollable_frame, tip, cmb_yil.get())
            if kayitli_pos:
                self.pencere.after(50, lambda: self.finans_canvas.yview_moveto(kayitli_pos[0]))

        ModernButton(frm_btns, text="🔄 YENİLE", command=lambda: yenile(None), width=100, height=35, bg_color="#64748b").pack(side="left", padx=5)
        cmb_yil.bind("<<ComboboxSelected>>", lambda e: yenile(None))

        canvas = tk.Canvas(self.content_area, bg=c.get_color("bg_main"), highlightthickness=0)
        self.finans_canvas = canvas
        scrollbar = ttk.Scrollbar(self.content_area, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=c.get_color("bg_main"))
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(window_id, width=e.width))
        canvas.configure(yscrollcommand=scrollbar.set); canvas.pack(side="left", fill="both", expand=True, pady=5)
        scrollbar.pack(side="right", fill="y")
        self.mouse_scroll_ekle(canvas, scrollable_frame)
        yenile(korunacak_yview)

    def _aylik_tablolari_doldur(self, parent_frame, tip, secilen_yil):
        for widget in parent_frame.winfo_children(): widget.destroy()
        c = TM # Temayı almayı unutmamak için (Hata vermemesi için gerekli)
        aylar_listesi = ["OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN", "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"]
        genel_yil_kalan = 0.0; kayit_var_mi = False
        bugun = date.today()
        
        try: secilen_yil_int = int(secilen_yil)
        except: secilen_yil_int = bugun.year

        for ay_index in range(1, 13):
            ay_str = f"{secilen_yil}-{ay_index:02d}"
            if tip == "GELİR":
                sorgu = """
                    SELECT id, strftime('%d', vade_tarihi), fatura_adi, aciklama, tutar, satir_notu, vade_tarihi, durum 
                    FROM odemeler WHERE strftime('%Y-%m', vade_tarihi) = ? AND (satir_notu LIKE '%KURUM%' OR satir_notu = 'MANUEL_GELIR') ORDER BY vade_tarihi
                """
            else:
                # GİDER SORGUSU GÜNCELLENDİ: Çift gider yazılmasını önlemek için Depo Kredi Kartı ödemelerini gizler.
                sorgu = """
                    SELECT id, strftime('%d', vade_tarihi), fatura_adi, aciklama, tutar, satir_notu, vade_tarihi, durum 
                    FROM odemeler 
                    WHERE strftime('%Y-%m', vade_tarihi) = ? 
                    AND (satir_notu IS NULL OR (NOT satir_notu LIKE '%KURUM%' AND satir_notu != 'MANUEL_GELIR'))
                    AND (
                        satir_notu = 'MANUEL_GIDER' 
                        OR aciklama IS NULL 
                        OR (aciklama NOT LIKE '%KK]%' AND aciklama NOT LIKE '%Kredi Kartı%' AND aciklama NOT LIKE '%KREDİ KARTI%')
                    )
                    ORDER BY vade_tarihi
                """
            self.imlec_finans.execute(sorgu, (ay_str,))
            ham_veriler = self.imlec_finans.fetchall()
            
            if not ham_veriler: continue
            kayit_var_mi = True
            
            ay_toplami = sum([r[4] for r in ham_veriler])
            ay_kalan = sum([r[4] for r in ham_veriler if r[7] == 'ODENMEDİ'])
            
            if secilen_yil_int > bugun.year: genel_yil_kalan += ay_kalan
            elif secilen_yil_int == bugun.year:
                if ay_index >= bugun.month: genel_yil_kalan += ay_kalan

            # relief="solid" kaldırıldı, yerine ince ve yumuşak bir sınır (border) eklendi.
            card = tk.Frame(parent_frame, bg=c.get_color("card_bg"), bd=0, highlightbackground=c.get_color("border"), highlightthickness=1)
            card.pack(fill="x", pady=10, padx=10)
            baslik_bg = "#f1f5f9"; baslik_fg = "#334155"
            tutar_renk = "#166534" if tip == "GELİR" else "#ef4444"
            
            header_frame = tk.Frame(card, bg=baslik_bg, padx=10, pady=10, cursor="hand2"); header_frame.pack(fill="x")
            lbl_icon = tk.Label(header_frame, text="▶", font=("Segoe UI", 12), bg=baslik_bg, fg=baslik_fg); lbl_icon.pack(side="left", padx=(0, 10))
            lbl_title = tk.Label(header_frame, text=f"{aylar_listesi[ay_index-1]} {secilen_yil}", font=("Segoe UI", 11, "bold"), bg=baslik_bg, fg=baslik_fg); lbl_title.pack(side="left")
            
            # --- HATA YARATAN KISIM BURASIYDI, HİZASI DÜZELTİLDİ ---
            # Toplam tutar için özel bir arka plan çerçevesi (input_bg renginde) oluşturduk
            f_toplam_kutu = tk.Frame(header_frame, bg=c.get_color("input_bg"), padx=15, pady=5, relief="flat", bd=0)
            f_toplam_kutu.pack(side="right", fill="y")
            
            tk.Label(f_toplam_kutu, text="Aylık Toplam:", font=("Segoe UI", 10), bg=c.get_color("input_bg"), fg="#64748b").pack(side="left", padx=(0, 10))
            lbl_total = tk.Label(f_toplam_kutu, text=f"{ay_toplami:,.2f} ₺", font=("Segoe UI", 14, "bold"), bg=c.get_color("input_bg"), fg=tutar_renk)
            lbl_total.pack(side="right")
            # --------------------------------------------------------
            
            content_frame = tk.Frame(card, bg=c.get_color("card_bg"), padx=10, pady=10)
            content_frame.pack(fill="both", expand=True)
            
            cols = ("ID", "GUN", "FIRMA", "KAYNAK", "ACIKLAMA", "TUTAR", "TUR")
            h = len(ham_veriler) if len(ham_veriler) < 10 else 10
            tree = ttk.Treeview(content_frame, columns=cols, show="headings", height=h)
            
            tree.heading("ID", text="ID"); tree.column("ID", width=0, stretch=False)
            tree.heading("GUN", text="Gün"); tree.column("GUN", width=50, anchor="center")
            tree.heading("FIRMA", text="Firma / İlgili Kişi"); tree.column("FIRMA", width=220)
            tree.heading("KAYNAK", text="İşlem Tipi / Kaynak"); tree.column("KAYNAK", width=180, anchor="center")
            tree.heading("ACIKLAMA", text="Açıklama / Not"); tree.column("ACIKLAMA", width=300)
            tree.heading("TUTAR", text="Tutar (TL)"); tree.column("TUTAR", width=130, anchor="e")
            tree.heading("TUR", text="Kayıt Türü"); tree.column("TUR", width=100, anchor="center")
            
            tree.tag_configure("yesil", background="#ecfdf5", foreground="#065f46") 
            tree.tag_configure("kirmizi", background="#fef2f2", foreground="#991b1b") 
            
            tree.pack(fill="both", expand=True)

            # --- EKLENEN KISIM: SCROLL (KAYDIRMA) ÇÖZÜMÜ ---
            # Fare tablonun (Treeview) üzerindeyken ana ekranın (Canvas) kaymasını sağlar.
            def tree_scroll(event, canvas=parent_frame.master):
                try: 
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                except: 
                    pass
            
            tree.bind('<MouseWheel>', tree_scroll)
            # -----------------------------------------------

            for row in ham_veriler:
                oid, gun, firma, aciklama, tutar, notu, tam_tarih_str, durum = row
                tur = "MANUEL" if notu in ["MANUEL_GELIR", "MANUEL_GIDER"] else "OTOMATİK"
                aciklama_goster = aciklama if aciklama else ""
                kaynak_bilgisi = "-"
                
                # --- KREDİLERİ VE KARTLARI YAKALAYIP DÜZENLEME ---
                if aciklama and str(aciklama).startswith("KREDI:"):
                    tur = "OTOMATİK"
                    kaynak_bilgisi = str(aciklama).replace("KREDI:", "").strip()
                    aciklama_goster = "Aylık Kredi Taksiti"
                elif notu and "KART:" in str(notu):
                    tur = "OTOMATİK"
                    # "KART:" yazısını temizle ve "[" işaretinden sonrasını ([İşlem: ...] kısmını) at
                    kaynak_bilgisi = str(notu).replace("KART:", "").split("[")[0].strip()
                    aciklama_goster = "Kredi Kartı Ödemesi"
                else:
                    # Kredi veya Kart değilse normal işlem
                    if tip == "GİDER":
                        if tur == "MANUEL": kaynak_bilgisi = "Elden / Manuel"
                        else: kaynak_bilgisi = "Cari / Nakit"
                    else: 
                        if "KURUM" in (notu or ""): kaynak_bilgisi = "Kurum Ödemesi"
                        else: kaynak_bilgisi = "Elden / Diğer"

                satir_renk_tag = ""
                try:
                    vade_dt = datetime.strptime(tam_tarih_str, "%Y-%m-%d").date()
                    if vade_dt <= bugun: satir_renk_tag = "yesil"
                    else: satir_renk_tag = "kirmizi"
                except: pass
                tree.insert("", "end", values=(oid, gun, firma, kaynak_bilgisi, aciklama_goster, f"{tutar:,.2f} ₺", tur), tags=(satir_renk_tag,))

            self._sag_tik_menu_bagla(tree, tip)
            def toggle_func(c_frame=content_frame, icon=lbl_icon):
                if c_frame.winfo_viewable(): c_frame.forget(); icon.config(text="▶")
                else: c_frame.pack(fill="x", expand=True); icon.config(text="▼")
            for widget in [header_frame, lbl_icon, lbl_title, lbl_total]: widget.bind("<Button-1>", lambda e, f=toggle_func: f())

        if not kayit_var_mi: tk.Label(parent_frame, text=f"{secilen_yil} yılına ait {tip.lower()} kaydı bulunamadı.", font=("Segoe UI", 12), fg="#64748b", bg=parent_frame["bg"]).pack(pady=20)
        else: tk.Label(parent_frame, text=f"{secilen_yil} YILI KALAN BORÇ (Gelecek): {genel_yil_kalan:,.2f} ₺", font=("Segoe UI", 16, "bold"), bg=parent_frame["bg"], fg="#ef4444", pady=20).pack(fill="x")

    def _sag_tik_menu_bagla(self, tree, tip):
        menu = tk.Menu(self.pencere, tearoff=0)
        menu.add_command(label="✏️ Düzenle (Sadece Manuel)", command=lambda: self._finans_satir_duzenle(tree, tip))
        menu.add_command(label="🗑️ Sil (Sadece Manuel)", command=lambda: self._finans_satir_sil(tree, tip))
        def on_right_click(event):
            item = tree.identify_row(event.y)
            if item:
                tree.selection_set(item)
                menu.post(event.x_root, event.y_root)
        tree.bind("<Button-3>", on_right_click)

    def _finans_satir_sil(self, tree, tip):
        sel = tree.selection()
        if not sel: return
        
        item = tree.item(sel[0])
        val = item['values']
        oid = val[0]
        tur = str(val[6]).strip() 
        
        # --- ZORLA SİLME (HAYALET/KALINTI KAYITLAR İÇİN) DÜZENLEMESİ ---
        if tur != "MANUEL":
            onay = messagebox.askyesno("Sistem Uyarısı", f"⚠️ DİKKAT: Bu kayıt sisteme OTOMATİK olarak eklenmiş.\n\nNormal şartlarda krediyi, kartı veya faturayı iptal ettiğinizde bunun da silinmesi/kalması gerekirdi. Ancak bunu muhasebeden kalıcı olarak temizlemek istiyorsanız ZORLA SİLEBİLİRSİNİZ.\n\nBu '{tip}' kaydını kalıcı olarak ZORLA SİLMEK istiyor musunuz?")
            if not onay: 
                return
        else:
            if not messagebox.askyesno("Sil", "Bu manuel kaydı silmek istiyor musunuz?"):
                return
                
        self.imlec_finans.execute("SELECT * FROM odemeler WHERE id=?", (oid,))
        eski_veri = self.imlec_finans.fetchone()
        
        self.imlec_finans.execute("DELETE FROM odemeler WHERE id=?", (oid,))
        self.baglanti_finans.commit()
        
        self.gecmise_kaydet("SIL", "odemeler", oid, eski_veri)
        
        mevcut_yview = self.finans_canvas.yview() if hasattr(self, 'finans_canvas') else None
        self._olustur_aylik_finans_sayfasi(tip, korunacak_yview=mevcut_yview)

    def _finans_satir_duzenle(self, tree, tip):
        sel = tree.selection()
        if not sel: return
        
        item = tree.item(sel[0])
        val = item['values']
        oid = val[0]
        tur = str(val[6]).strip() 
        
        if tur != "MANUEL":
            messagebox.showwarning("İzin Yok", "Otomatik kayıtlar buradan düzenlenemez.")
            return

        self._finans_duzenle_popup(oid, tip)

    def _finans_duzenle_popup(self, oid, tip):
        self.imlec_finans.execute("SELECT fatura_adi, vade_tarihi, tutar, aciklama FROM odemeler WHERE id=?", (oid,))
        kayit = self.imlec_finans.fetchone()
        if not kayit: return
        
        win = tk.Toplevel(self.pencere)
        win.title("Düzenle")
        win.geometry("400x400")
        
        tk.Label(win, text="Kaydı Düzenle", font=("Segoe UI", 12, "bold")).pack(pady=10)
        frm = tk.Frame(win, padx=20)
        frm.pack(fill="both")
        
        tk.Label(frm, text="Firma/Kaynak:", anchor="w").pack(fill="x")
        e_ad = tk.Entry(frm, relief="solid", bd=1)
        e_ad.pack(fill="x")
        e_ad.insert(0, kayit[0])
        
        tk.Label(frm, text="Tarih (YYYY-AA-GG):", anchor="w").pack(fill="x", pady=(5,0))
        e_tar = tk.Entry(frm, relief="solid", bd=1)
        e_tar.pack(fill="x")
        e_tar.insert(0, kayit[1])
        tarih_secici_bagla(frm, e_tar)
        
        tk.Label(frm, text="Tutar:", anchor="w").pack(fill="x", pady=(5,0))
        e_tut = tk.Entry(frm, relief="solid", bd=1)
        e_tut.pack(fill="x")
        e_tut.insert(0, f"{kayit[2]:.2f}")
        
        tk.Label(frm, text="Açıklama:", anchor="w").pack(fill="x", pady=(5,0))
        e_ack = tk.Entry(frm, relief="solid", bd=1)
        e_ack.pack(fill="x")
        e_ack.insert(0, kayit[3])
        
        def kaydet():
            try:
                yeni_tut = float(e_tut.get().replace(",", "."))
                self.imlec_finans.execute("UPDATE odemeler SET fatura_adi=?, vade_tarihi=?, tutar=?, aciklama=? WHERE id=?", 
                                   (e_ad.get(), e_tar.get(), yeni_tut, e_ack.get(), oid))
                self.baglanti_finans.commit()
                win.destroy()
                
                mevcut_yview = self.finans_canvas.yview() if hasattr(self, 'finans_canvas') else None
                self._olustur_aylik_finans_sayfasi(tip, korunacak_yview=mevcut_yview)
                
            except: 
                messagebox.showerror("Hata", "Tutar hatalı.")
                
        ModernButton(win, text="GÜNCELLE", command=kaydet, bg_color="#3b82f6").pack(pady=20)
    

    def manuel_finans_popup(self, tip):
        tema_renk = "#10b981" if tip == "GELİR" else "#ef4444"
        pencere = tk.Toplevel(self.pencere); pencere.title(f"Yeni {tip} Kaydı"); pencere.geometry("420x550"); pencere.configure(bg="white")
        pencere.transient(self.pencere); pencere.grab_set(); pencere.focus_force()
        x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 210
        y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 275
        pencere.geometry(f"+{x}+{y}")
        header_frame = tk.Frame(pencere, bg=tema_renk, height=80); header_frame.pack(fill="x"); header_frame.pack_propagate(False)
        icon = "💰" if tip == "GELİR" else "📉"
        tk.Label(header_frame, text=f"{icon} YENİ {tip} GİRİŞİ", font=("Segoe UI", 16, "bold"), bg=tema_renk, fg="white").pack(expand=True)
        container = tk.Frame(pencere, bg="white", padx=30, pady=20); container.pack(fill="both", expand=True)

        def create_modern_entry(label_text, default_val=None, mask_func=None, is_date=False):
            tk.Label(container, text=label_text, font=("Segoe UI", 10, "bold"), bg="white", fg="#475569").pack(anchor="w", pady=(10, 5))
            entry_frame = tk.Frame(container, bg="white", bd=1, relief="solid"); entry_frame.pack(fill="x")
            entry = tk.Entry(entry_frame, font=("Segoe UI", 11), bd=0, bg="#f8fafc", highlightthickness=0)
            entry.pack(side="left", fill="x", expand=True, ipady=8, padx=5)
            if default_val: entry.insert(0, default_val)
            if mask_func: entry.bind("<KeyRelease>", mask_func)
            if is_date: tarih_secici_bagla(entry_frame, entry)
            return entry

        e_tarih = create_modern_entry("📅 İşlem Tarihi", date.today().strftime("%Y-%m-%d"), mask_tarih_otomatik, is_date=True)
        lbl_kaynak = "Nereden Geldi? (Müşteri, Satış vb.)" if tip == "GELİR" else "Nereye Ödendi? (Elektrik, Yemek vb.)"
        e_kaynak = create_modern_entry(f"🏢 {lbl_kaynak}")
        e_tutar = create_modern_entry("💵 Tutar (TL)", mask_func=mask_para_birimi); e_tutar.configure(fg=tema_renk, font=("Segoe UI", 12, "bold"))
        e_aciklama = create_modern_entry("📝 Açıklama / Not")
        btn_frame = tk.Frame(pencere, bg="white", pady=20); btn_frame.pack(fill="x", padx=30, side="bottom")

        def kaydet():
            try:
                tutar = temizle_para(e_tutar.get())
                kaynak = e_kaynak.get().strip()
                not_kodu = "MANUEL_GELIR" if tip == "GELİR" else "MANUEL_GIDER"
                if not kaynak: messagebox.showwarning("Eksik", "Lütfen kaynak/yer belirtiniz.", parent=pencere); return
                if tutar <= 0: messagebox.showwarning("Eksik", "Lütfen geçerli bir tutar giriniz.", parent=pencere); return
                aciklama_raw = e_aciklama.get().strip()
                if not aciklama_raw: aciklama_raw = f"{tip} Girişi"

                self.imlec_finans.execute("INSERT INTO odemeler (fatura_adi, vade_tarihi, tutar, aciklama, durum, satir_notu) VALUES (?, ?, ?, ?, 'ODENMEDİ', ?)", 
                                   (kaynak, e_tarih.get(), tutar, aciklama_raw, not_kodu))
                self.baglanti_finans.commit()
                messagebox.showinfo("Başarılı", f"{tip} kaydı başarıyla eklendi.", parent=pencere)
                pencere.destroy()
                if tip == "GELİR": self.sayfa_gelir()
                else: self.sayfa_gider()
            except Exception as e: messagebox.showerror("Hata", str(e), parent=pencere)

        tk.Button(btn_frame, text="İPTAL", font=("Segoe UI", 10, "bold"), bg="#f1f5f9", fg="#64748b", relief="flat", cursor="hand2", width=12, pady=10, command=pencere.destroy).pack(side="left")
        tk.Button(btn_frame, text="KAYDET", font=("Segoe UI", 10, "bold"), bg=tema_renk, fg="white", relief="flat", cursor="hand2", width=18, pady=10, command=kaydet).pack(side="right")
        pencere.bind('<Return>', lambda e: kaydet()); pencere.bind('<Escape>', lambda e: pencere.destroy())

    def sayfa_kasa_defteri(self):
        try:
            sutunlar = [row[1] for row in self.imlec_finans.execute("PRAGMA table_info(kasa_defteri)")]
            if sutunlar:
                if "islem_turu" not in sutunlar: self.imlec_finans.execute("ALTER TABLE kasa_defteri ADD COLUMN islem_turu TEXT")
                if "kategori" not in sutunlar: self.imlec_finans.execute("ALTER TABLE kasa_defteri ADD COLUMN kategori TEXT")
                if "aciklama" not in sutunlar: self.imlec_finans.execute("ALTER TABLE kasa_defteri ADD COLUMN aciklama TEXT")
                if "tutar" not in sutunlar: self.imlec_finans.execute("ALTER TABLE kasa_defteri ADD COLUMN tutar REAL")
                if "odeme_yontemi" not in sutunlar: self.imlec_finans.execute("ALTER TABLE kasa_defteri ADD COLUMN odeme_yontemi TEXT")
                if "kasa_adi" not in sutunlar: self.imlec_finans.execute("ALTER TABLE kasa_defteri ADD COLUMN kasa_adi TEXT DEFAULT 'Kasa 1'")
                self.baglanti_finans.commit()
        except Exception as e: 
            print("Kasa Tablo tamir edilemedi:", e)

        # YENİ: Sınırsız Kasalar Tablosu ve Varsayılan Kasa Oluşturma
        try:
            self.imlec_finans.execute("CREATE TABLE IF NOT EXISTS kasalar (id INTEGER PRIMARY KEY AUTOINCREMENT, kasa_adi TEXT UNIQUE)")
            self.imlec_finans.execute("INSERT OR IGNORE INTO kasalar (kasa_adi) VALUES ('Kasa 1')")
            self.baglanti_finans.commit()
        except Exception as e: 
            print("Kasalar tablo hatası:", e)

        def kasalari_getir():
            self.imlec_finans.execute("SELECT kasa_adi FROM kasalar ORDER BY id ASC")
            return [r[0] for r in self.imlec_finans.fetchall()]

        for w in self.content_area.winfo_children(): 
            w.destroy()
        c = TM

        # =====================================================================
        # HATA ÖNLEYİCİ: YENİLEME FONKSİYONU VE DÖNEM KUTUSU EN ÜSTTE TANIMLANIYOR
        # =====================================================================
        self.var_donem = tk.StringVar(value=date.today().strftime("%Y-%m"))
        mevcut_kasalar = kasalari_getir()
        
        self.var_kasa_filtre = tk.StringVar(value="Tümü")
        self.secili_kasa_giris = tk.StringVar(value=mevcut_kasalar[0] if mevcut_kasalar else "Kasa 1")

        def guncelle_arayuz_kasalari():
            guncel_liste = kasalari_getir()
            if hasattr(self, 'cmb_kasa_giris'):
                self.cmb_kasa_giris['values'] = guncel_liste + ["+ Yeni Kasa Ekle", "⚙️ Kasaları Yönet"]
                if self.secili_kasa_giris.get() not in guncel_liste:
                    self.secili_kasa_giris.set(guncel_liste[0] if guncel_liste else "")
            if hasattr(self, 'cmb_kasa_filtre'):
                self.cmb_kasa_filtre['values'] = ["Tümü"] + guncel_liste
                if self.var_kasa_filtre.get() not in ["Tümü"] + guncel_liste:
                    self.var_kasa_filtre.set("Tümü")
            yenile_tetik()

        def kasa_yonetimi_popup():
            pop = tk.Toplevel(self.pencere)
            pop.title("Kasa Yönetimi")
            pop.geometry("400x420")
            pop.configure(bg="#f8fafc")
            pop.transient(self.pencere)
            pop.grab_set()

            # Ekranı ortala
            x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 200
            y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 210
            pop.geometry(f"+{x}+{y}")

            tk.Label(pop, text="Mevcut Kasalar", font=("Segoe UI", 12, "bold"), bg="#f8fafc", fg="#334155").pack(pady=15)
            
            listbox = tk.Listbox(pop, font=("Segoe UI", 11), relief="solid", bd=1)
            listbox.pack(fill="both", expand=True, padx=20, pady=5)
            
            def listeyi_doldur():
                listbox.delete(0, tk.END)
                for k in kasalari_getir():
                    listbox.insert(tk.END, k)
            listeyi_doldur()

            def ad_degistir():
                sel = listbox.curselection()
                if not sel: return
                eski_ad = listbox.get(sel[0])
                yeni_ad = simpledialog.askstring("İsim Değiştir", f"'{eski_ad}' kasanızın yeni adını giriniz:", initialvalue=eski_ad, parent=pop)
                
                if yeni_ad and yeni_ad.strip() and yeni_ad.strip() != eski_ad:
                    yeni_ad = yeni_ad.strip()
                    try:
                        self.imlec_finans.execute("UPDATE kasalar SET kasa_adi=? WHERE kasa_adi=?", (yeni_ad, eski_ad))
                        self.imlec_finans.execute("UPDATE kasa_defteri SET kasa_adi=? WHERE kasa_adi=?", (yeni_ad, eski_ad))
                        self.baglanti_finans.commit()
                        listeyi_doldur()
                        guncelle_arayuz_kasalari()
                        messagebox.showinfo("Başarılı", "Kasa adı ve ilgili tüm geçmiş işlemler güncellendi.", parent=pop)
                    except sqlite3.IntegrityError:
                        messagebox.showwarning("Uyarı", "Bu isimde bir kasa zaten var!", parent=pop)

            def kasa_sil():
                sel = listbox.curselection()
                if not sel: return
                silinecek = listbox.get(sel[0])
                
                if listbox.size() <= 1:
                    messagebox.showwarning("Hata", "Sistemde en az 1 kasa kalmak zorundadır. Bu kasayı silemezsiniz.", parent=pop)
                    return

                cevap = messagebox.askyesnocancel("Silme Seçenekleri", f"'{silinecek}' isimli kasayı siliyorsunuz.\n\nEVET: Kasayı ve içindeki TÜM İŞLEMLERİ sil.\nHAYIR: Kasayı sil ama geçmiş işlemleri koru.\nİPTAL: İşlemi durdur.", parent=pop)
                
                if cevap is None: return
                
                if cevap: # EVET dedi, o kasadaki tüm işlemleri temizle
                    self.imlec_finans.execute("DELETE FROM kasa_defteri WHERE kasa_adi=?", (silinecek,))
                
                self.imlec_finans.execute("DELETE FROM kasalar WHERE kasa_adi=?", (silinecek,))
                self.baglanti_finans.commit()
                listeyi_doldur()
                guncelle_arayuz_kasalari()
                messagebox.showinfo("Başarılı", "Kasa başarıyla silindi.", parent=pop)

            btn_frame = tk.Frame(pop, bg="#f8fafc")
            btn_frame.pack(fill="x", pady=15, padx=20)
            ModernButton(btn_frame, text="✏️ AD DEĞİŞTİR", command=ad_degistir, bg_color="#3b82f6", width=160, height=35).pack(side="left", padx=5)
            ModernButton(btn_frame, text="🗑️ SİL", command=kasa_sil, bg_color="#ef4444", width=160, height=35).pack(side="right", padx=5)

        def kasa_secim_kontrol(event, cmb_widget, is_filter=False):
            secim = cmb_widget.get()
            if secim == "+ Yeni Kasa Ekle":
                yeni_kasa = simpledialog.askstring("Yeni Kasa Ekle", "Oluşturulacak kasanın adını giriniz:\n(Örn: Kasa 2, Döviz Kasası)", parent=self.pencere)
                if yeni_kasa and yeni_kasa.strip():
                    yeni_kasa = yeni_kasa.strip()
                    try:
                        self.imlec_finans.execute("INSERT INTO kasalar (kasa_adi) VALUES (?)", (yeni_kasa,))
                        self.baglanti_finans.commit()
                        messagebox.showinfo("Başarılı", f"'{yeni_kasa}' sisteme başarıyla eklendi.")
                    except sqlite3.IntegrityError:
                        messagebox.showwarning("Uyarı", "Bu kasa adı zaten mevcut!")
                    
                    guncelle_arayuz_kasalari()
                    cmb_widget.set(yeni_kasa)
                else:
                    guncelle_arayuz_kasalari()
            
            elif secim == "⚙️ Kasaları Yönet":
                kasa_yonetimi_popup()
                guncelle_arayuz_kasalari()
                
            elif is_filter:
                yenile_tetik()

        def yenile_tetik():
            self._kasa_verilerini_yukle(self.var_donem.get())

        # =====================================================================
        # ANA KONTEYNER
        # =====================================================================
        main_container = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        main_container.pack(fill="both", expand=True, padx=15, pady=15)

        # ---------------------------------------------------------------------
        # 1. SOL PANEL: HIZLI POS EKRANI (Simetrik ve Orantılı)
        # ---------------------------------------------------------------------
        f_left = tk.Frame(main_container, bg=c.get_color("card_bg"), width=440, bd=1, relief="solid", highlightbackground=c.get_color("border"), highlightthickness=1)
        f_left.pack(side="left", fill="y", padx=(0, 20))
        f_left.pack_propagate(False)

        tk.Label(f_left, text="⚡ Hızlı İşlem (POS)", font=("Segoe UI", 16, "bold"), bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).pack(pady=(20, 10))

        f_form = tk.Frame(f_left, bg=c.get_color("card_bg"), padx=25)
        f_form.pack(fill="both", expand=True)

        lbl_style = {"bg": c.get_color("card_bg"), "fg": c.get_color("fg_text"), "font": ("Segoe UI", 11, "bold")}

        # --- A. Tutar Girişi ---
        tk.Label(f_form, text="Tutar (TL):", **lbl_style).pack(anchor="w", pady=(0, 5))
        e_tut = tk.Entry(f_form, font=("Segoe UI", 32, "bold"), bg="#eff6ff", fg="#2563eb", relief="solid", bd=1, insertbackground="#2563eb", justify="right")
        e_tut.pack(fill="x", ipady=8, pady=(0, 15))
        e_tut.bind("<KeyRelease>", mask_para_birimi)
        e_tut.focus_set()

        # --- B. Hızlı İşlem Butonları (KUSURSUZ SİMETRİ İÇİN GRID AYARLARI) ---
        tk.Label(f_form, text="Sık Kullanılan İşlemler:", **lbl_style).pack(anchor="w", pady=(0, 5))
        f_pos_grid = tk.Frame(f_form, bg=c.get_color("card_bg"))
        f_pos_grid.pack(fill="x", pady=(0, 10))
        
        # Sütunları ve Satırları Eşit Böl
        f_pos_grid.grid_columnconfigure(0, weight=1, uniform="pos")
        f_pos_grid.grid_columnconfigure(1, weight=1, uniform="pos")
        f_pos_grid.grid_rowconfigure(0, weight=1)
        f_pos_grid.grid_rowconfigure(1, weight=1)
        f_pos_grid.grid_rowconfigure(2, weight=1)

        # Seçilenleri Tutan Değişkenler
        self.secili_tur = tk.StringVar(value="GİRİŞ")
        self.secili_kat = tk.StringVar(value="Nakit Satış")
        self.secili_yon = tk.StringVar(value="NAKİT")
        
        lbl_secim = tk.Label(f_form, text="GİRİŞ | Nakit Satış (NAKİT)", font=("Segoe UI", 10, "bold"), bg="#dcfce7", fg="#059669", pady=8, relief="flat", wraplength=380)
        lbl_secim.pack(fill="x", pady=(0, 15))

        def hizli_secim_yap(kat, tur, yon, bg_renk, fg_renk):
            self.secili_kat.set(kat)
            self.secili_tur.set(tur)
            self.secili_yon.set(yon)
            lbl_secim.config(text=f"{tur} | {kat} ({yon})", bg=bg_renk, fg=fg_renk)

        def popup_diger_ac():
            pop = tk.Toplevel(self.pencere)
            pop.title("Manuel Giriş")
            pop.geometry("400x200")
            pop.configure(bg="white")
            pop.transient(self.pencere); pop.grab_set(); pop.focus_force()
            
            x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 200
            y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 100
            pop.geometry(f"+{x}+{y}")
            
            tk.Label(pop, text="İşlem Adı / Kategori Giriniz", font=("Segoe UI", 12, "bold"), bg="white", fg="#334155").pack(pady=(20, 10))
            e_isim = tk.Entry(pop, font=("Segoe UI", 12), relief="solid", bd=1, bg="#f8fafc")
            e_isim.pack(fill="x", padx=30, ipady=5)
            e_isim.focus_set()
            
            def onayla():
                val = e_isim.get().strip()
                if val:
                    # Diğer seçildiğinde varsayılan olarak Çıkış/Nakit ayarlanır
                    hizli_secim_yap(val, "ÇIKIŞ", "NAKİT", "#f1f5f9", "#475569")
                pop.destroy()
                
            pop.bind('<Return>', lambda e: onayla())
            ModernButton(pop, text="KAYDET", command=onayla, width=150, height=35, bg_color="#10b981").pack(pady=20)

        # Yeni Liste: Dünden Kalanlar DEVİR olarak ayarlandı
        pos_btns = [
            ("Nakit Satış", "GİRİŞ", "NAKİT", "#dcfce7", "#059669", None),
            ("Kartlı Satış", "GİRİŞ", "KREDİ KARTI", "#dbeafe", "#2563eb", None),
            ("SGK / Kurum", "GİRİŞ", "HAVALE/EFT", "#e0e7ff", "#4338ca", None),
            ("Kasa Mevcudu", "DEVİR", "NAKİT", "#fef08a", "#ca8a04", None), # GİDER OLMAKTAN ÇIKARILDI
            ("İade", "ÇIKIŞ", "NAKİT", "#ffedd5", "#d97706", None), 
            ("Yemek / Çay", "ÇIKIŞ", "NAKİT", "#fce7f3", "#db2777", None),
            ("DİĞER ✍️", "", "", "#f1f5f9", "#475569", popup_diger_ac),
        ]

        for i, (kat, tur, yon, bg_col, fg_col, cmd) in enumerate(pos_btns):
            r, c_idx = divmod(i, 2)
            if cmd: # Eğer "Diğer" butonu ise özel fonksiyon çalışsın
                komut = cmd
            else:   # Normal buton ise seçim yapsın
                komut = lambda k=kat, t=tur, y=yon, b=bg_col, f=fg_col: hizli_secim_yap(k, t, y, b, f)
                
            # wraplength=160 ve justify="center" eklenerek uzun metinlerin alt satıra taşması sağlandı
            btn = tk.Button(f_pos_grid, text=kat, font=("Segoe UI", 11, "bold"), bg=bg_col, fg=fg_col, relief="flat", cursor="hand2", activebackground=fg_col, activeforeground="white", command=komut, wraplength=160, justify="center")
            btn.grid(row=r, column=c_idx, sticky="nsew", padx=4, pady=4, ipady=8)

        # --- C. GİRİŞ/ÇIKIŞ VE DETAY SEÇENEKLERİ ---
        f_detay = tk.Frame(f_form, bg=c.get_color("card_bg"))
        f_detay.pack(fill="x", pady=(5, 5))

        f_ikili = tk.Frame(f_detay, bg=c.get_color("card_bg"))
        f_ikili.pack(fill="x")
        
        f_tur_icerik = tk.Frame(f_ikili, bg=c.get_color("card_bg"))
        f_tur_icerik.pack(side="left", fill="x", expand=True, padx=(0, 5))
        tk.Label(f_tur_icerik, text="İşlem Türü:", font=("Segoe UI", 9), bg=c.get_color("card_bg"), fg="#64748b").pack(anchor="w")
        
        # Seçeneklere DEVİR eklendi
        cmb_tur = ttk.Combobox(f_tur_icerik, values=["GİRİŞ", "ÇIKIŞ", "DEVİR"], textvariable=self.secili_tur, state="readonly", font=("Segoe UI", 11))
        cmb_tur.pack(fill="x", ipady=4)

        f_yon_icerik = tk.Frame(f_ikili, bg=c.get_color("card_bg"))
        f_yon_icerik.pack(side="right", fill="x", expand=True, padx=(5, 0))
        tk.Label(f_yon_icerik, text="Ödeme Yöntemi:", font=("Segoe UI", 9), bg=c.get_color("card_bg"), fg="#64748b").pack(anchor="w")
        # Stringvar ile bağlandı
        cmb_yon = ttk.Combobox(f_yon_icerik, values=["NAKİT", "KREDİ KARTI", "HAVALE/EFT", "ÇEK"], textvariable=self.secili_yon, state="readonly", font=("Segoe UI", 11))
        cmb_yon.pack(fill="x", ipady=4)

        # --- D. Tarih ve Kasa Seçimi ---
        f_ayarlar = tk.Frame(f_form, bg=c.get_color("card_bg"))
        f_ayarlar.pack(fill="x", pady=(15, 0))

        tk.Label(f_ayarlar, text="Tarih:", **lbl_style).pack(side="left")
        e_tar = tk.Entry(f_ayarlar, font=("Segoe UI", 11, "bold"), width=11, justify="center", bg=c.get_color("input_bg"), fg=c.get_color("fg_text"), relief="solid", bd=1)
        e_tar.pack(side="left", padx=(5, 10), ipady=4)
        e_tar.insert(0, date.today().strftime("%Y-%m-%d"))
        e_tar.bind("<Double-Button-1>", lambda e: TakvimPopup(e_tar.winfo_toplevel(), e_tar))

        tk.Label(f_ayarlar, text="Kasa:", **lbl_style).pack(side="left", padx=(10, 5))
        # Sınırsız Kasa Listesi, Yeni Ekle ve Yönet Seçenekleri
        self.cmb_kasa_giris = ttk.Combobox(f_ayarlar, values=kasalari_getir() + ["+ Yeni Kasa Ekle", "⚙️ Kasaları Yönet"], textvariable=self.secili_kasa_giris, state="readonly", width=15, font=("Segoe UI", 10, "bold"))
        self.cmb_kasa_giris.pack(side="left", ipady=4)
        self.cmb_kasa_giris.bind("<<ComboboxSelected>>", lambda e: kasa_secim_kontrol(e, self.cmb_kasa_giris, is_filter=False))

        # --- E. Dev Kaydet Butonu ---
        def islem_ekle():
            try:
                tutar = temizle_para(e_tut.get())
                
                if tutar <= 0: 
                    messagebox.showwarning("Hata", "Lütfen geçerli bir tutar girin.")
                    return
                    
                if not self.secili_kat.get(): 
                    messagebox.showwarning("Hata", "Lütfen işlem adını (Kategori) belirleyin.")
                    return

                self.imlec_finans.execute("""
                    INSERT INTO kasa_defteri (tarih, islem_turu, kategori, aciklama, tutar, odeme_yontemi, kasa_adi) 
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (e_tar.get(), self.secili_tur.get(), self.secili_kat.get(), "-", tutar, self.secili_yon.get(), self.secili_kasa_giris.get()))
                
                self.baglanti_finans.commit()
                
                # Başarılı kayıt sonrası tabloyu yenile ve tutar kutusunu temizle
                yenile_tetik()
                e_tut.delete(0, tk.END)
                e_tut.focus_set()
                
            except Exception as e: 
                messagebox.showerror("Hata", f"Kayıt hatası: {e}")

        ModernButton(f_form, text="✅ SİSTEME İŞLE", command=islem_ekle, bg_color="#0f172a", width=340, height=50).pack(side="bottom", pady=20)


        # ---------------------------------------------------------------------
        # 2. SAĞ PANEL: İSTATİSTİKLER VE AKORDİYON TABLO
        # ---------------------------------------------------------------------
        f_right = tk.Frame(main_container, bg=c.get_color("bg_main"))
        f_right.pack(side="right", fill="both", expand=True)

        header = tk.Frame(f_right, bg=c.get_color("bg_main"))
        header.pack(fill="x", pady=(0, 15))

        tk.Label(header, text="Kasa Hareketleri", font=("Segoe UI", 22, "bold"), bg=c.get_color("bg_main"), fg="#334155").pack(side="left")

        frm_filter = tk.Frame(header, bg=c.get_color("bg_main"))
        frm_filter.pack(side="right")

        tk.Label(frm_filter, text="Kasa:", font=("Segoe UI", 11, "bold"), bg=c.get_color("bg_main"), fg="#64748b").pack(side="left", padx=5)
        
        # Sınırsız Kasa Filtreleme Menüsü
        self.cmb_kasa_filtre = ttk.Combobox(frm_filter, values=["Tümü"] + kasalari_getir(), textvariable=self.var_kasa_filtre, state="readonly", width=10, font=("Segoe UI", 11))
        self.cmb_kasa_filtre.pack(side="left", padx=5, ipady=5)
        self.cmb_kasa_filtre.bind("<<ComboboxSelected>>", lambda e: kasa_secim_kontrol(e, self.cmb_kasa_filtre, is_filter=True))

        tk.Label(frm_filter, text="Dönem:", font=("Segoe UI", 11, "bold"), bg=c.get_color("bg_main"), fg="#64748b").pack(side="left", padx=5)
        
        # Geçmiş 2 yılın aylarını (24 ay) otomatik hesapla
        bugun_tarihi = date.today()
        gecmis_donemler = []
        for i in range(24):
            ay = bugun_tarihi.month - i
            yil = bugun_tarihi.year
            while ay <= 0:
                ay += 12
                yil -= 1
            gecmis_donemler.append(f"{yil}-{ay:02d}")

        # Açılır Menü (Combobox) oluştur ve listeyi içine ekle
        cmb_donem = ttk.Combobox(frm_filter, values=gecmis_donemler, textvariable=self.var_donem, state="readonly", width=10, font=("Segoe UI", 11))
        cmb_donem.pack(side="left", padx=5, ipady=5)
        cmb_donem.bind("<<ComboboxSelected>>", lambda e: yenile_tetik()) # Seçildiği an otomatik yeniler
        
        ModernButton(frm_filter, text="🔄 YENİLE", command=yenile_tetik, width=100, height=40, bg_color="#64748b").pack(side="left", padx=10)

        # --- İstatistik Kartları ---
        stats_frame = tk.Frame(f_right, bg=c.get_color("bg_main"))
        stats_frame.pack(fill="x", pady=(0, 15))
        
        def kart_yap(parent, baslik, renk):
            f = tk.Frame(parent, bg="white", padx=10, pady=20, bd=1, relief="solid"); f.pack(side="left", fill="x", expand=True, padx=3)
            tk.Label(f, text=baslik, font=("Segoe UI", 10, "bold"), bg="white", fg="#64748b").pack(anchor="center")
            lbl_deger = tk.Label(f, text="0.00 ₺", font=("Segoe UI", 20, "bold"), bg="white", fg=renk); lbl_deger.pack(anchor="center", pady=(5,0))
            return lbl_deger
            
        self.lbl_giris = kart_yap(stats_frame, "TOPLAM GİRİŞ (+)", "#059669")
        self.lbl_cikis = kart_yap(stats_frame, "GERÇEK GİDER (-)", "#e11d48")
        self.lbl_bakiye = kart_yap(stats_frame, "BUGÜNÜN KARI", "#2563eb")

        # --- TABLO ---
        tree_frame = tk.Frame(f_right, bg="white", bd=1, relief="solid")
        tree_frame.pack(fill="both", expand=True)
        
        style = ttk.Style()
        style.configure("Kasa.Treeview", rowheight=45, font=("Segoe UI", 12))
        style.configure("Kasa.Treeview.Heading", font=("Segoe UI", 12, "bold"))

        cols = ("ID", "TARIH", "KATEGORI", "TUTAR")
        self.tree_kasa = ttk.Treeview(tree_frame, columns=cols, show="tree headings", height=12, style="Kasa.Treeview")
        
        self.tree_kasa.heading("#0", text=" 📅 GÜN"); self.tree_kasa.column("#0", width=250, anchor="w")
        self.tree_kasa.heading("ID", text="ID"); self.tree_kasa.column("ID", width=0, stretch=False)
        self.tree_kasa.heading("TARIH", text="Tarih"); self.tree_kasa.column("TARIH", width=0, stretch=False)
        self.tree_kasa.heading("KATEGORI", text="İşlem Detayı (Kategori & Yöntem)"); self.tree_kasa.column("KATEGORI", width=600, anchor="w")
        self.tree_kasa.heading("TUTAR", text="Tutar"); self.tree_kasa.column("TUTAR", width=200, anchor="e")
        
        self.tree_kasa.tag_configure("giris", font=("Segoe UI", 13, "bold"), background="#ffffff", foreground="#059669") 
        self.tree_kasa.tag_configure("cikis", font=("Segoe UI", 13, "bold"), background="#fff1f2", foreground="#e11d48") 
        self.tree_kasa.tag_configure("devir", font=("Segoe UI", 13, "bold"), background="#fefce8", foreground="#ca8a04") # YENİ TABLO RENGİ (Sarımsı)
        self.tree_kasa.tag_configure("gun_baslik", font=("Segoe UI", 14, "bold"), background="#f8fafc", foreground="#0f172a") 
        
        sc = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_kasa.yview); self.tree_kasa.configure(yscroll=sc.set)
        sc.pack(side="right", fill="y"); self.tree_kasa.pack(fill="both", expand=True)

        menu = tk.Menu(self.pencere, tearoff=0, font=("Segoe UI", 11))
        menu.add_command(label="🗑️ Bu Kaydı Sil", command=self._kasa_kayit_sil)
        def on_right_click(event):
            item = self.tree_kasa.identify_row(event.y)
            if item:
                self.tree_kasa.selection_set(item)
                menu.post(event.x_root, event.y_root)
        self.tree_kasa.bind("<Button-3>", on_right_click)
        
        yenile_tetik()

    def _kasa_verilerini_yukle(self, donem_filtresi):
        for i in self.tree_kasa.get_children(): 
            self.tree_kasa.delete(i)
        
        secili_kasa = self.var_kasa_filtre.get() if hasattr(self, 'var_kasa_filtre') else "Tümü"
        
        if secili_kasa == "Tümü":
            sorgu = "SELECT id, tarih, islem_turu, kategori, aciklama, tutar, odeme_yontemi, kasa_adi FROM kasa_defteri WHERE tarih LIKE ? ORDER BY tarih DESC, id DESC"
            params = (f"{donem_filtresi}%",)
        else:
            sorgu = "SELECT id, tarih, islem_turu, kategori, aciklama, tutar, odeme_yontemi, kasa_adi FROM kasa_defteri WHERE tarih LIKE ? AND kasa_adi=? ORDER BY tarih DESC, id DESC"
            params = (f"{donem_filtresi}%", secili_kasa)

        try:
            self.imlec_finans.execute(sorgu, params)
            kayitlar = self.imlec_finans.fetchall()
        except: 
            kayitlar = []
        
        top_giris = 0.0
        top_cikis = 0.0
        top_devir = 0.0 # DEVİR (mevcut kasa) İÇİN YENİ DEĞİŞKEN
        gunluk_havuz = {}
        
        for row in kayitlar:
            rid, tar, tur, kat, ack, tut, yontem, k_adi = row
            yontem = yontem if yontem else "-"
            k_adi = k_adi if k_adi else "Kasa 1" 
            tur = str(tur).upper() if tur else "ÇIKIŞ"
            tut = float(tut) if tut else 0.0
            
            if tar not in gunluk_havuz:
                gunluk_havuz[tar] = {'liste': [], 'n_giris': 0, 'kk_giris': 0, 'diger_giris': 0, 'cikis': 0}
                
            gunluk_havuz[tar]['liste'].append(row)
            
            if tur == "GİRİŞ":
                top_giris += tut
                if "NAKİT" in yontem: 
                    gunluk_havuz[tar]['n_giris'] += tut
                elif "KART" in yontem: 
                    gunluk_havuz[tar]['kk_giris'] += tut
                else: 
                    gunluk_havuz[tar]['diger_giris'] += tut
            elif tur == "DEVİR" or kat == "Kasa Mevcudu":
                # GEÇMİŞTE ÇIKIŞ OLARAK KAYDEDİLENLERİ BİLE DEVİR KABUL ET Kİ GEÇMİŞ DÜZELSİN
                top_devir += tut
                # Kasada kalan para "Kasadan çıktığı için" net rakamı eksiltir
                gunluk_havuz[tar]['cikis'] += tut 
            else:
                # GERÇEK GİDERLER (Yemek, Fatura vb.)
                top_cikis += tut
                gunluk_havuz[tar]['cikis'] += tut

        sirali_tarihler = sorted(gunluk_havuz.keys(), reverse=True)
        
        for tar in sirali_tarihler:
            veri = gunluk_havuz[tar]
            gunluk_net = (veri['n_giris'] + veri['kk_giris'] + veri['diger_giris']) - veri['cikis']
            
            try: 
                tar_gosterim = datetime.strptime(tar, "%Y-%m-%d").strftime("%d.%m.%Y")
            except: 
                tar_gosterim = tar
            
            # ANA SATIR: Sadece Tarih ve Net Bakiye
            parent_id = self.tree_kasa.insert("", "end", text=f" 📅 {tar_gosterim}", 
                                              values=("", "", "", f"{gunluk_net:,.2f} ₺"), 
                                              tags=("gun_baslik",))
            
            # ALT SATIRLAR: İşlemler
            for row in veri['liste']:
                rid, r_tar, r_tur, kat, ack, tut, yontem, k_adi = row
                r_tur = str(r_tur).upper() if r_tur else "ÇIKIŞ"
                tut = float(tut) if tut else 0.0
                k_adi = k_adi if k_adi else "Kasa 1"
                
                # Dinamik Etiket Belirleme
                if r_tur == "GİRİŞ":
                    tag = "giris"
                elif r_tur == "DEVİR" or kat == "Kasa Mevcudu":
                    tag = "devir"
                else:
                    tag = "cikis"
                
                gosterilecek_kategori = f"[{k_adi}] {kat} ({yontem})" if yontem and yontem != "-" else f"[{k_adi}] {kat}"

                # Sıralama: ID, TARIH, KATEGORI, TUTAR
                self.tree_kasa.insert(parent_id, "end", values=(rid, r_tar, gosterilecek_kategori, f"{tut:,.2f} ₺"), tags=(tag,))

        # Üst istatistik kartlarını güncelle
        if hasattr(self, 'lbl_giris') and self.lbl_giris.winfo_exists():
            self.lbl_giris.config(text=f"{top_giris:,.2f} ₺")
            self.lbl_cikis.config(text=f"{top_cikis:,.2f} ₺")
            if hasattr(self, 'lbl_devir'):
                self.lbl_devir.config(text=f"{top_devir:,.2f} ₺")
                
            # Patronun cebine giren net bugünün karı = (Tüm Girişler) - (Gerçek Giderler) - (Kasada Yarına Bırakılanlar)
            bakiye = top_giris - (top_cikis + top_devir)
            self.lbl_bakiye.config(text=f"{bakiye:,.2f} ₺", fg="#10b981" if bakiye >= 0 else "#ef4444")

    def _kasa_kayit_sil(self):
        """Kasa defterinden sağ tık ile kayıt silme motoru"""
        sel = self.tree_kasa.selection()
        if not sel: return
        
        item = self.tree_kasa.item(sel[0])
        rid = item['values'][0]
        
        # Eğer tıklanan satır ana başlık satırıysa (Tarih yazan satırsa) hata vermesini engelle
        if not rid or str(rid).strip() == "": 
            messagebox.showwarning("Uyarı", "Günlük özet başlığını silemezsiniz.\nLütfen silmek istediğiniz işlemin (alt satırın) üzerine sağ tıklayın.")
            return
            
        if messagebox.askyesno("Onay", "Bu işlemi kalıcı olarak silmek istediğinize emin misiniz?"):
            try:
                self.imlec_finans.execute("DELETE FROM kasa_defteri WHERE id=?", (rid,))
                self.baglanti_finans.commit()
                
                # Listeyi silinmiş haliyle anında yenile
                if hasattr(self, 'var_donem'):
                    self._kasa_verilerini_yukle(self.var_donem.get())
                    
                messagebox.showinfo("Başarılı", "Kayıt başarıyla silindi.")
            except Exception as e:
                messagebox.showerror("Hata", f"Silme işlemi başarısız oldu:\n{e}")    

    # =========================================================================
    # MODÜL 15: KREDİ & FİNANSMAN SİMÜLATÖRÜ (SADECE HESAPLAMA)
    # =========================================================================
    def arayuz_kredi_hesaplama(self):
        c = TM
        for w in self.content_area.winfo_children(): w.destroy()

        # Bireysel (İhtiyaç/Taşıt) = %15 BSMV + %15 KKDF -> Çarpan 1.30
        # Ticari = %5 BSMV -> Çarpan 1.05
        # Konut = Vergisiz -> Çarpan 1.00
        kredi_veriler = { 
            "İhtiyaç": {"oran": 4.29, "vergi_carpani": 1.30}, 
            "Taşıt": {"oran": 3.49, "vergi_carpani": 1.30}, 
            "Konut": {"oran": 3.05, "vergi_carpani": 1.00}, 
            "Ticari": {"oran": 3.59, "vergi_carpani": 1.05} 
        }

        self.hesap_modu = tk.StringVar(value="TUTAR") 
        self.k_tutar_taksit = tk.StringVar()
        self.k_vade = tk.StringVar(value="12")
        self.k_oran = tk.StringVar()
        self.k_tur = tk.StringVar(value="İhtiyaç")

        # Başlık
        header = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        header.pack(fill="x", pady=(0, 20))
        tk.Label(header, text="🏦 Kredi Hesaplama Simülatörü", font=FONT_HEAD, bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left")

        main_body = tk.Frame(self.content_area, bg=c.get_color("bg_main"), padx=20)
        main_body.pack(fill="both", expand=True)

        # Mod Değiştirici
        toggle_frame = tk.Frame(main_body, bg=c.get_color("card_bg"), bd=1, relief="solid", highlightbackground=c.get_color("border"), highlightthickness=1)
        toggle_frame.pack(fill="x", pady=(0, 20))
        
        btn_tutar = tk.Button(toggle_frame, text="💰 Çekilecek Tutara Göre Hesapla", font=("Segoe UI", 11, "bold"), bd=0, cursor="hand2", pady=12)
        btn_taksit = tk.Button(toggle_frame, text="📅 Ödeyebileceğiniz Taksite Göre Hesapla", font=("Segoe UI", 11, "bold"), bd=0, cursor="hand2", pady=12)
        btn_tutar.pack(side="left", fill="x", expand=True)
        btn_taksit.pack(side="left", fill="x", expand=True)

        split_frame = tk.Frame(main_body, bg=c.get_color("bg_main"))
        split_frame.pack(fill="both", expand=True)

        f_form = tk.Frame(split_frame, bg=c.get_color("card_bg"), padx=40, pady=30, bd=1, relief="solid", highlightbackground=c.get_color("border"), highlightthickness=1)
        f_form.pack(side="left", fill="both", expand=True, padx=(0, 15))

        f_res = tk.Frame(split_frame, bg="#0f172a", width=450)
        f_res.pack(side="right", fill="y")
        f_res.pack_propagate(False)
        f_res_inner = tk.Frame(f_res, bg="#0f172a", padx=40, pady=40)
        f_res_inner.pack(fill="both", expand=True)

        # SAĞ EKRAN: SONUÇLAR
        tk.Label(f_res_inner, text="HESAP ÖZETİ", font=("Segoe UI", 14, "bold"), bg="#0f172a", fg="#94a3b8").pack(anchor="w", pady=(0, 20))
        lbl_res_main_title = tk.Label(f_res_inner, text="Ödenecek Aylık Taksit Tutarı:", font=("Segoe UI", 12), bg="#0f172a", fg="#cbd5e1")
        lbl_res_main_title.pack(anchor="w")
        lbl_res_main_val = tk.Label(f_res_inner, text="0.00 ₺", font=("Segoe UI", 36, "bold"), bg="#0f172a", fg="#34d399")
        lbl_res_main_val.pack(anchor="w", pady=(5, 30))

        tk.Frame(f_res_inner, height=1, bg="#334155").pack(fill="x", pady=10)

        def detail_row(parent, title):
            row = tk.Frame(parent, bg="#0f172a")
            row.pack(fill="x", pady=15)
            tk.Label(row, text=title, font=("Segoe UI", 11), bg="#0f172a", fg="#94a3b8").pack(side="left")
            val_lbl = tk.Label(row, text="0.00 ₺", font=("Segoe UI", 14, "bold"), bg="#0f172a", fg="white")
            val_lbl.pack(side="right")
            return val_lbl

        lbl_res_toplam = detail_row(f_res_inner, "Toplam Geri Ödeme:")
        lbl_res_faiz = detail_row(f_res_inner, "Toplam Faiz + Vergi:")

        lbl_info = tk.Label(f_res_inner, text="* Kredi maliyetleri tahmini olup, kuruma göre farklılık gösterebilir. BSMV/KKDF oranları seçilen türe göre (Ticari, Bireysel vb.) otomatik yansıtılır.", font=("Segoe UI", 9, "italic"), bg="#0f172a", fg="#64748b", wraplength=350, justify="left")
        lbl_info.pack(side="bottom", anchor="w", pady=20)

        # SOL EKRAN: GİRİŞ FORMU
        lbl_style = {"bg": c.get_color("card_bg"), "fg": c.get_color("fg_text"), "font": ("Segoe UI", 11, "bold")}
        
        tk.Label(f_form, text="Kredi Türü Seçimi:", **lbl_style).pack(anchor="w", pady=(0, 8))
        f_types = tk.Frame(f_form, bg=c.get_color("card_bg"))
        f_types.pack(fill="x", pady=(0, 20))
        
        type_btns = []
        def select_type(btn, tip, data):
            self.k_tur.set(tip)
            self.k_oran.set(str(data["oran"]))
            for b in type_btns:
                if b == btn: b.config(bg=c.get_color("btn_primary"), fg="white")
                else: b.config(bg=c.get_color("input_bg"), fg=c.get_color("fg_text"))
            hesapla()
            
        for tip, data in kredi_veriler.items():
            b = tk.Button(f_types, text=tip, font=("Segoe UI", 10, "bold"), bd=1, relief="solid", cursor="hand2", padx=15, pady=8)
            b.config(command=lambda btn=b, t=tip, d=data: select_type(btn, t, d))
            b.pack(side="left", padx=(0, 5))
            type_btns.append(b)

        lbl_dynamic_input = tk.Label(f_form, text="Çekilecek Kredi Tutarı (TL):", **lbl_style)
        lbl_dynamic_input.pack(anchor="w", pady=(10, 5))
        e_dynamic = tk.Entry(f_form, textvariable=self.k_tutar_taksit, font=("Segoe UI", 18, "bold"), bg=c.get_color("input_bg"), fg=c.get_color("input_fg"), relief="solid", bd=1, justify="right")
        e_dynamic.pack(fill="x", ipady=10, pady=(0, 20))
        e_dynamic.bind("<KeyRelease>", lambda e: [mask_para_birimi(e), hesapla()])

        tk.Label(f_form, text="Vade (Ay):", **lbl_style).pack(anchor="w", pady=(10, 5))
        f_vade = tk.Frame(f_form, bg=c.get_color("card_bg"))
        f_vade.pack(fill="x", pady=(0, 20))
        
        vade_btns = []
        def select_vade(btn, val):
            self.k_vade.set(str(val))
            for b in vade_btns:
                if b == btn: b.config(bg="#f59e0b", fg="white")
                else: b.config(bg=c.get_color("input_bg"), fg=c.get_color("fg_text"))
            hesapla()

        for m in [12, 24, 36, 48, 60]:
            b = tk.Button(f_vade, text=f"{m} Ay", font=("Segoe UI", 10, "bold"), bd=1, relief="solid", cursor="hand2", padx=15, pady=8)
            b.config(command=lambda btn=b, val=m: select_vade(btn, val))
            b.pack(side="left", padx=(0, 5))
            vade_btns.append(b)
            
        e_vade = tk.Entry(f_vade, textvariable=self.k_vade, font=("Segoe UI", 13, "bold"), bg=c.get_color("input_bg"), fg=c.get_color("input_fg"), relief="solid", bd=1, justify="center")
        e_vade.pack(side="left", fill="both", expand=True, ipady=8, padx=(10, 0))
        e_vade.bind("<KeyRelease>", lambda e: [b.config(bg=c.get_color("input_bg"), fg=c.get_color("fg_text")) for b in vade_btns] or hesapla())

        tk.Label(f_form, text="Aylık Faiz Oranı (%):", **lbl_style).pack(anchor="w", pady=(10, 5))
        e_oran = tk.Entry(f_form, textvariable=self.k_oran, font=("Segoe UI", 16, "bold"), bg=c.get_color("input_bg"), fg=c.get_color("input_fg"), relief="solid", bd=1, justify="right")
        e_oran.pack(fill="x", ipady=8, pady=(0, 20))
        e_oran.bind("<KeyRelease>", lambda e: hesapla())

        def reset_results():
            lbl_res_main_val.config(text="0.00 ₺")
            lbl_res_toplam.config(text="0.00 ₺")
            lbl_res_faiz.config(text="0.00 ₺")

        def hesapla(*args):
            try:
                val_str = self.k_tutar_taksit.get().replace(".", "").replace(",", ".")
                if not val_str:
                    reset_results(); return
                
                val = float(val_str)
                n = int(self.k_vade.get())
                tur = self.k_tur.get()
                ham_faiz = float(self.k_oran.get().replace(",", "."))
                
                if n <= 0 or ham_faiz < 0: raise ValueError
                
                vergi_carpani = kredi_veriler[tur]["vergi_carpani"]
                r = (ham_faiz / 100.0) * vergi_carpani
                
                if self.hesap_modu.get() == "TUTAR":
                    P = val
                    taksit = P * (r * (1 + r)**n) / ((1 + r)**n - 1) if r > 0 else P / n
                    toplam_odeme = taksit * n
                    toplam_faiz = toplam_odeme - P
                    
                    lbl_res_main_title.config(text="Ödenecek Aylık Taksit Tutarı:")
                    lbl_res_main_val.config(text=f"{taksit:,.2f} ₺")
                else:
                    A = val
                    P = A * ((1 + r)**n - 1) / (r * (1 + r)**n) if r > 0 else A * n
                    toplam_odeme = A * n
                    toplam_faiz = toplam_odeme - P
                    
                    lbl_res_main_title.config(text="Çekilebilecek Maksimum Kredi:")
                    lbl_res_main_val.config(text=f"{P:,.2f} ₺")
                
                lbl_res_toplam.config(text=f"{toplam_odeme:,.2f} ₺")
                lbl_res_faiz.config(text=f"{toplam_faiz:,.2f} ₺")
            except Exception:
                reset_results()

        def set_mode(mod):
            self.hesap_modu.set(mod)
            if mod == "TUTAR":
                btn_tutar.config(bg=c.get_color("btn_primary"), fg="white")
                btn_taksit.config(bg=c.get_color("card_bg"), fg=c.get_color("fg_text"))
                lbl_dynamic_input.config(text="İhtiyacınız Olan Kredi Tutarı (TL):")
            else:
                btn_taksit.config(bg=c.get_color("btn_primary"), fg="white")
                btn_tutar.config(bg=c.get_color("card_bg"), fg=c.get_color("fg_text"))
                lbl_dynamic_input.config(text="Aylık Ödeyebileceğiniz Maksimum Taksit (TL):")
            self.k_tutar_taksit.set("") 
            hesapla()

        btn_tutar.config(command=lambda: set_mode("TUTAR"))
        btn_taksit.config(command=lambda: set_mode("TAKSIT"))

        set_mode("TUTAR")
        if type_btns: select_type(type_btns[0], "İhtiyaç", kredi_veriler["İhtiyaç"])
        if vade_btns: select_vade(vade_btns[0], 12)


    def arayuz_mevcut_krediler(self):
        c = TM
        for w in self.content_area.winfo_children(): w.destroy()

        # =====================================================================
        # OTOMATİK DÜZELTME YAMASI (Eski kaydedilenleri Depo'dan Gider'e taşır)
        # =====================================================================
        try:
            self.imlec_finans.execute("UPDATE odemeler SET aciklama = satir_notu, satir_notu = 'MANUEL_GIDER' WHERE satir_notu LIKE 'KREDI:%'")
            self.baglanti_finans.commit()
            
            # Gider sekmesinde anında görünmesi için "Depo Ödemeleri" sekmesini yenilemek isteyebiliriz
            if hasattr(self, 'notebook'):
                self.sekmeleri_guncelle()
        except: pass

        kredi_veriler = {
            "İhtiyaç": {"oran": 4.29, "vergi_carpani": 1.30},
            "Taşıt": {"oran": 3.49, "vergi_carpani": 1.30},
            "Konut": {"oran": 3.05, "vergi_carpani": 1.00},
            "Ticari": {"oran": 3.59, "vergi_carpani": 1.05}
        }

        # --- ÜST BAŞLIK ---
        header = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        header.pack(fill="x", pady=(0, 15))
        tk.Label(header, text="📂 Mevcut Kredilerim & Sisteme İşle", font=FONT_HEAD, bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left")

        # --- ANA KONTEYNER (2 SÜTUNLU YAPI) ---
        main_container = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        main_container.pack(fill="both", expand=True, padx=10, pady=5)

        # =====================================================================
        # SOL PANEL: KREDİ EKLEME FORMU (SABİT GENİŞLİK, ASLA BOZULMAZ)
        # =====================================================================
        f_left = tk.Frame(main_container, bg=c.get_color("card_bg"), width=440, bd=1, relief="solid", highlightbackground=c.get_color("border"), highlightthickness=1)
        f_left.pack(side="left", fill="y", padx=(0, 15))
        f_left.pack_propagate(False) 

        tk.Label(f_left, text="Krediyi Sisteme İşle", font=("Segoe UI", 14, "bold"), bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).pack(pady=(25, 20))

        # --- FORM ELEMANLARI ---
        f_form = tk.Frame(f_left, bg=c.get_color("card_bg"), padx=25)
        f_form.pack(fill="both", expand=True)

        lbl_style = {"bg": c.get_color("card_bg"), "fg": c.get_color("fg_text"), "font": ("Segoe UI", 9, "bold")}
        entry_style = {"font": ("Segoe UI", 11), "bg": c.get_color("input_bg"), "fg": c.get_color("input_fg"), "insertbackground": c.get_color("input_fg"), "relief": "solid", "bd": 1}

        def form_satir(parent, label_text):
            f = tk.Frame(parent, bg=c.get_color("card_bg"))
            f.pack(fill="x", pady=(0, 12))
            tk.Label(f, text=label_text, **lbl_style).pack(anchor="w", pady=(0, 3))
            return f

        f_ad = form_satir(f_form, "Banka / Kredi Adı (Örn: İş Bankası İhtiyaç):")
        e_ad = tk.Entry(f_ad, **entry_style)
        e_ad.pack(fill="x", ipady=4)

        f_row2 = tk.Frame(f_form, bg=c.get_color("card_bg"))
        f_row2.pack(fill="x", pady=(0, 12))

        f_tur = tk.Frame(f_row2, bg=c.get_color("card_bg"))
        f_tur.pack(side="left", fill="x", expand=True, padx=(0, 5))
        tk.Label(f_tur, text="Kredi Türü:", **lbl_style).pack(anchor="w", pady=(0, 3))
        cmb_tur = ttk.Combobox(f_tur, values=list(kredi_veriler.keys()), state="readonly", font=("Segoe UI", 11))
        cmb_tur.pack(fill="x", ipady=3)
        cmb_tur.current(0)

        f_vade = tk.Frame(f_row2, bg=c.get_color("card_bg"))
        f_vade.pack(side="right", fill="x", expand=True, padx=(5, 0))
        tk.Label(f_vade, text="Vade (Ay):", **lbl_style).pack(anchor="w", pady=(0, 3))
        e_vade = tk.Entry(f_vade, **entry_style)
        e_vade.pack(fill="x", ipady=4)

        f_mod = form_satir(f_form, "Hesaplama Yöntemi:")
        cmb_mod = ttk.Combobox(f_mod, values=["Aylık Taksiti Biliyorum (Kendim Yazacağım)", "Çekilen Tutar ve Faizden Sistemi Hesaplasın"], state="readonly", font=("Segoe UI", 10))
        cmb_mod.pack(fill="x", ipady=3)
        cmb_mod.current(0)

        f_row3 = tk.Frame(f_form, bg=c.get_color("card_bg"))
        f_row3.pack(fill="x", pady=(0, 12))

        f_tutar = tk.Frame(f_row3, bg=c.get_color("card_bg"))
        f_tutar.pack(side="left", fill="x", expand=True, padx=(0, 5))
        lbl_tutar = tk.Label(f_tutar, text="Aylık Taksit Tutarı (TL):", **lbl_style)
        lbl_tutar.pack(anchor="w", pady=(0, 3))
        e_tutar = tk.Entry(f_tutar, font=("Segoe UI", 12, "bold"), bg="#eff6ff", fg="#1e3a8a", relief="solid", bd=1, insertbackground="#1e3a8a")
        e_tutar.pack(fill="x", ipady=4)
        e_tutar.bind("<KeyRelease>", mask_para_birimi)

        f_faiz = tk.Frame(f_row3, bg=c.get_color("card_bg"))
        f_faiz.pack(side="right", fill="x", expand=True, padx=(5, 0))
        lbl_faiz = tk.Label(f_faiz, text="Aylık Faiz (%):", **lbl_style)
        lbl_faiz.pack(anchor="w", pady=(0, 3))
        e_faiz = tk.Entry(f_faiz, **entry_style)
        e_faiz.pack(fill="x", ipady=4)

        def mod_degisti(e=None):
            if cmb_mod.current() == 0:
                lbl_tutar.config(text="Aylık Taksit Tutarı (TL):")
                lbl_faiz.config(text="-- Gerek Yok --", fg="gray")
                e_faiz.delete(0, tk.END)
                e_faiz.config(state="disabled")
            else:
                lbl_tutar.config(text="Çekilen Ana Para (TL):")
                lbl_faiz.config(text="Aylık Faiz Oranı (%):", fg=c.get_color("fg_text"))
                e_faiz.config(state="normal")

        cmb_mod.bind("<<ComboboxSelected>>", mod_degisti)
        mod_degisti()

        f_row4 = tk.Frame(f_form, bg=c.get_color("card_bg"))
        f_row4.pack(fill="x", pady=(0, 12))

        f_masraf = tk.Frame(f_row4, bg=c.get_color("card_bg"))
        f_masraf.pack(side="left", fill="x", expand=True, padx=(0, 5))
        tk.Label(f_masraf, text="Tahsis Masrafı (TL):", **lbl_style).pack(anchor="w", pady=(0, 3))
        e_masraf = tk.Entry(f_masraf, **entry_style)
        e_masraf.pack(fill="x", ipady=4)
        e_masraf.insert(0, "0")
        e_masraf.bind("<KeyRelease>", mask_para_birimi)

        f_sigorta = tk.Frame(f_row4, bg=c.get_color("card_bg"))
        f_sigorta.pack(side="right", fill="x", expand=True, padx=(5, 0))
        tk.Label(f_sigorta, text="Sigorta Ücreti (TL):", **lbl_style).pack(anchor="w", pady=(0, 3))
        e_sigorta = tk.Entry(f_sigorta, **entry_style)
        e_sigorta.pack(fill="x", ipady=4)
        e_sigorta.insert(0, "0")
        e_sigorta.bind("<KeyRelease>", mask_para_birimi)

        f_tarih = form_satir(f_form, "İlk Taksit Ödeme Tarihi (YYYY-AA-GG):")
        f_date_inner = tk.Frame(f_tarih, bg=c.get_color("card_bg"), bd=1, relief="solid")
        f_date_inner.pack(fill="x")
        e_tarih = tk.Entry(f_date_inner, font=("Segoe UI", 11), relief="flat", bg=c.get_color("input_bg"), fg=c.get_color("input_fg"), insertbackground=c.get_color("input_fg"))
        e_tarih.pack(side="left", fill="both", expand=True, ipady=4, padx=2)
        e_tarih.insert(0, (date.today() + timedelta(days=30)).strftime("%Y-%m-%d"))
        tarih_secici_bagla(f_date_inner, e_tarih)
        e_tarih.bind("<KeyRelease>", mask_tarih_otomatik)

        f_btn_wrapper = tk.Frame(f_left, bg=c.get_color("card_bg"), pady=25)
        f_btn_wrapper.pack(side="bottom", fill="x")
        ModernButton(f_btn_wrapper, text="✅ KREDİYİ SİSTEME KAYDET", command=lambda: krediyi_kaydet(), bg_color="#10b981", width=340, height=45).pack(anchor="center")

        # =====================================================================
        # SAĞ PANEL: SİSTEMDEKİ KREDİLER TABLOSU
        # =====================================================================
        f_right = tk.Frame(main_container, bg=c.get_color("card_bg"), bd=1, relief="solid", highlightbackground=c.get_color("border"), highlightthickness=1)
        f_right.pack(side="right", fill="both", expand=True)

        f_right_top = tk.Frame(f_right, bg=c.get_color("card_bg"), pady=15, padx=20)
        f_right_top.pack(fill="x")
        tk.Label(f_right_top, text="Sistemdeki Aktif Krediler", font=("Segoe UI", 14, "bold"), bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).pack(side="left")
        ModernButton(f_right_top, text="🗑️ Krediyi Sil", command=lambda: krediyi_sil(), bg_color=c.get_color("btn_danger"), width=150, height=35).pack(side="right")

        f_tree = tk.Frame(f_right, bg=c.get_color("card_bg"), padx=20, pady=5)
        f_tree.pack(fill="both", expand=True, pady=(0, 20))

        cols = ("KREDİ ADI", "KALAN İŞLEM", "AYLIK TUTAR", "KALAN TOPLAM BORÇ")
        tree_krediler = ttk.Treeview(f_tree, columns=cols, show="headings", height=15)

        tree_krediler.heading("KREDİ ADI", text="Banka / Kredi Adı")
        tree_krediler.heading("KALAN İŞLEM", text="Kalan Taksit")
        tree_krediler.heading("AYLIK TUTAR", text="Taksit Tutarı")
        tree_krediler.heading("KALAN TOPLAM BORÇ", text="Kalan Toplam Borç")

        tree_krediler.column("KREDİ ADI", width=250)
        tree_krediler.column("KALAN İŞLEM", width=120, anchor="center")
        tree_krediler.column("AYLIK TUTAR", width=130, anchor="e")
        tree_krediler.column("KALAN TOPLAM BORÇ", width=150, anchor="e")

        sc_kredi = ttk.Scrollbar(f_tree, orient="vertical", command=tree_krediler.yview)
        tree_krediler.configure(yscroll=sc_kredi.set)
        sc_kredi.pack(side="right", fill="y")
        tree_krediler.pack(side="left", fill="both", expand=True)

        # =====================================================================
        # HESAPLAMA VE VERİTABANI İŞLEMLERİ
        # =====================================================================
        def listeyi_guncelle():
            for i in tree_krediler.get_children(): tree_krediler.delete(i)
            
            # Hem toplam taksiti hem kalan taksiti hesaplayan gelişmiş SQL sorgusu
            sorgu = """
                SELECT 
                    fatura_adi, 
                    SUM(CASE WHEN durum='ODENMEDİ' THEN 1 ELSE 0 END) as kalan_taksit,
                    COUNT(id) as toplam_taksit,
                    MIN(tutar) as aylik_tutar,
                    SUM(CASE WHEN durum='ODENMEDİ' THEN tutar ELSE 0 END) as kalan_borc
                FROM odemeler 
                WHERE aciklama LIKE 'KREDI:%'
                GROUP BY fatura_adi
                HAVING kalan_taksit > 0
                ORDER BY kalan_borc DESC
            """
            try:
                self.imlec_finans.execute(sorgu)
                for row in self.imlec_finans.fetchall():
                    kredi_adi = row[0]
                    kalan_taksit = int(row[1]) if row[1] else 0
                    toplam_taksit = int(row[2]) if row[2] else 0
                    aylik_tutar = row[3] if row[3] else 0.0
                    kalan_borc = row[4] if row[4] else 0.0
                    
                    # 8 / 13 formatında yazdırıyoruz
                    taksit_metni = f"{kalan_taksit} / {toplam_taksit}"
                    
                    tree_krediler.insert("", "end", values=(kredi_adi, taksit_metni, f"{aylik_tutar:,.2f} ₺", f"{kalan_borc:,.2f} ₺"))
            except Exception as e:
                print("Kredi listeleme hatası:", e)

        def guvenli_para_cevir(val_str):
            val_str = str(val_str).strip()
            if not val_str: return 0.0
            if "," in val_str:
                val_str = val_str.replace(".", "").replace(",", ".")
            else:
                val_str = val_str.replace(".", "")
            return float(val_str)

        def krediyi_kaydet():
            ad = e_ad.get().strip()
            tur = cmb_tur.get()
            vade_str = e_vade.get().strip()
            tutar_str = e_tutar.get().strip()
            faiz_str = e_faiz.get().strip()
            masraf_str = e_masraf.get().strip()
            sigorta_str = e_sigorta.get().strip()
            tar_str = e_tarih.get().strip()

            if not ad or not vade_str or not tutar_str:
                messagebox.showwarning("Eksik Bilgi", "Lütfen Banka/Kredi Adı, Vade ve Tutar alanlarını eksiksiz doldurunuz.", parent=self.pencere)
                return

            try:
                n = int(vade_str)
                ilk_tarih = datetime.strptime(tar_str, "%Y-%m-%d").date()
                masraf = guvenli_para_cevir(masraf_str)
                sigorta = guvenli_para_cevir(sigorta_str)
                ana_giris = guvenli_para_cevir(tutar_str)
                taksit = 0.0

                if cmb_mod.current() == 0:
                    taksit = ana_giris
                else:
                    if not faiz_str:
                        messagebox.showwarning("Eksik", "Lütfen hesaplama için 'Aylık Faiz Oranını' giriniz.", parent=self.pencere)
                        return
                    ham_faiz = float(faiz_str.replace(",", "."))
                    vergi_carpani = kredi_veriler[tur]["vergi_carpani"]
                    r = (ham_faiz / 100.0) * vergi_carpani
                    
                    if r > 0:
                        taksit = ana_giris * (r * (1 + r)**n) / ((1 + r)**n - 1)
                    else:
                        taksit = ana_giris / n

                onay_metni = f"Banka: {ad}\nAylık Taksit: {taksit:,.2f} TL\nToplam Vade: {n} Ay\nMasraf: {masraf:,.2f} TL\nSigorta: {sigorta:,.2f} TL\nİlk Ödeme: {tar_str}\n\nBu kredi planını sisteme işlemek istiyor musunuz?"
                if not messagebox.askyesno("Sisteme İşle", onay_metni, parent=self.pencere):
                    return

                # Veritabanına Yazma İşlemi (satir_notu = MANUEL_GIDER yapılarak Depo'dan izole edildi)
                for i in range(n):
                    yil = ilk_tarih.year + (ilk_tarih.month + i - 1) // 12
                    ay = (ilk_tarih.month + i - 1) % 12 + 1
                    ayin_son_gunu = calendar.monthrange(yil, ay)[1]
                    gun = min(ilk_tarih.day, ayin_son_gunu)
                    
                    vade_tarihi = date(yil, ay, gun)
                    
                    if vade_tarihi.weekday() == 5: vade_tarihi += timedelta(days=2)
                    elif vade_tarihi.weekday() == 6: vade_tarihi += timedelta(days=1)
                    
                    vade_str_db = vade_tarihi.strftime("%Y-%m-%d")
                    aciklama_str = f"KREDI: {ad} ({i+1}/{n})"
                    
                    # --- AKILLI TARİH KONTROLÜ BAŞLANGICI ---
                    # Eğer taksitin vade tarihi bugünden küçük veya eşitse, 
                    # bu taksit geçmişte kalmıştır ve ödenmiş sayılmalıdır.
                    if vade_tarihi <= date.today():
                        taksit_durumu = 'ODENDİ'
                    else:
                        taksit_durumu = 'ODENMEDİ'
                    # -----------------------------------------
                    
                    # 'ODENMEDİ' yazan yeri değişkene (taksit_durumu) bağladık
                    self.imlec_finans.execute("""
                        INSERT INTO odemeler (fatura_adi, vade_tarihi, tutar, aciklama, durum, alim_tarihi, satir_notu) 
                        VALUES (?, ?, ?, ?, ?, ?, 'MANUEL_GIDER')
                    """, (ad, vade_str_db, taksit, aciklama_str, taksit_durumu, date.today().strftime("%Y-%m-%d")))

                # Masraflar
                if masraf > 0:
                    self.imlec_finans.execute("""
                        INSERT INTO odemeler (fatura_adi, vade_tarihi, tutar, aciklama, durum, alim_tarihi, satir_notu) 
                        VALUES (?, ?, ?, ?, 'ODENMEDİ', ?, 'MANUEL_GIDER')
                    """, (ad, tar_str, masraf, f"KREDI: {ad} (Masraf)", date.today().strftime("%Y-%m-%d")))
                
                if sigorta > 0:
                    self.imlec_finans.execute("""
                        INSERT INTO odemeler (fatura_adi, vade_tarihi, tutar, aciklama, durum, alim_tarihi, satir_notu) 
                        VALUES (?, ?, ?, ?, 'ODENMEDİ', ?, 'MANUEL_GIDER')
                    """, (ad, tar_str, sigorta, f"KREDI: {ad} (Sigorta)", date.today().strftime("%Y-%m-%d")))

                self.baglanti_finans.commit()
                messagebox.showinfo("Başarılı", "Kredi planı oluşturuldu ve Gider Takibi sayfasına aktarıldı.", parent=self.pencere)
                
                e_ad.delete(0, tk.END)
                e_vade.delete(0, tk.END)
                e_tutar.delete(0, tk.END)
                e_faiz.delete(0, tk.END)
                e_masraf.delete(0, tk.END); e_masraf.insert(0, "0")
                e_sigorta.delete(0, tk.END); e_sigorta.insert(0, "0")
                
                listeyi_guncelle()

            except ValueError:
                messagebox.showerror("Hatalı Giriş", "Lütfen Vade ve Tutar alanlarına sadece rakam giriniz.", parent=self.pencere)
            except Exception as e:
                messagebox.showerror("Hata", str(e), parent=self.pencere)

        def krediyi_sil():
            sel = tree_krediler.selection()
            if not sel:
                messagebox.showwarning("Uyarı", "Lütfen listeden iptal etmek istediğiniz kredinin üzerine tıklayınız.", parent=self.pencere)
                return
            
            item = tree_krediler.item(sel[0])
            kredi_adi = item['values'][0]

            # Mesaj güncellendi ve SQL sorgusundaki 'ODENMEDI' şartı KALDIRILDI!
            if messagebox.askyesno("Krediyi Sil", f"'{kredi_adi}' isimli krediyi sistemden tamamen silmek üzeresiniz.\n\nGeçmişte 'ÖDENDİ' olarak işaretlenenler dahil TÜM TAKSİTLER silinecektir.\nOnaylıyor musunuz?", parent=self.pencere):
                try:
                    self.imlec_finans.execute("DELETE FROM odemeler WHERE fatura_adi=? AND aciklama LIKE 'KREDI:%'", (kredi_adi,))
                    self.baglanti_finans.commit()
                    messagebox.showinfo("Silindi", "Krediye ait tüm taksitler sistemden başarıyla temizlendi.", parent=self.pencere)
                    listeyi_guncelle()
                except Exception as e:
                    messagebox.showerror("Hata", str(e), parent=self.pencere)

        listeyi_guncelle()



    def tema_uygula(self):
        c = TM
        bg_main = c.get_color("bg_main")
        self.pencere.configure(bg=bg_main)
        if hasattr(self, 'ana_container'):
            self.ana_container.configure(bg=bg_main)
        
        style = ttk.Style()
        style.theme_use('clam')
        
        # --- ORTAK TABLO TASARIMI ---
        style.configure("Treeview", 
                        background=c.get_color("card_bg"), 
                        foreground=c.get_color("fg_text"), 
                        fieldbackground=c.get_color("card_bg"), 
                        rowheight=35, 
                        borderwidth=0, 
                        font=FONT_NORM)
        
        style.configure("Treeview.Heading", 
                        background=c.get_color("header_bg"), 
                        foreground="white", 
                        relief="flat", 
                        font=FONT_BOLD)
        
        style.map("Treeview", background=[('selected', '#fee2e2')], foreground=[('selected', '#991b1b')])
        style.map("Treeview.Heading", background=[('active', c.get_color("btn_primary"))])
        
        # --- SEKMELER (NOTEBOOK) TASARIMI ---
        style.configure("TNotebook", background=bg_main, borderwidth=0)
        style.configure("TNotebook.Tab", 
                        background="#e5e7eb", 
                        foreground="#6b7280", 
                        padding=[20, 10], 
                        font=FONT_BOLD,
                        borderwidth=0)
        
        style.map("TNotebook.Tab", 
                  background=[("selected", "white")], 
                  foreground=[("selected", "#dc2626")],
                  expand=[("selected", [1, 1, 1, 0])])
                  
        if hasattr(self, 'content_area') and self.content_area.winfo_exists():
            self.content_area.configure(bg=bg_main)

    def başlangıç_ayarı_güncelle(self, durum):
        key_path = r"Software\Microsoft\Windows\CurrentVersion\Run"
        app_name = "EczaneAsistani"
        if getattr(sys, 'frozen', False): app_path = sys.executable
        else: app_path = f'"{sys.executable}" "{os.path.realpath(sys.argv[0])}"'
        try:
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path, 0, winreg.KEY_SET_VALUE)
            if durum: winreg.SetValueEx(key, app_name, 0, winreg.REG_SZ, app_path)
            else:
                try: winreg.DeleteValue(key, app_name)
                except FileNotFoundError: pass
            winreg.CloseKey(key)
        except Exception as e: print(f"Registry Hatası: {e}")

    def arayuz_sayim_modu(self):
        c = TM
        self.sayim_aktif_id = None
        self.sayim_aktif_adi = ""

        # --- BAŞLIK VE BUTONLAR ---
        header_frame = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        header_frame.pack(fill="x", pady=(0, 10))

        tk.Label(header_frame, text="📦 Canlı Stok Sayım ve Miad Kontrolü", font=FONT_HEAD, 
                 bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left")

        # Üst Butonlar
        btn_bar = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        btn_bar.pack(fill="x", pady=(0, 10))
        
        # Sol Grup
        ModernButton(btn_bar, text="TEMİZLE (YENİ)", command=self.sayim_temizle, bg_color="#64748b", width=120).pack(side="left", padx=2)
        ModernButton(btn_bar, text="📂 KAYITLI LİSTELER", command=self.kayitli_listeleri_ac_dialog, bg_color="#34495e", width=160).pack(side="left", padx=2)
        
        # Sağ Grup
        ModernButton(btn_bar, text="💾 KAYDET", command=self.sayim_kaydet_yoneticisi, bg_color=c.get_color("btn_success"), width=120).pack(side="right", padx=2)
        ModernButton(btn_bar, text="🚀 TOPLU EKLEME", command=self.toplu_sayim_penceresi, bg_color="#8e44ad", width=140).pack(side="right", padx=2)

        # --- EKRANI İKİYE BÖLME ---
        paned = tk.PanedWindow(self.content_area, orient=tk.HORIZONTAL, sashwidth=6, bg=c.get_color("border"))
        paned.pack(fill="both", expand=True)

        # ---------------- SOL TARAF: HAM LİSTE (AKIS) ----------------
        f_left = tk.Frame(paned, bg=c.get_color("bg_main"))
        paned.add(f_left, width=500) 

        # Okuyucu Input
        f_inp = tk.LabelFrame(f_left, text="Karekod Okutun", bg=c.get_color("card_bg"), fg=c.get_color("fg_text"), font=FONT_BOLD)
        f_inp.pack(fill="x", pady=(0, 10))
        self.ent_sayim = tk.Entry(f_inp, font=("Consolas", 14), bg=c.get_color("input_bg"), fg=c.get_color("fg_text"), insertbackground=c.get_color("fg_text"))
        self.ent_sayim.pack(fill="x", padx=10, pady=10, ipady=5)
        self.ent_sayim.bind("<Return>", self.sayim_ekle_satir)
        self.ent_sayim.focus_set()

        tk.Label(f_left, text="📄 Okutulan Karekod Listesi", font=("Segoe UI", 10, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(anchor="w")

        # Sol Tablo
        cols_raw = ("NO", "İLAÇ ADI", "KAREKOD")
        self.tree_sayim = ttk.Treeview(f_left, columns=cols_raw, show="headings", height=15)
        sc_l = ttk.Scrollbar(f_left, orient="vertical", command=self.tree_sayim.yview)
        self.tree_sayim.configure(yscroll=sc_l.set); sc_l.pack(side="right", fill="y")
        self.tree_sayim.pack(fill="both", expand=True)
        
        self.tree_sayim.heading("NO", text="NO"); self.tree_sayim.column("NO", width=40, anchor="center")
        self.tree_sayim.heading("İLAÇ ADI", text="İLAÇ ADI"); self.tree_sayim.column("İLAÇ ADI", width=250)
        self.tree_sayim.heading("KAREKOD", text="KAREKOD"); self.tree_sayim.column("KAREKOD", width=200)
        
        ModernButton(f_left, text="SEÇİLİ SATIRI SİL", command=self.sayim_satir_sil, bg_color=c.get_color("btn_danger"), width=150, height=30).pack(pady=5)

        # ---------------- SAĞ TARAF: DETAYLI MİAD ÖZETİ ----------------
        f_right = tk.Frame(paned, bg=c.get_color("card_bg"))
        paned.add(f_right)

        tk.Label(f_right, text="📊 Stok ve Miad Analizi", font=("Segoe UI", 12, "bold"), bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).pack(pady=10)

        # Sağ Tablo (Sütunlar Güncellendi)
        cols_sum = ("İLAÇ ADI", "SKT", "ADET")
        self.tree_ozet = ttk.Treeview(f_right, columns=cols_sum, show="headings")
        sc_r = ttk.Scrollbar(f_right, orient="vertical", command=self.tree_ozet.yview)
        self.tree_ozet.configure(yscroll=sc_r.set); sc_r.pack(side="right", fill="y")
        self.tree_ozet.pack(fill="both", expand=True, padx=10, pady=(0,10))

        self.tree_ozet.heading("İLAÇ ADI", text="İLAÇ ADI"); self.tree_ozet.column("İLAÇ ADI", width=300)
        self.tree_ozet.heading("SKT", text="MİAD"); self.tree_ozet.column("SKT", width=100, anchor="center")
        self.tree_ozet.heading("ADET", text="TOPLAM"); self.tree_ozet.column("ADET", width=80, anchor="center")

        # RENKLER (Kritik Miadlar İçin)
        self.tree_ozet.tag_configure("kirmizi", background="#ef4444", foreground="white") # < 3 Ay
        self.tree_ozet.tag_configure("turuncu", background="#f97316", foreground="black") # < 1 Yıl
        self.tree_ozet.tag_configure("normal", background=c.get_color("input_bg"), foreground=c.get_color("fg_text"))

        # Toplam Sayaç
        self.lbl_toplam_sayim = tk.Label(f_right, text="Toplam Okutulan: 0", font=("Segoe UI", 11, "bold"), bg="#f1c40f", fg="black", padx=10, pady=5)
        self.lbl_toplam_sayim.pack(pady=10, fill="x")

    def sayim_ekle_satir(self, event=None):
        raw = self.ent_sayim.get().strip()
        if not raw: return
        
        # --- MÜKERRER KONTROLÜ ---
        for item in self.tree_sayim.get_children():
            mevcut_karekod = self.tree_sayim.item(item, "values")[2] # 2. index Karekod
            if mevcut_karekod == raw:
                messagebox.showwarning("Mükerrer Kayıt", "⚠️ Bu karekod zaten listeye eklenmiş!\n\nAynı kutuyu iki kere sayıyorsunuz.")
                self.ent_sayim.delete(0, tk.END)
                return 

        # Karekoddan İsim Bul
        gtin = raw[2:16] if raw.startswith("01") else raw[:14]
        self.imlec.execute("SELECT ad FROM ilac_kartlari WHERE gtin=?", (gtin,))
        res = self.imlec.fetchone()
        ad = res[0] if res else "Bilinmeyen İlaç"

        # SOL TABLOYA EKLE (En üste)
        sira = len(self.tree_sayim.get_children()) + 1
        self.tree_sayim.insert("", 0, values=(sira, ad, raw)) 
        
        self.ent_sayim.delete(0, tk.END)
        
        # SAĞ TABLOYU GÜNCELLE
        self.ozeti_guncelle()

    def ozeti_guncelle(self):
        havuz = {} 
        total_items = 0
        bugun = date.today()
        
        # Sol tablodaki her satırı gez
        for item in self.tree_sayim.get_children():
            vals = self.tree_sayim.item(item, "values")
            ad = vals[1]
            raw_qr = vals[2] 
            
            skt_obj = datetime.max.date()
            skt_str = "Belirsiz"
            
            try:
                for i in range(16, len(raw_qr)-7):
                    if raw_qr[i:i+2] == "17":
                        aday = raw_qr[i+2:i+8]
                        if aday.isdigit():
                            try:
                                skt_obj = datetime.strptime(aday, "%y%m%d").date()
                                skt_str = skt_obj.strftime("%d.%m.%Y")
                                break
                            except: continue
            except: pass
            
            key = (ad, skt_str, skt_obj)
            havuz[key] = havuz.get(key, 0) + 1
            total_items += 1

        # 2. Sağ Tabloyu Temizle
        for i in self.tree_ozet.get_children():
            self.tree_ozet.delete(i)
            
        # 3. Sırala
        sirali_liste = sorted(havuz.items(), key=lambda x: x[0][2])
        
        # 4. Tabloya Ekle ve Renklendir
        kritik_sayac = 0
        
        for (ilac_adi, skt_gosterim, skt_tarih), adet in sirali_liste:
            tag = "normal"
            if skt_tarih != datetime.max.date():
                kalan_gun = (skt_tarih - bugun).days
                if kalan_gun < 90: tag = "kirmizi"; kritik_sayac += adet
                elif kalan_gun < 365: tag = "turuncu"
            
            self.tree_ozet.insert("", "end", values=(ilac_adi, skt_gosterim, f"{adet} Adet"), tags=(tag,))
            
        # 5. Alt Toplam Bilgisi
        msg = f"Genel Toplam: {total_items} Kutu"
        if kritik_sayac > 0:
            msg += f" | ⚠️ {kritik_sayac} KRİTİK ÜRÜN!"
            self.lbl_toplam_sayim.config(bg="#ef4444", fg="white", text=msg) 
        else:
            self.lbl_toplam_sayim.config(bg="#2ecc71", fg="white", text=msg) 

    def sayim_satir_sil(self):
        sel = self.tree_sayim.selection()
        if not sel: return
        for i in sel: self.tree_sayim.delete(i)
        self.ozeti_guncelle() 
        
    def sayim_temizle(self, onay_sorma=True):
        if onay_sorma and self.tree_sayim.get_children():
            if not messagebox.askyesno("Temizle", "Tüm sayım verileri silinecek?"): return
        for i in self.tree_sayim.get_children(): self.tree_sayim.delete(i)
        self.ozeti_guncelle() 
        self.sayim_aktif_id = None
        self.sayim_aktif_adi = ""

    def sayim_kaydet_yoneticisi(self):
        items = self.tree_sayim.get_children()
        if not items:
            messagebox.showwarning("Boş", "Kaydedilecek veri yok.")
            return
        
        icerik_listesi = []
        for i in items:
            vals = self.tree_sayim.item(i, "values")
            icerik_listesi.append(vals[2]) # Karekod index 2 (NO, AD, KAREKOD)
        
        icerik_str = "\n".join(icerik_listesi)

        if self.sayim_aktif_id:
            secim = messagebox.askyesnocancel("Kayıt Seçimi", 
                                              f"Şu an '{self.sayim_aktif_adi}' listesi üzerinde çalışıyorsunuz.\n\n"
                                              "EVET: Mevcut listeyi güncelle (Üzerine Yaz)\n"
                                              "HAYIR: Yeni bir isimle farklı kaydet\n"
                                              "İPTAL: İşlemi durdur")
            if secim is None: return
            
            if secim: # EVET -> Güncelle
                self.imlec.execute("UPDATE kayitli_listeler SET icerik=?, tarih=? WHERE id=?", 
                                   (icerik_str, str(datetime.now()), self.sayim_aktif_id))
                self.baglanti_skt.commit()
                messagebox.showinfo("Başarılı", "Liste güncellendi.")
                return 

        ad = simpledialog.askstring("Kaydet", "Liste Adı Giriniz:")
        if ad:
            self.imlec.execute("INSERT INTO kayitli_listeler (liste_adi, icerik, tarih) VALUES (?, ?, ?)", 
                               (ad, icerik_str, str(datetime.now())))
            self.baglanti_skt.commit()
            self.sayim_aktif_id = self.imlec.lastrowid
            self.sayim_aktif_adi = ad
            messagebox.showinfo("Başarılı", "Yeni liste kaydedildi.")

    def kayitli_listeleri_ac_dialog(self):
        top = tk.Toplevel(self.pencere)
        top.title("Kayıtlı Sayım Listeleri")
        top.geometry("400x500")
        top.configure(bg=TM.get_color("bg_main"))
        
        tk.Label(top, text="Geçmiş Kayıtlar", font=("Segoe UI", 12, "bold"), bg=TM.get_color("bg_main"), fg=TM.get_color("fg_text")).pack(pady=10)
        
        lb = tk.Listbox(top, font=("Segoe UI", 11), bg=TM.get_color("input_bg"), fg=TM.get_color("fg_text"))
        lb.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.imlec.execute("SELECT id, liste_adi, tarih FROM kayitli_listeler ORDER BY id DESC")
        cache = {}
        for r in self.imlec.fetchall():
            tarih = r[2].split()[0]
            display = f"{r[1]} ({tarih})"
            lb.insert(tk.END, display)
            cache[display] = r[0]

        def yukle():
            sel = lb.curselection()
            if not sel: return
            val = lb.get(sel[0])
            lid = cache[val]
            
            self.imlec.execute("SELECT liste_adi, icerik FROM kayitli_listeler WHERE id=?", (lid,))
            res = self.imlec.fetchone()
            
            if res:
                self.sayim_temizle(onay_sorma=False) 
                self.sayim_aktif_id = lid
                self.sayim_aktif_adi = res[0]
                
                raw_data = res[1].split('\n')
                for idx, kk in enumerate(raw_data, 1):
                    if not kk.strip(): continue
                    gtin = kk[2:16] if kk.startswith("01") else kk[:14]
                    self.imlec.execute("SELECT ad FROM ilac_kartlari WHERE gtin=?", (gtin,))
                    r_ad = self.imlec.fetchone()
                    ad = r_ad[0] if r_ad else "Bilinmeyen İlaç"
                    self.tree_sayim.insert("", "end", values=(idx, ad, kk))
                
                self.ozeti_guncelle()
                top.destroy()
                messagebox.showinfo("Yüklendi", f"'{res[0]}' listesi ekrana yüklendi.")

        def sil():
            sel = lb.curselection()
            if not sel: return
            if not messagebox.askyesno("Sil", "Bu kayıt silinsin mi?"): return
            val = lb.get(sel[0])
            lid = cache[val]
            self.imlec.execute("DELETE FROM kayitli_listeler WHERE id=?", (lid,))
            self.baglanti_skt.commit()
            lb.delete(sel[0])

        f_btn = tk.Frame(top, bg=TM.get_color("bg_main"))
        f_btn.pack(fill="x", pady=10)
        ModernButton(f_btn, text="SEÇİLENİ SİL", command=sil, bg_color=TM.get_color("btn_danger"), width=120).pack(side="left", padx=10)
        ModernButton(f_btn, text="EKRANA YÜKLE", command=yukle, bg_color=TM.get_color("btn_success"), width=150).pack(side="right", padx=10)

    def toplu_sayim_penceresi(self):
        top = tk.Toplevel(self.pencere)
        top.title("Toplu Karekod Aktarım ve Miad Analizi")
        top.geometry("1200x650") 
        top.configure(bg=TM.get_color("bg_main"))
        x = self.pencere.winfo_x() + 20; y = self.pencere.winfo_y() + 20
        top.geometry(f"+{x}+{y}")

        paned = tk.PanedWindow(top, orient=tk.HORIZONTAL, sashwidth=5, bg=TM.get_color("border"))
        paned.pack(fill="both", expand=True, padx=10, pady=10)

        f_left = tk.Frame(paned, bg=TM.get_color("bg_main"))
        paned.add(f_left, width=400)
        tk.Label(f_left, text="1. Karekodları Buraya Yapıştırın:", font=("Segoe UI", 11, "bold"), bg=TM.get_color("bg_main"), fg=TM.get_color("fg_text")).pack(anchor="w", pady=(0, 5))
        frame_txt = tk.Frame(f_left); frame_txt.pack(fill="both", expand=True)
        sc_txt = tk.Scrollbar(frame_txt); sc_txt.pack(side="right", fill="y")
        txt_input = tk.Text(frame_txt, bg=TM.get_color("input_bg"), fg=TM.get_color("fg_text"), font=("Consolas", 10), yscrollcommand=sc_txt.set)
        txt_input.pack(side="left", fill="both", expand=True); sc_txt.config(command=txt_input.yview)
        f_analiz_btn = tk.Frame(f_left, bg=TM.get_color("bg_main"), pady=10); f_analiz_btn.pack(fill="x")

        f_right = tk.Frame(paned, bg=TM.get_color("card_bg"))
        paned.add(f_right)
        tk.Label(f_right, text="2. Miad ve Stok Analizi:", font=("Segoe UI", 11, "bold"), bg=TM.get_color("card_bg"), fg=TM.get_color("fg_text")).pack(anchor="w", pady=(0, 5), padx=10)
        cols = ("İLAÇ ADI", "SKT", "ADET")
        tree_analiz = ttk.Treeview(f_right, columns=cols, show="headings")
        sc_tree = ttk.Scrollbar(f_right, orient="vertical", command=tree_analiz.yview)
        tree_analiz.configure(yscroll=sc_tree.set); sc_tree.pack(side="right", fill="y"); tree_analiz.pack(fill="both", expand=True, padx=10)
        tree_analiz.heading("İLAÇ ADI", text="İLAÇ ADI"); tree_analiz.column("İLAÇ ADI", width=350)
        tree_analiz.heading("SKT", text="MİAD (SKT)"); tree_analiz.column("SKT", width=120, anchor="center")
        tree_analiz.heading("ADET", text="ADET"); tree_analiz.column("ADET", width=80, anchor="center")
        tree_analiz.tag_configure("kirmizi", background="#ef4444", foreground="white")
        tree_analiz.tag_configure("turuncu", background="#f97316", foreground="black")
        tree_analiz.tag_configure("normal", background=TM.get_color("input_bg"), foreground=TM.get_color("fg_text"))
        lbl_bilgi = tk.Label(f_right, text="Analiz bekleniyor...", bg=TM.get_color("card_bg"), fg="#f39c12", font=("Segoe UI", 10, "bold")); lbl_bilgi.pack(pady=5)

        gecerli_karekodlar = []

        def analiz_et():
            nonlocal gecerli_karekodlar; gecerli_karekodlar = []
            for i in tree_analiz.get_children(): tree_analiz.delete(i)
            raw_text = txt_input.get("1.0", tk.END).strip()
            if not raw_text: return
            lines = raw_text.split('\n'); ozet_havuzu = {}; bugun = date.today()
            for line in lines:
                line = line.strip()
                if len(line) < 16: continue
                gecerli_karekodlar.append(line)
                gtin = line[2:16] if line.startswith("01") else line[:14]
                self.imlec.execute("SELECT ad FROM ilac_kartlari WHERE gtin=?", (gtin,))
                res = self.imlec.fetchone(); ad = res[0] if res else "Bilinmeyen İlaç"
                skt_obj = datetime.max.date(); skt_str = "Belirsiz"
                try:
                    for i in range(16, len(line)-7):
                        if line[i:i+2] == "17":
                            aday = line[i+2:i+8]
                            if aday.isdigit():
                                skt_obj = datetime.strptime(aday, "%y%m%d").date()
                                skt_str = skt_obj.strftime("%d.%m.%Y")
                                break
                except: pass
                key = (ad, skt_str, skt_obj)
                ozet_havuzu[key] = ozet_havuzu.get(key, 0) + 1
            sirali = sorted(ozet_havuzu.items(), key=lambda x: x[0][2]); kritik = 0
            for (ad, skt_s, skt_d), adet in sirali:
                tag = "normal"
                if skt_d != datetime.max.date():
                    gun = (skt_d - bugun).days
                    if gun < 90: tag = "kirmizi"; kritik += adet
                    elif gun < 365: tag = "turuncu"
                tree_analiz.insert("", "end", values=(ad, skt_s, f"{adet}"), tags=(tag,))
            msg = f"Toplam {len(gecerli_karekodlar)} karekod."
            if kritik > 0: msg += f" ⚠️ {kritik} KRİTİK MİAD!"
            lbl_bilgi.config(text=msg)

        def aktar_ve_kapat():
            if not gecerli_karekodlar: return
            mevcut_karekodlar = set()
            for item in self.tree_sayim.get_children(): mevcut_karekodlar.add(self.tree_sayim.item(item, "values")[2])
            eklenen = 0; atlanan = 0; mevcut_sayi = len(self.tree_sayim.get_children())
            for kk in gecerli_karekodlar:
                if kk in mevcut_karekodlar: atlanan += 1; continue
                gtin = kk[2:16] if kk.startswith("01") else kk[:14]
                self.imlec.execute("SELECT ad FROM ilac_kartlari WHERE gtin=?", (gtin,))
                res = self.imlec.fetchone(); ad = res[0] if res else "Bilinmeyen İlaç"
                mevcut_sayi += 1
                self.tree_sayim.insert("", 0, values=(mevcut_sayi, ad, kk))
                mevcut_karekodlar.add(kk); eklenen += 1
            self.ozeti_guncelle()
            msg = f"✅ {eklenen} karekod listeye eklendi."
            if atlanan > 0: msg += f"\n\n⚠️ {atlanan} karekod zaten listede olduğu için EKLENMEDİ."
            messagebox.showinfo("Aktarım Sonucu", msg, parent=top); top.destroy()

        ModernButton(f_analiz_btn, text="⚡ ANALİZ ET (Miad Kontrol)", command=analiz_et, bg_color="#e67e22", width=250).pack(pady=5)
        f_bot = tk.Frame(top, bg=TM.get_color("bg_main"), pady=15); f_bot.pack(fill="x")
        ModernButton(f_bot, text="✅ KAYDET VE LİSTEYE EKLE", command=aktar_ve_kapat, bg_color=TM.get_color("btn_success"), width=300).pack()

    # =========================================================================
    # 6. SON KULLANMA TARİHİ TAKİP 
    # =========================================================================
    
    # =========================================================================
    # YENİ ÖZELLİK: GEÇMİŞ YÜKLEMELERİ LİSTELEME VE FİLTRELEME
    # =========================================================================
    def yukleme_listesini_doldur(self):
        try:
            # Son 20 yüklemeyi getir
            self.imlec.execute("SELECT id, dosya_adi, tarih FROM yuklemeler ORDER BY id DESC LIMIT 20")
            veriler = self.imlec.fetchall()
        except: veriler = []
        
        liste_degerleri = []
        self.yukleme_map = {} 
        ilk_id = None 
        
        for i, v in enumerate(veriler):
            try: tarih_format = datetime.strptime(v[2], "%Y-%m-%d").strftime("%d.%m.%Y")
            except: tarih_format = v[2]
            
            etiket = f"{tarih_format} - {v[1]}"
            liste_degerleri.append(etiket)
            self.yukleme_map[etiket] = v[0]
            
            if i == 0: ilk_id = v[0] # En güncel olanın ID'si
            
        if hasattr(self, 'cb_yuklemeler'):
            self.cb_yuklemeler['values'] = liste_degerleri
            if liste_degerleri:
                self.cb_yuklemeler.current(0) # Görsel olarak seç
            else:
                self.cb_yuklemeler.set("Yüklenmiş Dosya Yok")
        
        return ilk_id

    def stok_filtre_degisti(self, event=None):
        secilen_etiket = self.cb_yuklemeler.get()
        secilen_id = self.yukleme_map.get(secilen_etiket)
        
        if secilen_id:
            # Seçilen ID ile listeyi güncelle
            self.listeyi_guncelle(yukleme_filtresi=secilen_id)
        else:
            # Seçim yoksa boşalt
            for i in self.tablo_stok.get_children(): self.tablo_stok.delete(i)

    # =========================================================================
    # SIKINTILI SKT ANALİZİ (ID KONTROLÜ + YEŞİL [ NEW ] KUTUSU)
    # =========================================================================
    def ac_sikintili_skt_penceresi(self):
        # 1. En Son Yüklenen Listeyi Bul
        self.imlec.execute("SELECT id, dosya_adi, tarih FROM yuklemeler ORDER BY id DESC LIMIT 1")
        yukleme = self.imlec.fetchone()
        
        if not yukleme:
            messagebox.showwarning("Veri Yok", "Analiz için sisteme stok dosyası yüklemelisiniz.")
            return

        aktif_id = yukleme[0]        
        aktif_isim = yukleme[1]
        
        # 3. Pencereyi Oluştur
        top = tk.Toplevel(self.pencere)
        top.title(f"⚠️ Akıllı İade Yönetimi - {aktif_isim}")
        top.geometry("1550x850")
        c = TM
        top.configure(bg=c.get_color("bg_main"))

        # --- ÜST BUTON PANELI ---
        f_top_btns = tk.Frame(top, bg=c.get_color("header_bg"), pady=10, padx=10)
        f_top_btns.pack(fill="x")

        # Bilgi Etiketi
        info_text = "GÖSTERGE: 🆕  yazanlar bu listeyle ilk kez stoğa girenlerdir.\nHESAP: Kısa girenler için 60 gün, normal girenler için 1 yıl iade kuralı uygulanır."
        tk.Label(f_top_btns, text=info_text, bg=c.get_color("header_bg"), fg="#f1c40f", 
                 font=("Segoe UI", 9), justify="left").pack(side="left")

        # --- AKILLI HESAPLAMA MOTORU ---
        def motor_hesapla(skt_str, kayit_tarihi_str):
            bugun = date.today()
            try: skt_dt = datetime.strptime(skt_str, '%Y-%m-%d').date()
            except: return None, 0, "Hata", False

            try: kayit_dt = datetime.strptime(kayit_tarihi_str, '%Y-%m-%d').date()
            except: kayit_dt = bugun 

            # İlaç eczaneye girdiğinde ömrü 1 yıldan az mıydı?
            giris_anindaki_omur = (skt_dt - kayit_dt).days
            
            if giris_anindaki_omur < 365:
                kural = "60 Gün (Kısa Giriş)"
                son_iade_tarihi = kayit_dt + timedelta(days=60)
                kalan_gun = (son_iade_tarihi - bugun).days
                mail_tetik = True if kalan_gun <= 35 else False
            else:
                kural = "Standart (1 Yıl)"
                son_iade_tarihi = skt_dt - timedelta(days=365)
                kalan_gun = (son_iade_tarihi - bugun).days
                mail_tetik = True if kalan_gun <= 10 else False
                
            return son_iade_tarihi, kalan_gun, kural, mail_tetik

        # --- MAİL VE KOPYALAMA ---
        tab_trees = {} 

        def aktif_tree_getir():
            current_tab = notebook.select() 
            return tab_trees.get(current_tab)

        def akilli_mail_gonder():
            kritik_liste = []
            # ID kontrolü için yukleme_id'yi de çekiyoruz
            self.imlec.execute("""SELECT ad, barkod, skt, kayit_tarihi, yukleme_id FROM ilaclar WHERE raf_yeri=? """, (str(aktif_id),))
            
            bugun = date.today()
            for r in self.imlec.fetchall():
                ad, gtin, skt, k_tar, y_id = r
                son_iade, kalan, kural, mail_durumu = motor_hesapla(skt, k_tar)
                if son_iade is None: continue

                # 15 Ay filtresi
                skt_dt = datetime.strptime(skt, '%Y-%m-%d').date()
                if (skt_dt - bugun).days > 455: continue 

                if mail_durumu:
                    tarih_fmt = son_iade.strftime("%d.%m.%Y")
                    durum_str = f"⚠️ {kalan} GÜN KALDI" if kalan >= 0 else "SÜRE GEÇTİ"
                    
                    # Mailde de belirtelim
                    yeni_mi = "🆕" if y_id == aktif_id else ""
                    
                    kritik_liste.append((ad, gtin, skt, tarih_fmt, kural, durum_str, yeni_mi))

            if not kritik_liste:
                detay = (
                    "Şu anda acil mail atılmasını gerektiren kritik seviyede bir ürün bulunmuyor.\n\n"
                    "📌 Mail Gönderme Kriterleri Nelerdir?\n\n"
                    "1. Kısa Miadlı Girenler (1 Yıldan az ömürlü gelenler):\n"
                    "   • Eczaneye girdiği günden itibaren 60 gün iade süresi başlar.\n"
                    "   • İade hakkının bitmesine 35 Gün veya daha az kalırsa mail atılır.\n\n"
                    "2. Normal Miadlı Girenler (1 Yıldan uzun ömürlü gelenler):\n"
                    "   • SKT'sine tam 1 Yıl (365 gün) kalana kadar iade edilebilir.\n"
                    "   • Bu son iade şansına 10 Gün veya daha az kalırsa mail atılır."
                )
                messagebox.showinfo("Temiz (Kritik Ürün Yok)", detay)
                return

            html_body = """
            <html><body><h3>🚨 İADE SÜRESİ UYARISI</h3>
            <table><tr><th>Durum</th><th>İlaç Adı</th><th>GTIN</th><th>SKT</th><th>Son İade</th><th>Kalan</th></tr>
            """
            for item in kritik_liste:
                renk = "green" if item[6] else "black"
                html_body += f"<tr><td style='color:{renk};font-weight:bold'>{item[6]}</td><td>{item[0]}</td><td>{item[1]}</td><td>{item[2]}</td><td>{item[3]}</td><td style='color:red'><b>{item[5]}</b></td></tr>"
            html_body += "</table></body></html>"

            self.basit_mail_gonder(f"🚨 İADE LİSTESİ ({len(kritik_liste)} Kalem)", html_body, is_html=True)
            messagebox.showinfo("Mail Gönderildi", f"{len(kritik_liste)} adet ürün için uyarı maili gönderildi.")

        def secilenleri_kopyala():
            tree = aktif_tree_getir()
            if not tree: return
            # İndeksler: 4:GTIN, 5:AD, 9:KAREKOD
            satirlar = [f"{tree.item(i,'values')[4]}\t{tree.item(i,'values')[5]}\t{tree.item(i,'values')[9]}" for i in tree.get_children() if tree.item(i,"values")[1]=="☑"]
            if satirlar:
                top.clipboard_clear()
                top.clipboard_append("\n".join(satirlar))
                messagebox.showinfo("Kopyalandı", f"{len(satirlar)} satır kopyalandı.")
            else: messagebox.showwarning("Uyarı", "Seçim yapınız.")

        ModernButton(f_top_btns, text="📋 SEÇİLİLERİ KOPYALA", command=secilenleri_kopyala, bg_color="#3498db", width=250, height=35).pack(side="right", padx=5)
        ModernButton(f_top_btns, text="🚨 KRİTİK ÜRÜNLERİ MAİL AT", command=akilli_mail_gonder, bg_color="#e74c3c", width=250, height=35).pack(side="right", padx=10)

        # --- SEKMELER ---
        notebook = ttk.Notebook(top)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        sekmeler = [
            ("🔴 12 AYDAN KISA (İadesi Yakın/Geçmiş)", -9999, 365), 
            ("🟠 12-15 AY ARASI (Takipte)", 365, 455)
        ]

        bugun = date.today()

        def treeview_sort_column(tv, col, reverse):
            l = [(tv.set(k, col), k) for k in tv.get_children('')]
            try: l.sort(key=lambda t: float(t[0].split()[0]) if t[0] and t[0][0].isdigit() else t[0], reverse=reverse)
            except: l.sort(key=lambda t: t[0].lower(), reverse=reverse)
            for index, (val, k) in enumerate(l): tv.move(k, '', index)
            for idx, item_id in enumerate(tv.get_children()): 
                vals = list(tv.item(item_id, "values")); vals[0] = idx + 1; tv.item(item_id, values=vals)
            tv.heading(col, command=lambda: treeview_sort_column(tv, col, not reverse))

        # --- TABLO DOLDURUCU ---
        def tablo_doldur(parent_frame, min_g, max_g):
            cols = ("NO", "TIK", "DURUM", "GTIN", "İLAÇ ADI", "SKT", "SON İADE TARİHİ", "KURAL", "İADEYE KALAN", "KAREKOD")
            tree = ttk.Treeview(parent_frame, columns=cols, show="headings", height=20, selectmode="extended")
            
            sc = ttk.Scrollbar(parent_frame, orient="vertical", command=tree.yview)
            tree.configure(yscroll=sc.set); sc.pack(side="right", fill="y")
            tree.pack(fill="both", expand=True)
            
            def tumunu_sec_toggle():
                durum = getattr(tree, "tumunu_sec_durum", False)
                yeni_durum = not durum
                setattr(tree, "tumunu_sec_durum", yeni_durum)
                ikon = "☑" if yeni_durum else "☐"
                tree.heading("TIK", text=ikon)
                for item in tree.get_children():
                    vals = list(tree.item(item, "values")); vals[1] = ikon; tree.item(item, values=vals)

            tree.heading("NO", text="NO"); tree.column("NO", width=40, anchor="center")
            tree.heading("TIK", text="☐", command=tumunu_sec_toggle); tree.column("TIK", width=20, anchor="center")
            
            # DURUM sütununa tıklayınca sıralama yapması için command parametresi eklendi
            tree.heading("DURUM", text="DURUM", command=lambda: treeview_sort_column(tree, "DURUM", False))
            tree.column("DURUM", width=60, anchor="center")
            
            tree.heading("GTIN", text="GTIN"); tree.column("GTIN", width=120)
            tree.heading("İLAÇ ADI", text="İLAÇ ADI", command=lambda: treeview_sort_column(tree, "İLAÇ ADI", False)); tree.column("İLAÇ ADI", width=400)
            tree.heading("SKT", text="SKT", command=lambda: treeview_sort_column(tree, "SKT", False)); tree.column("SKT", width=90, anchor="center")
            tree.heading("SON İADE TARİHİ", text="SON İADE", command=lambda: treeview_sort_column(tree, "SON İADE TARİHİ", False)); tree.column("SON İADE TARİHİ", width=100, anchor="center")
            tree.heading("KURAL", text="KURAL"); tree.column("KURAL", width=120, anchor="center")
            tree.heading("İADEYE KALAN", text="İADEYE KALAN", command=lambda: treeview_sort_column(tree, "İADEYE KALAN", False)); tree.column("İADEYE KALAN", width=100, anchor="center")
            tree.heading("KAREKOD", text="KAREKOD"); tree.column("KAREKOD", width=350)
            
            def satir_tikla(event):
                if tree.identify_column(event.x) == "#2":
                    item = tree.identify_row(event.y)
                    if item:
                        vals = list(tree.item(item, "values")); vals[1] = "☑" if vals[1] == "☐" else "☐"; tree.item(item, values=vals)
            tree.bind("<Button-1>", satir_tikla)

            # RENK TANIMLAMALARI
            tree.tag_configure("normal", background="white", foreground="black")
            tree.tag_configure("yeni_giris_style", background="#d1fae5", foreground="#065f46") # Açık Yeşil (New)
            tree.tag_configure("kritik", background="#fee2e2", foreground="#c0392b") # Kırmızı (Süre az)
            
            # Verileri Çek
            self.imlec.execute("""
                SELECT barkod, ad, skt, seri_no, parti_no, kayit_tarihi, yukleme_id 
                FROM ilaclar WHERE raf_yeri=? ORDER BY skt ASC
            """, (str(aktif_id),))
            
            count = 0
            for r in self.imlec.fetchall():
                gtin, ad, skt_str, sn, bn, k_tar, yuk_id = r
                
                try: skt_dt = datetime.strptime(skt_str, '%Y-%m-%d').date()
                except: continue
                
                skt_kalan_gun = (skt_dt - bugun).days
                if skt_kalan_gun > 455: continue 
                if not (min_g <= skt_kalan_gun < max_g): continue
                
                son_iade_dt, iadeye_kalan, kural_tipi, mail_tetik = motor_hesapla(skt_str, k_tar)
                son_iade_str = son_iade_dt.strftime("%d.%m.%Y")
                
                # --- ID KARŞILAŞTIRMA (YENİ Mİ?) ---
                # Eğer ilacın 'yukleme_id'si == 'aktif_id' ise bu yeni gelmiştir.
                if yuk_id == aktif_id:
                    yeni_ikon = "🆕"
                    satir_tag = "yeni_giris_style"
                else:
                    yeni_ikon = ""
                    satir_tag = "normal"

                # Süre Kritikse Rengi Kırmızıya Çevir (New olsa bile kritiklik daha önemlidir)
                # Ancak [NEW] yazısı kalmalı.
                if iadeye_kalan <= 10 or mail_tetik:
                    satir_tag = "kritik"

                sayac_str = f"{iadeye_kalan} Gün"
                if iadeye_kalan < 0: sayac_str = "SÜRE GEÇTİ!"
                elif mail_tetik: sayac_str = f"⚠️ {iadeye_kalan} GÜN!"

                g_pr = str(gtin).replace('.0','').zfill(14)
                try: qr = f"01{g_pr}21{str(sn).replace('.0','')}17{skt_dt.strftime('%y%m%d')}10{str(bn).replace('.0','')}"
                except: qr = ""
                skt_tr = datetime.strptime(skt_str, "%Y-%m-%d").strftime("%d.%m.%Y")

                count += 1
                tree.insert("", "end", values=(count, "☐", yeni_ikon, g_pr, ad, skt_tr, son_iade_str, kural_tipi, sayac_str, qr), tags=(satir_tag,))
            
            return tree

        for baslik, min_g, max_g in sekmeler:
            tab_frame = tk.Frame(notebook, bg=c.get_color("bg_main"))
            notebook.add(tab_frame, text=baslik)
            tree_ref = tablo_doldur(tab_frame, min_g, max_g)
            tab_trees[str(tab_frame)] = tree_ref
    

    # =========================================================================
    # YÜKLEME YÖNETİCİSİ (GÜNCELLENDİ: ÇOKLU SEÇİM VE SİLME)
    # =========================================================================
    def ac_yukleme_yoneticisi(self):
        top = tk.Toplevel(self.pencere)
        top.title("📂 Kayıtlı Stok Listeleri Yönetimi")
        top.geometry("650x550")
        c = TM
        top.configure(bg=c.get_color("bg_main"))
        
        x = self.pencere.winfo_x() + 100; y = self.pencere.winfo_y() + 100
        top.geometry(f"+{x}+{y}")

        tk.Label(top, text="Kayıtlı Excel Listeleri", font=("Segoe UI", 14, "bold"), 
                 bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(pady=15)

        # Tablo Alanı
        frame_table = tk.Frame(top, bg=c.get_color("card_bg"))
        frame_table.pack(fill="both", expand=True, padx=20, pady=10)

        # Sütunlar: TIK kutusu eklendi
        cols = ("ID", "TIK", "TARİH", "DOSYA ADI")
        tree = ttk.Treeview(frame_table, columns=cols, show="headings", selectmode="extended")
        
        sc = ttk.Scrollbar(frame_table, orient="vertical", command=tree.yview)
        tree.configure(yscroll=sc.set); sc.pack(side="right", fill="y"); tree.pack(fill="both", expand=True)

        tree.heading("ID", text="ID"); tree.column("ID", width=0, stretch=False)
        
        # TIK Sütunu (Başlığa basınca hepsi seçilir)
        self.yonetici_tumunu_sec_durum = False
        def tumunu_sec_yonetici():
            self.yonetici_tumunu_sec_durum = not self.yonetici_tumunu_sec_durum
            ikon = "☑" if self.yonetici_tumunu_sec_durum else "☐"
            tree.heading("TIK", text=ikon)
            for item in tree.get_children():
                vals = list(tree.item(item, "values"))
                vals[1] = ikon
                tree.item(item, values=vals)

        tree.heading("TIK", text="☐", command=tumunu_sec_yonetici); tree.column("TIK", width=40, anchor="center")
        tree.heading("TARİH", text="YÜKLEME TARİHİ"); tree.column("TARİH", width=120, anchor="center")
        tree.heading("DOSYA ADI", text="DOSYA ADI"); tree.column("DOSYA ADI", width=300)

        # Tıklama Olayı (Checkbox mantığı)
        def satir_tiklama_yonetici(event):
            region = tree.identify("region", event.x, event.y)
            if region == "cell":
                col = tree.identify_column(event.x)
                if col == "#2": # TIK sütunu
                    item = tree.identify_row(event.y)
                    if item:
                        vals = list(tree.item(item, "values"))
                        vals[1] = "☑" if vals[1] == "☐" else "☐"
                        tree.item(item, values=vals)
        
        tree.bind("<Button-1>", satir_tiklama_yonetici)

        def verileri_tazele():
            for i in tree.get_children(): tree.delete(i)
            self.imlec.execute("SELECT id, tarih, dosya_adi FROM yuklemeler ORDER BY id DESC")
            for r in self.imlec.fetchall():
                try: tarih_fmt = datetime.strptime(r[1], "%Y-%m-%d").strftime("%d.%m.%Y")
                except: tarih_fmt = r[1]
                tree.insert("", "end", values=(r[0], "☐", tarih_fmt, r[2]))

        verileri_tazele()

        # ÇOKLU SİLME İŞLEMİ
        def secilenleri_sil():
            silinecek_ids = []
            silinecek_isimler = []
            
            for item in tree.get_children():
                vals = tree.item(item, "values")
                if vals[1] == "☑":
                    silinecek_ids.append(vals[0])
                    silinecek_isimler.append(vals[3])
            
            if not silinecek_ids:
                messagebox.showwarning("Seçim Yok", "Lütfen silinecek listeleri kutucuklardan (☑) seçin.", parent=top)
                return

            onay = messagebox.askyesno("Çoklu Silme", 
                                       f"Seçili {len(silinecek_ids)} adet liste ve içindeki TÜM STOKLAR silinecek.\n\n"
                                       f"Silinecekler:\n" + "\n".join(silinecek_isimler[:5]) + ("\n..." if len(silinecek_isimler)>5 else "") + 
                                       "\n\nOnaylıyor musunuz?", parent=top)
            if not onay: return

            try:
                # Toplu Silme
                for yuk_id in silinecek_ids:
                    self.imlec.execute("DELETE FROM ilaclar WHERE yukleme_id=?", (yuk_id,))
                    self.imlec.execute("DELETE FROM yuklemeler WHERE id=?", (yuk_id,))
                
                self.baglanti_skt.commit()
                verileri_tazele()
                messagebox.showinfo("Başarılı", "Seçilen listeler silindi.", parent=top)
            except Exception as e:
                messagebox.showerror("Hata", str(e), parent=top)

        def kapat_ve_guncelle():
            self.cb_yuklemeler.set('') 
            yeni_id = self.yukleme_listesini_doldur()
            if yeni_id:
                self.cb_yuklemeler.current(0)
                self.stok_filtre_degisti()
            else:
                self.cb_yuklemeler.set("Yüklenmiş Dosya Yok")
                self.listeyi_guncelle()
            top.destroy()

        # Alt Butonlar
        f_btn = tk.Frame(top, bg=c.get_color("bg_main"))
        f_btn.pack(fill="x", padx=20, pady=20)

        ModernButton(f_btn, text="🗑️ SEÇİLENLERİ SİL", command=secilenleri_sil, bg_color="#e74c3c", width=200).pack(side="left")
        ModernButton(f_btn, text="💾 KAPAT", command=kapat_ve_guncelle, bg_color="#2ecc71", width=150).pack(side="right")

        top.protocol("WM_DELETE_WINDOW", kapat_ve_guncelle)

    # =========================================================================
    # ARAYÜZ (FİNAL: HIZLI EKLE KUTUSU DARALTILDI)
    # =========================================================================
    def arayuz_stok_takip(self):
        # Önce ekranı temizle
        for widget in self.content_area.winfo_children(): widget.destroy()
        c = TM
        
        # --- 1. SATIR: BAŞLIK (SOL) ve ANALİZ BUTONU (SAĞ) ---
        row1 = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        row1.pack(fill="x", pady=(0, 10))
        
        tk.Label(row1, text="Son Kullanma Tarihi Takip", font=FONT_HEAD, 
                 bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left")
        
        # Genişlik 320 yapıldı, metin tam sığacak
        ModernButton(row1, text="⚠️ TAKİP EDİLMESİ GEREKENLER", command=self.ac_sikintili_skt_penceresi, 
                     bg_color="#e74c3c", width=320, height=40).pack(side="right", padx=10)

        # --- 2. SATIR: BUTON GRUPLARI ---
        row2 = tk.Frame(self.content_area, bg=c.get_color("card_bg"), padx=5, pady=10)
        row2.pack(fill="x", pady=(0, 10))
        
        # Ortak Buton Boyutları
        BTN_W = 115
        BTN_H = 32

        # --- SOL GRUP ---
        f_left_group = tk.Frame(row2, bg=c.get_color("card_bg"))
        f_left_group.pack(side="left")

        tk.Label(f_left_group, text="📂 Liste:", font=FONT_BOLD, bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).pack(side="left")
        self.cb_yuklemeler = ttk.Combobox(f_left_group, state="readonly", width=25, font=FONT_NORM)
        self.cb_yuklemeler.pack(side="left", padx=(5, 5))
        
        ModernButton(f_left_group, text="📂 YÜKLE", command=self.excel_yukle_stok, 
                     bg_color=c.get_color("btn_success"), width=BTN_W, height=BTN_H).pack(side="left", padx=4)
        
        ModernButton(f_left_group, text="⚙️ DÜZENLE", command=self.ac_yukleme_yoneticisi, 
                     bg_color="#95a5a6", width=BTN_W, height=BTN_H).pack(side="left", padx=4)

        # --- YENİ EKLENEN: BİLGİLENDİRME LİNKİ VE POPUP ---
        def stok_nasil_yuklenir(event=None):
            bilgi = (
                "İTS'den Stok Yükleme Adımları:\n\n"
                "1. İlaç Takip Sistemi (İTS) portalına eczane şifrenizle giriş yapın.\n"
                "2. 'Stok İşlemleri' bölümünden güncel karekod/stok listenizi bulun.\n"
                "3. Listeyi bilgisayarınıza 'Excel (.xlsx)' formatında indirin.\n"
                "4. İndirdiğiniz dosyada GTIN, Seri No (SN), Parti No (BN) ve SKT sütunlarının olduğuna emin olun.\n"
                "5. Yandaki '📂 YÜKLE' butonuna basarak bilgisayarınıza inen bu dosyayı seçin.\n\n"
                "Sistem akıllı okuyucusuyla dosyayı tarayacak ve tüm miadları otomatik renklendirecektir."
            )
            messagebox.showinfo("İTS Stok Dosyası Nasıl Yüklenir?", bilgi)

        lbl_nasil = tk.Label(f_left_group, text="❓ Nasıl Yüklenir?", font=("Segoe UI", 9, "underline", "bold"), bg=c.get_color("card_bg"), fg="#3b82f6", cursor="hand2")
        lbl_nasil.pack(side="left", padx=(10, 5))
        lbl_nasil.bind("<Button-1>", stok_nasil_yuklenir)

        # --- SAĞ GRUP ---
        f_right_group = tk.Frame(row2, bg=c.get_color("card_bg"))
        f_right_group.pack(side="right")

        # Görünüm butonunun ilk halini belirliyoruz
        baslangic_txt = "👁️ GİZLE" if getattr(self, 'detayli_gorunum', True) else "🔲 GÖSTER"
        baslangic_renk = "#2c3e50" if getattr(self, 'detayli_gorunum', True) else "#95a5a6"

        self.btn_mod = ModernButton(f_right_group, text=baslangic_txt, command=self.mod_degistir, 
                                    bg_color=baslangic_renk, width=BTN_W, height=BTN_H)
        self.btn_mod.pack(side="left", padx=4)

        ModernButton(f_right_group, text="📋 KOPYALA", command=self.stok_karekodlari_kopyala, 
                     bg_color="#8b5cf6", width=BTN_W, height=BTN_H).pack(side="left", padx=4)
        
        ModernButton(f_right_group, text="🗑️ SİL", command=self.ilac_sil, 
                     bg_color=c.get_color("btn_danger"), width=BTN_W, height=BTN_H).pack(side="left", padx=4)
        
        self.menubtn_mail = tk.Menubutton(f_right_group, text="📧 MAİL ▼", width=11, bg="#e11d48", fg="white", 
                                          font=FONT_BOLD, relief="flat", activebackground="#be123c", 
                                          activeforeground="white", pady=5, padx=5)
        self.menu_mail = tk.Menu(self.menubtn_mail, tearoff=0, bg=c.get_color("card_bg"), fg=c.get_color("fg_text"))
        self.menu_mail.add_command(label="📄 Sadece Karekodları Gönder", command=lambda: self.ozel_mail_gonder("karekod"))
        self.menu_mail.add_command(label="📊 Tablo Olarak Gönder (Adet)", command=lambda: self.ozel_mail_gonder("tablo"))
        self.menubtn_mail["menu"] = self.menu_mail
        self.menubtn_mail.pack(side="left", padx=(0, 5))


        # --- 3. SATIR: ORTA BÖLME (ARAMA GENİŞ - HIZLI EKLE DAR) ---
        row3 = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        row3.pack(fill="x", pady=(0, 10))

        # SOL YARI: ARAMA (expand=True ile kalan tüm alanı kaplar)
        f_search_area = tk.LabelFrame(row3, text="🔍 İlaç Arama", bg=c.get_color("card_bg"), 
                                      fg=c.get_color("fg_text"), font=FONT_BOLD, padx=10, pady=5)
        f_search_area.pack(side="left", fill="both", expand=True, padx=(0, 10)) 

        self.ent_search = tk.Entry(f_search_area, font=FONT_NORM, bg=c.get_color("input_bg"), 
                                   fg=c.get_color("fg_text"), insertbackground=c.get_color("fg_text"))
        self.ent_search.pack(side="left", fill="x", expand=True, padx=5, ipady=3)
        self.ent_search.bind("<Return>", lambda e: self.listeyi_guncelle(query=self.ent_search.get()))
        
        ModernButton(f_search_area, text="BUL", command=lambda: self.listeyi_guncelle(query=self.ent_search.get()), 
                     width=70, height=20).pack(side="left", padx=2)
        ModernButton(f_search_area, text="SIFIRLA", command=lambda: [self.ent_search.delete(0,tk.END), self.stok_filtre_degisti()], 
                     bg_color="#64748b", width=70, height=20).pack(side="left", padx=2)

        # SAĞ YARI: HIZLI EKLEME (expand=False ve sabit width ile dar kalır)
        f_quick_add = tk.LabelFrame(row3, text="⚡ Hızlı Ekle", bg=c.get_color("card_bg"), 
                                    fg="#27ae60", font=FONT_BOLD, padx=10, pady=5)
        f_quick_add.pack(side="right", fill="y", expand=False) # Sağa yasla, genişleme

        # Entry genişliği kutuyu belirler (30 karakter genişliği yeterli)
        self.ent_qr_input = tk.Entry(f_quick_add, font=("Consolas", 11), width=30, bg=c.get_color("input_bg"), 
                                     fg=c.get_color("fg_text"), insertbackground=c.get_color("fg_text"))
        self.ent_qr_input.pack(fill="x", ipady=3, padx=5)
        self.ent_qr_input.bind("<Return>", self.karekod_parse_ekle)
        # Placeholder niyetine etiket
        tk.Label(f_quick_add, text="(Karekod Okutunuz)", font=("Segoe UI", 8), bg=c.get_color("card_bg"), fg="grey").pack()
        
        # --- 4. RENK FİLTRELERİ (ÇOKLU SEÇİM - CHECKBOX MANTIĞI) ---
        ops = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        ops.pack(fill="x", pady=(0,5))
        
        # Filtre durumlarını hafızada tut (İlk açılışta hepsi seçili gelsin)
        if not hasattr(self, 'aktif_renk_filtreleri'):
            self.aktif_renk_filtreleri = {"kirmizi": True, "turuncu": True, "sari": True, "yesil": True}

        filters = [("#ef4444", "kirmizi", "ACİL (0-3 Ay)"), 
                   ("#f97316", "turuncu", "RİSKLİ (3-6 Ay)"),
                   ("#eab308", "sari", "DİKKAT (6-12 Ay)"), 
                   ("#22c55e", "yesil", "GÜVENLİ (>1 Yıl)")]

        def filtre_kutusu_olustur(f_parent, color, tag, text):
            f_box = tk.Frame(f_parent, bg=c.get_color("bg_main"), cursor="hand2")
            f_box.pack(side="left", padx=(0, 25)) # Kutular arası boşluk

            # Başlangıç ikonunu belirle
            ikon = "☑" if self.aktif_renk_filtreleri[tag] else "☐"
            fg_ikon = "#10b981" if self.aktif_renk_filtreleri[tag] else c.get_color("fg_text")

            lbl_check = tk.Label(f_box, text=ikon, font=("Segoe UI", 13), bg=c.get_color("bg_main"), fg=fg_ikon, cursor="hand2")
            lbl_check.pack(side="left", padx=(0, 5))

            lbl_color = tk.Label(f_box, bg=color, width=2, height=1, bd=0, cursor="hand2")
            lbl_color.pack(side="left", padx=(0, 5))

            lbl_text = tk.Label(f_box, text=text, bg=c.get_color("bg_main"), fg=c.get_color("fg_text"), font=("Segoe UI", 9, "bold"), cursor="hand2")
            lbl_text.pack(side="left")

            def toggle_filtre(event=None):
                # Durumu tersine çevir
                mevcut_durum = self.aktif_renk_filtreleri[tag]
                self.aktif_renk_filtreleri[tag] = not mevcut_durum
                
                # İkonu ve rengi güncelle
                yeni_ikon = "☑" if self.aktif_renk_filtreleri[tag] else "☐"
                yeni_renk = "#10b981" if self.aktif_renk_filtreleri[tag] else c.get_color("fg_text")
                lbl_check.config(text=yeni_ikon, fg=yeni_renk)
                
                # Tabloyu yenile
                self.listeyi_guncelle()

            # Kutu içindeki herhangi bir şeye tıklayınca filtre değişsin
            for w in (f_box, lbl_check, lbl_color, lbl_text):
                w.bind("<Button-1>", toggle_filtre)

        # Filtreleri ekrana bas
        for color, tag, label_text in filters:
            filtre_kutusu_olustur(ops, color, tag, label_text)


        # --- 5. TABLO ---
        tf = tk.Frame(self.content_area, bg=c.get_color("card_bg")); tf.pack(fill="both", expand=True)
        self.tablo_stok = ttk.Treeview(tf, show="headings", selectmode="extended", height=15)
        sc = ttk.Scrollbar(tf, orient="vertical", command=self.tablo_stok.yview)
        self.tablo_stok.configure(yscroll=sc.set); sc.pack(side="right", fill="y"); self.tablo_stok.pack(fill="both", expand=True)
        self.tablo_stok.bind('<ButtonRelease-1>', self.satir_tiklama)
        # --- 3. MADDE: SAĞ TIK MENÜSÜ VE KLAVYE KISAYOLLARI ---
        
        # Sağ Tık Menüsünü Oluştur
        self.sag_tik_menusu = tk.Menu(self.pencere, tearoff=0, bg=c.get_color("card_bg"), fg=c.get_color("fg_text"), font=FONT_NORM)
        self.sag_tik_menusu.add_command(label="📋 Seçilileri Kopyala", command=self.stok_karekodlari_kopyala)
        self.sag_tik_menusu.add_command(label="🗑️ Seçilileri Sil", command=self.ilac_sil)

        def sag_tik_goster(event):
            # Tıklanan satırı bul
            iid = self.tablo_stok.identify_row(event.y)
            if iid:
                # Satırı seçili hale getir
                self.tablo_stok.selection_set(iid)
                
                # Eğer tıklanan satır tikli değilse, otomatik tik at (İşlem yapması kolay olsun diye)
                vals = list(self.tablo_stok.item(iid, "values"))
                if vals[1] == "☐":
                    vals[1] = "☑"
                    self.tablo_stok.item(iid, values=vals)
                    
                # Menüyü farenin ucunda aç
                self.sag_tik_menusu.tk_popup(event.x_root, event.y_root)

        # Tabloya sağ tık (<Button-3>) özelliğini bağla
        self.tablo_stok.bind("<Button-3>", sag_tik_goster)
        
        # Klavye Kısayolları
        # Tabloda "Delete" tuşuna basınca silme işlemi başlar
        self.tablo_stok.bind("<Delete>", lambda e: self.ilac_sil())
        
        # "Ctrl + F" basınca doğrudan arama kutusuna yazar
        self.pencere.bind_all("<Control-f>", lambda e: self.ent_search.focus_set())
        self.pencere.bind_all("<Control-F>", lambda e: self.ent_search.focus_set())
        # -------------------------------------------------------------------------
        
        for t, co in [("kirmizi", "#fee2e2"), ("turuncu", "#ffedd5"), ("sari", "#fef9c3"), ("yesil", "#dcfce7")]:
            self.tablo_stok.tag_configure(t, background=co, foreground="black") 
            
        # --- AÇILIŞ İŞLEMLERİ ---
        en_guncel_id = self.yukleme_listesini_doldur()
        self.cb_yuklemeler.bind("<<ComboboxSelected>>", self.stok_filtre_degisti)
        
        if en_guncel_id:
            self.cb_yuklemeler.current(0)
            self.listeyi_guncelle(yukleme_filtresi=en_guncel_id)
        else:
            self.listeyi_guncelle()

    # --- GÜNCELLENEN FONKSİYON: SADECE SEÇİLİLERİ KOPYALA ---
    def stok_karekodlari_kopyala(self):
        # Eğer detaylı görünümde değilse (karekodlar görünmüyorsa) uyar
        if not self.detayli_gorunum:
            messagebox.showwarning("Uyarı", "Karekodları kopyalamak için lütfen 'KAREKODLARI GÖSTER' moduna geçiniz.")
            return

        karekodlar = []
        # Tablodaki tüm satırları gez
        for child in self.tablo_stok.get_children():
            # Satır verisini al
            vals = self.tablo_stok.item(child, "values")
            # vals[1] -> TIK Sütunu ( "☐" veya "☑" )
            # vals[-1] -> KAREKOD Sütunu (En sonda)
            
            # SADECE TİKLİ OLANLARI AL
            if vals[1] == "☑":
                karekodlar.append(vals[-1])
        
        if karekodlar:
            self.pencere.clipboard_clear()
            self.pencere.clipboard_append("\n".join(karekodlar))
            messagebox.showinfo("Başarılı", f"Seçilen {len(karekodlar)} adet karekod panoya kopyalandı.")
        else:
            messagebox.showwarning("Seçim Yok", "Lütfen kopyalamak için listeden seçim (☑) yapınız.")

    def mod_degistir(self):
        self.detayli_gorunum = not self.detayli_gorunum
        txt = "👁️ GİZLE" if self.detayli_gorunum else "🔲 GÖSTER"
        col = "#2c3e50" if self.detayli_gorunum else "#95a5a6"
        self.btn_mod.update_color(col)
        self.btn_mod.itemconfig(self.btn_mod.find_withtag("text"), text=txt)
        self.listeyi_guncelle()

    def listeyi_guncelle(self, filtre="tumu", query=None, yukleme_filtresi=None):
        if not hasattr(self, 'tablo_stok'): return
        for i in self.tablo_stok.get_children(): self.tablo_stok.delete(i)
        
        kosullar = []; params = []
        
        # --- GÖRÜNÜM AYARLARI ---
        if self.detayli_gorunum:
            cols = ("NO", "TIK", "GTIN", "İLAÇ ADI", "SKT", "KAREKOD")
            widths = [50, 30, 135, 350, 200, 200]
            stretches = [False, False, False, True, False, True]
            base_sql = "SELECT id, barkod, ad, skt, seri_no, parti_no FROM ilaclar"
        else:
            cols = ("NO", "TIK", "GTIN", "İLAÇ ADI", "ADET", "SKT")
            widths = [50, 30, 200, 250, 100, 250]
            stretches = [False, False, False, True, False, False]
            base_sql = "SELECT MIN(id), barkod, ad, MIN(skt), COUNT(*), parti_no FROM ilaclar"

        # --- DÜZELTME BURADA: FİLTRELEME MANTIĞI ---
        if yukleme_filtresi is not None:
            # Artık 'yukleme_id' (Doğum) değil, 'raf_yeri' (Son Görülen Liste) sütununa bakıyoruz.
            # Böylece o listede güncellenen eski ilaçlar da görünür!
            kosullar.append("raf_yeri = ?")
            params.append(str(yukleme_filtresi))
            
        if query:
            kosullar.append("(ad LIKE ? OR barkod LIKE ?)")
            params.extend([f'%{query}%', f'%{query}%'])

        if kosullar: 
            if self.detayli_gorunum: base_sql += " WHERE " + " AND ".join(kosullar)
            else: 
                # Gruplu görünümde de aynı filtre geçerli
                base_sql += " WHERE " + " AND ".join(kosullar)

        if not self.detayli_gorunum: 
            base_sql += " GROUP BY barkod, parti_no, ad"
            if query: base_sql += " HAVING (ad LIKE ? OR barkod LIKE ?)"

        base_sql += " ORDER BY skt ASC"
        
        try: self.imlec.execute(base_sql, params)
        except Exception as e: print(e); return
        
        # Tabloyu Kur
        self.tablo_stok["columns"] = cols
        for c, w, s in zip(cols, widths, stretches):
            if c == "TIK": self.tablo_stok.heading(c, text="☐", command=self.tumunu_sec_stok)
            else: self.tablo_stok.heading(c, text=c)
            align = "center" if c in ["NO", "TIK", "GTIN", "SKT", "ADET"] else "w"
            self.tablo_stok.column(c, width=w, anchor=align, stretch=s)

        # Veriyi Dök
        for idx, r in enumerate(self.imlec.fetchall(), 1):
            try:
                if self.detayli_gorunum:
                    iid, gtin, ad, skt, sn, bn = r
                else:
                    iid, gtin, ad, skt, adet, bn = r
                
                try: skt_dt = datetime.strptime(skt, '%Y-%m-%d').date()
                except: skt_dt = date.today()
                kalan = (skt_dt - date.today()).days
                
                tag = "yesil"
                if kalan < 0: tag = "kirmizi"
                elif kalan < 90: tag = "kirmizi"
                elif kalan < 180: tag = "turuncu"
                elif kalan < 365: tag = "sari"
                
                # --- YENİ TİKLİ FİLTRE KONTROLÜ ---
                # Eğer renk filtresi oluşturulduysa ve bu renk kapalıysa (False), o satırı atla
                if hasattr(self, 'aktif_renk_filtreleri'):
                    if not self.aktif_renk_filtreleri.get(tag, True):
                        continue 
                
                skt_str = f"{skt} ({kalan} gün)"
                
                if self.detayli_gorunum:
                    ymd = skt_dt.strftime('%y%m%d')
                    # Karekod görseli temizliği
                    g_c = str(gtin).replace('.0','').zfill(14)
                    s_c = str(sn).replace('.0','')
                    b_c = str(bn).replace('.0','')
                    qr_code = f"01{g_c}21{s_c}17{ymd}10{b_c}"
                    vals = (idx, "☐", g_c, ad, skt_str, qr_code)
                else:
                    g_c = str(gtin).replace('.0','').zfill(14)
                    vals = (idx, "☐", g_c, ad, f"{adet} Adet", skt_str)
                
                self.tablo_stok.insert("", "end", iid=iid, values=vals, tags=(tag,))
            except: pass

    # 11. YENİLENMİŞ AYARLAR (SOLDA ALT MENÜLÜ SİSTEM)
    def arayuz_ayarlar(self):
        c = TM
        # Ana içerik alanını temizle
        for w in self.content_area.winfo_children(): w.destroy()

        # --- LAYOUT YAPISI ---
        # Sol Taraf: Alt Menü (Sub-Sidebar)
        # Sağ Taraf: İçerik Alanı (Content)
        
        # Ana Konteyner
        container = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        container.pack(fill="both", expand=True)

        # 1. SOL ALT MENÜ (SUB-SIDEBAR)
        sub_sidebar = tk.Frame(container, bg=c.get_color("card_bg"), width=200)
        sub_sidebar.pack(side="left", fill="y", padx=(0, 20))
        sub_sidebar.pack_propagate(False) # Genişliği sabitle

        # Alt Menü Başlığı
        tk.Label(sub_sidebar, text="AYARLAR", font=("Segoe UI", 12, "bold"), 
                 bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).pack(pady=(20, 20))

        # 2. SAĞ İÇERİK ALANI
        self.settings_content_frame = tk.Frame(container, bg=c.get_color("bg_main"))
        self.settings_content_frame.pack(side="left", fill="both", expand=True)

        # --- İÇERİK YÖNETİM FONKSİYONU ---
        def icerik_yukle(sayfa_kodu):
            # Sağ tarafı temizle
            for w in self.settings_content_frame.winfo_children(): w.destroy()
            
            # Seçili butonu vurgula (Basit bir görsel efekt)
            for btn in btn_list:
                if btn.text == sayfa_kodu: btn.update_color(c.get_color("btn_primary"))
                else: btn.update_color(c.get_color("bg_sidebar")) # Pasif renk

            if sayfa_kodu == "GENEL & MAİL": render_mail_ayarlari()
            elif sayfa_kodu == "ÖZET EKRANI": render_dashboard_ayarlari() 
            elif sayfa_kodu == "YEDEKLEME": render_yedekleme()
            elif sayfa_kodu == "KULLANICI YÖNETİMİ": self.render_kullanici_ayarlari()
            elif sayfa_kodu == "AĞ & VERİTABANI": self.render_ag_ayarlari()
            elif sayfa_kodu == "İŞLEM GEÇMİŞİ": self.render_log_ayarlari()
            elif sayfa_kodu == "AĞ & VERİTABANI": self.render_ag_ayarlari()

        # --- ALT MENÜ BUTONLARI ---
        btn_list = []
        
        # Buton 1: Mail
        b1 = ModernButton(sub_sidebar, text="GENEL & MAİL", width=180, height=40, 
                          command=lambda: icerik_yukle("GENEL & MAİL"), bg_color=c.get_color("bg_sidebar"))
        b1.pack(pady=5)
        btn_list.append(b1)
        # YENİ BUTON: Özet Ekranı
        b_ozet = ModernButton(sub_sidebar, text="ÖZET EKRANI", width=180, height=40, 
                              command=lambda: icerik_yukle("ÖZET EKRANI"), bg_color=c.get_color("bg_sidebar"))
        b_ozet.pack(pady=5)
        btn_list.append(b_ozet)

        # Buton 2: Yedekleme
        b2 = ModernButton(sub_sidebar, text="YEDEKLEME", width=180, height=40, 
                          command=lambda: icerik_yukle("YEDEKLEME"), bg_color=c.get_color("bg_sidebar"))
        b2.pack(pady=5)
        btn_list.append(b2)
        # Buton 3: Kullanıcı Yönetimi
        b_kullanici = ModernButton(sub_sidebar, text="KULLANICI YÖNETİMİ", width=180, height=40, 
                                   command=lambda: icerik_yukle("KULLANICI YÖNETİMİ"), bg_color=c.get_color("bg_sidebar"))
        b_kullanici.pack(pady=5)
        btn_list.append(b_kullanici)

        # Buton 4: Ağ ve Veritabanı
        b_ag = ModernButton(sub_sidebar, text="AĞ & VERİTABANI", width=180, height=40, 
                            command=lambda: icerik_yukle("AĞ & VERİTABANI"), bg_color=c.get_color("bg_sidebar"))
        b_ag.pack(pady=5)
        btn_list.append(b_ag)

        # Buton 5: İşlem Geçmişi
        b_log = ModernButton(sub_sidebar, text="İŞLEM GEÇMİŞİ (LOG)", width=180, height=40, 
                            command=lambda: icerik_yukle("İŞLEM GEÇMİŞİ"), bg_color=c.get_color("bg_sidebar"))
        b_log.pack(pady=5)
        btn_list.append(b_log)

        # --- SAĞ TARAFTAKİ EKRANLAR ---

        def render_mail_ayarlari():
            # Başlık
            tk.Label(self.settings_content_frame, text="Genel Sistem ve E-Posta Ayarları", font=FONT_HEAD, 
                     bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(anchor="w", pady=(0, 20))

            # --- 1. E-POSTA VE BAŞLANGIÇ KARTI ---
            card1 = tk.Frame(self.settings_content_frame, bg=c.get_color("card_bg"), padx=30, pady=20, 
                            highlightbackground=c.get_color("border"), highlightthickness=1)
            card1.pack(fill="x", anchor="n", pady=(0, 15))

            lbl_style = {"bg": c.get_color("card_bg"), "fg": c.get_color("fg_text"), "font": ("Segoe UI", 10, "bold")}
            entry_style = {"font": ("Segoe UI", 11), "bg": c.get_color("input_bg"), "fg": c.get_color("input_fg"), 
                           "insertbackground": c.get_color("input_fg"), "relief": "solid", "bd": 1}

            # Mail Formu
            tk.Label(card1, text="Gönderen Gmail Adresi:", **lbl_style).pack(anchor="w", pady=(5, 2))
            e1 = tk.Entry(card1, **entry_style); e1.pack(fill="x", pady=(0, 10), ipady=3)

            f_pass = tk.Frame(card1, bg=c.get_color("card_bg")); f_pass.pack(fill="x")
            tk.Label(f_pass, text="Gmail Uygulama Şifresi:", **lbl_style).pack(side="left")
            
            def sifre_nasil_alinir_goster(event=None):
                bilgi_metni = (
                    "Gmail Uygulama Şifresi Almak İçin Adımlar:\n\n"
                    "1. Gönderici olarak kullanacağınız Gmail hesabıyla Google'a giriş yapın.\n"
                    "2. Sağ üstten profil resminize tıklayıp 'Google Hesabınızı yönetin' seçeneğine girin.\n"
                    "3. Sol menüden 'Güvenlik' sekmesine tıklayın.\n"
                    "4. Eğer kapalıysa, '2 Adımlı Doğrulama'yı mutlaka açın. (Bu şarttır!)\n"
                    "5. Arama çubuğuna (sayfanın üstünde) 'Uygulama şifreleri' yazın ve tıklayın.\n"
                    "6. Uygulama adına 'Eczane Asistanı' yazın ve Oluştur butonuna basın.\n"
                    "7. Ekranda size 16 harfli (örn: abcd efgh ijkl mnop) bir şifre verecek.\n"
                    "8. O şifreyi kopyalayıp aşağıdaki alana yapıştırın ve KAYDET deyin."
                )
                messagebox.showinfo("Uygulama Şifresi Nasıl Alınır?", bilgi_metni)

            lbl_help = tk.Label(f_pass, text=" ❓ Nasıl Alınır? ", bg="#e67e22", fg="white", cursor="hand2", font=("Segoe UI", 9, "bold"))
            lbl_help.pack(side="left", padx=10)
            lbl_help.bind("<Button-1>", sifre_nasil_alinir_goster)

            e2 = tk.Entry(card1, show="*", **entry_style); e2.pack(fill="x", pady=(0, 10), ipady=3)

            tk.Label(card1, text="Alıcı Mail Adresi:", **lbl_style).pack(anchor="w", pady=(5, 2))
            e3 = tk.Entry(card1, **entry_style); e3.pack(fill="x", pady=(0, 10), ipady=3)

            # Başlangıç Ayarı
            tk.Frame(card1, height=1, bg=c.get_color("border")).pack(fill="x", pady=15)
            
            card1.var_startup = tk.BooleanVar()
            self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='pc_acilis_baslat'")
            res = self.imlec.fetchone()
            card1.var_startup.set(True if res and res[0] == '1' else False)

            chk = tk.Checkbutton(card1, text="Bilgisayar açıldığında programı otomatik başlat", 
                                 variable=card1.var_startup, **lbl_style, 
                                 selectcolor=c.get_color("input_bg"), activebackground=c.get_color("card_bg"))
            chk.pack(anchor="w")


            # --- 2. OTOMATİK ÖDEME ONAY KARTI (YENİ EKLENEN BÖLÜM) ---
            card2 = tk.Frame(self.settings_content_frame, bg=c.get_color("card_bg"), padx=30, pady=20, 
                            highlightbackground=c.get_color("border"), highlightthickness=1)
            card2.pack(fill="x", anchor="n", pady=(0, 15))

            tk.Label(card2, text="Otomatik Ödeme Onayı (Günü Geçenleri Kapat)", font=("Segoe UI", 12, "bold"), bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).pack(anchor="w", pady=(0, 5))
            tk.Label(card2, text="Bu ayarlar aktif edildiğinde; tarihi (vadesi) geçmiş olan kayıtlar, sistem tarafından otomatik olarak 'ÖDENDİ' yapılıp yeşile çevrilir.", font=("Segoe UI", 9), bg=c.get_color("card_bg"), fg="#64748b").pack(anchor="w", pady=(0, 15))

            card2.var_oto_depo = tk.BooleanVar()
            card2.var_oto_kart = tk.BooleanVar()

            # Veritabanından mevcut finans otomatik ödeme ayarlarını çek
            try:
                self.imlec_finans.execute("SELECT deger FROM program_ayarlari WHERE ayar_adi='oto_odeme_depo'")
                res_depo = self.imlec_finans.fetchone()
                card2.var_oto_depo.set(True if res_depo and res_depo[0] == '1' else False)

                self.imlec_finans.execute("SELECT deger FROM program_ayarlari WHERE ayar_adi='oto_odeme_kart'")
                res_kart = self.imlec_finans.fetchone()
                card2.var_oto_kart.set(True if res_kart and res_kart[0] == '1' else False)
            except Exception as e:
                pass

            chk_depo = tk.Checkbutton(card2, text="🚚 Depo Ödemeleri için günü geçenleri otomatik ÖDENDİ yap", 
                                 variable=card2.var_oto_depo, **lbl_style, 
                                 selectcolor=c.get_color("input_bg"), activebackground=c.get_color("card_bg"))
            chk_depo.pack(anchor="w", pady=2)

            chk_kart = tk.Checkbutton(card2, text="💳 Kredi Kartları için günü geçenleri otomatik ÖDENDİ yap", 
                                 variable=card2.var_oto_kart, **lbl_style, 
                                 selectcolor=c.get_color("input_bg"), activebackground=c.get_color("card_bg"))
            chk_kart.pack(anchor="w", pady=2)

            # Verileri Yükle (Mail)
            try:
                self.imlec.execute("SELECT anahtar, deger FROM ayarlar")
                d = dict(self.imlec.fetchall())
                e1.insert(0, d.get('gonderen_mail', ''))
                e2.insert(0, d.get('uygulama_sifresi', ''))
                e3.insert(0, d.get('alici_mail', ''))
            except: pass

            def kaydet():
                try:
                    self.başlangıç_ayarı_güncelle(card1.var_startup.get())
                    
                    # Mail ayarlarını SKT veritabanına kaydet
                    self.imlec.execute("UPDATE ayarlar SET deger=? WHERE anahtar='gonderen_mail'", (e1.get().strip(),))
                    self.imlec.execute("UPDATE ayarlar SET deger=? WHERE anahtar='uygulama_sifresi'", (e2.get().strip(),))
                    self.imlec.execute("UPDATE ayarlar SET deger=? WHERE anahtar='alici_mail'", (e3.get().strip(),))
                    self.imlec.execute("UPDATE ayarlar SET deger=? WHERE anahtar='pc_acilis_baslat'", ('1' if card1.var_startup.get() else '0',))
                    self.baglanti_skt.commit() 

                    # Oto Ödeme ayarlarını FİNANS veritabanına kaydet
                    val_depo = '1' if card2.var_oto_depo.get() else '0'
                    val_kart = '1' if card2.var_oto_kart.get() else '0'
                    self.imlec_finans.execute("INSERT OR REPLACE INTO program_ayarlari (ayar_adi, deger) VALUES ('oto_odeme_depo', ?)", (val_depo,))
                    self.imlec_finans.execute("INSERT OR REPLACE INTO program_ayarlari (ayar_adi, deger) VALUES ('oto_odeme_kart', ?)", (val_kart,))
                    self.baglanti_finans.commit()

                    # KAYDEDİLDİĞİ ANDA SİSTEMİ ÇALIŞTIR (Geçmişleri Anında Yeşile Çevirsin)
                    try:
                        self.otomatik_odeme_motoru()
                    except: pass

                    messagebox.showinfo("Başarılı", "Ayarlar başarıyla kaydedildi.\nAktif ettiğiniz otomatik sistemler hemen devreye girdi.")
                except Exception as e: 
                    messagebox.showerror("Hata", f"Kaydedilemedi: {e}")

            ModernButton(self.settings_content_frame, text="KAYDET", command=kaydet, bg_color=c.get_color("btn_success"), width=150).pack(pady=10, anchor="w")

        def render_yedekleme():
            # Başlık
            tk.Label(self.settings_content_frame, text="Yedekleme Merkezi", font=FONT_HEAD, 
                     bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(anchor="w", pady=(0, 20))

            # --- AYAR KARTI ---
            card = tk.Frame(self.settings_content_frame, bg=c.get_color("card_bg"), padx=30, pady=30, 
                            highlightbackground=c.get_color("border"), highlightthickness=1)
            card.pack(fill="x", anchor="n")

            tk.Label(card, text="Yedekleme Konumu Seçimi", font=("Segoe UI", 12, "bold"), 
                     bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).pack(anchor="w", pady=(0,10))

            # Konum Seçim Satırı
            f_path = tk.Frame(card, bg=c.get_color("card_bg"))
            f_path.pack(fill="x", pady=(0, 15))

            # Mevcut konumu veritabanından çek
            self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='yedekleme_konumu'")
            res = self.imlec.fetchone()
            mevcut_konum = res[0] if res else "Yedekler"

            entry_path = tk.Entry(f_path, font=("Segoe UI", 11), bg=c.get_color("input_bg"), 
                                  fg=c.get_color("fg_text"), relief="solid", bd=1)
            entry_path.insert(0, mevcut_konum)
            entry_path.pack(side="left", fill="x", expand=True, ipady=5)

            # Klasör Seçme Butonu
            def konum_degistir():
                klasor = filedialog.askdirectory(title="Yedeklerin Kaydedileceği Klasörü Seçin")
                if klasor:
                    entry_path.delete(0, tk.END)
                    entry_path.insert(0, klasor)
                    # Seçimi anında kaydet
                    self.imlec.execute("UPDATE ayarlar SET deger=? WHERE anahtar='yedekleme_konumu'", (klasor,))
                    self.baglanti_skt.commit()

            ModernButton(f_path, text="📂 SEÇ", command=konum_degistir, 
                         bg_color=c.get_color("btn_primary"), width=80, height=32).pack(side="left", padx=(10, 0))

            # YEDEKLEME İŞLEMİ
            def yedekle_ve_listele():
                target_dir = entry_path.get()
                
                # Klasör yoksa oluştur (Eğer varsayılan 'Yedekler' silindiyse vs.)
                if not os.path.exists(target_dir):
                    try:
                        os.makedirs(target_dir)
                    except OSError:
                        messagebox.showerror("Hata", "Belirtilen klasör oluşturulamadı veya erişilemiyor.")
                        return

                dosya_adi = f"EczaneYedek_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
                hedef_dosya = os.path.join(target_dir, dosya_adi)

                try:
                    shutil.copy2(self.db_skt_adi, hedef_dosya)
                    messagebox.showinfo("Başarılı", f"Yedek alındı:\n{hedef_dosya}")
                    gecmis_listesini_guncelle()
                except Exception as e:
                    messagebox.showerror("Hata", f"Yedekleme başarısız:\n{str(e)}")

            ModernButton(card, text="ŞİMDİ YEDEKLE", command=yedekle_ve_listele, 
                         bg_color=c.get_color("btn_warning"), width=200).pack(anchor="w")

            # --- GEÇMİŞ YEDEKLER LİSTESİ ---
            tk.Label(self.settings_content_frame, text="📂 Bu Konumdaki Yedekler", font=("Segoe UI", 12, "bold"), 
                     bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(anchor="w", pady=(30, 10))
            
            list_frame = tk.Frame(self.settings_content_frame, bg=c.get_color("card_bg"))
            list_frame.pack(fill="both", expand=True, pady=(0, 20))
            
            sc = ttk.Scrollbar(list_frame, orient="vertical")
            lb = tk.Listbox(list_frame, bg=c.get_color("input_bg"), fg=c.get_color("fg_text"), bd=0, 
                            highlightthickness=0, font=("Consolas", 10), yscrollcommand=sc.set)
            sc.config(command=lb.yview); sc.pack(side="right", fill="y")
            lb.pack(side="left", fill="both", expand=True, padx=5, pady=5)

            def gecmis_listesini_guncelle():
                lb.delete(0, tk.END)
                hedef = entry_path.get()
                if os.path.exists(hedef):
                    dosyalar = []
                    for f in os.listdir(hedef):
                        if f.endswith(".db"):
                            tam_yol = os.path.join(hedef, f)
                            dosyalar.append((tam_yol, os.path.getmtime(tam_yol), os.path.getsize(tam_yol)))
                    
                    # Tarihe göre tersten sırala (en yeni en üstte)
                    dosyalar.sort(key=lambda x: x[1], reverse=True)

                    if dosyalar:
                        for yol, mtime, size in dosyalar:
                            ad = os.path.basename(yol)
                            tarih = datetime.fromtimestamp(mtime).strftime('%d.%m.%Y %H:%M')
                            mb = size / 1024 / 1024
                            lb.insert(tk.END, f"📦 {ad:<35} | {tarih} | {mb:.2f} MB")
                    else:
                        lb.insert(tk.END, "Bu klasörde yedek dosyası bulunamadı.")
                else:
                    lb.insert(tk.END, "Seçili klasör mevcut değil.")

            # İlk açılışta listeyi doldur
            gecmis_listesini_guncelle()

        def render_dashboard_ayarlari():
            tk.Label(self.settings_content_frame, text="Özet Ekranı (Dashboard) Ayarları", font=FONT_HEAD, bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(anchor="w", pady=(0, 20))
            card = tk.Frame(self.settings_content_frame, bg=c.get_color("card_bg"), padx=30, pady=20, highlightbackground=c.get_color("border"), highlightthickness=1)
            card.pack(fill="x", anchor="n")

            TUM_METRIKLER = {
                "stok_toplam": "Toplam Stok (Kutu)",
                "stok_kritik": "Kritik Miad (0-3 Ay) 🔴",
                "stok_uyari": "Yaklaşan Miad (3-6 Ay) 🟠",
                "stok_dikkat": "Dikkat Miad (6-12 Ay) 🟡",
                "finans_borc": "Toplam Açık Borç",
                "finans_gelecek_ay": "Gelecek Ayki Toplam Borç",
                "finans_bakiye": "Aylık Net Bakiye",
                "kasa_bugun_giris": "Bugünkü Kasa Girişi",
                "kasa_bugun_cikis": "Bugünkü Kasa Çıkışı",
                "finans_giris": "Bu Ay Kasa Girişi",
                "finans_cikis": "Bu Ay Kasa Çıkışı",
                "kasa_kk_giris": "Bu Ay POS/Kart Çekimi",
                "finans_depo": "Bu Ayki Depo Ödemesi",
                "finans_kart": "Bu Ayki Kart Ödemesi",
                "finans_kredi": "Bu Ayki Kredi Taksitleri",
                "finans_sgk": "Bekleyen Kurum Geliri"
            }

            self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='dashboard_kard_ayarlari'")
            res = self.imlec.fetchone()
            mevcut_secimler = eval(res[0]) if res and res[0] else ["stok_toplam", "stok_kritik", "finans_bakiye", "finans_borc"]

            f_top = tk.Frame(card, bg=c.get_color("card_bg"))
            f_top.pack(fill="x", pady=(0, 20))
            tk.Label(f_top, text="Ekranda Kaç Kutu Görünsün?", font=("Segoe UI", 11, "bold"), bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).pack(side="left")
            
            var_count = tk.StringVar(value=str(len(mevcut_secimler)))
            cmb_count = ttk.Combobox(f_top, values=["2", "4", "6", "8", "10"], textvariable=var_count, state="readonly", width=5, font=("Segoe UI", 11))
            cmb_count.pack(side="left", padx=10)

            f_dropdowns = tk.Frame(card, bg=c.get_color("card_bg"))
            f_dropdowns.pack(fill="both", expand=True)

            secim_kutu_listesi = []

            def secenekleri_ciz(event=None):
                for w in f_dropdowns.winfo_children(): w.destroy()
                secim_kutu_listesi.clear()
                
                kutu_sayisi = int(var_count.get())
                options = list(TUM_METRIKLER.values())

                for i in range(kutu_sayisi):
                    row = i // 2
                    col = (i % 2) * 2
                    tk.Label(f_dropdowns, text=f"{i+1}. Kutu Verisi:", font=("Segoe UI", 10), bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).grid(row=row, column=col, sticky="e", padx=5, pady=10)
                    
                    cmb = ttk.Combobox(f_dropdowns, values=options, state="readonly", width=25, font=("Segoe UI", 10))
                    
                    if i < len(mevcut_secimler) and mevcut_secimler[i] in TUM_METRIKLER:
                        cmb.set(TUM_METRIKLER[mevcut_secimler[i]])
                    else:
                        cmb.current(min(i, len(options)-1))
                        
                    cmb.grid(row=row, column=col+1, sticky="w", padx=5, pady=10)
                    secim_kutu_listesi.append(cmb)

            cmb_count.bind("<<ComboboxSelected>>", secenekleri_ciz)
            secenekleri_ciz()

            def kaydet_ayarlar():
                yeni_liste = []
                ters_sozluk = {v: k for k, v in TUM_METRIKLER.items()}
                for cmb in secim_kutu_listesi:
                    yeni_liste.append(ters_sozluk[cmb.get()])
                
                try:
                    self.imlec.execute("INSERT OR REPLACE INTO ayarlar (anahtar, deger) VALUES ('dashboard_kard_ayarlari', ?)", (str(yeni_liste),))
                    self.baglanti_skt.commit() # DÜZELTİLDİ: baglanti_skt yapıldı
                    messagebox.showinfo("Başarılı", "Özet ekranı ayarlarınız kaydedildi!\nAna sayfaya döndüğünüzde değişiklikleri görebilirsiniz.")
                except Exception as e:
                    messagebox.showerror("Hata", str(e))

            ModernButton(self.settings_content_frame, text="KUTU DÜZENİNİ KAYDET", command=kaydet_ayarlar, bg_color=c.get_color("btn_success"), width=250).pack(pady=20, anchor="w")
       
        # Varsayılan olarak ilk sekmeyi aç
        icerik_yukle("GENEL & MAİL")

    def render_kullanici_ayarlari(self):
            c = TM 
            for widget in self.settings_content_frame.winfo_children(): widget.destroy()

            tk.Label(self.settings_content_frame, text="Kullanıcı ve Yetki Yönetimi", font=("Segoe UI", 18, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(anchor="w", pady=(0, 10))

            # =================================================================
            # PERSONEL YETKİ KONTROL PANELİ
            # =================================================================
            f_yetki = tk.LabelFrame(self.settings_content_frame, text="Personel Yetki Sınırlandırması (Genel)", font=("Segoe UI", 11, "bold"), bg=c.get_color("card_bg"), fg="#ef4444", padx=15, pady=10)
            f_yetki.pack(fill="x", pady=(0, 20))

            tk.Label(f_yetki, text="Aşağıdaki kutuları işaretleyerek personelin sisteme girdiğinde neleri görebileceğini seçebilirsiniz:", font=("Segoe UI", 9), bg=c.get_color("card_bg"), fg="#64748b").pack(anchor="w", pady=(0, 10))

            var_finans = tk.BooleanVar(value=self.personel_yetkileri.get("finans_gorsun", False))
            var_para = tk.BooleanVar(value=self.personel_yetkileri.get("para_gorsun", False))
            var_ayar = tk.BooleanVar(value=self.personel_yetkileri.get("ayarlar_gorsun", False))

            tk.Checkbutton(f_yetki, text="📊 Finans Modüllerine (Kasa, Depo, Kart) Girebilsin", variable=var_finans, bg=c.get_color("card_bg"), fg=c.get_color("fg_text"), selectcolor=c.get_color("input_bg"), font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=2)
            tk.Checkbutton(f_yetki, text="💰 Ana Sayfadaki Parasal Değerleri Görebilsin (Aksi halde ***** ₺ görünür)", variable=var_para, bg=c.get_color("card_bg"), fg=c.get_color("fg_text"), selectcolor=c.get_color("input_bg"), font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=2)
            tk.Checkbutton(f_yetki, text="⚙️ Ayarlar Sekmesine Girebilsin", variable=var_ayar, bg=c.get_color("card_bg"), fg=c.get_color("fg_text"), selectcolor=c.get_color("input_bg"), font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=2)

            def yetkileri_kaydet():
                yeni_yetki = {
                    "finans_gorsun": var_finans.get(),
                    "para_gorsun": var_para.get(),
                    "ayarlar_gorsun": var_ayar.get()
                }
                self.personel_yetkileri = yeni_yetki
                self.imlec.execute("INSERT OR REPLACE INTO ayarlar (anahtar, deger) VALUES ('personel_yetkileri', ?)", (str(yeni_yetki),))
                self.baglanti_skt.commit()
                messagebox.showinfo("Başarılı", "Personel yetkileri başarıyla güncellendi.")

            ModernButton(f_yetki, text="YETKİLERİ KAYDET", command=yetkileri_kaydet, bg_color="#10b981", width=180, height=30).pack(anchor="e", pady=(10, 0))

            # --- YENİ KULLANICI EKLEME FORMU ---
            form_frame = tk.Frame(self.settings_content_frame, bg=c.get_color("card_bg"), bd=1, relief="solid")
            form_frame.pack(fill="x", pady=(0, 20), ipady=10)

            tk.Label(form_frame, text="Yeni Kullanıcı Ekle", font=("Segoe UI", 11, "bold"), bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).grid(row=0, column=0, columnspan=6, pady=10, sticky="w", padx=15)

            tk.Label(form_frame, text="Kullanıcı Adı:", bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).grid(row=1, column=0, padx=(15, 5), pady=5)
            ent_kadi = ttk.Entry(form_frame, font=("Segoe UI", 10))
            ent_kadi.grid(row=1, column=1, padx=5, pady=5)

            tk.Label(form_frame, text="Şifre:", bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).grid(row=1, column=2, padx=5, pady=5)
            ent_sifre = ttk.Entry(form_frame, font=("Segoe UI", 10), show="*")
            ent_sifre.grid(row=1, column=3, padx=5, pady=5)

            tk.Label(form_frame, text="Rol/Yetki:", bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).grid(row=1, column=4, padx=5, pady=5)
            cmb_rol = ttk.Combobox(form_frame, values=["Yönetici", "Personel"], state="readonly", width=12, font=("Segoe UI", 10))
            cmb_rol.current(1) 
            cmb_rol.grid(row=1, column=5, padx=5, pady=5)

            def kullanici_ekle():
                k_adi = ent_kadi.get().strip()
                sifre = ent_sifre.get().strip()
                rol = cmb_rol.get()

                if not k_adi or not sifre:
                    messagebox.showwarning("Uyarı", "Kullanıcı adı ve şifre boş bırakılamaz!")
                    return
                try:
                    self.imlec_kullanici.execute("SELECT COUNT(*) FROM kullanicilar WHERE k_adi=?", (k_adi,))
                    if self.imlec_kullanici.fetchone()[0] > 0:
                        messagebox.showerror("Hata", "Bu kullanıcı adı zaten mevcut.")
                        return
                    self.imlec_kullanici.execute("INSERT INTO kullanicilar (k_adi, sifre, rol) VALUES (?, ?, ?)", (k_adi, sifre, rol))
                    self.vt_kullanici.commit()
                    ent_kadi.delete(0, 'end'); ent_sifre.delete(0, 'end')
                    liste_guncelle()
                    messagebox.showinfo("Başarılı", f"{k_adi} sisteme {rol} yetkisiyle eklendi.")
                except Exception as e:
                    messagebox.showerror("Hata", f"Kullanıcı eklenemedi: {e}")

            ModernButton(form_frame, text="KAYDET", command=kullanici_ekle, bg_color="#3b82f6", width=100, height=28).grid(row=1, column=6, padx=(15, 10))

            # --- KAYITLI KULLANICILAR LİSTESİ ---
            cols = ("ID", "Kullanıcı Adı", "Yetki (Rol)")
            tree = ttk.Treeview(self.settings_content_frame, columns=cols, show="headings", height=8)
            tree.heading("ID", text="ID"); tree.heading("Kullanıcı Adı", text="Kullanıcı Adı"); tree.heading("Yetki (Rol)", text="Yetki (Rol)")
            tree.column("ID", width=50, anchor="center"); tree.column("Kullanıcı Adı", width=200, anchor="w"); tree.column("Yetki (Rol)", width=150, anchor="center")
            tree.pack(fill="x", pady=5)

            def liste_guncelle():
                for row in tree.get_children(): tree.delete(row)
                self.imlec_kullanici.execute("SELECT id, k_adi, rol FROM kullanicilar")
                for veri in self.imlec_kullanici.fetchall():
                    tree.insert("", "end", values=veri)
            liste_guncelle()

            def kullanici_sil():
                secili = tree.selection()
                if not secili: return
                item = tree.item(secili[0])
                k_id = item['values'][0]
                rol = item['values'][2]

                if rol == "Yönetici":
                    self.imlec_kullanici.execute("SELECT COUNT(*) FROM kullanicilar WHERE rol='Yönetici'")
                    if self.imlec_kullanici.fetchone()[0] <= 1:
                        messagebox.showerror("Yetki Hatası", "Sistemdeki TEK yönetici hesabı silinemez!\nÖnce başka bir yönetici eklemelisiniz.")
                        return

                if messagebox.askyesno("Kullanıcı Sil", "Seçilen kullanıcı silinecektir. Onaylıyor musunuz?"):
                    self.imlec_kullanici.execute("DELETE FROM kullanicilar WHERE id=?", (k_id,))
                    self.vt_kullanici.commit()
                    liste_guncelle()

            ModernButton(self.settings_content_frame, text="Seçili Kullanıcıyı Sil", command=kullanici_sil, bg_color="#ef4444", width=180, height=35).pack(anchor="e", pady=10)

    def render_ag_ayarlari(self):
        c = TM  
        for w in self.settings_content_frame.winfo_children(): w.destroy()
        
        tk.Label(self.settings_content_frame, text="Ortak Veritabanı ve Ağ Ayarları", font=FONT_HEAD, bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(anchor="w", pady=(0, 20))
        
        card = tk.Frame(self.settings_content_frame, bg=c.get_color("card_bg"), padx=30, pady=20, highlightbackground=c.get_color("border"), highlightthickness=1)
        card.pack(fill="x", anchor="n")
        
        tk.Label(card, text="Ağ Yolu veya IP Adresi", font=("Segoe UI", 12, "bold"), bg=c.get_color("card_bg"), fg=c.get_color("fg_text")).pack(anchor="w", pady=(0, 5))
        
        bilgi_metni = (
            "📌 Eğer Ana Bilgisayardaysanız: Yandaki 'KLASÖR SEÇ' butonuna basarak veritabanı klasörünü seçin.\n"
            "📌 Eğer Diğer Bilgisayardaysanız: Kutunun içindekileri silip ana bilgisayarın IP adresini klavyeden yazın:\n"
            "👉 Örnek IP Formatı: \\\\192.168.1.50\\EczaneOrtakVeri"
        )
        tk.Label(card, text=bilgi_metni, font=("Segoe UI", 10), bg=c.get_color("card_bg"), fg="#64748b", justify="left").pack(anchor="w", pady=(0, 15))
        
        f_path = tk.Frame(card, bg=c.get_color("card_bg"))
        f_path.pack(fill="x", pady=(0, 15))
        
        mevcut_yol = yerel_db_yolunu_getir()
        
        # Kullanıcının buraya yazı yazabileceğini belli etmek için yazıyı büyüttük ve belirginleştirdik
        ent_path = tk.Entry(f_path, font=("Segoe UI", 13, "bold"), bg="#f8fafc", fg="#2563eb", relief="solid", bd=1, insertbackground="#2563eb")
        ent_path.insert(0, mevcut_yol)
        ent_path.pack(side="left", fill="x", expand=True, ipady=8)
        
        def sec_klasor():
            klasor = filedialog.askdirectory(title="Ortak Veritabanı Klasörünü Seçin")
            if klasor:
                ent_path.delete(0, tk.END)
                ent_path.insert(0, klasor)
        
        ModernButton(f_path, text="📂 KLASÖR SEÇ", command=sec_klasor, bg_color="#f59e0b", width=140, height=40).pack(side="left", padx=(10, 0))
        
        def kaydet():
            yeni_yol = ent_path.get().strip()
            if not yeni_yol:
                messagebox.showwarning("Uyarı", "Lütfen geçerli bir yol veya IP adresi girin.")
                return
            yerel_db_yolunu_kaydet(yeni_yol)
            messagebox.showinfo("Başarılı", "Ağ IP/Bağlantı yolu başarıyla kaydedildi!\n\nDeğişikliklerin aktif olması için lütfen programı KAPATIP YENİDEN AÇIN.", parent=self.pencere)
        
        ModernButton(card, text="✅ KAYDET VE UYGULA", command=kaydet, bg_color="#10b981", width=250, height=45).pack(anchor="w", pady=(10,0))
    
    # --- DİĞER MODÜLLER ---
    def veritabani_kur(self):
        self.imlec.execute("""CREATE TABLE IF NOT EXISTS ilaclar (id INTEGER PRIMARY KEY AUTOINCREMENT, barkod TEXT, ad TEXT, parti_no TEXT, seri_no TEXT, adet INTEGER, skt TEXT, raf_yeri TEXT, kayit_tarihi TEXT)""")
        self.imlec.execute("""CREATE TABLE IF NOT EXISTS ilac_kartlari (gtin TEXT PRIMARY KEY, ad TEXT, guncelleme_tarihi TEXT)""")
        self.imlec.execute("CREATE TABLE IF NOT EXISTS ayarlar (anahtar TEXT PRIMARY KEY, deger TEXT)")
        self.imlec.execute("CREATE TABLE IF NOT EXISTS kayitli_listeler (id INTEGER PRIMARY KEY AUTOINCREMENT, liste_adi TEXT, icerik TEXT, tarih TEXT)")
        
        # YENİ EKLENEN KISIM: 'yedekleme_konumu'
        defaults = {
            'gonderen_mail': '', 
            'uygulama_sifresi': '', 
            'alici_mail': '', 
            'son_stok_yukleme': '2000-01-01', 
            'son_mail_tarihi': '2000-01-01', 
            'pc_acilis_baslat': '0',
            'yedekleme_konumu': 'Yedekler' # Varsayılan olarak programın içindeki klasör
        }
        
        for k, v in defaults.items(): 
            self.imlec.execute("INSERT OR IGNORE INTO ayarlar (anahtar, deger) VALUES (?, ?)", (k, v))
        self.baglanti_skt.commit()

    def tablo_guncelleme_kontrol(self):
        # 1. Raf Yeri Kontrolü
        try: self.imlec.execute("SELECT raf_yeri FROM ilaclar LIMIT 1")
        except: self.imlec.execute("ALTER TABLE ilaclar ADD COLUMN raf_yeri TEXT"); self.baglanti_skt.commit()
        
        # 2. Kayıt Tarihi Kontrolü
        try: self.imlec.execute("SELECT kayit_tarihi FROM ilaclar LIMIT 1")
        except: self.imlec.execute("ALTER TABLE ilaclar ADD COLUMN kayit_tarihi TEXT"); self.baglanti_skt.commit()

        # 3. YÜKLEME ID KONTROLÜ (HATA VEREN KISIM BU - EKLENDİ)
        try: self.imlec.execute("SELECT yukleme_id FROM ilaclar LIMIT 1")
        except: self.imlec.execute("ALTER TABLE ilaclar ADD COLUMN yukleme_id INTEGER"); self.baglanti_skt.commit()

    def normalize_header(self, text):
        if not isinstance(text, str): return str(text)
        text = text.upper(); mapping = {'İ': 'I', 'Ğ': 'G', 'Ü': 'U', 'Ş': 'S', 'Ö': 'O', 'Ç': 'C', ' ': ''}
        for k, v in mapping.items(): text = text.replace(k, v)
        return text.strip()

    def evrensel_dosya_oku(self, dosya_yolu):
        # BU FONKSİYON DOSYA TÜRÜNÜ OTOMATİK ALGILAR (Excel mi CSV mi?)
        df = pd.DataFrame()
        try:
            # 1. Önce Gerçek Excel (.xlsx) gibi okumayı dene
            try:
                df = pd.read_excel(dosya_yolu, dtype=str).fillna("")
            except:
                # 2. Eğer açamazsa, bu "adı xlsx ama içi csv" olan bir dosyadır.
                # Senin dosyan virgül (,) ile ayrılmış. Bunu UTF-8 ile dene.
                try:
                    df = pd.read_csv(dosya_yolu, sep=',', engine='python', dtype=str, on_bad_lines='skip', encoding='utf-8-sig')
                except:
                    # 3. O da olmazsa noktalı virgül (;) veya Türkçe karakter seti (Latin5) dene
                    df = pd.read_csv(dosya_yolu, sep=None, engine='python', dtype=str, on_bad_lines='skip', encoding='latin5')

            # --- TEMİZLİK ROBOTU ---
            if not df.empty:
                # Başlıkları standartlaştır (Boşlukları at, BÜYÜK HARF yap)
                df.columns = [str(c).replace(';', '').strip().upper() for c in df.columns]
                
                # Sütun isimlerinde "UNNAMED" geçen bozuklukları at
                df = df.loc[:, ~df.columns.str.contains('^UNNAMED')]

                # Verilerin içindeki gereksiz ; ve boşlukları temizle
                for col in df.columns:
                    if df[col].dtype == 'object':
                        df[col] = df[col].str.replace(';', '').str.strip()
            
            return df
        except Exception as e:
            messagebox.showerror("Hata", f"Dosya okunamadı: {str(e)}")
            return pd.DataFrame()

    def excel_yukle_stok(self, otomatik_dosya_yolu=None): # <-- Parametre eklendi
        if otomatik_dosya_yolu:
            d = otomatik_dosya_yolu
        else:
            d = filedialog.askopenfilename(title="İTS Stok Dosyasını Seçin")
        
        if not d: return
        
        dosya_ismi = os.path.basename(d)
        bugun_str = date.today().strftime("%Y-%m-%d")

        try:
            self.imlec.execute("INSERT INTO yuklemeler (dosya_adi, tarih) VALUES (?, ?)", (dosya_ismi, bugun_str))
            self.baglanti_skt.commit()
            aktif_yukleme_id = self.imlec.lastrowid 
        except:
            self.imlec.execute("CREATE TABLE IF NOT EXISTS yuklemeler (id INTEGER PRIMARY KEY AUTOINCREMENT, dosya_adi TEXT, tarih TEXT)")
            self.imlec.execute("INSERT INTO yuklemeler (dosya_adi, tarih) VALUES (?, ?)", (dosya_ismi, bugun_str))
            self.baglanti_skt.commit()
            aktif_yukleme_id = self.imlec.lastrowid 

        df = self.evrensel_dosya_oku(d)
        count_eklenen = 0
        count_guncellenen = 0
        atlanan_satirlar = []
        
        if not df.empty:
            cols = df.columns
            c_g = self.find_column(cols, ['GTIN', 'BARKOD'])
            c_sn = self.find_column(cols, ['SN', 'SERI', 'SIRA'])
            c_bn = self.find_column(cols, ['BN', 'PARTI', 'LOT'])
            c_ad = self.find_column(cols, ['ADI', 'AD', 'İLAÇ ADI'])
            c_sk = self.find_column(cols, ['SON KULLANMA TARİHİ', 'SKT'])
            c_qr = self.find_column(cols, ['KAREKOD'])

            if c_g:
                for idx, r in df.iterrows():
                    try:
                        g = str(r[c_g]).replace(".0", "").strip().zfill(14)
                        if not g or len(g) < 13 or g == "00000000000000":
                            atlanan_satirlar.append(f"Satır {idx+2}: Geçersiz GTIN")
                            continue 

                        sn = str(r[c_sn]).replace(".0", "").strip() if c_sn else ""
                        bn = str(r[c_bn]).replace(".0", "").strip() if c_bn else ""
                        ad = str(r[c_ad]).strip() if c_ad else "Bilinmeyen İlaç"
                        raw_qr = str(r[c_qr]).strip() if c_qr else ""

                        skt_str = "2026-01-01" 
                        tarih_bulundu = False
                        if c_sk:
                            raw_skt = str(r[c_sk]).strip()
                            if len(raw_skt) > 5:
                                try:
                                    ts = pd.to_datetime(raw_skt)
                                    skt_str = ts.strftime('%Y-%m-%d')
                                    tarih_bulundu = True
                                except:
                                    try:
                                        ts = pd.to_datetime(raw_skt, dayfirst=True)
                                        skt_str = ts.strftime('%Y-%m-%d')
                                        tarih_bulundu = True
                                    except: pass
                        
                        if not tarih_bulundu and raw_qr:
                            try:
                                if "17" in raw_qr[-25:]:
                                    parcalar = raw_qr.split("17")
                                    for p in reversed(parcalar[1:]):
                                        if len(p) >= 6 and p[:6].isdigit():
                                            yy, mm, dd = p[:2], p[2:4], p[4:6]
                                            if 1 <= int(mm) <= 12:
                                                skt_str = datetime.strptime(f"{yy}{mm}{dd}", "%y%m%d").strftime("%Y-%m-%d")
                                                break
                            except: pass

                        self.ilac_karti_ekle_guncelle(g, ad)
                        
                        self.imlec.execute("SELECT id, yukleme_id FROM ilaclar WHERE barkod=? AND seri_no=?", (g, sn))
                        mevcut_kayit = self.imlec.fetchone()

                        # --- MANTIK: RAF YERİ = SON GÖRÜLEN LİSTE ID ---
                        if mevcut_kayit:
                            rec_id = mevcut_kayit[0]
                            self.imlec.execute("""
                                UPDATE ilaclar 
                                SET ad=?, parti_no=?, skt=?, raf_yeri=?
                                WHERE id=?
                            """, (ad, bn, skt_str, str(aktif_yukleme_id), rec_id))
                            count_guncellenen += 1
                        else:
                            # YENİ İSE: YUKLEME_ID = ŞİMDİKİ, RAF_YERI = ŞİMDİKİ
                            self.imlec.execute("""
                                INSERT INTO ilaclar (barkod, ad, parti_no, seri_no, adet, skt, raf_yeri, kayit_tarihi, yukleme_id) 
                                VALUES (?,?,?,?,1,?, ?, ?, ?)
                            """, (g, ad, bn, sn, skt_str, str(aktif_yukleme_id), bugun_str, aktif_yukleme_id))
                            count_eklenen += 1
                            
                    except Exception as e:
                        continue
                
                self.baglanti_skt.commit()
                
                # --- YENİ STOK YÜKLEME TARİHİNİ KAYDET (GÜNCELLENDİ) ---
                guncel_tarih = datetime.now().strftime("%d.%m.%Y - %H:%M")
                self.imlec.execute("INSERT OR REPLACE INTO ayarlar (anahtar, deger) VALUES ('son_stok_yukleme_tarihi', ?)", (guncel_tarih,))
                self.baglanti_skt.commit()
                
                self.yukleme_listesini_doldur()
                self.arayuz_stok_takip() # Listeyi yenile

                self.sisteme_bildirim_ekle(f"📦 İTS Stok güncellendi: {dosya_ismi}")
                
                messagebox.showinfo("Başarılı", f"Dosya: {dosya_ismi}\n\n✅ {count_eklenen} yeni ilaç eklendi.\n🔄 {count_guncellenen} ilaç güncellendi.")
            
            else:
                messagebox.showerror("Hata", f"Dosyada 'GTIN' sütunu bulunamadı.")

    
    def find_column(self, columns, keywords):
        for c in columns:
            for k in keywords:
                if self.normalize_header(k) in self.normalize_header(c): return c
        return None

    def treeview_sort(self, tv, col, reverse):
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        try: l.sort(key=lambda t: int(t[0]), reverse=reverse)
        except: l.sort(key=lambda t: t[0].lower(), reverse=reverse)
        for index, (val, k) in enumerate(l): tv.move(k, '', index)
        tv.heading(col, command=lambda: self.treeview_sort(tv, col, not reverse))

    def karekod_parse_ekle(self, event):
        raw = self.ent_qr_input.get().strip()
        if not raw: return
        
        try:
            # 1. KAREKOD AYRIŞTIRMA
            gtin = raw[2:16] if raw.startswith("01") else raw[:14]
            sn, bn, skt = "", "", ""
            
            if "21" in raw: sn = raw.split("21")[1].split("17")[0].split("10")[0][:20]
            if "10" in raw: bn = raw.split("10")[1][:20]
            
            # Tarih Ayrıştırma
            if "17" in raw: 
                skt_raw = raw.split("17")[1][:6]
                try: 
                    skt = datetime.strptime(skt_raw, "%y%m%d").strftime("%Y-%m-%d")
                except: 
                    skt = "2026-01-01"
            else: 
                skt = "2026-01-01"

            # === YENİ EKLENEN KISIM: ÇAKIŞMA KONTROLÜ ===
            # Veritabanında bu Barkod ve Seri No ikilisi var mı?
            self.imlec.execute("SELECT id, ad FROM ilaclar WHERE barkod=? AND seri_no=?", (gtin, sn))
            mevcut_kayit = self.imlec.fetchone()

            if mevcut_kayit:
                # EĞER KAYIT VARSA:
                mevcut_ad = mevcut_kayit[1]
                messagebox.showwarning("Çakışma Uyarısı", 
                                     f"⚠️ DİKKAT: Bu karekod zaten stokta var!\n\n"
                                     f"İlaç Adı: {mevcut_ad}\n"
                                     f"Seri No: {sn}\n\n"
                                     f"Ekleme işlemi İPTAL edildi.")
                
                self.ent_qr_input.delete(0, tk.END) # Kutuyu temizle
                return # Fonksiyondan çık, kaydetme!
            # ============================================

            # 2. İLAÇ ADINI BULMA (Kartlardan)
            self.imlec.execute("SELECT ad FROM ilac_kartlari WHERE gtin=?", (gtin,))
            res = self.imlec.fetchone()
            
            if res:
                ad = res[0] 
            else:
                ad = "Bilinmeyen İlaç" 

            # 3. VERİTABANINA KAYIT
            self.imlec.execute("""
                INSERT INTO ilaclar (barkod, ad, parti_no, seri_no, adet, skt, raf_yeri, kayit_tarihi) 
                VALUES (?,?,?,?,1,?, '', ?)
            """, (gtin, ad, bn, sn, skt, str(date.today())))
            
            self.baglanti_skt.commit()
            
            # Başarılı olduğunda giriş kutusunu temizle ve listeyi yenile
            self.ent_qr_input.delete(0, tk.END)
            self.listeyi_guncelle()
            
        except Exception as e: 
            messagebox.showerror("Hata", f"Karekod işlenirken hata oluştu:\n{str(e)}")

    def secili_kopyala_ozel(self, tip):
        s = [self.tablo_stok.item(i, "values") for i in self.tablo_stok.get_children() if self.tablo_stok.item(i, "values")[1] == "☑"]
        if not s: return
        txt = ""
        for v in s:
            if tip == "brkd": txt += v[2] + "\n"
            elif tip == "kk":
                if self.detayli_gorunum: txt += v[-1] + "\n"
                else: messagebox.showwarning("Hata", "Gruplu modda karekod kopyalanamaz."); return
        self.pencere.clipboard_clear(); self.pencere.clipboard_append(txt)
        messagebox.showinfo("Kopyalandı", "Panoya kopyalandı.")

    def ilac_sil(self):
        s = [i for i in self.tablo_stok.get_children() if self.tablo_stok.item(i, "values")[1] == "☑"]
        if not s: return
        if messagebox.askyesno("Sil", f"{len(s)} kayıt silinecek?"):
            for sid in s:
                if self.detayli_gorunum: self.imlec.execute("DELETE FROM ilaclar WHERE id=?", (sid,))
                else: 
                    vals = self.tablo_stok.item(sid, "values"); g, bn = vals[2], vals[3]
                    self.imlec.execute("DELETE FROM ilaclar WHERE barkod=? AND parti_no=?", (g, bn))
            self.baglanti_skt.commit(); self.listeyi_guncelle()

    def tumunu_sec_stok(self):
        self.secili_tumunu = not self.secili_tumunu; ic = "☑" if self.secili_tumunu else "☐"
        self.tablo_stok.heading("TIK", text=ic)
        for i in self.tablo_stok.get_children(): 
            vals = list(self.tablo_stok.item(i, "values")); vals[1]=ic
            self.tablo_stok.item(i, values=vals)
            
    def satir_tiklama(self, e):
        if self.tablo_stok.identify_column(e.x) == "#2":
            i = self.tablo_stok.identify_row(e.y)
            if i: 
                vals = list(self.tablo_stok.item(i, "values"))
                vals[1] = "☐" if vals[1]=="☑" else "☑"
                self.tablo_stok.item(i, values=vals)
            
    def cakisma_uyari_popup(self, urun_listesi):
        top = tk.Toplevel(self.pencere)
        top.title("⚠️ DİKKAT: '17' İKİLEMİ TESPİT EDİLDİ")
        top.geometry("800x500")
        top.configure(bg="#fff4e6") # Hafif turuncu uyarı rengi

        tk.Label(top, text=f"{len(urun_listesi)} Adet Üründe Tarih Karışıklığı Olabilir!", 
                 font=("Segoe UI", 12, "bold"), bg="#fff4e6", fg="#c0392b").pack(pady=10)
        
        tk.Label(top, text="Karekod içinde birden fazla '17' ve tarih formatı algılandı.\nSistem en sondaki tarihi kabul etti ama lütfen kontrol ediniz.", 
                 bg="#fff4e6", font=("Segoe UI", 10)).pack(pady=5)

        # Tablo Alanı
        frame_table = tk.Frame(top)
        frame_table.pack(fill="both", expand=True, padx=10, pady=10)

        cols = ("İLAÇ ADI", "KAREKOD", "BULUNAN ADAYLAR")
        tree = ttk.Treeview(frame_table, columns=cols, show="headings")
        
        sc = ttk.Scrollbar(frame_table, orient="vertical", command=tree.yview)
        tree.configure(yscroll=sc.set)
        sc.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True)

        tree.heading("İLAÇ ADI", text="İLAÇ ADI")
        tree.column("İLAÇ ADI", width=250)
        tree.heading("KAREKOD", text="KAREKOD (Sorunlu)")
        tree.column("KAREKOD", width=350)
        tree.heading("BULUNAN ADAYLAR", text="ALGILANAN TARİHLER")
        tree.column("BULUNAN ADAYLAR", width=150)

        for urun in urun_listesi:
            ad = urun[0]
            kod = urun[1]
            adaylar = ", ".join(urun[2]) # ['240501', '260101'] gibi listeyi stringe çevirir
            tree.insert("", "end", values=(ad, kod, adaylar))

        tk.Button(top, text="ANLAŞILDI, KAPAT", command=top.destroy, bg="#e67e22", fg="white", font=("Segoe UI", 10, "bold"), height=2).pack(fill="x", padx=20, pady=10)

    def arayuz_ilac_kartlari(self):
        c = TM
        # Başlık
        tk.Label(self.content_area, text="İlaç Kartları (Ana Veritabanı)", font=FONT_HEAD, 
                 bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(anchor="w", pady=(0, 10))

        # Arama Paneli
        fr_ara = tk.Frame(self.content_area, bg=c.get_color("card_bg"), padx=10, pady=10)
        fr_ara.pack(fill="x", pady=5)
        
        tk.Label(fr_ara, text="🔍 Kart Ara:", bg=c.get_color("card_bg"), fg=c.get_color("fg_text"), font=FONT_BOLD).pack(side="left")
        self.ent_ara = tk.Entry(fr_ara, font=FONT_NORM, bg=c.get_color("input_bg"), fg=c.get_color("fg_text"), insertbackground=c.get_color("fg_text"))
        self.ent_ara.pack(side="left", padx=10, fill="x", expand=True, ipady=3)
        self.ent_ara.bind("<KeyRelease>", self.kart_ara)

        # Üst Butonlar
        fr_btn = tk.Frame(self.content_area, bg=c.get_color("bg_main"), padx=10)
        fr_btn.pack(fill="x", pady=5)
        
        # Standart Buton Boyutları
        BTN_W = 155
        BTN_H = 32
        
        # 1. Kopyala Butonu
        ModernButton(fr_btn, text="📋 KOPYALA", command=self.kart_kopyala, 
                     bg_color="#8b5cf6", width=BTN_W, height=BTN_H).pack(side="left", padx=(0, 5))

        # 2. Manuel Ekle Butonu
        ModernButton(fr_btn, text="➕ MANUEL EKLE", command=self.manuel_kart_ekle_dialog, 
                     bg_color=c.get_color("btn_success"), width=BTN_W, height=BTN_H).pack(side="left", padx=5)

        # --- TABLO YAPISI ---
        cols = ("NO", "TİK", "GTIN", "İLAÇ ADI", "SON GÜNCELLEME")
        widths = [35, 30, 115, 500, 120]
        stretches = [False, False, False, True, False]

        self.tree_kart = ttk.Treeview(self.content_area, columns=cols, show="headings", height=25)
        
        sc = ttk.Scrollbar(self.content_area, orient="vertical", command=self.tree_kart.yview)
        self.tree_kart.configure(yscroll=sc.set)
        sc.pack(side="right", fill="y")
        self.tree_kart.pack(fill="both", expand=True)

        for col, w, s in zip(cols, widths, stretches):
            if col == "TİK":
                self.tree_kart.heading(col, text="☐", command=self.kart_tumu_sec)
                self.tree_kart.column(col, width=w, anchor="center", stretch=False)
            else:
                self.tree_kart.heading(col, text=col, anchor="w")
                align = "center" if col in ["NO", "GTIN", "SON GÜNCELLEME"] else "w"
                self.tree_kart.column(col, width=w, anchor=align, stretch=s)

        self.tree_kart.bind('<ButtonRelease-1>', self.kart_satir_tikla)
        self.kartlari_listele()

    def manuel_kart_ekle_dialog(self):
        # Yeni bir pencere (Popup) oluştur
        top = tk.Toplevel(self.pencere)
        top.title("Manuel İlaç Tanımla")
        top.geometry("400x300")
        top.configure(bg=TM.get_color("card_bg"))
        
        # Ekranın ortasında açılması için
        x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 200
        y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 150
        top.geometry(f"+{x}+{y}")

        # Başlık
        tk.Label(top, text="Yeni İlaç Kartı Oluştur", font=("Segoe UI", 12, "bold"), 
                 bg=TM.get_color("card_bg"), fg=TM.get_color("fg_text")).pack(pady=(20, 15))

        # Giriş Alanları
        lbl_style = {"bg": TM.get_color("card_bg"), "fg": TM.get_color("fg_text"), "font": ("Segoe UI", 10)}
        entry_style = {"bg": TM.get_color("input_bg"), "fg": TM.get_color("fg_text"), "relief": "solid", "bd": 1, "font": ("Segoe UI", 11)}

        # 1. Barkod/GTIN
        tk.Label(top, text="Barkod / GTIN (13 veya 14 Hane):", **lbl_style).pack(anchor="w", padx=30)
        e_gtin = tk.Entry(top, **entry_style)
        e_gtin.pack(fill="x", padx=30, pady=(0, 10), ipady=3)
        e_gtin.focus_set()

        # 2. İlaç Adı
        tk.Label(top, text="İlaç Adı:", **lbl_style).pack(anchor="w", padx=30)
        e_ad = tk.Entry(top, **entry_style)
        e_ad.pack(fill="x", padx=30, pady=(0, 20), ipady=3)

        # Kaydetme Mantığı
        def kaydet():
            gtin = e_gtin.get().strip()
            ad = e_ad.get().strip()

            if not gtin or not ad:
                messagebox.showwarning("Eksik Bilgi", "Lütfen barkod ve ilaç adını giriniz.", parent=top)
                return
            
            # Otomatik 14 haneye tamamlama (İsteğe bağlı, standart için iyi olur)
            if len(gtin) < 14 and gtin.isdigit():
                gtin = gtin.zfill(14)

            try:
                # Veritabanına Ekle (Varsa güncelle, yoksa ekle)
                self.ilac_karti_ekle_guncelle(gtin, ad)
                
                messagebox.showinfo("Başarılı", f"{ad}\nKartlara eklendi.", parent=top)
                self.kartlari_listele() # Listeyi yenile
                top.destroy() # Pencereyi kapat
            except Exception as e:
                messagebox.showerror("Hata", str(e), parent=top)

        # Kaydet Butonu
        tk.Button(top, text="KAYDET", command=kaydet, bg=TM.get_color("btn_success"), 
                  fg="white", font=("Segoe UI", 10, "bold"), relief="flat", pady=5).pack(fill="x", padx=30, pady=10)

    def kartlari_listele(self, filtre=""):
        for i in self.tree_kart.get_children(): self.tree_kart.delete(i)
        q = "SELECT gtin, ad, guncelleme_tarihi FROM ilac_kartlari"
        if filtre: q += f" WHERE ad LIKE '%{filtre}%' OR gtin LIKE '%{filtre}%'"
        self.imlec.execute(q + " ORDER BY ad ASC")
        for idx, r in enumerate(self.imlec.fetchall(), 1): self.tree_kart.insert("", "end", values=(idx, "☐", r[0], r[1], r[2]))
    
    def kart_ara(self, e): self.kartlari_listele(self.ent_ara.get())
    
    def kart_satir_tikla(self, e):
        if self.tree_kart.identify_column(e.x) == "#2":
            i = self.tree_kart.identify_row(e.y); 
            if i: v = list(self.tree_kart.item(i, "values")); v[1] = "☐" if v[1]=="☑" else "☑"; self.tree_kart.item(i, values=v)
    
    def kart_tumu_sec(self): pass
    
    def kart_kopyala(self):
        s = [self.tree_kart.item(i, "values") for i in self.tree_kart.get_children() if self.tree_kart.item(i, "values")[1] == "☑"]
        if s: self.pencere.clipboard_clear(); self.pencere.clipboard_append("\n".join(["\t".join(map(str, r[2:])) for r in s])); messagebox.showinfo("OK","Kopyalandı")

    def arayuz_karekod_olusturucu(self):
        c = TM
        # Başlık
        tk.Label(self.content_area, text="Karekod Üretici", font=FONT_HEAD, 
                 bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="top", anchor="w", pady=(0, 20))

        # Ana Panel
        f = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        f.pack(fill="both", expand=True)

        # --- ÜST BUTONLAR ---
        f_top = tk.Frame(f, bg=c.get_color("bg_main"))
        f_top.pack(fill="x", pady=(0, 10))

        ModernButton(f_top, text="📂 EXCEL LİSTESİ YÜKLE", command=self.karekod_dosya_yukle, 
                     bg_color=c.get_color("btn_warning"), width=250, height=35).pack(side="left", padx=(0, 10))

        # --- TABLO (TREEVIEW) ---
        # Sütunlar SKT Modülüyle eşitlendi
        cols = ("NO", "TIK", "GTIN", "SN", "BN", "İLAÇ ADI", "SKT", "KAREKOD")
        widths = [35, 30, 115, 110, 80, 350, 100, 250]
        
        self.tree_qr = ttk.Treeview(f, columns=cols, show="headings", height=20)
        
        # Scrollbar
        sc = ttk.Scrollbar(f, orient="vertical", command=self.tree_qr.yview)
        self.tree_qr.configure(yscroll=sc.set)
        sc.pack(side="right", fill="y")
        self.tree_qr.pack(fill="both", expand=True)

        # Başlık Ayarları
        for col, w in zip(cols, widths):
            if col == "TIK":
                self.tree_qr.heading(col, text="☐", command=self.qr_tumu_sec)
                self.tree_qr.column(col, width=w, anchor="center", stretch=False)
            else:
                self.tree_qr.heading(col, text=col, anchor="w")
                align = "center" if col in ["NO", "GTIN", "SKT"] else "w"
                # İlaç adı ve Karekod esnek olsun, diğerleri sabit
                stretch_col = True if col in ["İLAÇ ADI", "KAREKOD"] else False
                self.tree_qr.column(col, width=w, anchor=align, stretch=stretch_col)

        # Satır Tıklama (Tikleme) Olayı
        self.tree_qr.bind('<ButtonRelease-1>', self.qr_satir_tiklama)

        # --- ALT İŞLEM BUTONLARI ---
        f_bot = tk.Frame(f, bg=c.get_color("bg_main"), pady=15, padx=5)
        f_bot.pack(fill="x")

        ModernButton(f_bot, text="📋 KOPYALA", command=self.qr_secili_kopyala, bg_color="#3b82f6", width=160, height=35).pack(side="left", padx=5)
        ModernButton(f_bot, text="💾 STOĞA KAYDET", command=self.qr_secili_kaydet, bg_color=c.get_color("btn_success"), width=180, height=35).pack(side="left", padx=5)
        ModernButton(f_bot, text="📧 MAİL GÖNDER", command=lambda: self.mail_gonder_tetikle(True, tip="karekod"), bg_color="#8b5cf6", width=180, height=35).pack(side="right", padx=5)
        
        self.karekod_temp_data = []

    # --- YARDIMCI SEÇİM FONKSİYONLARI ---
    def qr_tumu_sec(self):
        # Basit tümünü seçme mantığı
        for item in self.tree_qr.get_children():
            vals = list(self.tree_qr.item(item, "values"))
            vals[1] = "☑" if vals[1] == "☐" else "☐" # Tersi yapma yerine direkt hepsini seçebiliriz ama bu da toggle yapar
            self.tree_qr.item(item, values=vals)

    def qr_satir_tiklama(self, e):
        if self.tree_qr.identify_column(e.x) == "#2": # TIK sütunu
            item = self.tree_qr.identify_row(e.y)
            if item:
                vals = list(self.tree_qr.item(item, "values"))
                vals[1] = "☐" if vals[1]=="☑" else "☑"
                self.tree_qr.item(item, values=vals)

    def karekod_dosya_yukle(self, otomatik_dosya_yolu=None): # <-- Parametre eklendi
        if otomatik_dosya_yolu:
            d = otomatik_dosya_yolu
        else:
            d = filedialog.askopenfilename(title="Karekod Üretilecek Excel Dosyasını Seçin", 
                                         filetypes=[("Excel Dosyaları", "*.xlsx"), ("CSV Dosyaları", "*.csv")])
        if not d: return

        try:
            # 1. Akıllı Dosya Okuma (CSV/Excel ayrımı ve temizlik yapar)
            df = self.evrensel_dosya_oku(d)
            if df.empty: return

            cols = df.columns
            
            # 2. Akıllı Sütun Eşleştirme (Senin başlıklarını tanır)
            c_gtin = self.find_column(cols, ['GTIN', 'BARKOD', 'EAN', 'URUN KODU'])
            c_sn = self.find_column(cols, ['SN', 'SERI', 'SERIAL', 'SIRA'])
            c_bn = self.find_column(cols, ['BN', 'PARTI', 'LOT', 'BATCH'])
            c_ad = self.find_column(cols, ['AD', 'ISIM', 'URUN', 'MALZEME', 'ILAC ADI'])
            c_sk = self.find_column(cols, ['SKT', 'S.K.T', 'MIAD', 'MİAD', 'TARIH', 'EXP', 'SON KUL'])

            # Eksik Sütun Kontrolü
            if not (c_gtin and c_sn and c_bn and c_sk):
                bulunanlar = [c for c in [c_gtin, c_sn, c_bn, c_sk] if c]
                messagebox.showerror("Eksik Sütun", f"Gerekli sütunlar bulunamadı.\nBulunanlar: {bulunanlar}\nArananlar: GTIN, SN, BN, SKT")
                return

            # Tabloyu Temizle
            for i in self.tree_qr.get_children(): self.tree_qr.delete(i)
            self.karekod_temp_data = []
            
            basarili_sayisi = 0

            # 3. Verileri İşle ve Temizle
            for idx, row in df.iterrows():
                try:
                    # GTIN Temizliği (;; ve .0 gibi hataları siler)
                    raw_gtin = str(row[c_gtin]).replace(".0", "").replace(";", "").strip()
                    if not raw_gtin or raw_gtin.lower() == "nan": continue
                    gtin = raw_gtin.zfill(14)
                    
                    # Diğer Verilerin Temizliği
                    sn = str(row[c_sn]).replace(".0", "").replace(";", "").strip()
                    bn = str(row[c_bn]).replace(".0", "").replace(";", "").strip()
                    ad = str(row[c_ad]).replace(";", "").strip() if c_ad else "Bilinmeyen İlaç"
                    
                    # Tarih (SKT) İşleme
                    raw_skt = str(row[c_sk]).replace(";", "").strip()
                    try:
                        # Pandas ile tarihi akıllıca tanı (2028-01-20 veya 20.01.2028 fark etmez)
                        ts = pd.to_datetime(raw_skt, dayfirst=True)
                        skt_formatted = ts.strftime('%y%m%d') # Karekod için (YYAAGG)
                        skt_display = ts.strftime('%d.%m.%Y') # Ekran için (GG.AA.YYYY)
                    except:
                        # Tarih bozuksa varsayılan
                        skt_formatted = "000000"
                        skt_display = raw_skt

                    # 4. Karekod Metnini Oluştur
                    # Format: 01 + GTIN + 21 + SN + 17 + SKT + 10 + BN
                    qr_text = f"01{gtin}21{sn}17{skt_formatted}10{bn}"

                    # Listeye Ekle
                    self.tree_qr.insert("", "end", values=(idx+1, "☐", gtin, sn, bn, ad, skt_display, qr_text))
                    
                    self.karekod_temp_data.append(qr_text)
                    basarili_sayisi += 1

                except Exception as e: continue
            
            if basarili_sayisi > 0:
                messagebox.showinfo("Tamamlandı", f"{basarili_sayisi} adet karekod başarıyla üretildi.")
            else:
                messagebox.showwarning("Uyarı", "Dosyada işlenecek geçerli veri bulunamadı.")

        except Exception as global_e:
            messagebox.showerror("Hata", f"Dosya okuma hatası:\n{str(global_e)}")

    def qr_secili_kopyala(self):
        # Tablodan ☑ işaretli olanları bul
        secili_karekodlar = []
        for item in self.tree_qr.get_children():
            vals = self.tree_qr.item(item, "values")
            # Vals sırası: NO(0), TIK(1), GTIN(2), SN(3), BN(4), AD(5), SKT(6), KAREKOD(7)
            if vals[1] == "☑":
                secili_karekodlar.append(vals[7]) # En sondaki Karekod sütunu
        
        if secili_karekodlar:
            text_to_copy = "\n".join(secili_karekodlar)
            self.pencere.clipboard_clear()
            self.pencere.clipboard_append(text_to_copy)
            messagebox.showinfo("Kopyalandı", f"{len(secili_karekodlar)} adet karekod panoya kopyalandı.")
        else:
            messagebox.showwarning("Uyarı", "Lütfen kopyalamak için listeden seçim yapınız (☑).") 
    
    # EKSİK OLAN FONKSİYON EKLENDİ
    def qr_secili_kaydet(self):
        # 1. Tablodan ☑ işaretli olan satırları bul
        secili_satirlar = []
        for item in self.tree_qr.get_children():
            vals = self.tree_qr.item(item, "values")
            # Sütun Sırası: NO(0), TIK(1), GTIN(2), SN(3), BN(4), AD(5), SKT(6), KAREKOD(7)
            if vals[1] == "☑":
                secili_satirlar.append(vals)
        
        # 2. Seçim Kontrolü
        if not secili_satirlar:
            messagebox.showwarning("Seçim Yok", "Lütfen stoğa kaydetmek istediğiniz karekodları listeden seçiniz (☑).")
            return

        # 3. Veritabanına Kayıt İşlemi
        kayit_sayisi = 0
        mevcut_sayisi = 0
        
        try:
            for v in secili_satirlar:
                gtin = v[2]
                sn = v[3]
                bn = v[4]
                ad = v[5]
                skt_display = v[6] # Ekranda dd.mm.yyyy formatında
                
                # Tarihi veritabanı formatına (yyyy-mm-dd) çevir
                try:
                    skt_db = datetime.strptime(skt_display, '%d.%m.%Y').strftime('%Y-%m-%d')
                except:
                    skt_db = date.today().strftime('%Y-%m-%d') # Hata olursa bugün
                
                # Mükerrer Kontrolü (Aynı karekod zaten var mı?)
                self.imlec.execute("SELECT id FROM ilaclar WHERE barkod=? AND seri_no=?", (gtin, sn))
                if not self.imlec.fetchone():
                    self.imlec.execute("""
                        INSERT INTO ilaclar (barkod, ad, parti_no, seri_no, adet, skt, raf_yeri, kayit_tarihi) 
                        VALUES (?, ?, ?, ?, 1, ?, '', ?)
                    """, (gtin, ad, bn, sn, skt_db, str(date.today())))
                    kayit_sayisi += 1
                else:
                    mevcut_sayisi += 1 # Zaten varsa atla
            
            self.baglanti_skt.commit()
            
            # 5. Sonuç Raporu
            sonuc_mesaji = f"✅ {kayit_sayisi} adet karekod başarıyla stoğa eklendi."
            if mevcut_sayisi > 0:
                sonuc_mesaji += f"\n⚠️ {mevcut_sayisi} adet karekod zaten stokta olduğu için tekrar eklenmedi."
                
            messagebox.showinfo("İşlem Tamamlandı", sonuc_mesaji)
                
        except Exception as e:
            messagebox.showerror("Hata", f"Kayıt sırasında hata oluştu:\n{str(e)}")

    def yedek_al(self):
        if not os.path.exists("Yedekler"): os.makedirs("Yedekler")
        shutil.copy2(self.db_adi, f"Yedekler/yedek_{datetime.now().strftime('%Y%m%d_%H%M')}.db")
        messagebox.showinfo("Yedek", "Yedek başarıyla alındı.")

    def mail_gonder_tetikle(self, manuel, tip="stok"):
        self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='gonderen_mail'"); g = self.imlec.fetchone()[0]
        self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='uygulama_sifresi'"); p = self.imlec.fetchone()[0]
        self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='alici_mail'"); a = self.imlec.fetchone()[0]
        if not g or not p or not a:
            if manuel: messagebox.showwarning("Hata", "Mail ayarları eksik."); return
        try:
            msg = MIMEMultipart(); msg['From'] = g; msg['To'] = a; msg['Subject'] = f"Rapor {tip}"; msg.attach(MIMEText("Ek.", "plain"))
            data = "Test verisi"
            part = MIMEBase('application', "octet-stream"); part.set_payload(data.encode('utf-8')); encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{tip}.txt"'); msg.attach(part)
            s = smtplib.SMTP('smtp.gmail.com', 587); s.starttls(); s.login(g, p); s.sendmail(g, a, msg.as_string()); s.quit()
            if manuel: messagebox.showinfo("Başarılı", "Mail gönderildi.")
        except Exception as e: messagebox.showerror("Hata", str(e))
    
    # =========================================================================
    # OTOMATİK KONTROL (AYDA 2 KERE MAİL SİSTEMİ) - THREAD SAFE (GÜVENLİ)
    # =========================================================================
    def otomatik_aylik_isler(self):
        """
        Program açılışında çalışır.
        Ayda 2 kere (1-15 arası ve 16-Sonu arası) otomatik rapor gönderir.
        Daha önce gönderdiyse tekrar göndermez.
        """
        import sqlite3
        from datetime import date
        
        # --- 1. ARKA PLAN İÇİN ÖZEL VERİTABANI KANALI AÇIYORUZ ---
        yerel_vt = sqlite3.connect(self.db_skt_adi, timeout=15)
        yerel_imlec = yerel_vt.cursor()
        
        bugun = date.today()
        donem_kodu = "1" if bugun.day <= 15 else "2"
        su_anki_donem_anahtari = f"{bugun.strftime('%Y-%m')}-{donem_kodu}"

        try:
            # self.imlec YERİNE ARTIK yerel_imlec KULLANIYORUZ
            yerel_imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='son_otomatik_mail_donemi'")
            res = yerel_imlec.fetchone()
            son_gonderilen_anahtar = res[0] if res else ""

            # --- EĞER BU DÖNEM HENÜZ GÖNDERİLMEDİYSE ---
            if su_anki_donem_anahtari != son_gonderilen_anahtar:
                print("Otomatik mail gönderimi tetiklendi...")
                
                # 1. Stok Hatırlatma Maili
                if donem_kodu == "1":
                    konu_hatirlatma = f"⚠️ Stok Güncelleme Hatırlatması - {bugun.strftime('%B %Y')}"
                    mesaj_hatirlatma = "Sayın Eczacı,\n\nYeni dönem başladı. Lütfen İTS üzerindeki güncel stok listenizi sisteme yükleyip analiz ediniz."
                    try: self.basit_mail_gonder(konu_hatirlatma, mesaj_hatirlatma)
                    except: pass

                # 2. Miad Raporu
                try:
                    self.otomatik_miat_raporu_gonder()
                except Exception as e:
                    print(f"Miat raporu maili hatası: {e}")
                
                # --- VERİTABANINA "GÖNDERİLDİ" DİYE İŞLE ---
                yerel_imlec.execute("INSERT OR REPLACE INTO ayarlar (anahtar, deger) VALUES (?, ?)", 
                                   ('son_otomatik_mail_donemi', su_anki_donem_anahtari))
                yerel_vt.commit() # self.baglanti_skt yerine yerel_vt kullanıyoruz
                print("Otomatik mail başarıyla gönderildi ve kaydedildi.")
                
            else:
                print("Bu dönem için zaten mail gönderilmiş. İşlem atlandı.")
                
        except Exception as e:
            print(f"Aylık kontrol sisteminde hata oluştu: {e}")
            
        finally:
            # İşlem bitince veya hata verse bile kanalı mutlaka güvenlice kapatıyoruz!
            yerel_vt.close()

    # =========================================================================
    # ÇOKLU BİLGİSAYAR AĞ BEKÇİSİ VE SENKRONİZASYON MONİTÖRÜ
    # =========================================================================
    def ag_ve_senkronizasyon_bekcisi(self):
        import os
        
        # Eğer henüz referans zamanları oluşturulmadıysa ilk kez oluştur
        if not hasattr(self, 'son_skt_guncelleme_zamani'):
            try: self.son_skt_guncelleme_zamani = os.path.getmtime(self.db_skt_adi)
            except: self.son_skt_guncelleme_zamani = 0
            
        if not hasattr(self, 'son_finans_guncelleme_zamani'):
            try: self.son_finans_guncelleme_zamani = os.path.getmtime(self.db_finans_adi)
            except: self.son_finans_guncelleme_zamani = 0

        # 1. BAĞLANTI KOPMA KONTROLÜ
        if not os.path.exists(self.db_skt_adi) or not os.path.exists(self.db_finans_adi):
            # Arka arkaya bildirim atmaması için kontrol
            if not getattr(self, 'baglanti_kopuk_uyarisi_verildi', False):
                self.goster_bildirim("⚠️ BAĞLANTI KOPTU", "Ana bilgisayara (Server) ulaşılamıyor!\nLütfen ağınızı veya ana bilgisayarı kontrol edin.")
                self.baglanti_kopuk_uyarisi_verildi = True
        else:
            # Bağlantı geri geldiyse bayrağı sıfırla
            self.baglanti_kopuk_uyarisi_verildi = False

            # 2. YENİ VERİ GİRİŞİ KONTROLÜ (Mofifiye Zamanı Değişti mi?)
            try:
                guncel_skt_zamani = os.path.getmtime(self.db_skt_adi)
                guncel_finans_zamani = os.path.getmtime(self.db_finans_adi)
                
                veri_degisti = False
                if guncel_skt_zamani > self.son_skt_guncelleme_zamani:
                    self.son_skt_guncelleme_zamani = guncel_skt_zamani
                    veri_degisti = True
                    
                if guncel_finans_zamani > self.son_finans_guncelleme_zamani:
                    self.son_finans_guncelleme_zamani = guncel_finans_zamani
                    veri_degisti = True
                    
                if veri_degisti:
                    self.goster_bildirim("🔄 Senkronizasyon", "Ağdaki başka bir kullanıcı sisteme yeni bir kayıt girdi/sildi. Görmek için bulunduğunuz sekmeyi yenileyin.")
                    
            except Exception as e:
                pass

        # Bekçiyi her 3 saniyede bir (3000 ms) arka planda yormadan tekrar çağır
        self.pencere.after(3000, self.ag_ve_senkronizasyon_bekcisi)

    def otomatik_miat_raporu_gonder(self):
        bugun = date.today()
        sinir_3 = (bugun + timedelta(days=90)).strftime('%Y-%m-%d')
        sinir_6 = (bugun + timedelta(days=180)).strftime('%Y-%m-%d')
        def veri_getir(tarih_siniri):
            self.imlec.execute("""SELECT ad, skt, COUNT(*) as adet FROM ilaclar WHERE skt <= ? AND skt > ? GROUP BY ad, skt ORDER BY skt ASC""", (tarih_siniri, bugun.strftime('%Y-%m-%d')))
            return self.imlec.fetchall()
        rapor_3 = veri_getir(sinir_3)
        rapor_6 = veri_getir(sinir_6)
        html_icerik = "<h3>🚨 Kritik Miad Raporu (Aylık Otomatik)</h3>"
        for baslik, veriler in [("🔴 3 Aydan Az Kalanlar", rapor_3), ("🟠 3-6 Ay Arası Kalanlar", rapor_6)]:
            html_icerik += f"<h4>{baslik}</h4><table border='1' cellpadding='5' style='border-collapse:collapse;'><tr><th>İlaç Adı</th><th>SKT</th><th>Adet</th></tr>"
            for r in veriler: html_icerik += f"<tr><td>{r[0]}</td><td>{r[1]}</td><td>{r[2]} Kutu</td></tr>"
            if not veriler: html_icerik += "<tr><td colspan='3'>Bu kategoride ürün bulunmamaktadır.</td></tr>"
            html_icerik += "</table><br>"
        self.basit_mail_gonder(f"📋 Aylık Miad Takip Raporu - {bugun.strftime('%d.%m.%Y')}", html_icerik, is_html=True)

    def basit_mail_gonder(self, konu, icerik, is_html=False):
        self.imlec.execute("SELECT anahtar, deger FROM ayarlar"); ayar_dict = dict(self.imlec.fetchall())
        g = ayar_dict.get('gonderen_mail'); p = ayar_dict.get('uygulama_sifresi'); a = ayar_dict.get('alici_mail')
        if not g or not p or not a: return
        try:
            msg = MIMEMultipart()
            msg['From'], msg['To'], msg['Subject'] = g, a, konu
            msg.attach(MIMEText(icerik, 'html' if is_html else 'plain'))
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls(); server.login(g, p); server.send_message(msg)
                self.sisteme_bildirim_ekle(f"📧 E-Posta gönderildi: {konu}")
        except Exception as e: print(f"Otomatik mail hatası: {e}")

               
    
    def ozel_mail_gonder(self, tur):
        # 1. Mail Ayarlarını Kontrol Et
        self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='gonderen_mail'"); g = self.imlec.fetchone()[0]
        self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='uygulama_sifresi'"); p = self.imlec.fetchone()[0]
        self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='alici_mail'"); a = self.imlec.fetchone()[0]
        
        if not g or not p or not a:
            messagebox.showwarning("Eksik Ayar", "Lütfen önce Ayarlar menüsünden mail bilgilerini girin.")
            return

        # 2. Listeden Sadece SEÇİLİ (☑) Olanları Çek
        secili_veriler = []
        for i in self.tablo_stok.get_children():
            vals = self.tablo_stok.item(i, "values")
            if vals[1] == "☑": # TIK sütunu
                secili_veriler.append(vals)

        if not secili_veriler:
            messagebox.showwarning("Seçim Yapılmadı", "Lütfen listeden göndermek istediğiniz ilaçları tikleyin (☑).")
            return

        try:
            msg = MIMEMultipart()
            msg['From'] = g
            msg['To'] = a
            tarih = datetime.now().strftime("%d.%m.%Y %H:%M")
            
            # --- SENARYO A: KAREKOD LİSTESİ (TXT) ---
            if tur == "karekod":
                if not self.detayli_gorunum:
                    messagebox.showwarning("Mod Hatası", "Karekod göndermek için lütfen 'KAREKODLARI GÖSTER' moduna geçiniz.")
                    return

                msg['Subject'] = f"Karekod Listesi - {tarih}"
                txt_icerik = ""
                for satir in secili_veriler:
                    try: txt_icerik += satir[-1] + "\n" # En son sütun karekoddur
                    except: pass
                
                msg.attach(MIMEText(f"{len(secili_veriler)} adet karekod ektedir.", "plain"))
                part = MIMEBase('application', "octet-stream")
                part.set_payload(txt_icerik.encode('utf-8')); encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename="karekodlar_{datetime.now().strftime("%H%M")}.txt"')
                msg.attach(part)

            # --- SENARYO B: GRUPLANDIRILMIŞ TABLO (GTIN, BN, AD BAZLI) ---
            elif tur == "tablo":
                msg['Subject'] = f"Toplu Stok Transfer Listesi - {tarih}"
                
                # GRUPLAMA MANTIĞI (Aggregation)
                # Sözlük yapısı: { (GTIN, BN, AD, SKT) : Toplam_Adet }
                stok_havuzu = {}

                for vals in secili_veriler:
                    if self.detayli_gorunum:
                        # Detaylı Mod İndeksleri: GTIN=2, SN=3, BN=4, AD=5, SKT=6
                        # SN (Seri No) hariç alıyoruz
                        anahtar = (vals[2], vals[4], vals[5], vals[6]) # (GTIN, BN, AD, SKT)
                        adet = 1
                    else:
                        # Özet Mod İndeksleri: GTIN=2, BN=3, AD=4, SKT=5, ADET=6
                        anahtar = (vals[2], vals[3], vals[4], vals[5]) # (GTIN, BN, AD, SKT)
                        try:
                            # "5 Adet" stringinden 5'i al
                            adet = int(vals[6].split()[0])
                        except: adet = 1
                    
                    # Havuza Ekle
                    if anahtar in stok_havuzu:
                        stok_havuzu[anahtar] += adet
                    else:
                        stok_havuzu[anahtar] = adet

                # HTML Oluştur
                html = """
                <html>
                <head>
                <style>
                    table {border-collapse: collapse; width: 100%; font-family: Arial, sans-serif;}
                    th, td {border: 1px solid #dddddd; text-align: left; padding: 8px;}
                    th {background-color: #3b82f6; color: white;}
                    tr:nth-child(even) {background-color: #f9f9f9;}
                </style>
                </head>
                <body>
                <h3>İlaç Teslim Listesi</h3>
                <p>Aşağıdaki ürünler GTIN ve Parti Numarasına göre gruplandırılmıştır.</p>
                <table>
                  <tr>
                    <th>İlaç Adı</th>
                    <th>GTIN</th>
                    <th>Parti No (BN)</th>
                    <th>SKT</th>
                    <th>Toplam Adet</th>
                  </tr>
                """
                
                # Havuzdaki verileri tabloya dök
                for (gtin, bn, ad, skt), toplam_adet in stok_havuzu.items():
                    html += f"<tr><td>{ad}</td><td>{gtin}</td><td>{bn}</td><td>{skt}</td><td style='font-weight:bold;'>{toplam_adet}</td></tr>"
                
                html += "</table><br><p>Eczane Asistanı Sistemi</p></body></html>"
                msg.attach(MIMEText(html, 'html'))

            # Gönderim
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls(); server.login(g, p); server.send_message(msg); server.quit()
            
            self.sisteme_bildirim_ekle(f"📧 Rapor Maili İletildi: {msg['Subject']}")
            messagebox.showinfo("Başarılı", "Mail başarıyla gönderildi.")

        except Exception as e:
            messagebox.showerror("Hata", f"Mail gönderilemedi:\n{str(e)}")


        # --- BURADAN İTİBAREN YAPIŞTIR ---
    def ilac_karti_ekle_guncelle(self, gtin, ad):
        """İlaç kartları tablosuna yeni ilaç ekler veya adını günceller."""
        bugun = datetime.now().strftime("%d.%m.%Y %H:%M")
        try:
            self.imlec.execute("""
                INSERT INTO ilac_kartlari (gtin, ad, guncelleme_tarihi) 
                VALUES (?, ?, ?)
                ON CONFLICT(gtin) DO UPDATE SET 
                ad = excluded.ad, 
                guncelleme_tarihi = excluded.guncelleme_tarihi
            """, (gtin, ad, bugun))
            self.baglanti_skt.commit()
        except Exception as e:
            print(f"Kart güncelleme hatası: {e}")

    # =========================================================================
    # 7. TOPLU İŞLEM VE KARŞILAŞTIRMA MODÜLÜ (KARŞILAŞTIRMA ÖZELLİĞİ EKLENDİ)
    # =========================================================================
    def arayuz_toplu_karekod(self):
        c = TM
        for w in self.content_area.winfo_children(): w.destroy()

        # --- ÜST BAŞLIK VE BUTON ---
        header = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        header.pack(fill="x", pady=(0, 10))
        tk.Label(header, text="🚀 Toplu Karekod Yönetimi", font=FONT_HEAD, 
                 bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left")
        
        # ✨ YENİ EKLENEN KARŞILAŞTIR BUTONU ✨
        ModernButton(header, text="⚖️ LİSTELERİ KARŞILAŞTIR", command=self.karekodlari_karsilastir, 
                     bg_color="#9b59b6", width=250, height=40).pack(side="right", padx=10)

        # --- ANA YATAY PANEL ---
        paned_main = tk.PanedWindow(self.content_area, orient=tk.HORIZONTAL, sashwidth=8, bg=c.get_color("border"))
        paned_main.pack(fill="both", expand=True)

        # SOL BLOK (ESKİ | YENİ)
        paned_left = tk.PanedWindow(paned_main, orient=tk.HORIZONTAL, sashwidth=8, bg=c.get_color("border"))
        paned_main.add(paned_left, stretch="always") 

        # 1. Eski
        self.frame_eski = tk.Frame(paned_left, bg=c.get_color("bg_main"))
        paned_left.add(self.frame_eski, width=650) 
        self.tree_eski = self._panel_olustur(self.frame_eski, "📄 ESKİ KAREKODLAR (Referans)", "eski")

        # 2. Yeni
        self.frame_yeni = tk.Frame(paned_left, bg=c.get_color("bg_main"))
        paned_left.add(self.frame_yeni, width=550) 
        self.tree_yeni = self._panel_olustur(self.frame_yeni, "✨ YENİ KAREKODLAR (Çalışma)", "yeni")

        # SAĞ BLOK (KAYITLI LİSTELER)
        frame_kayitlar = tk.Frame(paned_main, bg=c.get_color("card_bg"))
        paned_main.add(frame_kayitlar, width=280, stretch="never") 

        lbl_baslik = tk.Label(frame_kayitlar, text="💾 KAYITLI LİSTELER", font=("Segoe UI", 11, "bold"), 
                 bg="#2c3e50", fg="white", pady=8)
        lbl_baslik.pack(fill="x")
        
        self.lb_kayitlar = tk.Listbox(frame_kayitlar, bg=c.get_color("input_bg"), fg=c.get_color("fg_text"), 
                                      font=("Segoe UI", 10), bd=0, highlightthickness=0, selectmode="single")
        self.lb_kayitlar.pack(fill="both", expand=True, padx=5, pady=5)
        
        f_sag_btn = tk.Frame(frame_kayitlar, bg=c.get_color("card_bg"))
        f_sag_btn.pack(fill="x", pady=10, padx=10)

        def listeyi_yukle(hedef_tree):
            sel = self.lb_kayitlar.curselection()
            if not sel: return
            secilen_metin = self.lb_kayitlar.get(sel[0])
            
            if not hasattr(self, 'kayit_cache'): return
            kayit_id = self.kayit_cache.get(secilen_metin)
            
            self.imlec.execute("SELECT icerik FROM kayitli_listeler WHERE id=?", (kayit_id,))
            res = self.imlec.fetchone()
            if res:
                for i in hedef_tree.get_children(): hedef_tree.delete(i)
                icerik = res[0]
                satirlar = icerik.splitlines()
                yuklenen = 0
                for karekod in satirlar:
                    if not karekod.strip(): continue
                    gtin = karekod[2:16] if karekod.startswith("01") else karekod[:14]
                    self.imlec.execute("SELECT ad FROM ilac_kartlari WHERE gtin=?", (gtin,))
                    r_ad = self.imlec.fetchone()
                    ad = r_ad[0] if r_ad else "Bilinmeyen İlaç"
                    
                    hedef_tree.insert("", "end", values=("☐", karekod, ad))
                    yuklenen += 1
                messagebox.showinfo("Yüklendi", f"'{secilen_metin}' listesi ({yuklenen} satır) yüklendi.")

        def listeyi_sil():
            sel = self.lb_kayitlar.curselection()
            if not sel: return
            secilen_metin = self.lb_kayitlar.get(sel[0])
            kayit_id = self.kayit_cache.get(secilen_metin)
            if messagebox.askyesno("Sil", "Bu liste silinsin mi?"):
                self.imlec.execute("DELETE FROM kayitli_listeler WHERE id=?", (kayit_id,))
                self.baglanti_skt.commit()
                self.kayitlari_tazele()

        # Standart Boyutlar
        BTN_W = 230
        BTN_H = 35

        ModernButton(f_sag_btn, text="⬅️ ESKİ'YE YÜKLE", command=lambda: listeyi_yukle(self.tree_eski), bg_color="#64748b", width=BTN_W, height=BTN_H).pack(pady=5)
        ModernButton(f_sag_btn, text="⬅️ YENİ'YE YÜKLE", command=lambda: listeyi_yukle(self.tree_yeni), bg_color="#f59e0b", width=BTN_W, height=BTN_H).pack(pady=5)
        tk.Frame(f_sag_btn, height=10, bg=c.get_color("card_bg")).pack()
        ModernButton(f_sag_btn, text="🗑️ LİSTEYİ SİL", command=listeyi_sil, bg_color="#ef4444", width=BTN_W, height=BTN_H).pack(pady=5)

        self.kayitlari_tazele() 

    # --- YENİ EKLENEN KARŞILAŞTIRMA BEYNİ ---
    def karekodlari_karsilastir(self):
        # 1. Eski listeyi sözlüğe al (Hızlı arama için)
        eski_karekodlar = {}
        for i in self.tree_eski.get_children():
            vals = self.tree_eski.item(i, "values")
            eski_karekodlar[vals[1]] = vals[2] # {karekod: ilaç_adı}
            
        # 2. Yeni listeyi sözlüğe al
        yeni_karekodlar = {}
        for i in self.tree_yeni.get_children():
            vals = self.tree_yeni.item(i, "values")
            yeni_karekodlar[vals[1]] = vals[2]

        if not eski_karekodlar and not yeni_karekodlar:
            messagebox.showwarning("Uyarı", "Karşılaştırma yapabilmek için tablolarda veri olmalıdır.")
            return

        # 3. Kümelerle (Set) matematiksel fark bulma
        eski_kume = set(eski_karekodlar.keys())
        yeni_kume = set(yeni_karekodlar.keys())

        eksikler = eski_kume - yeni_kume    # Eski'de olup Yeni'de olmayanlar
        fazlalar = yeni_kume - eski_kume    # Yeni'de olup Eski'de olmayanlar
        eslesenler = eski_kume & yeni_kume  # Her ikisinde de olanlar

        # 4. Sonuç Penceresi (Popup) Oluştur
        top = tk.Toplevel(self.pencere)
        top.title("⚖️ Akıllı Karşılaştırma Sonuçları")
        top.geometry("1000x650")
        top.configure(bg=TM.get_color("bg_main"))
        
        # Özet Bilgi Çubuğu
        f_ozet = tk.Frame(top, bg="#2c3e50", pady=15)
        f_ozet.pack(fill="x")
        
        ozet_metin = f"Eski Liste: {len(eski_kume)} Karekod  |  Yeni Liste: {len(yeni_kume)} Karekod"
        tk.Label(f_ozet, text=ozet_metin, bg="#2c3e50", fg="white", font=("Segoe UI", 12, "bold")).pack()
        
        # Sekme Alanı
        nb = ttk.Notebook(top)
        nb.pack(fill="both", expand=True, padx=15, pady=15)
        
        def sonuc_sekmesi_olustur(baslik, liste, renk, dict_referans):
            f = tk.Frame(nb, bg=TM.get_color("bg_main"))
            nb.add(f, text=f"{baslik} ({len(liste)})")
            
            tree = ttk.Treeview(f, columns=("DURUM", "KAREKOD", "İLAÇ ADI"), show="headings")
            sc = ttk.Scrollbar(f, orient="vertical", command=tree.yview)
            tree.configure(yscroll=sc.set)
            sc.pack(side="right", fill="y")
            tree.pack(fill="both", expand=True, pady=5)
            
            tree.heading("DURUM", text="DURUM"); tree.column("DURUM", width=120, anchor="center")
            tree.heading("KAREKOD", text="KAREKOD"); tree.column("KAREKOD", width=250)
            tree.heading("İLAÇ ADI", text="İLAÇ ADI"); tree.column("İLAÇ ADI", width=400)
            
            tree.tag_configure("satir_rengi", background=renk, foreground="black")
            
            durum_yazisi = baslik.split(" ")[1] # İkon sonrasındaki kelime (EKSİKLER vs)
            
            for kk in liste:
                ad = dict_referans.get(kk, "Bilinmeyen İlaç")
                tree.insert("", "end", values=(durum_yazisi, kk, ad), tags=("satir_rengi",))
            
            def listeyi_kopyala():
                metin = "\n".join(liste)
                if metin:
                    self.pencere.clipboard_clear()
                    self.pencere.clipboard_append(metin)
                    messagebox.showinfo("Başarılı", f"{len(liste)} adet karekod panoya kopyalandı.", parent=top)
            
            tk.Button(f, text="📋 BU LİSTEYİ KOPYALA", command=listeyi_kopyala, 
                      bg="#3b82f6", fg="white", font=("Segoe UI", 10, "bold"), relief="flat", pady=5).pack(fill="x")

        # Sekmeleri Verilerle Doldur
        # Eksiklerin adını Eski listeden, Fazlalıkların adını Yeni listeden alırız
        sonuc_sekmesi_olustur("❌ EKSİKLER (Okutulmayanlar)", eksikler, "#fee2e2", eski_karekodlar)
        sonuc_sekmesi_olustur("⚠️ FAZLALIKLAR (Fazla Okutulanlar)", fazlalar, "#fef3c7", yeni_karekodlar)
        sonuc_sekmesi_olustur("✅ EŞLEŞENLER (Sorunsuzlar)", eslesenler, "#d1fae5", eski_karekodlar)

    # -------------------------------------------------------------------------
    def kayitlari_tazele(self):
        self.lb_kayitlar.delete(0, tk.END)
        self.kayit_cache = {}
        self.imlec.execute("SELECT id, liste_adi, tarih FROM kayitli_listeler ORDER BY id DESC")
        for r in self.imlec.fetchall():
            tarih_kisa = r[2].split()[0] if r[2] else ""
            gorunum = f"{r[1]} ({tarih_kisa})"
            self.lb_kayitlar.insert(tk.END, gorunum)
            self.kayit_cache[gorunum] = r[0]

    def _panel_olustur(self, parent, baslik, tip):
        c = TM
        frame_border = tk.LabelFrame(parent, text=baslik, font=("Segoe UI", 10, "bold"), 
                                     bg=c.get_color("bg_main"), fg=c.get_color("fg_text"))
        frame_border.pack(fill="both", expand=True, padx=5, pady=5)

        f_txt = tk.Frame(frame_border, bg=c.get_color("bg_main"), pady=0)
        f_txt.pack(side="top", fill="x")
        
        txt_input = tk.Text(f_txt, height=2, font=("Consolas", 9), bg=c.get_color("input_bg"), 
                            fg=c.get_color("fg_text"), insertbackground=c.get_color("fg_text"))
        txt_input.pack(fill="x", padx=5)
        
        placeholder_text = "Karekod yapıştır (Ctrl+V)..."
        txt_input.insert("1.0", placeholder_text)
        def on_focus_in(event):
            if txt_input.get("1.0", tk.END).strip() == placeholder_text:
                txt_input.delete("1.0", tk.END)
        txt_input.bind("<FocusIn>", on_focus_in)

        f_btn = tk.Frame(frame_border, bg=c.get_color("bg_main"), pady=5)
        f_btn.pack(side="bottom", fill="x", padx=5)

        lbl_info = tk.Label(frame_border, text="Liste Boş", font=("Segoe UI", 9), bg=c.get_color("bg_main"), fg="gray")
        lbl_info.pack(side="bottom", anchor="w", padx=5)

        cols = ("TIK", "KAREKOD", "İLAÇ ADI")
        tree = ttk.Treeview(frame_border, columns=cols, show="headings", selectmode="none")
        
        sc = ttk.Scrollbar(frame_border, orient="vertical", command=tree.yview)
        tree.configure(yscroll=sc.set)
        sc.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        def tumunu_sec_toggle():
            durum = getattr(tree, "tumunu_sec_durum", False)
            yeni_durum = not durum
            setattr(tree, "tumunu_sec_durum", yeni_durum)
            ikon = "☑" if yeni_durum else "☐"
            tree.heading("TIK", text=ikon)
            for item in tree.get_children():
                vals = list(tree.item(item, "values"))
                vals[0] = ikon
                tree.item(item, values=vals)

        tree.heading("TIK", text="☐", command=tumunu_sec_toggle)
        tree.column("TIK", width=40, anchor="center", stretch=False)
        tree.heading("KAREKOD", text="KAREKOD"); tree.column("KAREKOD", width=150)
        tree.heading("İLAÇ ADI", text="İLAÇ ADI"); tree.column("İLAÇ ADI", width=250)

        def satir_tikla(event):
            region = tree.identify("region", event.x, event.y)
            if region == "cell":
                col = tree.identify_column(event.x)
                if col == "#1":
                    item = tree.identify_row(event.y)
                    if item:
                        vals = list(tree.item(item, "values"))
                        vals[0] = "☑" if vals[0] == "☐" else "☐"
                        tree.item(item, values=vals)
        tree.bind("<Button-1>", satir_tikla)

        # === AKILLI AYIKLAYICI ===
        def analiz_et(event=None):
            raw_data = txt_input.get("1.0", tk.END).strip()
            txt_input.delete("1.0", tk.END)
            if not raw_data or "yapıştır" in raw_data: return

            satirlar = [s.strip() for s in raw_data.splitlines() if s.strip()]
            for satir in satirlar:
                karekod = satir
                
                if "\t" in satir:
                    for p in satir.split("\t"):
                        if p.strip().startswith("01") and len(p.strip()) > 20:
                            karekod = p.strip()
                            break
                elif " " in satir:
                    for p in satir.split(" "):
                        if p.strip().startswith("01") and len(p.strip()) > 20:
                            karekod = p.strip()
                            break

                if len(karekod) < 10: continue
                
                gtin = karekod[2:16] if karekod.startswith("01") else karekod[:14]
                self.imlec.execute("SELECT ad FROM ilac_kartlari WHERE gtin=?", (gtin,))
                r = self.imlec.fetchone()
                ad = r[0] if r else "Bilinmeyen İlaç"
                
                tree.insert("", 0, values=("☐", karekod, ad))
            
            lbl_info.config(text=f"Toplam: {len(tree.get_children())} Satır")

        def on_paste(event): txt_input.after(50, analiz_et)
        txt_input.bind("<<Paste>>", on_paste)
        txt_input.bind("<Return>", analiz_et)

        def secili_kopyala():
            kopyalanacak = [tree.item(i, "values")[1] for i in tree.get_children() if tree.item(i, "values")[0] == "☑"]
            if kopyalanacak:
                self.pencere.clipboard_clear()
                self.pencere.clipboard_append("\n".join(kopyalanacak))
                messagebox.showinfo("Kopyalandı", f"{len(kopyalanacak)} karekod kopyalandı.")
            else: messagebox.showwarning("Uyarı", "Seçim yapınız.")

        def secili_sil():
            silinecek = [i for i in tree.get_children() if tree.item(i, "values")[0] == "☑"]
            if silinecek:
                for i in silinecek: tree.delete(i)
                lbl_info.config(text=f"Toplam: {len(tree.get_children())} Satır")
            else: messagebox.showwarning("Uyarı", "Seçim yapınız.")

        def listeyi_kaydet():
            items = tree.get_children()
            if not items: return
            ad = simpledialog.askstring("Kaydet", "Liste Adı:")
            if ad:
                icerik = "\n".join([tree.item(i, "values")[1] for i in items])
                self.imlec.execute("INSERT INTO kayitli_listeler (liste_adi, icerik, tarih) VALUES (?,?,?)", 
                                   (ad, icerik, str(datetime.now())))
                self.baglanti_skt.commit()
                self.kayitlari_tazele()
                messagebox.showinfo("Başarılı", "Liste sağ panele kaydedildi.")

        # Standart Boyutlar
        BTN_W = 160
        BTN_H = 45

        ModernButton(f_btn, text="📋 KOPYALA", command=secili_kopyala, bg_color="#3b82f6", width=BTN_W, height=BTN_H).pack(side="left", padx=4)
        ModernButton(f_btn, text="🗑️ SİL", command=secili_sil, bg_color="#ef4444", width=BTN_W, height=BTN_H).pack(side="left", padx=4)
        ModernButton(f_btn, text="💾 KAYDET", command=listeyi_kaydet, bg_color="#10b981", width=BTN_W, height=BTN_H).pack(side="right", padx=4)
        ModernButton(f_btn, text="🧹 TEMİZLE", command=lambda: [tree.delete(i) for i in tree.get_children()], bg_color="#64748b", width=BTN_W, height=BTN_H).pack(side="right", padx=4)

        return tree

    def kayitlari_tazele(self):
        self.lb_kayitlar.delete(0, tk.END)
        self.kayit_cache = {}
        self.imlec.execute("SELECT id, liste_adi, tarih FROM kayitli_listeler ORDER BY id DESC")
        for r in self.imlec.fetchall():
            tarih_kisa = r[2].split()[0] if r[2] else ""
            gorunum = f"{r[1]} ({tarih_kisa})"
            self.lb_kayitlar.insert(tk.END, gorunum)
            self.kayit_cache[gorunum] = r[0]

    def _panel_olustur(self, parent, baslik, tip):
        """
        Panel oluşturucu: 
        Sütun Sırası: TİK | KAREKOD | İLAÇ ADI
        Veri Yazma Sırası: ("☐", karekod, ad) -> DÜZELTİLDİ
        """
        c = TM
        
        # Panel Çerçevesi
        frame_border = tk.LabelFrame(parent, text=baslik, font=("Segoe UI", 10, "bold"), 
                                     bg=c.get_color("bg_main"), fg=c.get_color("fg_text"))
        frame_border.pack(fill="both", expand=True, padx=5, pady=5)

        # 1. ÜST: Text Giriş
        f_txt = tk.Frame(frame_border, bg=c.get_color("bg_main"), pady=0)
        f_txt.pack(side="top", fill="x")
        
        txt_input = tk.Text(f_txt, height=2, font=("Consolas", 9), bg=c.get_color("input_bg"), 
                            fg=c.get_color("fg_text"), insertbackground=c.get_color("fg_text"))
        txt_input.pack(fill="x", padx=5)
        
        placeholder_text = "Karekod yapıştır (Ctrl+V)..."
        txt_input.insert("1.0", placeholder_text)
        def on_focus_in(event):
            if txt_input.get("1.0", tk.END).strip() == placeholder_text:
                txt_input.delete("1.0", tk.END)
        txt_input.bind("<FocusIn>", on_focus_in)

        # 2. ALT: Butonlar (En altta sabit)
        f_btn = tk.Frame(frame_border, bg=c.get_color("bg_main"), pady=5)
        f_btn.pack(side="bottom", fill="x", padx=5)

        # 3. ALT: Bilgi Label
        lbl_info = tk.Label(frame_border, text="Liste Boş", font=("Segoe UI", 9), bg=c.get_color("bg_main"), fg="gray")
        lbl_info.pack(side="bottom", anchor="w", padx=5)

        # 4. ORTA: Tablo
        # Sütun Tanımları: TİK - KAREKOD - İLAÇ ADI
        cols = ("TIK", "KAREKOD", "İLAÇ ADI")
        tree = ttk.Treeview(frame_border, columns=cols, show="headings", selectmode="none")
        
        sc = ttk.Scrollbar(frame_border, orient="vertical", command=tree.yview)
        tree.configure(yscroll=sc.set)
        sc.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        # Sütun Başlıkları ve Genişlikleri
        def tumunu_sec_toggle():
            durum = getattr(tree, "tumunu_sec_durum", False)
            yeni_durum = not durum
            setattr(tree, "tumunu_sec_durum", yeni_durum)
            ikon = "☑" if yeni_durum else "☐"
            tree.heading("TIK", text=ikon)
            for item in tree.get_children():
                vals = list(tree.item(item, "values"))
                vals[0] = ikon
                tree.item(item, values=vals)

        tree.heading("TIK", text="☐", command=tumunu_sec_toggle)
        tree.column("TIK", width=40, anchor="center", stretch=False)
        
        tree.heading("KAREKOD", text="KAREKOD"); tree.column("KAREKOD", width=150)
        tree.heading("İLAÇ ADI", text="İLAÇ ADI"); tree.column("İLAÇ ADI", width=250)

        # Tıklama ile Tikleme
        def satir_tikla(event):
            region = tree.identify("region", event.x, event.y)
            if region == "cell":
                col = tree.identify_column(event.x)
                if col == "#1":
                    item = tree.identify_row(event.y)
                    if item:
                        vals = list(tree.item(item, "values"))
                        vals[0] = "☑" if vals[0] == "☐" else "☐"
                        tree.item(item, values=vals)
        tree.bind("<Button-1>", satir_tikla)

        # --- FONKSİYONLAR ---
        def analiz_et(event=None):
            raw_data = txt_input.get("1.0", tk.END).strip()
            txt_input.delete("1.0", tk.END)
            if not raw_data or "yapıştır" in raw_data: return

            satirlar = [s.strip() for s in raw_data.splitlines() if s.strip()]
            for satir in satirlar:
                
                # --- YENİ EKLENEN AKILLI AYIKLAMA KISMI ---
                karekod = satir
                # Eğer Excel'den vs. kopyalandıysa arada TAB (\t) veya BOŞLUK olabilir
                if "\t" in satir:
                    parcalar = satir.split("\t")
                    # İçindeki parçalardan "01" ile başlayan ve karekod uzunluğunda olanı bul
                    for p in parcalar:
                        if p.strip().startswith("01") and len(p.strip()) > 20:
                            karekod = p.strip()
                            break
                elif " " in satir:
                    parcalar = satir.split(" ")
                    for p in parcalar:
                        if p.strip().startswith("01") and len(p.strip()) > 20:
                            karekod = p.strip()
                            break
                # ------------------------------------------

                if len(karekod) < 10: continue
                
                # Saf karekoddan GTIN'i çek
                gtin = karekod[2:16] if karekod.startswith("01") else karekod[:14]
                
                # İlaç adını veritabanından sorgula
                self.imlec.execute("SELECT ad FROM ilac_kartlari WHERE gtin=?", (gtin,))
                r = self.imlec.fetchone()
                ad = r[0] if r else "Bilinmeyen İlaç"
                
                # Veriyi tabloya TİK - KAREKOD - AD sırasıyla ekle
                tree.insert("", 0, values=("☐", karekod, ad))
            
            lbl_info.config(text=f"Toplam: {len(tree.get_children())} Satır")

        def on_paste(event): txt_input.after(50, analiz_et)
        txt_input.bind("<<Paste>>", on_paste)
        txt_input.bind("<Return>", analiz_et)

        def secili_kopyala():
            # Karekod 2. sütunda (indeks 1)
            kopyalanacak = [tree.item(i, "values")[1] for i in tree.get_children() if tree.item(i, "values")[0] == "☑"]
            if kopyalanacak:
                self.pencere.clipboard_clear()
                self.pencere.clipboard_append("\n".join(kopyalanacak))
                messagebox.showinfo("Kopyalandı", f"{len(kopyalanacak)} karekod kopyalandı.")
            else: messagebox.showwarning("Uyarı", "Seçim yapınız.")

        def secili_sil():
            silinecek = [i for i in tree.get_children() if tree.item(i, "values")[0] == "☑"]
            if silinecek:
                for i in silinecek: tree.delete(i)
                lbl_info.config(text=f"Toplam: {len(tree.get_children())} Satır")
            else: messagebox.showwarning("Uyarı", "Seçim yapınız.")

        def listeyi_kaydet():
            items = tree.get_children()
            if not items: return
            ad = simpledialog.askstring("Kaydet", "Liste Adı:")
            if ad:
                # Karekod 2. sütunda (indeks 1)
                icerik = "\n".join([tree.item(i, "values")[1] for i in items])
                self.imlec.execute("INSERT INTO kayitli_listeler (liste_adi, icerik, tarih) VALUES (?,?,?)", 
                                   (ad, icerik, str(datetime.now())))
                self.baglanti_skt.commit()
                self.kayitlari_tazele()
                messagebox.showinfo("Başarılı", "Liste sağ panele kaydedildi.")

        # --- ESNEK VE KESİLMEYEN BUTON YAPISI ---
        # fill="x" ve expand=True sayesinde butonlar alana eşit dağılır ve yazılar asla kesilmez.
        btn_style = {"font": ("Segoe UI", 9, "bold"), "fg": "white", "relief": "flat", "cursor": "hand2", "pady": 6}

        tk.Button(f_btn, text="📋 KOPYALA", bg="#3b82f6", activebackground="#2563eb", activeforeground="white", 
                  command=secili_kopyala, **btn_style).pack(side="left", fill="x", expand=True, padx=3)
                  
        tk.Button(f_btn, text="🗑️ SİL", bg="#ef4444", activebackground="#dc2626", activeforeground="white", 
                  command=secili_sil, **btn_style).pack(side="left", fill="x", expand=True, padx=3)
                  
        tk.Button(f_btn, text="🧹 TEMİZLE", bg="#64748b", activebackground="#475569", activeforeground="white", 
                  command=lambda: [tree.delete(i) for i in tree.get_children()], **btn_style).pack(side="left", fill="x", expand=True, padx=3)
                  
        tk.Button(f_btn, text="💾 KAYDET", bg="#10b981", activebackground="#059669", activeforeground="white", 
                  command=listeyi_kaydet, **btn_style).pack(side="left", fill="x", expand=True, padx=3)

        return tree

    def goster_bildirim(self, baslik, mesaj):
        """Ekranın sağ alt köşesinde şık bir bildirim kutusu gösterir."""
        toast = tk.Toplevel(self.pencere)
        toast.wm_overrideredirect(True) # Kenarlıkları kaldır
        toast.attributes("-topmost", True) # Her zaman en üstte dur
        toast.configure(bg="#2c3e50", highlightbackground="#34495e", highlightthickness=2) # İnce şık bir çerçeve
        
        # --- ÜST KISIM: Başlık ve Kapatma Butonu ---
        header = tk.Frame(toast, bg="#2c3e50")
        header.pack(fill="x", padx=15, pady=(15, 5))
        
        tk.Label(header, text=baslik, font=("Segoe UI", 11, "bold"), bg="#2c3e50", fg="#f1c40f").pack(side="left")
        
        # Kapatma (X) Tuşu
        btn_kapat = tk.Label(header, text="✖", font=("Segoe UI", 11, "bold"), bg="#2c3e50", fg="#ef4444", cursor="hand2")
        btn_kapat.pack(side="right")
        btn_kapat.bind("<Button-1>", lambda e: toast.destroy())
        
        # --- MESAJ KISMI (wraplength ile genişliği kısıtlayıp alt satıra atıyoruz) ---
        tk.Label(toast, text=mesaj, font=("Segoe UI", 10), bg="#2c3e50", fg="white", justify="left", wraplength=280).pack(padx=15, pady=(0, 20), anchor="w")
        
        # Ekranın sağ alt köşesini hesapla ve oraya yerleştir
        toast.update_idletasks()
        genislik = toast.winfo_width()
        yukseklik = toast.winfo_height()
        ekran_genislik = toast.winfo_screenwidth()
        ekran_yukseklik = toast.winfo_screenheight()
        
        x = ekran_genislik - genislik - 20
        y = ekran_yukseklik - yukseklik - 60 # Görev çubuğunun hemen üstü
        toast.geometry(f"+{x}+{y}")
        
        # 7 saniye (7000 ms) sonra kendi kendini yok et
        toast.after(7000, toast.destroy)

    def otomatik_kontrol_dongusu(self):
        """Programın arkada çalışmaya devam etmesini ve günlük bildirimler göndermesini sağlar"""
        bugun_str = date.today().strftime('%Y-%m-%d')
        
        # Veritabanından son bildirim tarihini çek (Uygulama kapanıp açılsa bile hatırlar)
        try:
            self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='son_gunluk_bildirim'")
            res = self.imlec.fetchone()
            kayitli_tarih = res[0] if res else ""
        except:
            kayitli_tarih = ""
            
        # Her gün sadece 1 kere bildirim verip rahatsız etmemesi için kontrol
        if kayitli_tarih != bugun_str:
            try:
                bugun = date.today()
                sinir_15_gun = (bugun + timedelta(days=15)).strftime('%Y-%m-%d')
                yarin_str = (bugun + timedelta(days=1)).strftime('%Y-%m-%d')
                
                # =========================================================
                # 1. MİAD (SKT) KONTROLÜ
                # =========================================================
                self.imlec.execute("SELECT COUNT(*) FROM ilaclar WHERE skt <= ?", (sinir_15_gun,))
                yakin_miad = self.imlec.fetchone()[0]
                
                if yakin_miad > 0:
                    self.goster_bildirim("📦 Kritik Stok", f"Dikkat! {yakin_miad} adet ürünün miadı dolmak üzere veya doldu.")
                    self.sisteme_bildirim_ekle(f"⚠️ {yakin_miad} adet ürünün miadı kritik seviyede!")

                # =========================================================
                # 2. KREDİ KARTI HESAP KESİM BİLDİRİMİ (Bugün)
                # =========================================================
                suanki_gun = bugun.day
                self.imlec_finans.execute("SELECT isim FROM kredi_kartlari WHERE kesim_gunu=?", (suanki_gun,))
                kesilen_kartlar = self.imlec_finans.fetchall()
                for (k_isim,) in kesilen_kartlar:
                    mesaj = f"Bugün '{k_isim}' kartınızın HESAP KESİM tarihidir."
                    self.goster_bildirim("✂️ Hesap Kesimi", mesaj)
                    self.sisteme_bildirim_ekle(f"💳 {mesaj}")

                # =========================================================
                # 3. KREDİ KARTI SON ÖDEME BİLDİRİMİ (Bugün ve Yarın) - GRUPLU MANTIK
                # =========================================================
                self.imlec_finans.execute("""
                    SELECT satir_notu, vade_tarihi, tutar FROM odemeler 
                    WHERE (aciklama LIKE '%Kredi Kartı%' OR satir_notu LIKE 'KART:%') 
                    AND durum='ODENMEDİ' AND (vade_tarihi=? OR vade_tarihi=?)
                """, (bugun_str, yarin_str))
                kart_odemeleri = self.imlec_finans.fetchall()
                
                # Kart isimlerine ve tarihe göre grupla (Aynı karta ait harcamaları topla)
                kart_gruplu = {}
                for (notu, v_tar, tutar) in kart_odemeleri:
                    kart_adi = "Kredi Kartı"
                    if notu and "KART:" in str(notu):
                        temp = str(notu).split("KART:")[1].strip()
                        kart_adi = temp.split("(")[0].strip() if "(" in temp else temp
                    
                    key = (kart_adi, v_tar)
                    kart_gruplu[key] = kart_gruplu.get(key, 0.0) + (tutar or 0.0)
                
                # Gruplanmış verileri tek bildirim olarak yolla
                for (k_adi, v_tar), top_tutar in kart_gruplu.items():
                    zaman = "BUGÜN" if v_tar == bugun_str else "YARIN"
                    
                    # Vade tarihinden yılı ve ayı çek (Örn: "2026-03-16" -> "2026-03")
                    v_tar_ay = v_tar[:7]
                    
                    # Veritabanında o aya ve o karta ait manuel girilmiş "Ekstre" değeri var mı kontrol et
                    guncel_tutar = top_tutar # Varsayılan: Sistemdeki Eczane Gideri (Taksitlerin toplamı)
                    try:
                        self.imlec_finans.execute("SELECT deger FROM program_ayarlari WHERE ayar_adi=?", (f"ekstre_toplam_{k_adi}_{v_tar_ay}",))
                        res_ekstre = self.imlec_finans.fetchone()
                        if res_ekstre and res_ekstre[0]:
                            ekstre_val = float(res_ekstre[0])
                            if ekstre_val > 0:
                                guncel_tutar = ekstre_val # Eğer ekstre girilmişse, ekstre tutarını kullan
                    except: pass

                    mesaj = f"'{k_adi}' kartınızın toplam ödemesi {zaman}!\nToplam Tutar: {guncel_tutar:,.2f} ₺"
                    self.goster_bildirim("🚨 Kart Ödemesi", mesaj)
                    self.sisteme_bildirim_ekle(f"💳 {mesaj}")

                # =========================================================
                # 4. KREDİ TAKSİT BİLDİRİMİ (Bugün ve Yarın)
                # =========================================================
                self.imlec_finans.execute("""
                    SELECT aciklama, vade_tarihi, tutar FROM odemeler 
                    WHERE aciklama LIKE 'KREDI:%' 
                    AND durum='ODENMEDİ' AND (vade_tarihi=? OR vade_tarihi=?)
                """, (bugun_str, yarin_str))
                kredi_odemeleri = self.imlec_finans.fetchall()
                
                for (ack, v_tar, tutar) in kredi_odemeleri:
                    # 'KREDI: İş Bankası İhtiyaç (1/12)' yazısından sadece banka adını ayıklıyoruz
                    kredi_adi = ack.split('(')[0].replace('KREDI:', '').strip()
                    zaman = "BUGÜN" if v_tar == bugun_str else "YARIN"
                    mesaj = f"'{kredi_adi}' taksit ödemeniz {zaman}!\nTutar: {tutar:,.2f} ₺"
                    self.goster_bildirim("🏦 Kredi Taksiti", mesaj)
                    self.sisteme_bildirim_ekle(f"🏦 {mesaj}")

                # Bugün bildirim atıldı olarak işaretle ve KALICI olarak veritabanına kaydet
                self.son_bildirim_tarihi = bugun_str
                self.imlec.execute("INSERT OR REPLACE INTO ayarlar (anahtar, deger) VALUES ('son_gunluk_bildirim', ?)", (bugun_str,))
                self.baglanti_skt.commit()
                
            except Exception as e:
                print(f"Bildirim Hatası: {e}")

        # Döngüyü 1 saatte bir (3600000 ms) tekrar çalışacak şekilde ayarla
        self.pencere.after(3600000, self.otomatik_kontrol_dongusu)

    # =========================================================================
    # 8. GERİ BİLDİRİM MODÜLÜ
    # =========================================================================
    def arayuz_geribildirim(self):
        c = TM
        for w in self.content_area.winfo_children(): w.destroy()

        # Başlık Alanı
        header = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        header.pack(fill="x", pady=(0, 10))
        tk.Label(header, text="📢 Geri Bildirim ve Destek", font=FONT_HEAD, 
                 bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(anchor="w")

        tk.Label(self.content_area, text="Uygulama ile ilgili karşılaştığınız hataları, eksiklikleri veya yeni özellik önerilerinizi doğrudan geliştiriciye iletebilirsiniz.", 
                 font=FONT_NORM, bg=c.get_color("bg_main"), fg=c.get_color("fg_text"), justify="left").pack(anchor="w", pady=(0, 20))

        # Tasarım Kartı
        card = tk.Frame(self.content_area, bg=c.get_color("card_bg"), padx=30, pady=30, highlightbackground=c.get_color("border"), highlightthickness=1)
        card.pack(fill="both", expand=True)

        lbl_style = {"bg": c.get_color("card_bg"), "fg": c.get_color("fg_text"), "font": ("Segoe UI", 11, "bold")}
        entry_style = {"font": ("Segoe UI", 11), "bg": c.get_color("input_bg"), "fg": c.get_color("fg_text"), "insertbackground": c.get_color("fg_text"), "relief": "solid", "bd": 1}
        
        # Konu
        tk.Label(card, text="Konu (Kısaca özetleyiniz):", **lbl_style).pack(anchor="w", pady=(0, 5))
        ent_konu = tk.Entry(card, **entry_style)
        ent_konu.pack(fill="x", ipady=5, pady=(0, 20))

        # Mesaj Alanı
        tk.Label(card, text="Mesajınız (Detaylı açıklama):", **lbl_style).pack(anchor="w", pady=(0, 5))
        txt_mesaj = tk.Text(card, **entry_style, height=12)
        txt_mesaj.pack(fill="both", expand=True, pady=(0, 20))

        # Mail Gönderme Fonksiyonu
        def mesaj_gonder():
            konu = ent_konu.get().strip()
            mesaj = txt_mesaj.get("1.0", tk.END).strip()

            if not konu or not mesaj:
                messagebox.showwarning("Eksik Bilgi", "Lütfen konu ve mesaj alanlarını doldurunuz.")
                return

            # Ayarlardan eczanenin mail giriş bilgilerini çekiyoruz
            self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='gonderen_mail'")
            res_g = self.imlec.fetchone()
            gonderen = res_g[0] if res_g else ""

            self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='uygulama_sifresi'")
            res_p = self.imlec.fetchone()
            sifre = res_p[0] if res_p else ""

            if not gonderen or not sifre:
                messagebox.showwarning("Ayarlar Eksik", "Mail gönderebilmek için önce 'Ayarlar' sekmesinden E-Posta ve Uygulama Şifresi bilgilerinizi kaydetmelisiniz.")
                return

            # ---------------------------------------------------------
            # BURAYA KENDİ MAİL ADRESİNİ YAZMALISIN
            hedef_mail = "knleneshalit@gmail.com"
            # ---------------------------------------------------------

            try:
                ModernButton_Gonder.itemconfig(ModernButton_Gonder.find_withtag("text"), text="⏳ GÖNDERİLİYOR...")
                self.pencere.update()

                msg = MIMEMultipart()
                msg['From'] = gonderen
                msg['To'] = hedef_mail
                msg['Subject'] = f"Uygulama Geri Bildirimi: {konu}"
                
                full_mesaj = f"Kullanıcı Maili: {gonderen}\nTarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\nMesaj:\n{mesaj}"
                msg.attach(MIMEText(full_mesaj, 'plain'))

                server = smtplib.SMTP('smtp.gmail.com', 587)
                server.starttls()
                server.login(gonderen, sifre)
                server.send_message(msg)
                server.quit()

                messagebox.showinfo("Başarılı", "Geri bildiriminiz başarıyla bize ulaştı. Teşekkür ederiz!")
                ent_konu.delete(0, tk.END)
                txt_mesaj.delete("1.0", tk.END)

            except Exception as e:
                messagebox.showerror("Hata", f"Mail gönderilirken bir hata oluştu:\n{str(e)}\n\nİnternet bağlantınızı veya Uygulama Şifrenizi kontrol ediniz.")
            finally:
                ModernButton_Gonder.itemconfig(ModernButton_Gonder.find_withtag("text"), text="📤 GERİ BİLDİRİMİ GÖNDER")

        # Buton
        ModernButton_Gonder = ModernButton(card, text="📤 GERİ BİLDİRİMİ GÖNDER", command=mesaj_gonder, 
                                           bg_color=c.get_color("btn_primary"), width=280, height=45)
        ModernButton_Gonder.pack(anchor="e")

    # =========================================================================
    # 9. KULLANIM KILAVUZU (SLAYT GÖSTERİSİ)
    # =========================================================================
    def arayuz_kullanim_kilavuzu(self):
        c = TM
        for w in self.content_area.winfo_children(): w.destroy()

        # Slayt Verileri (Buraya istediğin kadar adım ekleyebilirsin)
        self.slaytlar = [
            {
                "resim": "kilavuz_1.png",
                "baslik": "1. İTS Stok Yükleme ve Miad Takibi",
                "metin": "Öncelikle 'Son Kullanma Tarihi' sekmesinden İTS'den indirdiğiniz Excel veya CSV formatındaki stok dosyanızı sisteme yükleyin. Yüklenen stoklar otomatik olarak SKT'lerine göre sıralanır ve kritik ürünler renklendirilir."
            },
            {
                "resim": "kilavuz_2.png",
                "baslik": "2. Toplu İşlem ve Karşılaştırma",
                "metin": "Toplu İşlem sekmesinde eski ve yeni karekod listelerinizi yan yana yapıştırarak eksik, fazla ve eşleşen ürünleri saniyeler içinde tespit edip analiz edebilirsiniz."
            },
            {
                "resim": "kilavuz_3.png",
                "baslik": "3. Canlı Stok Sayımı",
                "metin": "Elinizdeki barkod okuyucu ile ürünleri okutarak canlı sayım listeleri oluşturabilir, sayım listelerini Excel'e aktarabilir veya kayıtlı listelerden çağırabilirsiniz."
            },
            {
                "resim": "kilavuz_4.png",
                "baslik": "4. E-Posta ile Raporlama",
                "metin": "Ayarlar sekmesinden E-Posta ayarlarınızı yaptıktan sonra, miadı yaklaşan veya sayımını yaptığınız ürünleri tek tıkla Excel tablosu gibi mail atabilirsiniz."
            },
            {
                # YENİ EKLENEN KAREKOD ÜRET SLAYTI
                "resim": "kilavuz_5.png",
                "baslik": "5. Karekod Üretici",
                "metin": "İçerisinde GTIN, Seri No, Parti No ve SKT (Tarih) sütunları bulunan Excel dosyalarınızı bu sekmeye yükleyerek saniyeler içinde binlerce standart İTS karekodu metni üretebilir, bu karekodları kopyalayabilir veya stoğunuza kaydedebilirsiniz."
            }
        ]
        self.aktif_slayt_index = 0

        # Başlık Alanı
        header = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        header.pack(fill="x", pady=(0, 10))
        tk.Label(header, text="📖 Uygulama Kullanım Kılavuzu", font=FONT_HEAD, 
                 bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(side="left")

        # Slayt Çerçevesi (Ortadaki büyük kart)
        self.slide_frame = tk.Frame(self.content_area, bg=c.get_color("card_bg"), highlightbackground=c.get_color("border"), highlightthickness=1)
        self.slide_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # Resim Alanı
        self.lbl_slide_img = tk.Label(self.slide_frame, bg=c.get_color("card_bg"))
        self.lbl_slide_img.pack(expand=True, pady=(20, 10))

        # Metin Alanı
        self.lbl_slide_title = tk.Label(self.slide_frame, text="", font=("Segoe UI", 16, "bold"), bg=c.get_color("card_bg"), fg=c.get_color("btn_primary"))
        self.lbl_slide_title.pack(pady=(0,10))

        self.lbl_slide_text = tk.Label(self.slide_frame, text="", font=("Segoe UI", 12), bg=c.get_color("card_bg"), fg=c.get_color("fg_text"), wraplength=850, justify="center")
        self.lbl_slide_text.pack(pady=(0, 20))

        # Alt Kontrol Butonları
        controls = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        controls.pack(fill="x", pady=(0, 10))

        self.btn_prev = ModernButton(controls, text="⬅ ÖNCEKİ ADIM", command=self.slayt_onceki, bg_color="#64748b", width=180, height=40)
        self.btn_prev.pack(side="left", padx=20)

        self.lbl_counter = tk.Label(controls, text="", font=("Segoe UI", 12, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text"))
        self.lbl_counter.pack(side="left", expand=True)

        self.btn_next = ModernButton(controls, text="SONRAKİ ADIM ➡", command=self.slayt_sonraki, bg_color="#3b82f6", width=180, height=40)
        self.btn_next.pack(side="right", padx=20)

        # İlk slaytı yükle
        self.slayt_guncelle()

    def slayt_guncelle(self):
        slayt = self.slaytlar[self.aktif_slayt_index]
        
        # Metinleri Güncelle
        self.lbl_slide_title.config(text=slayt["baslik"])
        self.lbl_slide_text.config(text=slayt["metin"])
        self.lbl_counter.config(text=f"Adım {self.aktif_slayt_index + 1} / {len(self.slaytlar)}")

        # Resmi Güncelle (Pillow Kütüphanesi ile)
        try:
            from PIL import Image, ImageTk
            if getattr(sys, 'frozen', False):
                base_dir = os.path.dirname(sys.executable)
            else:
                base_dir = os.path.dirname(os.path.abspath(__file__))
                
            img_yolu = os.path.join(base_dir, slayt["resim"])
            
            if os.path.exists(img_yolu):
                # Resim varsa küçült ve göster
                img = Image.open(img_yolu)
                img.thumbnail((800, 450), Image.Resampling.LANCZOS)
                self.guncel_slayt_resmi = ImageTk.PhotoImage(img)
                self.lbl_slide_img.config(image=self.guncel_slayt_resmi, text="", relief="flat")
            else:
                # Resim klasörde yoksa şık bir yer tutucu göster
                yer_tutucu = f"📷 Ekran Görüntüsü Bekleniyor\n\nLütfen uygulamanın bulunduğu klasöre\n'{slayt['resim']}'\nadında bir resim ekleyiniz."
                self.lbl_slide_img.config(image="", text=yer_tutucu, font=("Segoe UI", 13), fg="#94a3b8", width=60, height=12, bg="#e2e8f0", relief="sunken")
        except Exception as e:
            self.lbl_slide_img.config(image="", text="Görsel yüklenemedi.", font=("Segoe UI", 12), width=50, height=10)

        # Butonların Aktiflik Durumu
        if self.aktif_slayt_index == 0:
            self.btn_prev.update_color("#cbd5e1") # Pasif gri renk
        else:
            self.btn_prev.update_color("#64748b") # Aktif renk

        if self.aktif_slayt_index == len(self.slaytlar) - 1:
            self.btn_next.update_color("#cbd5e1")
            self.btn_next.itemconfig(self.btn_next.find_withtag("text"), text="BİTİR ✔")
        else:
            self.btn_next.update_color("#3b82f6")
            self.btn_next.itemconfig(self.btn_next.find_withtag("text"), text="SONRAKİ ADIM ➡")

    def slayt_sonraki(self):
        if self.aktif_slayt_index < len(self.slaytlar) - 1:
            self.aktif_slayt_index += 1
            self.slayt_guncelle()

    def slayt_onceki(self):
        if self.aktif_slayt_index > 0:
            self.aktif_slayt_index -= 1
            self.slayt_guncelle()  

    # =========================================================================
    # BİLDİRİM SİSTEMİ (ZİL VE PANEL - HATASIZ SENKRONİZASYON)
    # =========================================================================
    def sisteme_bildirim_ekle(self, mesaj, kategori="Finans"):
        """Sisteme yeni bildirim ekler ve kategorize eder."""
        tarih_str = datetime.now().strftime("%d.%m.%Y %H:%M")
        
        # Mesaj içeriğinden otomatik kategori belirleme (Akıllı Yönlendirme)
        mesaj_kucuk = mesaj.lower()
        if "stok" in mesaj_kucuk or "miad" in mesaj_kucuk or "📦" in mesaj or "⚠️" in mesaj:
            kategori = "Stok"

        try:
            # Kategori sütunlu yeni tablo yapısı
            self.imlec.execute("CREATE TABLE IF NOT EXISTS sistem_bildirimleri (id INTEGER PRIMARY KEY AUTOINCREMENT, mesaj TEXT, tarih TEXT, okundu INTEGER DEFAULT 0, kategori TEXT DEFAULT 'Finans')")
            try: self.imlec.execute("ALTER TABLE sistem_bildirimleri ADD COLUMN okundu INTEGER DEFAULT 0")
            except: pass
            try: self.imlec.execute("ALTER TABLE sistem_bildirimleri ADD COLUMN kategori TEXT DEFAULT 'Finans'")
            except: pass
            
            self.imlec.execute("INSERT INTO sistem_bildirimleri (mesaj, tarih, okundu, kategori) VALUES (?, ?, 0, ?)", (mesaj, tarih_str, kategori))
            self.baglanti_skt.commit()
            
            # Kapasiteyi iki sekme olduğu için 30'a çıkardık
            self.imlec.execute("SELECT COUNT(*) FROM sistem_bildirimleri")
            sayi = self.imlec.fetchone()[0]
            if sayi > 30:
                self.imlec.execute("DELETE FROM sistem_bildirimleri WHERE id IN (SELECT id FROM sistem_bildirimleri ORDER BY id ASC LIMIT ?)", (sayi - 30,))
                self.baglanti_skt.commit()
                
            self.zil_guncelle()
        except Exception as e: print("Bildirim ekleme hatası:", e)

    def bildirim_panelini_ac(self):
        c = TM
        
        # 1. Tıklanma anı (Görsel sıfırlama)
        if hasattr(self, 'btn_zil'):
            self.btn_zil.config(text="🔔", fg="white")
            self.btn_zil.update()
            
        # 2. Arka Planda Okundu İşaretleme
        try:
            self.imlec.execute("CREATE TABLE IF NOT EXISTS sistem_bildirimleri (id INTEGER PRIMARY KEY AUTOINCREMENT, mesaj TEXT, tarih TEXT, okundu INTEGER DEFAULT 0)")
            try: self.imlec.execute("ALTER TABLE sistem_bildirimleri ADD COLUMN okundu INTEGER DEFAULT 0")
            except: pass
            
            self.imlec.execute("UPDATE sistem_bildirimleri SET okundu=1 WHERE okundu=0")
            self.baglanti_skt.commit()
            if hasattr(self, 'zil_guncelle'): self.zil_guncelle()
        except: pass

        if hasattr(self, 'bildirim_win') and self.bildirim_win.winfo_exists():
            self.bildirim_win.destroy()
            return
            
        self.bildirim_win = tk.Toplevel(self.pencere)
        self.bildirim_win.overrideredirect(True) 
        self.bildirim_win.attributes("-topmost", True)
        self.bildirim_win.geometry("380x500")
        
        ana_kutu = tk.Frame(self.bildirim_win, bg=c.get_color("bg_main"), highlightbackground="#cbd5e1", highlightthickness=2)
        ana_kutu.pack(fill="both", expand=True)

        x = self.pencere.winfo_rootx() + self.pencere.winfo_width() - 390
        y = self.pencere.winfo_rooty() + 65
        self.bildirim_win.geometry(f"+{x}+{y}")

        self.bildirim_win.focus_force()
        def odak_kayboldu(event):
            odaklanan = self.pencere.focus_displayof()
            if odaklanan and str(odaklanan).startswith(str(self.bildirim_win)): return
            try: self.bildirim_win.destroy()
            except: pass
        self.bildirim_win.bind("<FocusOut>", odak_kayboldu)

        try:
            self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='son_stok_yukleme_tarihi'")
            res = self.imlec.fetchone()
            son_yukleme = res[0] if res else "Henüz stok yüklenmedi (Bilinmiyor)"
        except: son_yukleme = "Bilinmiyor"
        
        f_pin = tk.Frame(ana_kutu, bg="#dcfce7", padx=15, pady=12, highlightbackground="#22c55e", highlightthickness=2)
        f_pin.pack(fill="x", padx=10, pady=(10, 5))
        
        tk.Label(f_pin, text="📌 GÜNCEL İTS STOK LİSTESİ", font=("Segoe UI", 9, "bold"), bg="#dcfce7", fg="#166534", wraplength=330, justify="left").pack(anchor="w")
        tk.Label(f_pin, text=f"Son Yükleme: {son_yukleme}", font=("Segoe UI", 11, "bold"), bg="#dcfce7", fg="#14532d", wraplength=330, justify="left").pack(anchor="w", pady=(2,0))

        f_liste = tk.Frame(ana_kutu, bg=c.get_color("card_bg"))
        f_liste.pack(fill="both", expand=True, padx=10, pady=(5, 10))
        
        # --- SEKMELİ BİLDİRİM PANELİ ---
        nb_bildirim = ttk.Notebook(f_liste)
        nb_bildirim.pack(fill="both", expand=True)
        
        tab_finans = tk.Frame(nb_bildirim, bg=c.get_color("card_bg"))
        tab_stok = tk.Frame(nb_bildirim, bg=c.get_color("card_bg"))
        
        nb_bildirim.add(tab_finans, text="📊 Finans")
        nb_bildirim.add(tab_stok, text="📦 Stok")
        
        try:
            self.imlec.execute("ALTER TABLE sistem_bildirimleri ADD COLUMN kategori TEXT DEFAULT 'Finans'")
            self.baglanti_skt.commit()
        except: pass

        def sekme_olustur(parent_tab, kategori_adi):
            txt = tk.Text(parent_tab, font=("Segoe UI", 10), bg=c.get_color("input_bg"), fg=c.get_color("fg_text"), state="normal", wrap="word", bd=0, padx=10, pady=10)
            txt.pack(side="left", fill="both", expand=True)
            sc = ttk.Scrollbar(parent_tab, command=txt.yview)
            txt.configure(yscrollcommand=sc.set)
            sc.pack(side="right", fill="y")
            
            def icerigi_yukle():
                txt.config(state="normal")
                txt.delete("1.0", tk.END) # Listeyi temizle
                try:
                    # Bildirimlerin ID'sini de çekiyoruz ki hangisini sileceğimizi bilelim
                    self.imlec.execute("SELECT id, tarih, mesaj FROM sistem_bildirimleri WHERE kategori=? ORDER BY id DESC LIMIT 20", (kategori_adi,))
                    bildirimler = self.imlec.fetchall()
                    
                    if not bildirimler:
                        txt.insert(tk.END, f"🔔 {kategori_adi} alanında yeni bildiriminiz yok.\n")
                    else:
                        for b in bildirimler:
                            b_id = b[0]
                            tarih_metni = b[1] if b[1] else "Tarih Yok"
                            mesaj_metni = b[2] if b[2] else "Boş Mesaj"
                            
                            # Metnin içine tıklanabilir Çöp Kutusu (Label) yerleştiriyoruz
                            lbl_sil = tk.Label(txt, text="🗑️", fg="#ef4444", bg=c.get_color("input_bg"), cursor="hand2", font=("Segoe UI", 11))
                            lbl_sil.bind("<Button-1>", lambda e, bid=b_id: tekli_sil(bid))
                            
                            txt.window_create(tk.END, window=lbl_sil)
                            txt.insert(tk.END, f"  [{tarih_metni}]\n", "tarih")
                            txt.insert(tk.END, f"{mesaj_metni}\n", "mesaj")
                            
                            # Bildirimleri ayırmak için ince bir çizgi ekleyelim
                            txt.insert(tk.END, "----------------------------------------\n", "ayirac")
                            
                    txt.tag_config("tarih", foreground="#3b82f6", font=("Segoe UI", 8, "bold")) 
                    txt.tag_config("mesaj", foreground=c.get_color("fg_text")) 
                    txt.tag_config("ayirac", foreground="#cbd5e1")
                except Exception as e:
                    txt.insert(tk.END, f"⚠️ HATA OLUŞTU:\n{str(e)}\n")
                
                txt.config(state="disabled")

            def tekli_sil(bildirim_id):
                try:
                    self.imlec.execute("DELETE FROM sistem_bildirimleri WHERE id=?", (bildirim_id,))
                    self.baglanti_skt.commit()
                    self.zil_guncelle() # Zildeki sayıyı düşür
                    icerigi_yukle() # Ekranda tıklanan bildirimi anında yok et
                except Exception as e:
                    print("Bildirim silme hatası:", e)

            # İlk açılışta listeyi doldur
            icerigi_yukle()

        sekme_olustur(tab_finans, "Finans")
        sekme_olustur(tab_stok, "Stok")

        # --- YENİ EKLENEN: BİLDİRİMLERİ SİLME BUTONU ---
        f_temizle = tk.Frame(ana_kutu, bg=c.get_color("bg_main"))
        f_temizle.pack(fill="x", padx=10, pady=(0, 10))

        def bildirimleri_sil():
            cevap = messagebox.askyesno("Temizle", "Geçmiş tüm bildirim kayıtları (Stok ve Finans) kalıcı olarak silinecek.\nOnaylıyor musunuz?", parent=self.bildirim_win)
            if cevap:
                try:
                    # Veritabanından tüm bildirimleri sil
                    self.imlec.execute("DELETE FROM sistem_bildirimleri")
                    self.baglanti_skt.commit()
                    
                    # Zili sıfırla ve pencereyi kapat
                    self.zil_guncelle()
                    self.bildirim_win.destroy()
                    
                    # Kullanıcıya sağ alttan bilgi ver
                    self.goster_bildirim("Temizlendi", "Tüm bildirim geçmişi başarıyla silindi.")
                except Exception as e:
                    messagebox.showerror("Hata", f"Silinemedi: {str(e)}", parent=self.bildirim_win)

        ModernButton(f_temizle, text="🧹 TÜM BİLDİRİMLERİ TEMİZLE", command=bildirimleri_sil, bg_color="#ef4444", width=360, height=35).pack(pady=5)

    def zil_guncelle(self):
        """Zil üzerindeki okunmamış bildirim sayısını günceller."""
        if not hasattr(self, 'btn_zil'): return
        try:
            self.imlec.execute("CREATE TABLE IF NOT EXISTS sistem_bildirimleri (id INTEGER PRIMARY KEY AUTOINCREMENT, mesaj TEXT, tarih TEXT, okundu INTEGER DEFAULT 0)")
            try: self.imlec.execute("ALTER TABLE sistem_bildirimleri ADD COLUMN okundu INTEGER DEFAULT 0")
            except: pass
            
            # SADECE OKUNMAMIŞ OLANLARI SAY
            self.imlec.execute("SELECT COUNT(*) FROM sistem_bildirimleri WHERE okundu=0")
            sayi = self.imlec.fetchone()[0]
        except:
            sayi = 0
            
        if sayi > 0:
            self.btn_zil.config(text=f"🔔 ({sayi})", fg="#f1c40f")
        else:
            self.btn_zil.config(text="🔔", fg="white")

    # =========================================================================
    # GLOBAL ARAMA (SPOTLIGHT) MOTORU
    # =========================================================================
    def global_arama_calistir(self, event=None):
        query = self.ent_global_search.get().strip()
        placeholder = "İlaç, Depo, Fatura veya Menü Ara... (Enter'a bas)"
        if not query or query == placeholder: return

        pop = tk.Toplevel(self.pencere)
        pop.title(f"Arama Sonuçları: '{query}'")
        pop.geometry("800x500")
        pop.configure(bg=TM.get_color("bg_main"))
        pop.transient(self.pencere)
        pop.grab_set()

        # Ekranı ortala
        pop.geometry(f"+{self.pencere.winfo_x() + int(self.pencere.winfo_width()/2) - 400}+{self.pencere.winfo_y() + 100}")

        tk.Label(pop, text=f"'{query}' İçin Bulunan Sonuçlar", font=("Segoe UI", 14, "bold"), bg=TM.get_color("bg_main"), fg=TM.get_color("fg_text")).pack(pady=10)

        tree = ttk.Treeview(pop, columns=("KATEGORİ", "BULUNAN", "DETAY"), show="headings", height=15)
        tree.heading("KATEGORİ", text="Kategori")
        tree.heading("BULUNAN", text="Kayıt Adı")
        tree.heading("DETAY", text="Detay / Bilgi")
        
        tree.column("KATEGORİ", width=140, anchor="center")
        tree.column("BULUNAN", width=300)
        tree.column("DETAY", width=320)
        
        sc = ttk.Scrollbar(pop, orient="vertical", command=tree.yview)
        tree.configure(yscroll=sc.set)
        sc.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=10, pady=5)

        sonuc_var = False

        # 1. Stok Araması
        try:
            self.imlec.execute("SELECT ad, skt FROM ilaclar WHERE ad LIKE ? OR barkod LIKE ? LIMIT 20", (f"%{query}%", f"%{query}%"))
            for r in self.imlec.fetchall():
                tree.insert("", "end", values=("📦 Stok Kaydı", r[0], f"SKT: {r[1]}"))
                sonuc_var = True
        except: pass

        # 2. Depo Araması
        try:
            self.imlec_finans.execute("SELECT fatura_adi, SUM(tutar) FROM odemeler WHERE durum='ODENMEDİ' AND fatura_adi LIKE ? GROUP BY fatura_adi", (f"%{query}%",))
            for r in self.imlec_finans.fetchall():
                tree.insert("", "end", values=("🏢 Depo / Firma", r[0], f"Açık Borç: {r[1]:,.2f} ₺"))
                sonuc_var = True
        except: pass

        # 3. Kasa Araması
        try:
            self.imlec_finans.execute("SELECT kategori, tutar, tarih FROM kasa_defteri WHERE kategori LIKE ? OR aciklama LIKE ? LIMIT 20", (f"%{query}%", f"%{query}%"))
            for r in self.imlec_finans.fetchall():
                tree.insert("", "end", values=("📒 Kasa İşlemi", r[0], f"{r[2]} | {r[1]:,.2f} ₺"))
                sonuc_var = True
        except: pass

        if not sonuc_var:
            tree.insert("", "end", values=("Sonuç Yok", "Aradığınız kriterde kayıt bulunamadı.", "-"))

        def git_modul(e):
            sel = tree.selection()
            if not sel: return
            kat = tree.item(sel[0])["values"][0]
            pop.destroy()
            if "Stok" in kat: self.sekme_degistir(1)
            elif "Depo" in kat: self.sekme_degistir(6)
            elif "Kasa" in kat: self.sekme_degistir(11)

        tree.bind("<Double-1>", git_modul)
        tk.Label(pop, text="İlgili modüle gitmek için satıra çift tıklayın.", font=("Segoe UI", 9, "italic"), bg=TM.get_color("bg_main"), fg="#64748b").pack(pady=5)

    def arayuz_anasayfa(self):
        c = TM
        for w in self.content_area.winfo_children(): w.destroy()

        self.bakiye_gizli = getattr(self, 'bakiye_gizli', False)

        def toggle_bakiye():
            # Bakiye durumunu tersine çevir
            self.bakiye_gizli = not self.bakiye_gizli
            
            # 1. Butonun metnini belirle (Açık veya Kapalı göz)
            yeni_ikon = "🙈    ₺" if self.bakiye_gizli else "👁️₺"
            
            # 2. Buton varsa içindeki 'text' özelliğini değiştirip yeniden çizdiriyoruz
            if hasattr(self, 'btn_bakiye_ozel'):
                self.btn_bakiye_ozel.text = yeni_ikon
                self.btn_bakiye_ozel._draw(self.btn_bakiye_ozel.bg_color)
            
            # 3. Rakamları maskele veya göster
            guncelle_ozet_kartlari()


        # =====================================================================
        # 1. KAYDIRMA (SWIPE) ALTYAPISI
        # =====================================================================
        self.current_page = getattr(self, 'last_home_page', 0)
        self.is_dragging = False
        self.animating = False
        self.drag_start_x = 0

        viewport = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        viewport.pack(fill="both", expand=True)

        self.page_ozet = tk.Frame(viewport, bg=c.get_color("bg_main"))
        self.page_menu = tk.Frame(viewport, bg=c.get_color("bg_main"))

        if self.current_page == 0:
            self.page_ozet.place(relx=0, rely=0, relwidth=1, relheight=1)
            self.page_menu.place(relx=1, rely=0, relwidth=1, relheight=1)
        else:
            self.page_ozet.place(relx=-1, rely=0, relwidth=1, relheight=1)
            self.page_menu.place(relx=0, rely=0, relwidth=1, relheight=1)

        def goto_page(page_num, from_drag=False):
            if self.animating: return
            if page_num == self.current_page and not from_drag: return
            self.animating = True
            
            try: ozet_start = float(self.page_ozet.place_info()['relx'])
            except: ozet_start = 0.0 if self.current_page == 0 else -1.0
            try: menu_start = float(self.page_menu.place_info()['relx'])
            except: menu_start = 1.0 if self.current_page == 0 else 0.0

            ozet_target = 0.0 if page_num == 0 else -1.0
            menu_target = 1.0 if page_num == 0 else 0.0

            def step(count=0):
                if count > 20:
                    self.page_ozet.place(relx=ozet_target)
                    self.page_menu.place(relx=menu_target)
                    self.current_page = page_num
                    self.last_home_page = page_num
                    self.animating = False
                    guncelle_noktalar()
                    return
                progress = count / 20.0
                ease = 1 - (1 - progress) * (1 - progress)
                self.page_ozet.place(relx=ozet_start + (ozet_target - ozet_start) * ease)
                self.page_menu.place(relx=menu_start + (menu_target - menu_start) * ease)
                self.pencere.after(10, lambda: step(count + 1))
            step()

        def on_press(e):
            if isinstance(e.widget, (tk.Entry, tk.Text, ttk.Combobox, tk.Listbox)): return
            if self.animating: return
            self.drag_start_x = e.x_root
            self.is_dragging = False

        def on_motion(e):
            # --- GÜVENLİK KONTROLÜ BAŞLANGICI ---
            if not viewport.winfo_exists(): return
            # --- GÜVENLİK KONTROLÜ BİTİŞİ ---
            
            if isinstance(e.widget, (tk.Entry, tk.Text, ttk.Combobox, tk.Listbox)): return
            if self.animating: return
            dx = e.x_root - self.drag_start_x
            if abs(dx) > 15: self.is_dragging = True
            
            if self.is_dragging:
                w = viewport.winfo_width()
                shift = dx / w
                if self.current_page == 0:
                    if shift > 0: shift *= 0.2
                    self.page_ozet.place(relx=shift)
                    self.page_menu.place(relx=1 + shift)
                else:
                    if shift < 0: shift *= 0.2
                    self.page_ozet.place(relx=-1 + shift)
                    self.page_menu.place(relx=shift)

        def on_release(e):
            # --- GÜVENLİK KONTROLÜ: Ana sayfada değilsek bu işlemi iptal et ---
            if not viewport.winfo_exists(): return
            
            if isinstance(e.widget, (tk.Entry, tk.Text, ttk.Combobox, tk.Listbox)): return
            if self.animating or not self.is_dragging: return
            dx = e.x_root - self.drag_start_x
            w = viewport.winfo_width()
            if dx < -(w * 0.15) and self.current_page == 0: goto_page(1, from_drag=True)
            elif dx > (w * 0.15) and self.current_page == 1: goto_page(0, from_drag=True)
            else: goto_page(self.current_page, from_drag=True)

        self.pencere.bind_all("<ButtonPress-1>", on_press)
        self.pencere.bind_all("<B1-Motion>", on_motion)
        self.pencere.bind_all("<ButtonRelease-1>", on_release)

        # Tıklama sorunu çözüldü (is_dragging koruması butonlardan kaldırıldı)
        def wrap_cmd(idx):
            def cmd():
                self.pencere.unbind_all("<MouseWheel>")
                self.pencere.unbind_all("<Left>")
                self.pencere.unbind_all("<Right>")
                self.sekme_degistir(idx)
            return cmd

        # =====================================================================
        # 2. SAYFA 0: ÖZET EKRANI (ECZACI DEFTERİ | DASHBOARD)
        # =====================================================================
        header_ozet = tk.Frame(self.page_ozet, bg=c.get_color("bg_main"))
        header_ozet.pack(fill="x", pady=(25, 10), padx=40)
        
        # Logo ve Metinleri yan yana tutmak için çerçeve
        f_baslik_icerik = tk.Frame(header_ozet, bg=c.get_color("bg_main"))
        f_baslik_icerik.pack(side="left", fill="both")

        # Dashboard Logosu
        if hasattr(self, 'logo_dashboard'):
            lbl_dash_logo = tk.Label(f_baslik_icerik, image=self.logo_dashboard, bg=c.get_color("bg_main"))
            lbl_dash_logo.pack(side="left", padx=(0, 25))

        # Metinleri tutan çerçeve
        f_baslik_metin = tk.Frame(f_baslik_icerik, bg=c.get_color("bg_main"))
        f_baslik_metin.pack(side="left")

        # Marka yerine sayfanın amacı vurgulandı
        # Ekranda kimin yetkisiyle girdiğinizi parantez içinde görebileceksiniz
        tk.Label(f_baslik_metin, text=f"Genel Durum Özeti ({self.aktif_rol})", font=("Segoe UI", 34, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(anchor="w")
        tk.Label(f_baslik_metin, text="Eczanenizin güncel SKT, stok ve finansal özet durumu", font=("Segoe UI", 13), bg=c.get_color("bg_main"), fg="#94a3b8").pack(anchor="w", pady=(2,0))
        
        # Bakiye Gizle/Göster Butonu (Sadece yetkisi olanlar görebilir)
        ilk_ikon = "🙈₺" if self.bakiye_gizli else "👁️₺"
        if self.aktif_rol == "Yönetici" or self.personel_yetkileri.get("para_gorsun", False):
            self.btn_bakiye_ozel = ModernButton(header_ozet, text=ilk_ikon, command=toggle_bakiye, bg_color="#64748b", width=70, height=28)
            self.btn_bakiye_ozel.place(relx=1.0, rely=0, anchor="ne")

        # --- TÜM OLASI VERİLERİ HESAPLA (TAM LİSTE) ---
        bugun = date.today()
        bugun_str = bugun.strftime("%Y-%m-%d")
        bu_ay_str = bugun.strftime("%Y-%m")
        sinir_90_gun = (bugun + timedelta(days=90)).strftime('%Y-%m-%d')
        sinir_180_gun = (bugun + timedelta(days=180)).strftime('%Y-%m-%d')
        sinir_365_gun = (bugun + timedelta(days=365)).strftime('%Y-%m-%d') # <--- YENİ EKLENDİ
        
        if bugun.month == 12: gelecek_ay_str = f"{bugun.year + 1}-01"
        else: gelecek_ay_str = f"{bugun.year}-{bugun.month + 1:02d}"

        
        # ===== 1. BLOK: STOK HESAPLAMALARI =====
        try:
            self.imlec.execute("SELECT id FROM yuklemeler ORDER BY id DESC LIMIT 1")
            son_liste = self.imlec.fetchone()
            aktif_id = str(son_liste[0]) if son_liste else None
            
            if aktif_id:
                self.imlec.execute("SELECT COUNT(*) FROM ilaclar WHERE raf_yeri=? OR raf_yeri=''", (aktif_id,))
                toplam_stok = self.imlec.fetchone()[0]
                self.imlec.execute("SELECT COUNT(*) FROM ilaclar WHERE (raf_yeri=? OR raf_yeri='') AND skt <= ?", (aktif_id, sinir_90_gun))
                kritik_stok = self.imlec.fetchone()[0]
                self.imlec.execute("SELECT COUNT(*) FROM ilaclar WHERE (raf_yeri=? OR raf_yeri='') AND skt > ? AND skt <= ?", (aktif_id, sinir_90_gun, sinir_180_gun))
                yaklasan_stok = self.imlec.fetchone()[0]
                self.imlec.execute("SELECT COUNT(*) FROM ilaclar WHERE (raf_yeri=? OR raf_yeri='') AND skt > ? AND skt <= ?", (aktif_id, sinir_180_gun, sinir_365_gun))
                dikkat_stok = self.imlec.fetchone()[0]
            else:
                self.imlec.execute("SELECT COUNT(*) FROM ilaclar")
                toplam_stok = self.imlec.fetchone()[0]
                self.imlec.execute("SELECT COUNT(*) FROM ilaclar WHERE skt <= ?", (sinir_90_gun,))
                kritik_stok = self.imlec.fetchone()[0]
                self.imlec.execute("SELECT COUNT(*) FROM ilaclar WHERE skt > ? AND skt <= ?", (sinir_90_gun, sinir_180_gun))
                yaklasan_stok = self.imlec.fetchone()[0]
                self.imlec.execute("SELECT COUNT(*) FROM ilaclar WHERE skt > ? AND skt <= ?", (sinir_180_gun, sinir_365_gun))
                dikkat_stok = self.imlec.fetchone()[0]
        except:
            toplam_stok, kritik_stok, yaklasan_stok, dikkat_stok = 0, 0, 0, 0

        # ===== 2. BLOK: FİNANS HESAPLAMALARI =====
        try:
            self.imlec_finans.execute("SELECT SUM(tutar) FROM kasa_defteri WHERE islem_turu='GİRİŞ' AND tarih LIKE ?", (f"{bu_ay_str}%",))
            aylik_kasa_giris = self.imlec_finans.fetchone()[0] or 0.0

            self.imlec_finans.execute("SELECT SUM(tutar) FROM odemeler WHERE satir_notu='KURUM_GELIR' AND strftime('%Y-%m', vade_tarihi)=?", (bu_ay_str,))
            aylik_kurum_geliri = self.imlec_finans.fetchone()[0] or 0.0

            self.imlec_finans.execute("SELECT SUM(tutar) FROM odemeler WHERE satir_notu='MANUEL_GELIR' AND strftime('%Y-%m', vade_tarihi)=?", (bu_ay_str,))
            aylik_diger_gelir = self.imlec_finans.fetchone()[0] or 0.0

            toplam_gelir = aylik_kasa_giris + aylik_kurum_geliri + aylik_diger_gelir

            self.imlec_finans.execute("SELECT SUM(tutar) FROM kasa_defteri WHERE islem_turu='ÇIKIŞ' AND tarih LIKE ?", (f"{bu_ay_str}%",))
            aylik_kasa_cikis = self.imlec_finans.fetchone()[0] or 0.0

            # --- KART GİDERİ: Ekstre tabanlı, kart başına 1 kez ---
            imlec2 = self.baglanti_finans.cursor()
            imlec2.execute("""
                SELECT satir_notu, SUM(tutar) FROM odemeler 
                WHERE (aciklama LIKE '%Kredi Kartı%' OR satir_notu LIKE 'KART:%') 
                AND strftime('%Y-%m', vade_tarihi)=?
                GROUP BY satir_notu
            """, (bu_ay_str,))
            _kart_satirlar = imlec2.fetchall()

            # Önce kart adına göre sistem toplamlarını birleştir
            _kart_sistem = {}
            for (_notu, _tut) in _kart_satirlar:
                _k_adi = "Diğer Kartlar"
                if _notu and "KART:" in str(_notu):
                    _temp = str(_notu).split("KART:")[1].strip()
                    _k_adi = _temp.split("(")[0].rstrip(" -").strip()
                _kart_sistem[_k_adi] = _kart_sistem.get(_k_adi, 0.0) + (_tut or 0.0)

            # Her unique kart için ekstre kontrolü (1 kez!)
            aylik_kart_gideri = 0.0
            for _k_adi, _sistem_tut in _kart_sistem.items():
                imlec2.execute(
                    "SELECT deger FROM program_ayarlari WHERE ayar_adi=?",
                    (f"ekstre_toplam_{_k_adi}_{bu_ay_str}",)
                )
                _res = imlec2.fetchone()
                if _res and _res[0]:
                    try:
                        _ev = float(_res[0])
                        aylik_kart_gideri += _ev if _ev > 0 else _sistem_tut
                    except:
                        aylik_kart_gideri += _sistem_tut
                else:
                    aylik_kart_gideri += _sistem_tut
            # -------------------------------------------------------

            self.imlec_finans.execute("SELECT SUM(tutar) FROM odemeler WHERE strftime('%Y-%m', vade_tarihi) = ? AND (satir_notu IS NULL OR (NOT satir_notu LIKE '%KART:%' AND NOT satir_notu LIKE '%KURUM%' AND satir_notu NOT IN ('MANUEL_GIDER', 'MANUEL_GELIR', 'KURUM_GELIR'))) AND (aciklama IS NULL OR aciklama NOT LIKE '%[SENET-KK]%')", (bu_ay_str,))
            aylik_depo_havale = self.imlec_finans.fetchone()[0] or 0.0

            self.imlec_finans.execute("SELECT SUM(tutar) FROM odemeler WHERE aciklama LIKE 'KREDI:%' AND strftime('%Y-%m', vade_tarihi)=?", (bu_ay_str,))
            aylik_kredi_gideri = self.imlec_finans.fetchone()[0] or 0.0

            self.imlec_finans.execute("SELECT SUM(tutar) FROM odemeler WHERE satir_notu='MANUEL_GIDER' AND (aciklama IS NULL OR aciklama NOT LIKE 'KREDI:%') AND strftime('%Y-%m', vade_tarihi)=?", (bu_ay_str,))
            aylik_diger_gider = self.imlec_finans.fetchone()[0] or 0.0

            toplam_gider = aylik_kasa_cikis + aylik_kart_gideri + aylik_depo_havale + aylik_kredi_gideri + aylik_diger_gider
            bakiye = toplam_gelir - toplam_gider

            self.bakiye_detay_verisi = {
                "gelirler": {"Kasa Girişleri (Satış vb.)": aylik_kasa_giris, "Kurum Gelirleri (Bu Ay)": aylik_kurum_geliri, "Diğer Manuel Gelirler": aylik_diger_gelir},
                "giderler": {"Kredi Kartı Taksitleri": aylik_kart_gideri, "Depo (Havale/EFT/Çek)": aylik_depo_havale, "Kasa Çıkışları (Nakit)": aylik_kasa_cikis, "Banka Kredileri": aylik_kredi_gideri, "Sabit/Diğer Giderler": aylik_diger_gider},
                "toplam_gelir": toplam_gelir, "toplam_gider": toplam_gider, "net": bakiye, "donem": bu_ay_str
            }

            self.imlec_finans.execute("SELECT SUM(tutar) FROM odemeler WHERE durum='ODENMEDİ' AND satir_notu NOT IN ('MANUEL_GELIR', 'KURUM_GELIR') AND (aciklama IS NULL OR aciklama NOT LIKE '%[SENET-KK]%')")
            toplam_borc = self.imlec_finans.fetchone()[0] or 0.0

            self.imlec_finans.execute("SELECT SUM(tutar) FROM odemeler WHERE strftime('%Y-%m', vade_tarihi) = ? AND (satir_notu IS NULL OR (NOT satir_notu LIKE '%KART:%' AND NOT satir_notu LIKE '%KURUM%' AND satir_notu NOT IN ('MANUEL_GIDER', 'MANUEL_GELIR', 'KURUM_GELIR')))", (bu_ay_str,))
            bu_ay_depo = self.imlec_finans.fetchone()[0] or 0.0

            self.imlec_finans.execute("SELECT SUM(tutar) FROM odemeler WHERE durum='ODENMEDİ' AND satir_notu='KURUM_GELIR'")
            bekleyen_sgk = self.imlec_finans.fetchone()[0] or 0.0

            self.imlec_finans.execute("SELECT SUM(tutar) FROM kasa_defteri WHERE islem_turu='GİRİŞ' AND tarih=?", (bugun_str,))
            bugun_giris = self.imlec_finans.fetchone()[0] or 0.0

            self.imlec_finans.execute("SELECT SUM(tutar) FROM kasa_defteri WHERE islem_turu='ÇIKIŞ' AND tarih=?", (bugun_str,))
            bugun_cikis = self.imlec_finans.fetchone()[0] or 0.0

            self.imlec_finans.execute("SELECT SUM(tutar) FROM kasa_defteri WHERE islem_turu='GİRİŞ' AND tarih LIKE ? AND odeme_yontemi='KREDİ KARTI'", (f"{bu_ay_str}%",))
            bu_ay_kk_giris = self.imlec_finans.fetchone()[0] or 0.0

            self.imlec_finans.execute("SELECT SUM(tutar) FROM odemeler WHERE durum='ODENMEDİ' AND strftime('%Y-%m', vade_tarihi) = ? AND aciklama LIKE 'KREDI:%'", (bu_ay_str,))
            bu_ay_kredi = self.imlec_finans.fetchone()[0] or 0.0

            self.imlec_finans.execute("SELECT SUM(tutar) FROM odemeler WHERE durum='ODENMEDİ' AND strftime('%Y-%m', vade_tarihi) = ? AND satir_notu NOT IN ('MANUEL_GELIR', 'KURUM_GELIR') AND (aciklama IS NULL OR aciklama NOT LIKE '%[SENET-KK]%')", (gelecek_ay_str,))
            gelecek_ay_borc = self.imlec_finans.fetchone()[0] or 0.0

            aylik_giris = aylik_kasa_giris
            aylik_cikis = aylik_kasa_cikis
            bu_ay_kart = aylik_kart_gideri
            bu_ay_kredi = aylik_kredi_gideri

        except Exception as _ex:
            print("Dashboard finans hatası:", _ex)
            toplam_borc = bakiye = aylik_giris = aylik_cikis = bu_ay_depo = bu_ay_kart = bekleyen_sgk = 0.0
            bugun_giris = bugun_cikis = bu_ay_kk_giris = bu_ay_kredi = gelecek_ay_borc = 0.0
            aylik_kart_gideri = aylik_depo_havale = aylik_kredi_gideri = 0.0
            aylik_kasa_giris = aylik_kurum_geliri = aylik_diger_gelir = toplam_gelir = toplam_gider = 0.0
            self.bakiye_detay_verisi = {"gelirler": {}, "giderler": {}, "toplam_gelir": 0, "toplam_gider": 0, "net": 0, "donem": bu_ay_str}    
            

        # --- DİNAMİK KART VERİ SÖZLÜĞÜ (BU KISIM KUTULARI ÇİZER) ---
        kart_datalari = {
            "stok_toplam": {"baslik": "Toplam Stok (Kutu)", "deger": f"{toplam_stok:,}", "ikon": "📦", "bg": "#3b82f6", "is_money": False},
            "stok_kritik": {"baslik": "Kritik Miad (0-3 Ay)", "deger": f"{kritik_stok:,}", "ikon": "🚨", "bg": "#ef4444", "is_money": False},
            "stok_uyari": {"baslik": "Yaklaşan Miad (3-6 Ay)", "deger": f"{yaklasan_stok:,}", "ikon": "⚠️", "bg": "#f59e0b", "is_money": False},
            "stok_dikkat": {"baslik": "Dikkat Miad (6-12 Ay)", "deger": f"{dikkat_stok:,}", "ikon": "🟡", "bg": "#eab308", "is_money": False},
            "finans_borc": {"baslik": "Toplam Açık Borç", "deger": f"{toplam_borc:,.2f} ₺", "ikon": "🏢", "bg": "#f97316", "is_money": True},
            "finans_gelecek_ay": {"baslik": "Gelecek Ayki Borç", "deger": f"{gelecek_ay_borc:,.2f} ₺", "ikon": "🗓️", "bg": "#8b5cf6", "is_money": True},
            "finans_bakiye": {"baslik": "Aylık Net Bakiye", "deger": f"{bakiye:,.2f} ₺", "ikon": "💰", "bg": "#10b981" if bakiye>=0 else "#be123c", "is_money": True},
            "kasa_bugun_giris": {"baslik": "Bugünkü Kasa Girişi", "deger": f"{bugun_giris:,.2f} ₺", "ikon": "💵", "bg": "#0ea5e9", "is_money": True},
            "kasa_bugun_cikis": {"baslik": "Bugünkü Kasa Çıkışı", "deger": f"{bugun_cikis:,.2f} ₺", "ikon": "💸", "bg": "#f43f5e", "is_money": True},
            "finans_giris": {"baslik": "Bu Ay Kasa Girişi", "deger": f"{aylik_giris:,.2f} ₺", "ikon": "📈", "bg": "#059669", "is_money": True},
            "finans_cikis": {"baslik": "Bu Ay Kasa Çıkışı", "deger": f"{aylik_cikis:,.2f} ₺", "ikon": "📉", "bg": "#e11d48", "is_money": True},
            "kasa_kk_giris": {"baslik": "Bu Ay POS/Kart Satışı", "deger": f"{bu_ay_kk_giris:,.2f} ₺", "ikon": "💳", "bg": "#0284c7", "is_money": True},
            "finans_depo": {"baslik": "Bu Ayki Depo Ödemesi", "deger": f"{bu_ay_depo:,.2f} ₺", "ikon": "🚚", "bg": "#4f46e5", "is_money": True},
            "finans_kart": {"baslik": "Bu Ayki Kart Ödemesi", "deger": f"{bu_ay_kart:,.2f} ₺", "ikon": "💳", "bg": "#a855f7", "is_money": True},
            "finans_kredi": {"baslik": "Bu Ayki Kredi Taksitleri", "deger": f"{bu_ay_kredi:,.2f} ₺", "ikon": "🏦", "bg": "#6366f1", "is_money": True},
            "finans_sgk": {"baslik": "Bekleyen Kurum Geliri", "deger": f"{bekleyen_sgk:,.2f} ₺", "ikon": "🏥", "bg": "#2563eb", "is_money": True}
        }

        # Veritabanından Ayarları Çek
        try:
            self.imlec.execute("SELECT deger FROM ayarlar WHERE anahtar='dashboard_kard_ayarlari'")
            res = self.imlec.fetchone()
            secili_kartlar = eval(res[0]) if res and res[0] else ["stok_toplam", "stok_kritik", "finans_bakiye", "finans_borc"]
        except: secili_kartlar = ["stok_toplam", "stok_kritik", "finans_bakiye", "finans_borc"]

        # Yetki Kontrolü
        if str(self.aktif_rol).strip() != "Yönetici" and not self.personel_yetkileri.get("para_gorsun", False):
            secili_kartlar = [k for k in secili_kartlar if k in kart_datalari and not kart_datalari[k]["is_money"]]

        f_cards = tk.Frame(self.page_ozet, bg=c.get_color("bg_main"))
        f_cards.pack(fill="both", expand=True, padx=40, pady=10)
        f_cards.grid_columnconfigure(0, weight=1); f_cards.grid_columnconfigure(1, weight=1)

        kutu_sayisi = len(secili_kartlar)
        satir_sayisi = (kutu_sayisi + 1) // 2
        for i in range(satir_sayisi): f_cards.grid_rowconfigure(i, weight=1)

    # --- YENİ EKLENEN: DETAYLI BAKİYE POPUP FONKSİYONU ---
        def goster_bakiye_detayi(event=None):
            if str(self.aktif_rol).strip() != "Yönetici" and not self.personel_yetkileri.get("para_gorsun", False): return
            if not hasattr(self, 'bakiye_detay_verisi'): return
            
            d = self.bakiye_detay_verisi
            win = tk.Toplevel(self.pencere)
            win.title("Aylık Net Bakiye Özeti")
            win.geometry("550x850") 
            win.configure(bg="#f8fafc")
            win.transient(self.pencere)
            win.grab_set()
            
            x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 275
            y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 425
            win.geometry(f"+{x}+{y}")
            
            f_net = tk.Frame(win, bg="#1e293b", pady=15)
            f_net.pack(fill="x", side="bottom")
            net_renk = "#34d399" if d["net"] >= 0 else "#f87171"
            tk.Label(f_net, text="BU AYKİ NET BAKİYE", font=("Segoe UI", 12, "bold"), bg="#1e293b", fg="white").pack()
            tk.Label(f_net, text=f"{d['net']:,.2f} ₺", font=("Segoe UI", 24, "bold"), bg="#1e293b", fg=net_renk).pack()

            tk.Label(win, text="📊 AYLIK KASA VE FİNANS ÖZETİ", font=("Segoe UI", 14, "bold"), bg="#f8fafc", fg="#334155").pack(pady=(15, 5))
            tk.Label(win, text=f"Dönem: {d['donem']}", font=("Segoe UI", 11), bg="#f8fafc", fg="#64748b").pack(pady=(0, 10))
            
            f_main_container = tk.Frame(win, bg="#f8fafc")
            f_main_container.pack(fill="both", expand=True, padx=20)
            
            canvas_bakiye = tk.Canvas(f_main_container, bg="#f8fafc", highlightthickness=0)
            sc_bakiye = ttk.Scrollbar(f_main_container, orient="vertical", command=canvas_bakiye.yview)
            f_main = tk.Frame(canvas_bakiye, bg="#f8fafc")
            f_main.bind("<Configure>", lambda e: canvas_bakiye.configure(scrollregion=canvas_bakiye.bbox("all")))
            w_id = canvas_bakiye.create_window((0, 0), window=f_main, anchor="nw")
            canvas_bakiye.bind("<Configure>", lambda e: canvas_bakiye.itemconfig(w_id, width=e.width))
            canvas_bakiye.configure(yscrollcommand=sc_bakiye.set)
            canvas_bakiye.pack(side="left", fill="both", expand=True)
            sc_bakiye.pack(side="right", fill="y")
            self.mouse_scroll_ekle(canvas_bakiye, f_main)
            
            def liste_ciz(parent, baslik, veri_dict, toplam, renk):
                f_kutu = tk.Frame(parent, bg="white", bd=1, relief="solid", padx=15, pady=8)
                f_kutu.pack(fill="x", pady=6)
                tk.Label(f_kutu, text=baslik, font=("Segoe UI", 11, "bold"), bg="white", fg=renk).pack(anchor="w", pady=(0, 5))
                for ad, tutar in veri_dict.items():
                    r = tk.Frame(f_kutu, bg="white"); r.pack(fill="x", pady=2)
                    tk.Label(r, text=ad, font=("Segoe UI", 10), bg="white", fg="#475569").pack(side="left")
                    tk.Label(r, text=f"{tutar:,.2f} ₺", font=("Segoe UI", 10, "bold"), bg="white", fg="#0f172a").pack(side="right")
                tk.Frame(f_kutu, height=1, bg="#e2e8f0").pack(fill="x", pady=6)
                tk.Label(f_kutu, text=f"TOPLAM: {toplam:,.2f} ₺", font=("Segoe UI", 11, "bold"), bg="white", fg=renk).pack(anchor="e")

            liste_ciz(f_main, "GELİRLER (+)", d["gelirler"], d["toplam_gelir"], "#10b981")
            liste_ciz(f_main, "GİDERLER (-)", d["giderler"], d["toplam_gider"], "#ef4444")


        # --- YENİ EKLENEN: DİĞER KARTLAR İÇİN AKILLI DETAY POPUP'I ---
        def goster_kutu_detayi(event, secili_k_id):
            if str(self.aktif_rol).strip() != "Yönetici" and not self.personel_yetkileri.get("para_gorsun", False): return
            
            if secili_k_id == "finans_bakiye":
                goster_bakiye_detayi()
                return
                
            win = tk.Toplevel(self.pencere)
            baslik_metni = kart_datalari[secili_k_id]["baslik"]
            win.title(f"{baslik_metni} Detayı")
            win.geometry("600x750") 
            win.configure(bg="#f8fafc")
            win.transient(self.pencere)
            win.grab_set()
            
            x = self.pencere.winfo_x() + (self.pencere.winfo_width() // 2) - 300
            y = self.pencere.winfo_y() + (self.pencere.winfo_height() // 2) - 375
            win.geometry(f"+{x}+{y}")
            
            # 1. EN ALT TOPLAM KISMINI ÖNCE YAPIŞTIRIYORUZ (Kesilmemesi İçin)
            f_bot = tk.Frame(win, bg="#1e293b", pady=15)
            f_bot.pack(fill="x", side="bottom")
            tk.Label(f_bot, text=f"GENEL TOPLAM {baslik_metni.upper()}", font=("Segoe UI", 12, "bold"), bg="#1e293b", fg="white").pack()
            lbl_toplam_val = tk.Label(f_bot, text="0.00 ₺", font=("Segoe UI", 24, "bold"), bg="#1e293b", fg="#34d399")
            lbl_toplam_val.pack()

            # 2. ÜST BAŞLIK VE DÖNEM BELİRTECİ
            tk.Label(win, text=f"📊 {baslik_metni.upper()}", font=("Segoe UI", 14, "bold"), bg="#f8fafc", fg="#334155").pack(pady=(20, 2))
            
            # Hangi kutunun hangi döneme ait olduğunu tespit et
            if secili_k_id in ["finans_depo", "finans_kart", "finans_kredi", "finans_giris", "finans_cikis", "kasa_kk_giris"]:
                donem_metni = f"Dönem: {bu_ay_str}"
            elif secili_k_id in ["kasa_bugun_giris", "kasa_bugun_cikis"]:
                donem_metni = f"Dönem: {bugun_str} (Bugün)"
            elif secili_k_id == "finans_gelecek_ay":
                donem_metni = f"Dönem: {gelecek_ay_str} (Gelecek Ay)"
            else:
                donem_metni = "Dönem: Tüm Zamanlar (Açık/Bekleyen Kayıtlar)"
                
            tk.Label(win, text=donem_metni, font=("Segoe UI", 11), bg="#f8fafc", fg="#64748b").pack(pady=(0, 10))

            # --- NASIL HESAPLANDIĞINI ANLATAN ÖZEL BİLGİLENDİRME METİNLER ---
            if secili_k_id == "finans_borc":
                info = "ℹ️ Bu rakam gelirlerden giderlerin çıkarılmasıyla HESAPLANMAZ (O işlem 'Aylık Net Bakiye'de yapılır).\n\nToplam Açık Borç; cebinizden çıkacak olan, sistemde kayıtlı ve henüz 'ÖDENDİ' olarak işaretlenmemiş TÜM geçmiş ve gelecek borç kalemlerinizin (Depo, Kart, Kredi) toplanmasıyla elde edilir."
                tk.Label(win, text=info, font=("Segoe UI", 9, "bold"), bg="#f8fafc", fg="#ef4444", justify="center", wraplength=550).pack(pady=(0,10))
            
            elif secili_k_id == "finans_sgk":
                info = "ℹ️ 'Kurum İşlemleri' sekmesinden sisteme eklediğiniz ve henüz tahsilatı yapılmamış (hesaba geçmemiş) tüm bekleyen kurum alacaklarınızın dökümüdür. Aylar bazında gruplandırılmıştır."
                tk.Label(win, text=info, font=("Segoe UI", 9, "bold"), bg="#f8fafc", fg="#10b981", justify="center", wraplength=550).pack(pady=(0,10))

            # 3. ORTA SCROLL ALANI VE LİSTE ÇİZİCİ
            f_main_container = tk.Frame(win, bg="#f8fafc")
            f_main_container.pack(fill="both", expand=True, padx=10, pady=5)
            
            canvas = tk.Canvas(f_main_container, bg="#f8fafc", highlightthickness=0)
            scrollbar = ttk.Scrollbar(f_main_container, orient="vertical", command=canvas.yview)
            f_main = tk.Frame(canvas, bg="#f8fafc")
            
            f_main.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            win_id = canvas.create_window((0, 0), window=f_main, anchor="nw")
            canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))
            
            canvas.configure(yscrollcommand=scrollbar.set)
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            self.mouse_scroll_ekle(canvas, f_main)
            
            def liste_ciz(parent, baslik, veri_dict, toplam, renk):
                # Çift yazmayı önleyen temizlik kodu
                for widget in parent.winfo_children(): widget.destroy()
                
                f_kutu = tk.Frame(parent, bg="white", bd=1, relief="solid", padx=15, pady=8)
                f_kutu.pack(fill="x", pady=6, padx=10)
                tk.Label(f_kutu, text=baslik, font=("Segoe UI", 11, "bold"), bg="white", fg=renk).pack(anchor="w", pady=(0, 5))
                for ad, tutar in sorted(veri_dict.items(), key=lambda x: x[1], reverse=True):
                    r = tk.Frame(f_kutu, bg="white"); r.pack(fill="x", pady=2)
                    tk.Label(r, text=ad, font=("Segoe UI", 10), bg="white", fg="#475569").pack(side="left")
                    tk.Label(r, text=f"{tutar:,.2f} ₺", font=("Segoe UI", 10, "bold"), bg="white", fg="#0f172a").pack(side="right")
                if not veri_dict:
                    r = tk.Frame(f_kutu, bg="white"); r.pack(fill="x", pady=2)
                    tk.Label(r, text="Bu kalemde kayıt bulunamadı.", font=("Segoe UI", 10), bg="white", fg="#475569").pack(side="left")
                    tk.Label(r, text="0.00 ₺", font=("Segoe UI", 10, "bold"), bg="white", fg="#0f172a").pack(side="right")
                tk.Frame(f_kutu, height=1, bg="#e2e8f0").pack(fill="x", pady=6)
                tk.Label(f_kutu, text=f"ALT TOPLAM: {toplam:,.2f} ₺", font=("Segoe UI", 11, "bold"), bg="white", fg=renk).pack(anchor="e")

            toplam_tutar = 0.0
            veriler = {} 
            kutu_renk = "#3b82f6"
            kutu_baslik = "TÜM KALEMLER"
            
            try:
                if secili_k_id == "finans_depo":
                    self.imlec_finans.execute("SELECT fatura_adi, SUM(tutar) FROM odemeler WHERE strftime('%Y-%m', vade_tarihi) = ? AND (satir_notu IS NULL OR (NOT satir_notu LIKE '%KART:%' AND NOT satir_notu LIKE '%KURUM%' AND satir_notu NOT IN ('MANUEL_GIDER', 'MANUEL_GELIR', 'KURUM_GELIR'))) GROUP BY fatura_adi", (bu_ay_str,))
                    for r in self.imlec_finans.fetchall(): veriler[r[0]] = r[1]
                    kutu_baslik = "DEPO ÖDEMELERİ (-)"
                    kutu_renk = "#ef4444"
                
                elif secili_k_id == "finans_kart":
                    # Bu ayki tüm taksitleri çek
                    self.imlec_finans.execute("""
                        SELECT satir_notu, SUM(tutar) FROM odemeler 
                        WHERE strftime('%Y-%m', vade_tarihi) = ? 
                        AND (aciklama LIKE '%Kredi Kartı%' OR satir_notu LIKE 'KART:%')
                        GROUP BY satir_notu
                    """, (bu_ay_str,))
                    
                    kart_sistem_toplamlar = {}
                    for (notu, tut) in self.imlec_finans.fetchall():
                        k_adi = "Diğer Kartlar"
                        if notu and "KART:" in str(notu):
                            temp = str(notu).split("KART:")[1].strip()
                            # "(1/12)" kısmından önce al
                            if "(" in temp:
                                k_adi = temp.split("(")[0].strip()
                            else:
                                k_adi = temp
                            # Trailing " -" veya "-" temizle (iki farklı format var)
                            k_adi = k_adi.rstrip(" -").strip()
                        kart_sistem_toplamlar[k_adi] = kart_sistem_toplamlar.get(k_adi, 0.0) + (tut or 0.0)
                    
                    # Her kart için ekstre girilmişse onu kullan, yoksa sistem toplamını kullan
                    for k_adi, sistem_tut in kart_sistem_toplamlar.items():
                        self.imlec_finans.execute(
                            "SELECT deger FROM program_ayarlari WHERE ayar_adi=?",
                            (f"ekstre_toplam_{k_adi}_{bu_ay_str}",)
                        )
                        res_ekstre = self.imlec_finans.fetchone()
                        if res_ekstre and res_ekstre[0]:
                            try:
                                ekstre_val = float(res_ekstre[0])
                                veriler[k_adi] = ekstre_val if ekstre_val > 0 else sistem_tut
                            except:
                                veriler[k_adi] = sistem_tut
                        else:
                            veriler[k_adi] = sistem_tut
                    kutu_baslik = "KART ÖDEMELERİ (Ekstre / Sistem)"
                    kutu_renk = "#ef4444"
                        
                elif secili_k_id == "finans_kredi":
                    # DÜZELTME: durum='ODENMEDİ' filtresi kaldırıldı, o ayki tüm taksitler hesaplanacak
                    self.imlec_finans.execute("SELECT aciklama, SUM(tutar) FROM odemeler WHERE strftime('%Y-%m', vade_tarihi) = ? AND aciklama LIKE 'KREDI:%' GROUP BY aciklama", (bu_ay_str,))
                    for r in self.imlec_finans.fetchall():
                        k_adi = str(r[0]).split("(")[0].replace("KREDI:", "").strip()
                        veriler[k_adi] = veriler.get(k_adi, 0.0) + r[1]
                    kutu_baslik = "KREDİ TAKSİTLERİ (-)"
                    kutu_renk = "#ef4444"
                        
                elif secili_k_id == "finans_sgk":
                    self.imlec_finans.execute("SELECT fatura_adi, strftime('%Y-%m', vade_tarihi), SUM(tutar) FROM odemeler WHERE durum='ODENMEDİ' AND satir_notu='KURUM_GELIR' GROUP BY fatura_adi, strftime('%Y-%m', vade_tarihi)")
                    aylar_tr = {"01":"Ocak", "02":"Şubat", "03":"Mart", "04":"Nisan", "05":"Mayıs", "06":"Haziran", "07":"Temmuz", "08":"Ağustos", "09":"Eylül", "10":"Ekim", "11":"Kasım", "12":"Aralık"}
                    for r in self.imlec_finans.fetchall():
                        kurum = r[0]
                        yil_ay = r[1]
                        tutar = r[2]
                        try:
                            yil, ay = yil_ay.split("-")
                            etiket = f"{kurum} ({aylar_tr.get(ay, ay)} {yil})"
                        except:
                            etiket = f"{kurum} ({yil_ay})"
                        veriler[etiket] = veriler.get(etiket, 0.0) + tutar
                    kutu_baslik = "BEKLEYEN KURUM GELİRLERİ (+)"
                    kutu_renk = "#10b981"
                    
                elif secili_k_id in ["finans_giris", "kasa_bugun_giris", "kasa_kk_giris"]:
                    zaman = bugun_str if secili_k_id == "kasa_bugun_giris" else f"{bu_ay_str}%"
                    yontem = "AND odeme_yontemi='KREDİ KARTI'" if secili_k_id == "kasa_kk_giris" else ""
                    self.imlec_finans.execute(f"SELECT kategori, SUM(tutar) FROM kasa_defteri WHERE islem_turu='GİRİŞ' AND tarih LIKE ? {yontem} GROUP BY kategori", (zaman,))
                    for r in self.imlec_finans.fetchall(): veriler[r[0]] = r[1]
                    kutu_baslik = "GELİR KALEMLERİ (+)"
                    kutu_renk = "#10b981"
                    
                elif secili_k_id in ["finans_cikis", "kasa_bugun_cikis"]:
                    zaman = bugun_str if secili_k_id == "kasa_bugun_cikis" else f"{bu_ay_str}%"
                    self.imlec_finans.execute("SELECT kategori, SUM(tutar) FROM kasa_defteri WHERE islem_turu='ÇIKIŞ' AND tarih LIKE ? GROUP BY kategori", (zaman,))
                    for r in self.imlec_finans.fetchall(): veriler[r[0]] = r[1]
                    kutu_baslik = "GİDER KALEMLERİ (-)"
                    kutu_renk = "#ef4444"
                    
                elif secili_k_id in ["finans_borc", "finans_gelecek_ay"]:
                    zaman_filtresi = "" if secili_k_id == "finans_borc" else f"AND strftime('%Y-%m', vade_tarihi) = '{gelecek_ay_str}'"
                    self.imlec_finans.execute(f"SELECT satir_notu, aciklama, tutar FROM odemeler WHERE durum='ODENMEDİ' AND satir_notu NOT IN ('MANUEL_GELIR', 'KURUM_GELIR') AND (aciklama IS NULL OR aciklama NOT LIKE '%[SENET-KK]%') {zaman_filtresi}")
                    for r in self.imlec_finans.fetchall():
                        notu, ack, tut = r
                        if ack and "KREDI:" in ack: kat = "Banka Kredileri"
                        elif (ack and "Kredi Kartı" in ack) or (notu and "KART:" in str(notu)): kat = "Kredi Kartı Taksitleri"
                        elif notu == "MANUEL_GIDER": kat = "Manuel Eklenen Giderler"
                        else: kat = "Depo / Firma Ödemeleri"
                        veriler[kat] = veriler.get(kat, 0.0) + tut
                    kutu_baslik = "ÖDENMEMİŞ BORÇ KALEMLERİ (-)"
                    kutu_renk = "#ef4444"

            except Exception as e:
                veriler = {"Listeleme Hatası": 0.0}
            
            toplam_tutar = sum(veriler.values())
            liste_ciz(f_main, kutu_baslik, veriler, toplam_tutar, kutu_renk)
            lbl_toplam_val.config(text=f"{toplam_tutar:,.2f} ₺", fg=kutu_renk)


        # --- KUTULARI ÇİZEN DÖNGÜ ---
        self.ozet_label_referanslari = []
        for idx, k_id in enumerate(secili_kartlar):
            if k_id not in kart_datalari: continue
            veri = kart_datalari[k_id]
            row, col = idx // 2, idx % 2
            
            kart = tk.Frame(f_cards, bg=veri["bg"])
            kart.grid(row=row, column=col, sticky="nsew", padx=8, pady=5)
            
            f_ust = tk.Frame(kart, bg=veri["bg"])
            f_ust.pack(fill="x", padx=15, pady=(10, 0)) 
            
            lbl_cap = tk.Label(f_ust, text=veri["baslik"], bg=veri["bg"], fg="white", justify="left")
            lbl_cap.pack(side="left", anchor="nw")
            
            lbl_ikon = tk.Label(f_ust, text=veri["ikon"], bg=veri["bg"], fg="white")
            lbl_ikon.pack(side="right", anchor="ne")
            
            lbl_val = tk.Label(kart, text="", bg=veri["bg"], fg="white", anchor="w")
            lbl_val.pack(fill="both", expand=True, padx=15, pady=(0, 10)) 
            
            self.ozet_label_referanslari.append((lbl_val, veri["deger"], veri["is_money"]))

            # --- DİNAMİK ÖLÇEKLEME ---
            def resize_fonts(event, l_v=lbl_val, l_c=lbl_cap, l_i=lbl_ikon):
                w = event.width
                h = event.height
                if h < 20: return 
                v_size = max(8, min(28, int(h * 0.28), int(w * 0.08))) 
                c_size = max(6,  min(10, int(h * 0.15)))                
                i_size = max(10, min(20, int(h * 0.25)))                
                
                l_v.config(font=("Segoe UI", v_size, "bold"))
                l_c.config(font=("Segoe UI", c_size, "bold"), wraplength=int(w * 0.70))
                l_i.config(font=("Segoe UI", i_size))

            kart.bind("<Configure>", resize_fonts)

            # --- FİNANSAL KARTLARA TIKLAMA OLAYI ---
            if veri.get("is_money", False):
                # İmleci el (tıklanabilir) şekline getir
                kart.config(cursor="hand2")
                f_ust.config(cursor="hand2")
                lbl_cap.config(cursor="hand2")
                lbl_ikon.config(cursor="hand2")
                lbl_val.config(cursor="hand2")
                
                # Tıklanan kutunun ID'sini komuta bağla
                cmd = lambda e, cid=k_id: goster_kutu_detayi(e, cid)
                
                kart.bind("<Button-1>", cmd)
                f_ust.bind("<Button-1>", cmd)
                lbl_cap.bind("<Button-1>", cmd)
                lbl_ikon.bind("<Button-1>", cmd)
                lbl_val.bind("<Button-1>", cmd)

        def guncelle_ozet_kartlari():
            for lbl, gercek_deger, is_money in self.ozet_label_referanslari:
                if is_money:
                    # GÜÇLÜ SANSÜR KONTROLÜ
                    if self.bakiye_gizli or (self.aktif_rol == "Personel" and not self.personel_yetkileri.get("para_gorsun", False)):
                        lbl.config(text="****** ₺")
                    else:
                        lbl.config(text=gercek_deger)
                else:
                    lbl.config(text=gercek_deger)
        guncelle_ozet_kartlari()

        # =====================================================================
        # 3. SAYFA 1: MENÜ EKRANI (SIĞMAMA SORUNU DÜZELTİLDİ)
        # =====================================================================
        
        # Üst boşluklar (pady) kısıldı
        tk.Label(self.page_menu, text="Uygulama Modülleri", font=("Segoe UI", 26, "bold"), bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(pady=(10, 5))

        main_pane = tk.Frame(self.page_menu, bg=c.get_color("bg_main"))
        main_pane.pack(fill="both", expand=True, padx=40, pady=5)

        # Ana kapsayıcıyı ızgara (grid) sistemine geçiriyoruz
        main_pane.grid_rowconfigure(0, weight=1) # 0. Satır (Üst menüler) esneyebilir
        main_pane.grid_rowconfigure(1, weight=0) # 1. Satır (Sistem Ayarları) ezilemez, sabit kalır
        main_pane.grid_columnconfigure(0, weight=1)

        f_top_split = tk.Frame(main_pane, bg=c.get_color("bg_main"))
        # pack yerine grid kullanıyoruz
        f_top_split.grid(row=0, column=0, sticky="nsew")

        def create_grid_button(parent, text, bg_col, row, col, cmd, colspan=1):
            btn = tk.Button(parent, text=text, command=cmd, 
                            bg=bg_col, fg="white", font=("Segoe UI", 13, "bold"), 
                            relief="flat", cursor="hand2", 
                            activebackground=bg_col, activeforeground="white")
            btn.grid(row=row, column=col, columnspan=colspan, sticky="nsew", padx=5, pady=4) 
            return btn

        # --- SOL YARI: ASİSTAN İŞLEMLERİ ---
        f_asistan = tk.LabelFrame(f_top_split, text="💊 ASİSTAN İŞLEMLERİ", font=("Segoe UI", 12, "bold"), bg=c.get_color("bg_main"), fg="#10b981", padx=10, pady=5)
        f_asistan.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        f_asistan.grid_columnconfigure(0, weight=1); f_asistan.grid_columnconfigure(1, weight=1)
        for i in range(3): f_asistan.grid_rowconfigure(i, weight=1)

        create_grid_button(f_asistan, "📅\n\nSKT TAKİBİ", "#0ea5e9", 0, 0, wrap_cmd(1))
        create_grid_button(f_asistan, "💊\n\nİLAÇ KARTLARI", "#10b981", 0, 1, wrap_cmd(2))
        create_grid_button(f_asistan, "🚀\n\nTOPLU İŞLEM", "#f59e0b", 1, 0, wrap_cmd(3))
        create_grid_button(f_asistan, "🔲\n\nKAREKOD ÜRET", "#8b5cf6", 1, 1, wrap_cmd(4))
        create_grid_button(f_asistan, "📦\n\nSTOK SAYIM", "#f97316", 2, 0, wrap_cmd(5), colspan=2)

        # --- SAĞ YARI: FİNANS YÖNETİMİ ---
        if self.aktif_rol == "Yönetici" or self.personel_yetkileri.get("finans_gorsun", False):
            f_finans = tk.LabelFrame(f_top_split, text="📊 FİNANS YÖNETİMİ", font=("Segoe UI", 12, "bold"), bg=c.get_color("bg_main"), fg="#f59e0b", padx=10, pady=5)
            f_finans.pack(side="right", fill="both", expand=True, padx=(10, 0))

            f_finans.grid_columnconfigure(0, weight=1); f_finans.grid_columnconfigure(1, weight=1)
            for i in range(4): f_finans.grid_rowconfigure(i, weight=1)

            create_grid_button(f_finans, "🏢\n\nDEPO ÖDEMELERİ", "#3b82f6", 0, 0, wrap_cmd(6))
            create_grid_button(f_finans, "💳\n\nKREDİ KARTLARI", "#ec4899", 0, 1, wrap_cmd(7))
            create_grid_button(f_finans, "🏛️\n\nKURUM İŞLEMLERİ", "#14b8a6", 1, 0, wrap_cmd(8))
            create_grid_button(f_finans, "💰\n\nGELİR TAKİBİ", "#22c55e", 1, 1, wrap_cmd(9))
            create_grid_button(f_finans, "📉\n\nGİDER TAKİBİ", "#ef4444", 2, 0, wrap_cmd(10))
            create_grid_button(f_finans, "📒\n\nKASA DEFTERİ", "#eab308", 2, 1, wrap_cmd(11))
            create_grid_button(f_finans, "🧮\n\nKREDİ HESAPLAMA", "#0284c7", 3, 0, wrap_cmd(15))
            create_grid_button(f_finans, "📂\n\nMEVCUT KREDİLER", "#8b5cf6", 3, 1, wrap_cmd(16))
            
        else:
            f_finans = tk.LabelFrame(f_top_split, text="🔒 FİNANS YÖNETİMİ", font=("Segoe UI", 12, "bold"), bg=c.get_color("bg_main"), fg="#64748b", padx=10, pady=5)
            f_finans.pack(side="right", fill="both", expand=True, padx=(10, 0))
            
            f_finans.grid_columnconfigure(0, weight=1)
            f_finans.grid_rowconfigure(0, weight=1)
            
            tk.Label(f_finans, text="Bu modüle erişim yetkiniz bulunmamaktadır.\n\nYetki almak için yöneticinizle görüşün.", font=("Segoe UI", 11, "italic"), bg=c.get_color("bg_main"), fg="#94a3b8").grid(row=0, column=0, sticky="nsew")

        # --- ALT: SİSTEM & AYARLAR ---
        f_sistem = tk.LabelFrame(main_pane, text="🛠️ SİSTEM & AYARLAR", font=("Segoe UI", 12, "bold"), bg=c.get_color("bg_main"), fg="#8b5cf6", padx=10, pady=5)
        
        f_sistem.grid(row=1, column=0, sticky="ew", pady=(10, 0), ipady=5) 
        
        f_sistem.grid_columnconfigure(0, weight=1)
        f_sistem.grid_columnconfigure(1, weight=1)
        f_sistem.grid_columnconfigure(2, weight=1)
        f_sistem.grid_rowconfigure(0, weight=1) 

        if self.aktif_rol == "Yönetici" or self.personel_yetkileri.get("ayarlar_gorsun", False):
            create_grid_button(f_sistem, "⚙️\nAYARLAR", "#64748b", 0, 0, wrap_cmd(12))
        else:
            tk.Button(f_sistem, text="🔒\nYETKİ YOK", bg="#cbd5e1", fg="white", font=("Segoe UI", 13, "bold"), relief="flat", state="disabled").grid(row=0, column=0, sticky="nsew", padx=5, pady=4)

        create_grid_button(f_sistem, "📢\n\nGERİ BİLDİRİM", "#6366f1", 0, 1, wrap_cmd(13))
        create_grid_button(f_sistem, "📖\n\nKILAVUZ", "#a855f7", 0, 2, wrap_cmd(14))
        
        

        # =====================================================================
        # 4. ALT NOKTA GÖSTERGESİ VE YENİ KAYDIRMA KONTROLLERİ
        # =====================================================================
        f_dots = tk.Frame(self.content_area, bg=c.get_color("bg_main"))
        f_dots.pack(side="bottom", fill="x", pady=(5, 15))
        
        # YENİ: Aşağıya alınan uyarı metni
        tk.Label(f_dots, text="Sayfalar arası geçiş yapmak için fare tekerleğini (Scroll) aşağı/yukarı yapabilir, sağ/sol ok tuşlarını (← →) kullanabilir veya ekranı kaydırabilirsiniz.", font=("Segoe UI", 7, "italic"), bg=c.get_color("bg_main"), fg="#94a3b8").pack(side="top", pady=(0, 5))
        
        canvas_dots = tk.Canvas(f_dots, bg=c.get_color("bg_main"), highlightthickness=0, height=20, width=60)
        canvas_dots.pack(anchor="center")
        
        # --- YENİ: SCROLL VE OK TUŞU BAĞLANTILARI ---
        def on_scroll_swipe(e):
            if self.animating: return
            # Eğer fare bir listenin veya tablonun üzerindeyse sayfa değiştirmesin (kendi içinde kaysın)
            w_class = e.widget.winfo_class()
            if w_class in ("Canvas", "Treeview", "Scrollbar", "Listbox", "Text", "Entry"): return
            
            # Tekerlek aşağı (delta negatif) -> Menüye (Sağa) geç
            if e.delta < 0 and self.current_page == 0: goto_page(1)
            # Tekerlek yukarı (delta pozitif) -> Özete (Sola) dön
            elif e.delta > 0 and self.current_page == 1: goto_page(0)

        def on_key_left(e):
            if self.current_page == 1: goto_page(0)
            
        def on_key_right(e):
            if self.current_page == 0: goto_page(1)

        self.pencere.bind_all("<MouseWheel>", on_scroll_swipe)
        self.pencere.bind_all("<Left>", on_key_left)
        self.pencere.bind_all("<Right>", on_key_right)

        def guncelle_noktalar():
            canvas_dots.delete("all")
            c0 = "#3b82f6" if self.current_page == 0 else "#cbd5e1"
            canvas_dots.create_oval(10, 5, 25, 20, fill=c0, outline="")
            c1 = "#3b82f6" if self.current_page == 1 else "#cbd5e1"
            canvas_dots.create_oval(35, 5, 50, 20, fill=c1, outline="")

        guncelle_noktalar()

    # =========================================================================
    # ARKA PLAN (SİSTEM TEPSİSİ - SYSTEM TRAY) İŞLEMLERİ
    # =========================================================================
    def kapatma_istegi(self):
        """X tuşuna basıldığında çalışır."""
        # Değişkenin tanımlı olup olmadığını kontrol eden ekstra güvenlik katmanı
        pystray_mevcut_mu = globals().get('PYSTRAY_VAR', False)
        
        if pystray_mevcut_mu:
            try:
                self.pencere.withdraw() # Ekranı gizle
                self.sistem_tepsisine_gonder()
            except Exception as e:
                print(f"Sistem tepsisi hatası: {e}")
                self.pencere.destroy()
                os._exit(0)
        else:
            # Eğer pystray yüklü değilse veya hata verdiyse normal kapat
            if messagebox.askyesno("Çıkış", "Uygulama arka planda çalışma modunda değil. Tamamen kapatılsın mı?"):
                self.pencere.destroy()
                os._exit(0)

    def sistem_tepsisine_gonder(self):
        """Gizli simgeler alanında (sağ altta) ikon oluşturur."""
        try:
            image = Image.open(self.logo_yolu_tray)
            
            # İkona sağ tıklanınca açılacak menü
            menu = pystray.Menu(
                item('Aç / Göster', self.tepsiden_cikar, default=True),
                item('Tamamen Kapat', self.tamamen_kapat)
            )
            
            # İkonu oluştur ve programı dondurmaması için ayrı bir Thread'de (iş parçacığı) çalıştır
            self.tray_icon = pystray.Icon("EczaciDefteri", image, "Eczacı Defteri", menu)
            threading.Thread(target=self.tray_icon.run, daemon=True).start()
            
            # Kullanıcıya sağ alttan bildirim ver (Programın açık kaldığını bilsin)
            self.goster_bildirim("Arka Plana Alındı", "Eczacı Defteri arka planda çalışmaya ve otomatik işlemleri yapmaya devam ediyor.")
        except Exception as e:
            print("Tepsi hatası:", e)

    def tepsiden_cikar(self, icon, item):
        """Kullanıcı tepsideki ikona çift tıkladığında veya Göster dediğinde çalışır."""
        icon.stop() # Tepsideki ikonu yok et
        self.pencere.after(0, self.pencere.deiconify) # Pencereyi tekrar görünür yap

    def tamamen_kapat(self, icon, item):
        """Kullanıcı menüden Tamamen Kapat'a bastığında çalışır."""
        icon.stop()
        self.pencere.destroy()
        os._exit(0) # Tüm arka plan işlemlerini (Thread'leri) zorla sonlandır

    # =========================================================================
    # 1. ECZANE İÇİ CANLI SOHBET MOTORU (KÜÇÜK EKRAN SORUNU ÇÖZÜLDÜ)
    # =========================================================================
    def chat_panelini_ac(self):
        # Tabloyu güvenceye al
        try:
            self.imlec.execute("CREATE TABLE IF NOT EXISTS eczane_chat (id INTEGER PRIMARY KEY AUTOINCREMENT, tarih TEXT, gonderen TEXT, mesaj TEXT)")
            self.baglanti_skt.commit()
        except: pass
        
        if hasattr(self, 'chat_win') and self.chat_win.winfo_exists():
            self.chat_win.destroy()
            return
            
        self.chat_win = tk.Toplevel(self.pencere)
        self.chat_win.title("💬 Eczane İçi Canlı Mesajlaşma")
        self.chat_win.geometry("350x500")
        self.chat_win.attributes("-topmost", True)
        self.chat_win.configure(bg=TM.get_color("bg_main"))
        
        # Sağ alta hizala
        x = self.pencere.winfo_rootx() + self.pencere.winfo_width() - 380
        y = self.pencere.winfo_rooty() + 150
        self.chat_win.geometry(f"+{x}+{y}")
        
        # 1. ÖNCE YAZMA ALANINI EN ALTA SABİTLİYORUZ (Pencere küçülse de kaybolmaz)
        f_yaz = tk.Frame(self.chat_win, bg=TM.get_color("bg_main"))
        f_yaz.pack(side="bottom", fill="x", padx=10, pady=(0, 10))
        
        def mesaj_gonder(e=None):
            m = ent_mesaj.get().strip()
            if not m: return
            saat = datetime.now().strftime("%H:%M")
            self.imlec.execute("INSERT INTO eczane_chat (tarih, gonderen, mesaj) VALUES (?, ?, ?)", (saat, self.aktif_rol, m))
            self.baglanti_skt.commit()
            ent_mesaj.delete(0, tk.END)
            mesajlari_yukle()
            
        btn_gonder = tk.Button(f_yaz, text="GÖNDER", command=mesaj_gonder, bg="#3b82f6", fg="white", font=("Segoe UI", 9, "bold"), relief="flat")
        btn_gonder.pack(side="right", padx=(5,0), ipady=3, ipadx=5)
        
        ent_mesaj = tk.Entry(f_yaz, font=("Segoe UI", 11))
        ent_mesaj.pack(side="left", fill="x", expand=True, ipady=5)
        
        self.chat_win.bind("<Return>", mesaj_gonder)
        
        # 2. SONRA MESAJLAR KISMINI EKLİYORUZ (Kalan boşluğu doldurur)
        f_mesajlar = tk.Frame(self.chat_win, bg=TM.get_color("card_bg"))
        f_mesajlar.pack(side="top", fill="both", expand=True, padx=10, pady=10)
        
        txt_chat = tk.Text(f_mesajlar, state="disabled", font=("Segoe UI", 10), bg=TM.get_color("input_bg"), fg=TM.get_color("fg_text"), wrap="word")
        txt_chat.pack(fill="both", expand=True)
        
        def mesajlari_yukle():
            self.imlec.execute("SELECT tarih, gonderen, mesaj FROM eczane_chat ORDER BY id DESC LIMIT 50")
            gecmis = reversed(self.imlec.fetchall())
            
            txt_chat.config(state="normal")
            txt_chat.delete("1.0", tk.END)
            for r in gecmis:
                renk = "blue" if r[1] == "Yönetici" else "green"
                txt_chat.insert(tk.END, f"[{r[0]}] {r[1]}: ", renk)
                txt_chat.insert(tk.END, f"{r[2]}\n")
            txt_chat.see(tk.END)
            txt_chat.config(state="disabled")
            
        mesajlari_yukle()
        
        # Mesajlaşma penceresi açıkken veritabanını 2 saniyede bir kontrol et (Canlı Chat hissi için)
        def canli_yenile():
            if self.chat_win.winfo_exists():
                mesajlari_yukle()
                self.chat_win.after(2000, canli_yenile)
        canli_yenile()

    # =========================================================================
    # 2. KİM NE YAPTI? (GİZLİ LOG/KAYIT MOTORU)
    # =========================================================================
    def islem_kaydet(self, islem_detayi):
        """Bu fonksiyonu kodun neresine yazarsan yaz, o işlemi kimin yaptığını kaydeder."""
        try:
            self.imlec.execute("CREATE TABLE IF NOT EXISTS sistem_loglari (id INTEGER PRIMARY KEY AUTOINCREMENT, tarih TEXT, kullanici TEXT, islem TEXT)")
            zaman = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
            self.imlec.execute("INSERT INTO sistem_loglari (tarih, kullanici, islem) VALUES (?, ?, ?)", (zaman, self.aktif_rol, islem_detayi))
            self.baglanti_skt.commit()
        except Exception as e:
            print("Log kaydedilemedi:", e)

    def render_log_ayarlari(self):
        c = TM
        for w in self.settings_content_frame.winfo_children(): w.destroy()
        
        tk.Label(self.settings_content_frame, text="Sistem İşlem Geçmişi (Log Kayıtları)", font=FONT_HEAD, bg=c.get_color("bg_main"), fg=c.get_color("fg_text")).pack(anchor="w", pady=(0, 20))
        tk.Label(self.settings_content_frame, text="Bu ekranda sisteme giriş yapan kullanıcıların yaptığı silme, ekleme gibi işlemlerin dökümünü görebilirsiniz.", font=("Segoe UI", 10), bg=c.get_color("bg_main"), fg="#64748b").pack(anchor="w", pady=(0, 10))
        
        f_liste = tk.Frame(self.settings_content_frame, bg=c.get_color("card_bg"), padx=10, pady=10)
        f_liste.pack(fill="both", expand=True)
        
        cols = ("ID", "TARİH VE SAAT", "YETKİ / KULLANICI", "İŞLEM DETAYI")
        tree = ttk.Treeview(f_liste, columns=cols, show="headings", height=20)
        tree.heading("ID", text="ID"); tree.column("ID", width=50, anchor="center")
        tree.heading("TARİH VE SAAT", text="Tarih & Saat"); tree.column("TARİH VE SAAT", width=150, anchor="center")
        tree.heading("YETKİ / KULLANICI", text="Kullanıcı Rolü"); tree.column("YETKİ / KULLANICI", width=150, anchor="center")
        tree.heading("İŞLEM DETAYI", text="Yapılan İşlem"); tree.column("İŞLEM DETAYI", width=500, anchor="w")
        
        sc = ttk.Scrollbar(f_liste, orient="vertical", command=tree.yview)
        tree.configure(yscroll=sc.set); sc.pack(side="right", fill="y"); tree.pack(fill="both", expand=True)
        
        try:
            self.imlec.execute("CREATE TABLE IF NOT EXISTS sistem_loglari (id INTEGER PRIMARY KEY AUTOINCREMENT, tarih TEXT, kullanici TEXT, islem TEXT)")
            self.imlec.execute("SELECT * FROM sistem_loglari ORDER BY id DESC LIMIT 200")
            for r in self.imlec.fetchall():
                tree.insert("", "end", values=r)
        except Exception as e:
            tree.insert("", "end", values=("HATA", "-", "-", str(e)))

# =========================================================================
# SADE VE MODERN (YEŞİL VURGULU) KULLANICI GİRİŞ EKRANI
# =========================================================================
class GirisEkrani:
    def __init__(self, root, basari_tetikleyici):
        self.root = root
        self.basari_tetikleyici = basari_tetikleyici
        
        import sqlite3
        import os
        
        # Veritabanı Yolu Ayarları
        ortak_yol = yerel_db_yolunu_getir()
        if not os.path.exists(ortak_yol):
            try: os.makedirs(ortak_yol)
            except: ortak_yol = ""
            
        db_yolu = os.path.join(ortak_yol, "kullanicilar.db") if ortak_yol else "kullanicilar.db"
        
        self.vt = sqlite3.connect(db_yolu, timeout=15, check_same_thread=False)
        self.imlec = self.vt.cursor()
        self.imlec.execute("CREATE TABLE IF NOT EXISTS kullanicilar (id INTEGER PRIMARY KEY, k_adi TEXT, sifre TEXT, rol TEXT)")
        self.imlec.execute("CREATE TABLE IF NOT EXISTS hatirla (id INTEGER PRIMARY KEY, k_adi TEXT, sifre TEXT, durum INTEGER)")
        
        self.imlec.execute("SELECT COUNT(*) FROM kullanicilar")
        if self.imlec.fetchone()[0] == 0:
            self.imlec.execute("INSERT INTO kullanicilar (k_adi, sifre, rol) VALUES ('admin', '1234', 'Yönetici')")
            self.vt.commit()

        # --- 1. FERAHLATICI VE NÖTR ARKA PLAN ---
        # Göz yormayan, açık ve profesyonel bir açık gri tonu
        self.bg_frame = tk.Frame(root, bg="#f1f5f9") 
        self.bg_frame.place(relx=0, rely=0, relwidth=1.0, relheight=1.0)
        
        # --- KESİLMEYEN, GENİŞ GİRİŞ KARTI ---
        self.card = tk.Frame(self.bg_frame, bg="#ffffff", highlightthickness=0, bd=0)
        self.card.place(relx=0.5, rely=0.5, anchor="center", width=460, height=800)
        self.card.pack_propagate(False)

        # --- 2. ŞIK YEŞİL ÜST VURGU ÇİZGİSİ ---
        tk.Frame(self.card, bg="#10b981", height=6).pack(fill="x", side="top")

        # --- LOGO ALANI ---
        try:
            import os, sys
            from PIL import Image, ImageTk
            base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
            logo_yolu = os.path.join(base_dir, "logo_hd_transparan.png")
            img = Image.open(logo_yolu).resize((250, 225), Image.Resampling.LANCZOS)
            self.logo_img = ImageTk.PhotoImage(img)
            tk.Label(self.card, image=self.logo_img, bg="#ffffff").pack(pady=(40, 10))
        except:
            # Logo yoksa yeşil bir ikon göster
            tk.Label(self.card, text="⚕️", font=("Segoe UI", 70), bg="#ffffff", fg="#10b981").pack(pady=(40, 10))

        # --- KARŞILAMA METİNLERİ ---
        tk.Label(self.card, text="Hoş Geldiniz", font=("Segoe UI", 15, "bold"), bg="#ffffff", fg="#1e293b").pack(pady=(0, 5))
        tk.Label(self.card, text="Sisteme erişmek için bilgilerinizi giriniz", font=("Segoe UI", 11), bg="#ffffff", fg="#64748b").pack(pady=(0, 35))

        pad_x_val = 45

        # --- KULLANICI ADI GİRİŞİ ---
        tk.Label(self.card, text="Kullanıcı Adı", font=("Segoe UI", 10, "bold"), bg="#ffffff", fg="#475569").pack(anchor="w", padx=pad_x_val, pady=(0, 5))
        
        f_user = tk.Frame(self.card, bg="#f8fafc", bd=0, highlightbackground="#e2e8f0", highlightthickness=1)
        f_user.pack(fill="x", padx=pad_x_val, pady=(0, 20), ipady=5)
        
        # İkon Zümrüt Yeşili yapıldı
        tk.Label(f_user, text="👤", font=("Segoe UI", 14), bg="#f8fafc", fg="#10b981").pack(side="left", padx=(15, 5))
        self.ent_kadi = tk.Entry(f_user, font=("Segoe UI", 13), bg="#f8fafc", fg="#0f172a", relief="flat", insertbackground="#10b981")
        self.ent_kadi.pack(side="left", fill="x", expand=True, padx=(0, 10))

        # --- ŞİFRE GİRİŞİ ---
        tk.Label(self.card, text="Şifre", font=("Segoe UI", 10, "bold"), bg="#ffffff", fg="#475569").pack(anchor="w", padx=pad_x_val, pady=(0, 5))
        
        f_pass = tk.Frame(self.card, bg="#f8fafc", bd=0, highlightbackground="#e2e8f0", highlightthickness=1)
        f_pass.pack(fill="x", padx=pad_x_val, pady=(0, 15), ipady=5) 
        
        # İkon Zümrüt Yeşili yapıldı
        tk.Label(f_pass, text="🔒", font=("Segoe UI", 14), bg="#f8fafc", fg="#10b981").pack(side="left", padx=(15, 5))
        self.ent_sifre = tk.Entry(f_pass, font=("Segoe UI", 13), bg="#f8fafc", fg="#0f172a", relief="flat", show="*", insertbackground="#10b981")
        self.ent_sifre.pack(side="left", fill="x", expand=True, padx=(0, 10))

        # --- ALT SEÇENEKLER ---
        f_actions = tk.Frame(self.card, bg="#ffffff")
        f_actions.pack(fill="x", padx=pad_x_val, pady=(5, 30))

        self.var_hatirla = tk.IntVar()
        self.chk_hatirla = tk.Checkbutton(f_actions, text=" Beni Hatırla", variable=self.var_hatirla, bg="#ffffff", fg="#475569", font=("Segoe UI", 10, "bold"), activebackground="#ffffff", activeforeground="#0f172a", selectcolor="#ffffff", cursor="hand2", relief="flat", bd=0)
        self.chk_hatirla.pack(side="left")

        # Şifremi unuttum yazısı Zümrüt Yeşili yapıldı
        self.btn_unuttum = tk.Button(f_actions, text="Şifremi Unuttum?", font=("Segoe UI", 10, "underline", "bold"), bg="#ffffff", fg="#10b981", relief="flat", cursor="hand2", activebackground="#ffffff", activeforeground="#059669", bd=0, command=self.sifremi_unuttum)
        self.btn_unuttum.pack(side="right")

        # --- 3. CANLI YEŞİL GİRİŞ BUTONU ---
        f_btn = tk.Frame(self.card, bg="#ffffff")
        f_btn.pack(fill="x", padx=pad_x_val)
        
        # Buton ana rengi Zümrüt Yeşili (#10b981), üzerine gelince koyu yeşil (#059669) olacak
        self.btn_giris_modern = ModernButton(f_btn, text="SİSTEME GİRİŞ YAP", command=self.giris_yap, bg_color="#10b981", hover_color="#059669", corner_radius=12, width=370, height=55, font=("Segoe UI", 12, "bold"))
        self.btn_giris_modern.pack(anchor="center")

        # Enter Tuşu Bağlantısı
        self.root.bind('<Return>', lambda e: self.giris_yap())
        
        # Hafızadaki kullanıcıyı kontrol et
        self.imlec.execute("SELECT k_adi, sifre, durum FROM hatirla WHERE id=1")
        hatirla_res = self.imlec.fetchone()
        
        if hatirla_res and hatirla_res[2] == 1:
            self.ent_kadi.insert(0, hatirla_res[0])
            self.ent_sifre.insert(0, hatirla_res[1])
            self.var_hatirla.set(1)
            self.btn_giris_modern.focus() 
        else:
            self.ent_kadi.focus()

    def giris_yap(self):
        k_adi = self.ent_kadi.get().strip()
        sifre = self.ent_sifre.get().strip()
        import tkinter.messagebox as messagebox
        
        if not k_adi or not sifre:
            messagebox.showwarning("Eksik Bilgi", "Lütfen kullanıcı adı ve şifrenizi eksiksiz girin.")
            return

        self.imlec.execute("SELECT rol FROM kullanicilar WHERE k_adi=? AND sifre=?", (k_adi, sifre))
        sonuc = self.imlec.fetchone()
        
        if sonuc:
            if self.var_hatirla.get() == 1:
                self.imlec.execute("INSERT OR REPLACE INTO hatirla (id, k_adi, sifre, durum) VALUES (1, ?, ?, 1)", (k_adi, sifre))
            else:
                self.imlec.execute("INSERT OR REPLACE INTO hatirla (id, k_adi, sifre, durum) VALUES (1, '', '', 0)")
            
            self.vt.commit()
            self.vt.close()
            self.root.unbind('<Return>')
            
            # --- BAŞARILI GİRİŞ EFEKTİ ---
            # Butonu başarı durumunda daha tok bir orman yeşiline çeviriyoruz
            self.btn_giris_modern.update_color("#047857") 
            self.btn_giris_modern.itemconfig(self.btn_giris_modern.find_withtag("text"), text="GİRİŞ BAŞARILI ✔")
            self.root.config(cursor="watch")
            self.root.update() 
            
            def aninda_gec():
                self.basari_tetikleyici(sonuc[0])
                try: self.bg_frame.destroy()
                except: pass
                self.root.config(cursor="")
                
            self.root.after(400, aninda_gec)
            
        else:
            messagebox.showerror("Hatalı Giriş", "Kullanıcı adı veya şifreniz yanlış. Lütfen tekrar deneyin.")

    def sifremi_unuttum(self):
        k_adi = self.ent_kadi.get().strip()
        import tkinter.messagebox as messagebox

        if k_adi == "sifirla1234":
            import sqlite3
            vt_kurtar = sqlite3.connect("kullanicilar.db")
            imlec_kurtar = vt_kurtar.cursor()
            imlec_kurtar.execute("UPDATE kullanicilar SET sifre='1234' WHERE k_adi='admin'")
            vt_kurtar.commit()
            vt_kurtar.close()
            messagebox.showinfo("Acil Kurtarma", "Yönetici (admin) şifresi '1234' olarak sıfırlandı!")
            self.ent_kadi.delete(0, 'end')
            self.ent_kadi.insert(0, 'admin')
            return

        messagebox.showinfo("Bilgi", "Şifrenizi sıfırlamak için kullanıcı adına 'sifirla1234' yazıp bu butona tekrar basabilirsiniz.")


# =========================================================================
# ANA ÇALIŞTIRMA BLOĞU
# =========================================================================
if __name__ == "__main__":
    try:
        import ctypes
        from ctypes import windll
        myappid = 'eczacidefteri.uygulama.v1' 
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception as e: pass

    try:
        from tkinterdnd2 import TkinterDnD, DND_FILES
        root = TkinterDnD.Tk()
        surukle_birak_aktif = True
    except ImportError:
        import tkinter as tk
        root = tk.Tk()
        surukle_birak_aktif = False

    root.title("Eczacı Defteri")
    root.configure(bg="#f8fafc") 
    
    window_width = 1400
    window_height = 900
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)
    root.geometry(f"{window_width}x{window_height}+{x}+{y}")

    def uygulamayi_baslat(aktif_rol="Yönetici"):
        app = EczaneAsistani(root, aktif_rol)
        if surukle_birak_aktif:
            root.drop_target_register(DND_FILES)
            root.dnd_bind('<<Drop>>', app.surukle_birak_yoneticisi)

    GirisEkrani(root, uygulamayi_baslat)

    root.mainloop()